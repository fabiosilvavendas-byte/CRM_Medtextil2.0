import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
import requests
from github import Github
import json
import hashlib
import uuid as _uuid_mod

# ====================== SUPABASE — CLIENTE CENTRALIZADO ======================
# Credenciais lidas dos secrets do Streamlit.
# Configure em .streamlit/secrets.toml:
#   [supabase]
#   url = "https://XXXXXXXX.supabase.co"
#   key = "eyJ..."
# No Streamlit Cloud: Settings → Secrets

def _supa_headers():
    """Retorna headers HTTP para chamadas à API REST do Supabase."""
    try:
        key = st.secrets["supabase"]["key"]
    except Exception:
        return None
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

def _supa_url(tabela):
    """Retorna a URL REST para uma tabela do Supabase."""
    try:
        base = st.secrets["supabase"]["url"].rstrip("/")
        return f"{base}/rest/v1/{tabela}"
    except Exception:
        return None

def supa_disponivel():
    """Verifica se as credenciais do Supabase estão configuradas."""
    try:
        _ = st.secrets["supabase"]["url"]
        _ = st.secrets["supabase"]["key"]
        return True
    except Exception:
        return False

def supa_select(tabela, filtros=None, ordem=None, limite=1000):
    """
    Lê registros de uma tabela.
    filtros: dict {coluna: valor} → igualdade simples
    ordem:   string ex: "criado_em.desc"
    Retorna lista de dicts ou [] em caso de erro.
    """
    url = _supa_url(tabela)
    headers = _supa_headers()
    if not url or not headers:
        return []
    params = {"limit": limite}
    if filtros:
        for col, val in filtros.items():
            params[col] = f"eq.{val}"
    if ordem:
        params["order"] = ordem
    try:
        r = requests.get(url, headers=headers, params=params, timeout=10)
        if r.status_code == 200:
            return r.json()
        return []
    except Exception:
        return []

def supa_insert(tabela, dados):
    """
    Insere um registro.
    dados: dict com os campos.
    Retorna o registro inserido (dict) ou None em caso de erro.
    """
    url = _supa_url(tabela)
    headers = _supa_headers()
    if not url or not headers:
        return None
    try:
        r = requests.post(url, headers=headers,
                          data=json.dumps(dados), timeout=10)
        if r.status_code in (200, 201):
            resultado = r.json()
            return resultado[0] if isinstance(resultado, list) else resultado
        return None
    except Exception:
        return None

def supa_update(tabela, id_valor, dados, id_col="id"):
    """
    Atualiza um registro pelo id.
    Retorna True em caso de sucesso.
    """
    url = _supa_url(tabela)
    headers = _supa_headers()
    if not url or not headers:
        return False
    try:
        r = requests.patch(
            url,
            headers=headers,
            params={id_col: f"eq.{id_valor}"},
            data=json.dumps(dados),
            timeout=10,
        )
        return r.status_code in (200, 204)
    except Exception:
        return False

def supa_delete(tabela, id_valor, id_col="id"):
    """Deleta um registro pelo id. Retorna True em caso de sucesso."""
    url = _supa_url(tabela)
    headers = _supa_headers()
    if not url or not headers:
        return False
    try:
        r = requests.delete(
            url,
            headers=headers,
            params={id_col: f"eq.{id_valor}"},
            timeout=10,
        )
        return r.status_code in (200, 204)
    except Exception:
        return False

# ── Funções específicas de pedidos ──────────────────────────────────────

def gerar_numero_pedido():
    """
    Gera o próximo número sequencial no padrão MED-AAAA-NNNNNN.
    Consulta o maior número do ano atual e incrementa.
    """
    ano = datetime.now().year
    prefixo = f"MED-{ano}-"
    registros = supa_select("pedidos", ordem="numero.desc", limite=1)
    ultimo_seq = 0
    for reg in registros:
        num = reg.get("numero", "")
        if num.startswith(prefixo):
            try:
                ultimo_seq = int(num.replace(prefixo, ""))
            except Exception:
                pass
    return f"{prefixo}{str(ultimo_seq + 1).zfill(6)}"

def salvar_pedido(dados_cliente, dados_pedido, itens, obs_pedido,
                  status="rascunho", usuario_id=None, usuario_nome=None,
                  pedido_id=None):
    """
    Cria ou atualiza um pedido completo (cabeçalho + itens).
    Se pedido_id for fornecido, atualiza; caso contrário, cria novo.
    Retorna (pedido_id, numero_pedido) ou (None, None) em caso de erro.
    """
    agora = datetime.now().isoformat()
    total = sum(i.get("total", 0) for i in itens)

    cabecalho = {
        "status":               status,
        "cliente_razao_social": dados_cliente.get("razao_social", ""),
        "cliente_cpf_cnpj":     dados_cliente.get("cpf_cnpj", ""),
        "cliente_ie":           dados_cliente.get("ie", ""),
        "cliente_cidade":       dados_cliente.get("cidade", ""),
        "cliente_estado":       dados_cliente.get("estado", ""),
        "cliente_telefone":     dados_cliente.get("telefone", ""),
        "cliente_email_nfe":    dados_cliente.get("email", ""),
        "cliente_endereco":     dados_cliente.get("endereco", ""),
        "representante":        dados_cliente.get("representante", ""),
        "obs_cliente":          dados_cliente.get("obs_cliente", ""),
        "tabela_preco":         dados_pedido.get("tabela_preco", ""),
        "tipo_frete":           dados_pedido.get("tipo_frete", "CIF"),
        "data_venda":           dados_pedido.get("data_venda", ""),
        "cond_pagto":           dados_pedido.get("cond_pagto", ""),
        "estado_comissao":      dados_pedido.get("estado_comissao", ""),
        "obs_pedido":           obs_pedido or "",
        "valor_total":          total,
        "atualizado_em":        agora,
    }

    if pedido_id:
        # Atualização
        ok = supa_update("pedidos", pedido_id, cabecalho)
        if not ok:
            return None, None
        # Apagar itens antigos e reinserir
        supa_delete("itens_pedido", pedido_id, id_col="pedido_id")
        numero = dados_pedido.get("numero", "")
    else:
        # Criação
        numero = gerar_numero_pedido()
        pedido_id = str(_uuid_mod.uuid4())
        cabecalho.update({
            "id":           pedido_id,
            "numero":       numero,
            "criado_por_id": usuario_id or "",
            "criado_por_nome": usuario_nome or "",
            "criado_em":    agora,
        })
        resultado = supa_insert("pedidos", cabecalho)
        if not resultado:
            return None, None

    # Inserir itens
    for item in itens:
        item_row = {
            "id":              str(_uuid_mod.uuid4()),
            "pedido_id":       pedido_id,
            "codigo_produto":  item.get("codigo", ""),
            "descricao":       item.get("descricao", ""),
            "gramatura":       item.get("peso", ""),
            "cx_embarque":     item.get("cx_embarque", ""),
            "quantidade":      item.get("quantidade", 0),
            "valor_unit":      item.get("valor_unit", 0),
            "preco_ref":       item.get("preco_ref", 0),
            "preco_historico": item.get("preco_historico", 0),
            "comissao_perc":   item.get("comissao", ""),
            "total":           item.get("total", 0),
            "alerta_preco_baixo": item.get("alerta_preco_baixo", False),
            "criado_em":       agora,
        }
        supa_insert("itens_pedido", item_row)

    # Registrar no histórico de status
    hist = {
        "id":              str(_uuid_mod.uuid4()),
        "pedido_id":       pedido_id,
        "status_anterior": "",
        "status_novo":     status,
        "usuario_id":      usuario_id or "",
        "usuario_nome":    usuario_nome or "",
        "observacao":      f"Pedido {'criado' if not pedido_id else 'atualizado'}",
        "criado_em":       agora,
    }
    supa_insert("historico_status", hist)

    return pedido_id, numero

def mudar_status_pedido(pedido_id, status_novo, usuario_id,
                        usuario_nome, observacao="", status_anterior=""):
    """Muda o status de um pedido e registra no histórico."""
    agora = datetime.now().isoformat()
    ok = supa_update("pedidos", pedido_id, {
        "status": status_novo,
        "atualizado_em": agora,
    })
    if ok:
        supa_insert("historico_status", {
            "id":              str(_uuid_mod.uuid4()),
            "pedido_id":       pedido_id,
            "status_anterior": status_anterior,
            "status_novo":     status_novo,
            "usuario_id":      usuario_id,
            "usuario_nome":    usuario_nome,
            "observacao":      observacao,
            "criado_em":       agora,
        })
    return ok

# ── Funções de autenticação via Supabase ────────────────────────────────

def _hash_senha(senha):
    """Hash SHA-256 simples para senhas."""
    return hashlib.sha256(senha.encode()).hexdigest()

def autenticar_usuario(email, senha):
    """
    Autentica pelo Supabase (tabela 'usuarios').
    Retorna dict do usuário ou None.
    """
    if not supa_disponivel():
        return None
    registros = supa_select("usuarios", filtros={"email": email, "ativo": "true"})
    for reg in registros:
        if reg.get("senha_hash") == _hash_senha(senha):
            return reg
    return None

# ====================== FUNÇÃO KPI CARD PROFISSIONAL ======================
def render_kpi_card(label, value, delta=None, icon="📊", color="#1F4788"):
    """Renderiza card KPI profissional com HTML/CSS — mobile-first, sem height fixo"""
    delta_html = ""
    if delta:
        delta_val = str(delta).replace("%","").replace(",","").replace("+","").strip()
        try:
            delta_color = "#10B981" if float(delta_val) >= 0 else "#EF4444"
        except Exception:
            delta_color = "#10B981" if "+" in str(delta) else "#EF4444"
        delta_html = f'<div style="color:{delta_color};font-size:0.78rem;font-weight:600;margin-top:6px;">{delta}</div>'
    st.markdown(f"""
    <div class="kpi-card" style="border-left:4px solid {color};">
        <div class="kpi-top">
            <span class="kpi-label">{label}</span>
            <span class="kpi-icon">{icon}</span>
        </div>
        <div class="kpi-value">{value}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)

# Configuração da página
st.set_page_config(
    page_title="Dashboard BI Medtextil", 
    layout="wide", 
    initial_sidebar_state="expanded",
    page_icon="https://i.imgur.com/gt3rgyL.png"  # Logo Medtextil
)

# ====================== CONFIGURAÇÃO DO ÍCONE ======================
# O Streamlit gerencia automaticamente o ícone via page_icon
# Nenhuma configuração adicional é necessária

# ====================== CSS CUSTOMIZADO - UX/UI PROFISSIONAL ======================
st.markdown("""
<style>
/* ═══════════════════════════════════════════════════════════════
   MEDTEXTIL BI — CSS UNIFICADO MOBILE-FIRST
   Um único bloco, sem duplicatas, responsivo por design.
   ═══════════════════════════════════════════════════════════════ */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ── Base ─────────────────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: 'Inter','Segoe UI',Roboto,sans-serif !important;
    -webkit-font-smoothing: antialiased;
}
.stApp { background-color: var(--background-color) !important; }

/* ── Remove padding desnecessário das colunas ─────────────────── */
div[data-testid="stHorizontalBlock"] { gap: 10px !important; }
div[data-testid="stHorizontalBlock"] > div[data-testid="stVerticalBlock"],
div[data-testid="column"] { padding: 0 !important; margin: 0 !important; min-width: 0 !important; }

/* ── Sidebar ──────────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: var(--secondary-background-color) !important;
    border-right: 1px solid rgba(128,128,128,0.15) !important;
    box-shadow: 2px 0 8px rgba(0,0,0,0.04) !important;
}
section[data-testid="stSidebar"] .stMarkdown h1,
section[data-testid="stSidebar"] .stMarkdown h2,
section[data-testid="stSidebar"] .stMarkdown h3 { color: #4A7BC8 !important; }

/* Navegação sidebar */
[data-testid="stSidebar"] [data-testid="stRadio"] { gap: 4px !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] label {
    background: #F8F9FA !important; border: 1px solid #E9ECEF !important;
    border-radius: 9px !important; padding: 10px 14px !important;
    margin: 0 !important; cursor: pointer !important;
    transition: all 0.18s ease !important; font-size: 0.875rem !important;
    font-weight: 500 !important; display: block !important; width: 100% !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
    background: #EEF2F7 !important; border-color: #4A7BC8 !important;
    transform: translateX(2px) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-checked="true"] {
    background: linear-gradient(135deg,#1F4788 0%,#2D5AA0 100%) !important;
    border-color: #4A7BC8 !important; color: white !important;
    font-weight: 600 !important; box-shadow: 0 2px 8px rgba(31,71,136,0.2) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] input[type="radio"] { display: none !important; }

/* Sidebar logo e badge */
.sidebar-logo-container {
    display: flex; flex-direction: column; align-items: center;
    padding: 14px 8px 8px; border-bottom: 1px solid #E9ECEF; margin-bottom: 10px;
}
.sidebar-user-badge {
    background: #F0F4FF; border: 1px solid #C5D5F0; border-radius: 8px;
    padding: 7px 10px; font-size: 0.82rem; color: #4A7BC8;
    font-weight: 600; width: 100%; text-align: center; margin-top: 6px;
}
section[data-testid="stSidebar"] .sidebar-cat-label {
    font-size: 0.60rem !important; font-weight: 700 !important;
    color: #8A96A8 !important; letter-spacing: 0.09em !important;
    text-transform: uppercase !important; margin: 8px 0 3px 4px !important;
}

/* ── KPI cards customizados (render_kpi_card) ─────────────────── */
.kpi-card {
    background: var(--secondary-background-color);
    border-radius: 12px;
    padding: clamp(12px, 2vw, 18px) clamp(12px, 2vw, 16px);
    box-shadow: 0 2px 8px rgba(31,71,136,0.07);
    transition: box-shadow 0.18s, transform 0.18s;
    box-sizing: border-box;
    width: 100%;
}
.kpi-card:hover {
    box-shadow: 0 6px 20px rgba(31,71,136,0.13) !important;
    transform: translateY(-2px);
}
.kpi-top {
    display: flex; justify-content: space-between;
    align-items: flex-start; margin-bottom: 8px;
}
.kpi-label {
    font-size: clamp(0.65rem, 1.2vw, 0.75rem);
    color: #6B7280; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.05em;
    line-height: 1.3;
}
.kpi-icon { font-size: clamp(1.2rem, 2.5vw, 1.6rem); line-height: 1; }
.kpi-value {
    font-size: clamp(1.1rem, 2.8vw, 1.65rem);
    font-weight: 700; color: #1F2937; line-height: 1.2;
    word-break: break-word;
}

/* ── st.metric ────────────────────────────────────────────────── */
[data-testid="stMetric"] {
    background: var(--secondary-background-color) !important;
    border-radius: 12px !important;
    padding: clamp(12px,2vw,18px) clamp(12px,2vw,20px) !important;
    border-left: 4px solid #1F4788 !important;
    box-shadow: 0 1px 6px rgba(31,71,136,0.07) !important;
    transition: box-shadow 0.18s,transform 0.18s !important;
}
[data-testid="stMetric"]:hover {
    box-shadow: 0 5px 18px rgba(31,71,136,0.13) !important;
    transform: translateY(-2px) !important;
}
[data-testid="stMetricLabel"] p {
    font-size: 0.71rem !important; font-weight: 700 !important;
    text-transform: uppercase !important; letter-spacing: 0.06em !important;
    color: #8A96A8 !important;
}
[data-testid="stMetricValue"] {
    font-size: clamp(1.1rem,2.5vw,1.45rem) !important;
    font-weight: 700 !important; color: #2C5AA0 !important;
}
div[data-testid="column"]:nth-child(2) [data-testid="stMetric"] { border-left-color: #2E86AB !important; }
div[data-testid="column"]:nth-child(3) [data-testid="stMetric"] { border-left-color: #28A745 !important; }
div[data-testid="column"]:nth-child(4) [data-testid="stMetric"] { border-left-color: #F4A261 !important; }

/* ── Botões ───────────────────────────────────────────────────── */
.stButton > button {
    border-radius: 8px !important; font-weight: 600 !important;
    font-size: 0.875rem !important; padding: 0.45rem 1.1rem !important;
    border: 1.5px solid #1F4788 !important; color: #4A7BC8 !important;
    background: var(--secondary-background-color) !important;
    transition: all 0.18s ease !important;
    box-shadow: 0 1px 4px rgba(31,71,136,0.08) !important;
}
.stButton > button:hover {
    background: #1F4788 !important; color: #FFFFFF !important;
    box-shadow: 0 4px 12px rgba(31,71,136,0.22) !important;
    transform: translateY(-1px) !important;
}
.stButton > button[kind="primary"] {
    background: #1F4788 !important; color: #FFFFFF !important; border-color: #4A7BC8 !important;
}
.stButton > button[kind="primary"]:hover {
    background: #163561 !important; border-color: #163561 !important;
}
[data-testid="stDownloadButton"] > button {
    border-radius: 8px !important; font-weight: 600 !important;
    background: var(--secondary-background-color) !important;
    color: #4A7BC8 !important; border: 1.5px solid #1F4788 !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #1F4788 !important; color: #FFFFFF !important;
}

/* ── Títulos ──────────────────────────────────────────────────── */
.stApp h1 {
    color: #2C5AA0 !important; font-weight: 700 !important;
    font-size: clamp(1.3rem,3vw,1.75rem) !important; letter-spacing: -0.02em !important;
}
.stApp h2 { color: #4A7BC8 !important; font-weight: 600 !important; font-size: clamp(1.05rem,2.5vw,1.3rem) !important; }
.stApp h3 { color: #2E4A7A !important; font-weight: 600 !important; }
h2, h3 { font-weight: 600 !important; letter-spacing: -0.01em !important; }

/* ── Inputs e selects ─────────────────────────────────────────── */
.stSelectbox > div > div,
.stTextInput > div > div > input,
.stDateInput > div > div > input,
.stNumberInput > div > div > input {
    border-radius: 8px !important; border-color: #DEE2E6 !important; font-size: 0.875rem !important;
}
.stSelectbox > div > div:focus-within,
.stTextInput > div > div:focus-within {
    border-color: #4A7BC8 !important; box-shadow: 0 0 0 2px rgba(31,71,136,0.13) !important;
}

/* ── Tabs ─────────────────────────────────────────────────────── */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    background: var(--secondary-background-color) !important;
    border-radius: 10px !important; padding: 4px !important;
    border-bottom: none !important; gap: 4px !important;
    flex-wrap: wrap !important;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    border-radius: 8px !important; font-weight: 500 !important;
    font-size: clamp(0.75rem,1.5vw,0.875rem) !important;
    color: #6C757D !important; padding: 5px 12px !important;
}
[data-testid="stTabs"] [aria-selected="true"] {
    background: #1F4788 !important; color: #FFFFFF !important; font-weight: 600 !important;
}

/* ── Dataframes ───────────────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border-radius: 10px !important; overflow: hidden !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
    max-width: 100% !important;
}

/* ── Alertas, divisores, expanders ───────────────────────────── */
[data-testid="stAlert"] { border-radius: 10px !important; }
hr { border-color: #E9ECEF !important; margin: 0.75rem 0 !important; }
div[data-testid="stExpander"] details {
    border: 1px solid rgba(128,128,128,0.15) !important;
    border-radius: 8px !important; background: transparent !important;
    margin-bottom: 6px !important;
}
div[data-testid="stExpander"] details summary {
    font-size: 0.78rem !important; font-weight: 600 !important;
    color: #8A96A8 !important; padding: 6px 12px !important;
}
div[data-testid="stExpander"] details summary:hover { color: #4A7BC8 !important; }
[data-testid="stExpander"] summary {
    font-size: 0.82rem !important; font-weight: 600 !important; color: #6C757D !important;
}

/* ── Barra de filtros ─────────────────────────────────────────── */
.filter-bar {
    background: #FFFFFF; border-radius: 10px;
    padding: 12px 16px; margin-bottom: 14px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05); border: 1px solid #E9ECEF;
}

/* ── Captions ─────────────────────────────────────────────────── */
.stCaption p {
    font-size: 0.7rem !important; color: #8A96A8 !important;
    margin-top: -4px !important; font-weight: 500 !important; letter-spacing: 0.03em !important;
}

/* ── HOME: cards med-card ─────────────────────────────────────── */
div.med-card {
    background: var(--secondary-background-color);
    border: 1px solid #E4E9F0; border-radius: 14px;
    padding: clamp(14px,2.5vw,20px) clamp(12px,2vw,18px) clamp(12px,2vw,16px);
    min-height: clamp(110px,15vw,138px);
    box-shadow: 0 1px 5px rgba(31,71,136,0.06);
    transition: box-shadow 0.18s,transform 0.18s,border-color 0.18s;
    cursor: pointer; position: relative; box-sizing: border-box;
}
div.med-card:hover {
    border-color: #B8CDF0 !important;
    box-shadow: 0 7px 22px rgba(31,71,136,0.14) !important;
    transform: translateY(-3px);
}
div.med-card-col div[data-testid="stButton"] > button {
    position: relative !important; display: block !important;
    width: 100% !important; height: clamp(110px,15vw,138px) !important;
    margin-top: calc(-1 * clamp(110px,15vw,138px) - 10px) !important;
    opacity: 0 !important; cursor: pointer !important;
    border: none !important; background: transparent !important;
    z-index: 99 !important; padding: 0 !important;
}
div.med-card .mc-icon {
    width: 36px; height: 36px; background: #EEF3FC; border-radius: 9px;
    display: flex; align-items: center; justify-content: center;
    margin-bottom: 9px; color: #4A7BC8; font-size: 15px;
}
div.med-card .mc-title {
    font-size: clamp(0.82rem,1.8vw,0.94rem); font-weight: 700;
    color: #2C5AA0; margin-bottom: 3px; letter-spacing: -0.01em;
}
div.med-card .mc-desc { font-size: 0.72rem; color: #6C757D; line-height: 1.4; margin-bottom: 7px; }
div.med-card .mc-info {
    font-size: 0.68rem; color: #ADB5BD;
    border-top: 1px solid #F0F2F5; padding-top: 6px;
}

/* ── Gráficos Plotly: scroll horizontal no mobile ─────────────── */
div[data-testid="stPlotlyChart"] {
    overflow-x: auto !important; max-width: 100% !important;
}

/* ═══════════════════════════════════════════════════════════════
   DARK MODE
   ═══════════════════════════════════════════════════════════════ */
@media (prefers-color-scheme: dark) {
    .stApp { background-color: #0E1117 !important; }
    .stApp h1,.stApp h2,.stApp h3 { color: #7EB3F7 !important; }
    .stMarkdown h2 { color: #7EB3F7 !important; }
    .stMarkdown p  { color: #C4CDD9 !important; }
    [data-testid="stMetric"] { background: #1A1D24 !important; box-shadow: 0 1px 6px rgba(0,0,0,0.3) !important; }
    [data-testid="stMetricLabel"] p { color: #8A96A8 !important; }
    [data-testid="stMetricValue"]   { color: #E8EDF5 !important; }
    [data-testid="stTabs"] [data-baseweb="tab-list"] { background: #1A1D24 !important; }
    [data-testid="stTabs"] [data-baseweb="tab"]      { color: #8A96A8 !important; }
    [data-testid="stDataFrame"]  { background: #1A1D24 !important; }
    [data-testid="stExpander"]   { background: #1A1D24 !important; border-color: #2D3139 !important; }
    section[data-testid="stSidebar"]             { background: #1A1D24 !important; border-right-color: #2D3139 !important; }
    .stButton > button           { background: #1A1D24 !important; color: #A8C4E8 !important; border-color: #2D5AA0 !important; }
    .stButton > button:hover     { background: #1F4788 !important; color: #FFFFFF !important; }
    [data-testid="stDownloadButton"] > button        { background: #1A2A45 !important; color: #7EB3F7 !important; border-color: #2D5AA0 !important; }
    [data-testid="stDownloadButton"] > button:hover  { background: #1F4788 !important; color: #FFFFFF !important; }
    .stSelectbox > div > div,
    .stTextInput > div > div > input,
    .stDateInput > div > div > input { background: #1A1D24 !important; border-color: #2D3139 !important; color: #E0E6EF !important; }
    div[data-testid="stHorizontalBlock"].filter-bar { background: #1A1D24 !important; border-color: #2D3139 !important; }
    div.med-card { background: #1A1D24 !important; border-color: #2D3139 !important; }
    div.med-card .mc-title { color: #A8C4E8 !important; }
    div.med-card .mc-desc  { color: #8A96A8 !important; }
    div.med-card .mc-info  { color: #5A6375 !important; border-top-color: #2D3139 !important; }
    .kpi-card { background: #1A1D24 !important; }
    .kpi-value { color: #E8EDF5 !important; }
    .kpi-label { color: #8A96A8 !important; }
    [data-testid="stSidebar"] .stRadio label[aria-checked="true"] p { color: #7EB3F7 !important; }
    section[data-testid="stSidebar"],
    section[data-testid="stSidebar"] > div { background-color: #1A1D24 !important; }
}
[data-theme="dark"] .stApp                         { background-color: #0E1117 !important; }
[data-theme="dark"] section[data-testid="stSidebar"]{ background: #1A1D24 !important; border-right-color: #2D3139 !important; }
[data-theme="dark"] [data-testid="stMetric"]        { background: #1A1D24 !important; }
[data-theme="dark"] [data-testid="stMetricValue"]   { color: #E8EDF5 !important; }
[data-theme="dark"] .kpi-card                       { background: #1A1D24 !important; }
[data-theme="dark"] .kpi-value                      { color: #E8EDF5 !important; }
[data-theme="dark"] div.med-card                    { background: #1A1D24 !important; border-color: #2D3139 !important; }
[data-theme="dark"] .stApp h1,[data-theme="dark"] .stApp h2,[data-theme="dark"] .stApp h3 { color: #E8EDF5 !important; }

/* ═══════════════════════════════════════════════════════════════
   MOBILE-FIRST  ≤ 768px
   ═══════════════════════════════════════════════════════════════ */
@media (max-width: 768px) {

    /* Sidebar sólida no mobile */
    section[data-testid="stSidebar"],
    section[data-testid="stSidebar"] > div { background-color: #FFFFFF !important; background: #FFFFFF !important; }

    /* Home grid: 2 cards por linha */
    div.home-grid > div[data-testid="stHorizontalBlock"] {
        flex-wrap: wrap !important; display: flex !important; gap: 6px !important;
    }
    div.home-grid > div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
        width: calc(50% - 3px) !important; min-width: calc(50% - 3px) !important;
        max-width: calc(50% - 3px) !important; flex: 0 0 calc(50% - 3px) !important;
        box-sizing: border-box !important;
    }

    /* KPI row: 2 por linha */
    div[data-testid="stHorizontalBlock"]:has(.kpi-card) {
        flex-wrap: wrap !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.kpi-card) > div[data-testid="column"] {
        width: calc(50% - 5px) !important; min-width: calc(50% - 5px) !important;
        flex: 0 0 calc(50% - 5px) !important;
    }

    /* Filtros: 2 por linha */
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stSelectbox"]) { flex-wrap: wrap !important; }
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stSelectbox"]) > div[data-testid="stVerticalBlock"] {
        width: calc(50% - 4px) !important; min-width: calc(50% - 4px) !important;
        flex: 0 0 calc(50% - 4px) !important;
    }

    /* Esconder captions de filtro no mobile */
    .filter-header-bar { display: none !important; }

    /* Tabs: scroll horizontal */
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        overflow-x: auto !important; flex-wrap: nowrap !important;
        -webkit-overflow-scrolling: touch !important;
    }

    /* Gráficos: scroll horizontal */
    div[data-testid="stPlotlyChart"] { overflow-x: auto !important; }

    /* Reduzir padding da área principal */
    .main .block-container { padding: 0.75rem 0.75rem 2rem !important; }

    /* Títulos menores no mobile */
    .stApp h1 { font-size: 1.25rem !important; }
    .stApp h2 { font-size: 1.05rem !important; }
}

@media (max-width: 480px) {

    /* KPI cards: 1 por linha em telas muito pequenas */
    div[data-testid="stHorizontalBlock"]:has(.kpi-card) > div[data-testid="column"] {
        width: 100% !important; min-width: 100% !important; flex: 0 0 100% !important;
    }

    /* Home cards: 1 por linha */
    div.home-grid > div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
        width: 100% !important; min-width: 100% !important;
        max-width: 100% !important; flex: 0 0 100% !important;
    }

    /* Tabs: fonte menor */
    [data-testid="stTabs"] [data-baseweb="tab"] { font-size: 0.72rem !important; padding: 4px 8px !important; }
}

/* ── Migrados do bloco legado (seletores exclusivos preservados) ── */
.page-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 4px;
}

.page-subtitle {
    color: #6C757D;
    font-size: 0.9rem;
    margin-bottom: 20px;
    margin-top: -4px;
}

.stButton > button p {
    font-size: 0.9rem !important;
}

.stRadio > div {
    flex-direction: row !important;
    gap: 8px !important;
}

.stRadio > div label {
    border-radius: 20px !important;
    padding: 4px 14px !important;
    border: 1.5px solid #DEE2E6 !important;
    font-size: 0.875rem !important;
    cursor: pointer !important;
    transition: all 0.15s !important;
}

.stSpinner > div {
    border-top-color: #4A7BC8 !important;
}

</style>
""", unsafe_allow_html=True)

# ====================== CONFIGURAÇÕES GITHUB ======================
GITHUB_REPO = "fabiosilvavendas-byte/CRM_Medtextil2.0"
GITHUB_FOLDER = "dados"  # ⭐ PASTA ONDE ESTÃO AS PLANILHAS
GITHUB_TOKEN = None  # Opcional: adicione token se repositório for privado

@st.cache_data(ttl=3600)
def listar_planilhas_github():
    """Lista todos os arquivos Excel da pasta 'dados' no repositório GitHub"""
    try:
        if GITHUB_TOKEN:
            g = Github(GITHUB_TOKEN, timeout=15)
        else:
            g = Github(timeout=15)
        
        repo = g.get_repo(GITHUB_REPO)
        # ⭐ BUSCAR NA PASTA 'dados'
        contents = repo.get_contents(GITHUB_FOLDER)
        
        planilhas = {
            'vendas': None,
            'inadimplencia': None,
            'vendas_produto': None,
            'produtos_agrupados': None,
            'pedidos_pendentes': None,
            'tabela_ne': None,
            'contrato': None,
            'todas': []
        }
        
        for content in contents:
            if content.name.endswith(('.xlsx', '.xls')):
                info = {
                    'nome': content.name,
                    'url': content.download_url,
                    'path': content.path
                }
                planilhas['todas'].append(info)
                
                # Identificar planilha de vendas
                if 'CONSULTA_VENDEDORES' in content.name.upper():
                    planilhas['vendas'] = info
                
                # Identificar planilha de inadimplência
                if 'LANCAMENTO A RECEBER' in content.name.upper() or 'LANCAMENTO_A_RECEBER' in content.name.upper():
                    planilhas['inadimplencia'] = info
                
                # Identificar planilha de vendas por produto
                if 'VENDAS POR PRODUTO' in content.name.upper() and 'GERAL' in content.name.upper():
                    planilhas['vendas_produto'] = info
                
                # Identificar planilha de produtos agrupados
                if 'PRODUTOS_AGRUPADOS_COMPLETOS_CONCILIADOS' in content.name.upper():
                    planilhas['produtos_agrupados'] = info
                
                # Identificar planilha de pedidos pendentes
                if 'PEDIDOSPENDENTES' in content.name.upper().replace(' ', '').replace('_', ''):
                    planilhas['pedidos_pendentes'] = info

                # Identificar tabela NE
                if 'TABELA_NE' in content.name.upper().replace(' ', '_'):
                    planilhas['tabela_ne'] = info
                # Identificar planilha de contratos (Grid Contrato Consulta)
                if 'CONTRATO' in content.name.upper() and 'CONSULTA' in content.name.upper():
                    planilhas['contrato'] = info

        
        if not planilhas['todas']:
            st.warning(f"⚠️ Nenhuma planilha Excel encontrada na pasta '{GITHUB_FOLDER}'")
        
        return planilhas
    except Exception as e:
        st.error(f"❌ Erro ao conectar ao GitHub: {str(e)}")
        st.info(f"💡 Verificando: {GITHUB_REPO}/{GITHUB_FOLDER}")
        return {'vendas': None, 'inadimplencia': None, 'vendas_produto': None, 'produtos_agrupados': None, 'pedidos_pendentes': None, 'tabela_ne': None, 'contrato': None, 'todas': []}

@st.cache_data(ttl=3600)
def carregar_planilha_github(url):
    """Carrega planilha diretamente do GitHub"""
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        df = pd.read_excel(io.BytesIO(response.content))
        return df
    except requests.exceptions.Timeout:
        st.error("⏱️ Timeout ao carregar planilha. Tente novamente.")
        return None
    except requests.exceptions.RequestException as e:
        st.error(f"❌ Erro ao carregar planilha: {str(e)}")
        return None
    except Exception as e:
        st.error(f"❌ Erro ao processar planilha: {str(e)}")
        return None

# ====================== AUTENTICAÇÃO — SISTEMA DUAL ======================
# Prioridade 1: Supabase (usuários individuais com e-mail + senha)
# Prioridade 2: Fallback legado (senhas compartilhadas) — ativo enquanto
#               o Supabase não estiver configurado ou o usuário não migrar.

_MODULOS_ADMIN = [
    "Dashboard", "Positivação", "Inadimplência", "Clientes sem Compra",
    "Histórico", "Pedidos Pendentes", "Rankings",
    "Performance de Vendedores", "Consulta Clientes",
    "Meus Pedidos", "Fila de Aprovação", "Todos os Pedidos",
]
_MODULOS_GESTOR = [
    "Dashboard", "Positivação", "Inadimplência", "Clientes sem Compra",
    "Histórico", "Pedidos Pendentes", "Rankings",
    "Performance de Vendedores", "Consulta Clientes",
    "Meus Pedidos", "Fila de Aprovação", "Todos os Pedidos",
]
_MODULOS_VENDEDOR = [
    "Histórico", "Consulta Clientes", "Meus Pedidos",
]
_MODULOS_COLABORADOR = [
    "Inadimplência", "Histórico", "Pedidos Pendentes", "Consulta Clientes",
]

_PERFIL_MODULOS = {
    "admin":        _MODULOS_ADMIN,
    "administrador":_MODULOS_ADMIN,
    "gestor":       _MODULOS_GESTOR,
    "vendedor":     _MODULOS_VENDEDOR,
    "colaborador":  _MODULOS_COLABORADOR,
}

# Fallback legado — removido quando todos estiverem no Supabase
_USUARIOS_LEGADO = {
    "admin123": {
        "tipo": "administrador", "nome": "Administrador",
        "modulos": _MODULOS_ADMIN,
        "id": "legacy-admin", "email": "admin@medtextil.local",
    },
    "colaborador123": {
        "tipo": "colaborador", "nome": "Colaborador",
        "modulos": _MODULOS_COLABORADOR,
        "id": "legacy-colab", "email": "colaborador@medtextil.local",
    },
}

def check_password():
    """
    Sistema de autenticação simples e direto.
    — Supabase configurado: login com e-mail + senha individual.
    — Sem Supabase: senha compartilhada legada (admin123 / colaborador123).
    """
    # Já autenticado — retorna imediatamente
    if st.session_state.get("password_correct"):
        return True

    # ── Cabeçalho visual ─────────────────────────────────────────────────
    st.markdown("""
    <div style="max-width:420px;margin:40px auto 0 auto;background:#FFFFFF;
                border-radius:16px;padding:36px;
                box-shadow:0 8px 32px rgba(31,71,136,0.12);
                border-top:4px solid #1F4788;text-align:center;">
        <img src="https://i.imgur.com/gt3rgyL.png" height="52"
             style="border-radius:8px;margin-bottom:10px;"
             onerror="this.style.display='none'"/>
        <div style="font-size:1.4rem;font-weight:700;color:#4A7BC8;">
            Medtextil ERP</div>
        <div style="font-size:0.85rem;color:#6C757D;margin-top:4px;">
            Dashboard Comercial 2.0</div>
    </div>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("<br>", unsafe_allow_html=True)

        if supa_disponivel():
            # ── MODO SUPABASE: e-mail + senha individual ──────────────────
            email = st.text_input("✉️ E-mail", placeholder="seu@email.com.br",
                                  key="sb_email")
            senha = st.text_input("🔑 Senha", type="password",
                                  placeholder="Digite sua senha...",
                                  key="sb_senha")
            if st.button("Entrar →", use_container_width=True,
                         type="primary", key="btn_entrar_sb"):
                if not email or not senha:
                    st.error("Preencha e-mail e senha.")
                else:
                    
                    reg = autenticar_usuario(email.strip().lower(), senha)
                    if reg:
                        perfil = reg.get("perfil", "vendedor")
                        st.session_state["password_correct"] = True
                        st.session_state["usuario"] = {
                            "id":      reg.get("id", ""),
                            "email":   reg.get("email", email),
                            "nome":    reg.get("nome", email),
                            "tipo":    perfil,
                            "modulos": _PERFIL_MODULOS.get(
                                perfil, _MODULOS_COLABORADOR),
                        }
                        st.rerun()
                    else:
                        st.error("❌ E-mail ou senha incorretos.")
        else:
            # ── MODO LEGADO: senha compartilhada ──────────────────────────
            senha = st.text_input("🔑 Senha de acesso", type="password",
                                  placeholder="Digite sua senha...",
                                  key="leg_senha")
            if st.button("Entrar →", use_container_width=True,
                         type="primary", key="btn_entrar_leg"):
                if not senha:
                    st.error("Digite a senha.")
                elif senha in _USUARIOS_LEGADO:
                    st.session_state["password_correct"] = True
                    st.session_state["usuario"] = _USUARIOS_LEGADO[senha]
                    st.rerun()
                else:
                    st.error("❌ Senha incorreta.")

    return False

# ====================== PROCESSAMENTO DE DADOS ======================
def calcular_prazo_historico(data_emissao, data_vencimento_str):
    """
    Calcula o prazo histórico CUMULATIVO em dias.
    NÃO conta o dia da emissão (começa a contar do dia seguinte).
    
    Exemplo:
    - Data Emissão: 18/12/2025 (não conta este dia)
    - Vencimentos: "15/01/2026; 22/01/2026; 29/01/2026; 05/02/2026"
    - Cálculo:
      * De 19/12 até 15/01 = 28 dias
      * De 19/12 até 22/01 = 35 dias
      * De 19/12 até 29/01 = 42 dias
      * De 19/12 até 05/02 = 49 dias
    - Resultado: "28/35/42/49"
    """
    try:
        if pd.isna(data_vencimento_str) or data_vencimento_str == '':
            return ''
        
        if pd.isna(data_emissao):
            return ''
        
        # Converter string para garantir que é texto e limpar
        data_vencimento_str = str(data_vencimento_str).strip()
        
        # Separar as datas por ponto e vírgula
        datas_vencimento = data_vencimento_str.split(';')
        
        prazos = []
        for data_venc_str in datas_vencimento:
            data_venc_str = data_venc_str.strip()
            if not data_venc_str:
                continue
            
            try:
                # Tentar converter para datetime com dayfirst=True (padrão brasileiro)
                data_venc = pd.to_datetime(data_venc_str, errors='coerce', dayfirst=True)
                
                # Normalizar para meia-noite (remove qualquer componente de hora)
                if pd.notna(data_venc):
                    data_venc = data_venc.normalize() if hasattr(data_venc, 'normalize') else data_venc.replace(hour=0, minute=0, second=0, microsecond=0)
                
                # Validar se a data é válida e razoável
                if pd.notna(data_venc):
                    # Verificar se a data está dentro de um intervalo razoável (ano entre 2020 e 2030)
                    if 2020 <= data_venc.year <= 2030:
                        # Calcular dias a partir do DIA SEGUINTE à emissão
                        # O Python já calcula correto: não conta o dia da emissão
                        diferenca = (data_venc - data_emissao).days
                        
                        # Validar prazo razoável (entre 1 e 365 dias)
                        if 1 <= diferenca <= 365:
                            prazos.append(str(diferenca))
            except Exception:
                # Ignorar datas inválidas silenciosamente
                continue
        
        # Retornar prazos separados por "/"
        if prazos:
            return '/'.join(prazos)
        else:
            return ''
    except Exception:
        return ''

def gerar_pdf_pedido(dados_cliente, dados_pedido, itens_pedido, observacao=''):
    """Gera PDF do pedido no formato da Medtextil"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo customizado
    style_title = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    style_normal = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica'
    )
    
    style_small = ParagraphStyle(
        'CustomSmall',
        parent=styles['Normal'],
        fontSize=8,
        fontName='Helvetica'
    )
    
    # CABEÇALHO - SOLUÇÃO COM CONTROLE TOTAL DE TAMANHO
    logo_adicionado = False
    try:
        # Buscar logo do GitHub
        logo_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{GITHUB_FOLDER}/logo.png"
        response = requests.get(logo_url)
        if response.status_code == 200:
            from PIL import Image as PILImage
            
            # Carregar imagem com PIL para ter controle total
            logo_buffer = io.BytesIO(response.content)
            pil_img = PILImage.open(logo_buffer)
            
            # Obter dimensões originais
            largura_original, altura_original = pil_img.size
            proporcao = largura_original / altura_original
            
            # DIMENSÕES FIXAS DO LOGO NO PDF
            altura_desejada_mm = 15  # Altura em milímetros
            largura_desejada_mm = altura_desejada_mm * proporcao
            
            # Criar Image do ReportLab com dimensões exatas
            logo_buffer.seek(0)  # Voltar ao início do buffer
            logo_img = Image(logo_buffer, width=largura_desejada_mm*mm, height=altura_desejada_mm*mm)
            
            # Texto ao lado - com width definido para não vazar
            texto_empresa = Paragraph(
                "<b>MEDTEXTIL PRODUTOS TEXTIL HOSPITALARES</b><br/>"
                "<font size=8>CNPJ: 40.357.820/0001-50  Inscrição Estadual: 16.390.286-0</font>",
                style_small
            )
            
            # Calcular colWidths baseado no logo real
            espaco_logo = largura_desejada_mm + 5  # Logo + margem de 5mm
            espaco_texto = 190 - espaco_logo  # Resto da página
            
            # Criar tabela com dimensões calculadas
            cabecalho_data = [[logo_img, texto_empresa]]
            cabecalho_table = Table(cabecalho_data, colWidths=[espaco_logo*mm, espaco_texto*mm])
            cabecalho_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (0, 0), 'TOP'),
                ('VALIGN', (1, 0), (1, 0), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(cabecalho_table)
            elements.append(Spacer(1, 5*mm))
            logo_adicionado = True
            
    except Exception as e:
        # Se PIL não disponível ou erro, fallback simples
        try:
            logo_buffer = io.BytesIO(response.content)
            # Fallback: tamanho fixo conservador
            logo_img = Image(logo_buffer, width=30*mm, height=15*mm)
            
            texto_empresa = Paragraph(
                "<b>MEDTEXTIL PRODUTOS TEXTIL HOSPITALARES</b><br/>"
                "<font size=8>CNPJ: 40.357.820/0001-50  Inscrição Estadual: 16.390.286-0</font>",
                style_small
            )
            
            cabecalho_data = [[logo_img, texto_empresa]]
            cabecalho_table = Table(cabecalho_data, colWidths=[35*mm, 155*mm])
            cabecalho_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(cabecalho_table)
            elements.append(Spacer(1, 5*mm))
            logo_adicionado = True
        except Exception:
            pass
    
    if not logo_adicionado:
        # Fallback final para texto
        elements.append(Paragraph("<b>MEDTEXTIL PRODUTOS TEXTIL HOSPITALARES</b><br/>"
                                 "CNPJ: 40.357.820/0001-50 | Inscrição Estadual: 16.390.286-0", style_title))
        elements.append(Spacer(1, 5*mm))
    
    # REPRESENTANTE (tabela simples como no modelo)
    data_repr = [
        ['Representante', dados_cliente.get('representante', '')],
        ['CNPJ', dados_cliente.get('cnpj', '')]
    ]
    table_repr = Table(data_repr, colWidths=[40*mm, 150*mm])
    table_repr.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    elements.append(table_repr)
    elements.append(Spacer(1, 2*mm))
    
    # INFORMAÇÕES DO CLIENTE (layout simplificado)
    data_cliente = [
        [Paragraph("<b>Informações sobre o Cliente</b>", style_normal), '', '', ''],
        ['Cliente', dados_cliente.get('razao_social', ''), '', Paragraph("<b>Nome</b><br/>Fantasia:<br/>Inscr.<br/>Estadual:", style_small)],
        ['CNPJ:', dados_cliente.get('cnpj', ''), '', dados_cliente.get('nome_fantasia', '')],
        ['Telefone:', dados_cliente.get('telefone', ''), 'Email NF-e:', dados_cliente.get('email', '')],
        ['Endereço:', dados_cliente.get('endereco', ''), '', dados_cliente.get('ie', '')],
        ['Observação:', dados_cliente.get('obs_cliente', ''), '', '']
    ]
    
    table_cliente = Table(data_cliente, colWidths=[25*mm, 85*mm, 25*mm, 55*mm])
    table_cliente.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 1), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('SPAN', (0, 0), (-1, 0)),
        ('SPAN', (1, 1), (2, 1)),
        ('SPAN', (1, 5), (-1, 5)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table_cliente)
    elements.append(Spacer(1, 2*mm))
    
    # INFORMAÇÕES DO PEDIDO (layout compacto)
    data_pedido_info = [
        ['Informações sobre Pedido Nº', dados_pedido.get('numero', ''), '', Paragraph("<b>Tabela de<br/>Preço:</b>", style_small)],
        ['Condições de Pagto:', dados_pedido.get('condicoes_pagto', ''), '', dados_pedido.get('tabela_preco', '')],
        ['Data da Venda:', dados_pedido.get('data_venda', ''), 'Tipo de<br/>Frete:', dados_pedido.get('tipo_frete', 'CIF')]
    ]
    
    table_pedido = Table(data_pedido_info, colWidths=[45*mm, 90*mm, 20*mm, 35*mm])
    table_pedido.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table_pedido)
    elements.append(Spacer(1, 3*mm))
    
    # DETALHE DO PEDIDO
    elements.append(Paragraph("<b>Detalhe do Pedido</b>", style_normal))
    elements.append(Spacer(1, 1*mm))
    
    # Cabeçalho da tabela de itens (cor azul igual ao modelo)
    header_itens = ['COD.', 'PRODUTO', 'PESO', 'CAIXA DE\nEMBARQUE', 'QTDE', 'VALOR', 'TOTAL']
    data_itens = [header_itens]
    
    # Adicionar itens
    for item in itens_pedido:
        descricao = str(item.get('descricao', ''))
        # Quebrar descrição se muito longa
        if len(descricao) > 45:
            descricao = descricao[:45] + '...'
        
        data_itens.append([
            str(item.get('codigo', '')),
            descricao,
            str(item.get('peso', '')),
            str(item.get('cx_embarque', '')),
            f"{item.get('quantidade', 0):.0f}",
            f"R$ {item.get('valor_unit', 0):.2f}",
            f"R$ {item.get('total', 0):.2f}"
        ])
    
    # Calcular totais
    total_qtde = sum([item.get('quantidade', 0) for item in itens_pedido])
    total_valor = sum([item.get('total', 0) for item in itens_pedido])
    
    # Linha de total (sem bordas superiores, fundo cinza)
    data_itens.append(['', '', '', '', f"{total_qtde:.0f}", '', f"R$ {formatar_numero_br(total_valor, 2)}"])
    
    col_widths = [12*mm, 76*mm, 15*mm, 18*mm, 12*mm, 22*mm, 25*mm]
    table_itens = Table(data_itens, colWidths=col_widths)
    
    # Estilo da tabela de itens (azul igual ao modelo)
    num_rows = len(data_itens)
    style_itens = [
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Azul escuro
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Centralizar QTDE
        ('ALIGN', (5, 1), (7, -1), 'RIGHT'),   # Alinhar valores à direita
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ]
    
    table_itens.setStyle(TableStyle(style_itens))
    elements.append(table_itens)
    elements.append(Spacer(1, 3*mm))
    
    # RESUMO FINAL
    data_resumo = [
        ['Qtde Itens', 'Frete', 'Total Final'],
        [f"{total_qtde:.0f}", dados_pedido.get('tipo_frete', 'CIF'), f"R$ {formatar_numero_br(total_valor, 2)}"]
    ]
    
    table_resumo = Table(data_resumo, colWidths=[60*mm, 60*mm, 70*mm])
    table_resumo.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table_resumo)
    elements.append(Spacer(1, 3*mm))
    
    # OBSERVAÇÃO
    if observacao:
        elements.append(Paragraph("<b>Observação</b>", style_normal))
        elements.append(Paragraph(observacao, style_small))
    
    # Construir PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer.getvalue()

def calcular_comissao(preco_unit, preco_ref):
    """
    Calcula o percentual de comissão com base no desvio do PrecoUnit em relação ao preço de referência.
    
    Regras:
    - PrecoUnit >= 6% ACIMA  do preço ref → 4%
    - PrecoUnit igual ao preço ref (0% desconto) → 3%
    - PrecoUnit até 3% ABAIXO do preço ref → 2,5%
    - PrecoUnit 3% ou mais ABAIXO do preço ref → 2%
    """
    try:
        if pd.isna(preco_unit) or pd.isna(preco_ref) or preco_ref == 0:
            return ''
        
        preco_unit = float(preco_unit)
        preco_ref  = float(preco_ref)
        
        # Arredondar para 2 casas decimais (centavos) antes de comparar
        # Elimina ruído de ponto flutuante sem distorcer valores reais
        preco_unit = round(preco_unit, 2)
        preco_ref  = round(preco_ref,  2)
        
        variacao = round(((preco_unit - preco_ref) / preco_ref) * 100, 4)
        
        if variacao >= 6:     # >= 6% acima → 4%
            return '4%'
        elif variacao >= 0:   # >= 0% (igual ou acima) → 3%
            return '3%'
        elif variacao > -3:   # Entre 0% e -3% → 2,5%
            return '2,5%'
        else:                 # -3% ou mais abaixo → 2%
            return '2%'
    except Exception:
        return ''


# ── Paleta institucional e helper de layout de gráficos ──────────────────
CORES_INST = ['#1F4788', '#2E86AB', '#28A745', '#F4A261', '#6C757D',
              '#163561', '#1B5E8A', '#1E7B34', '#C97A3A', '#495057']

def aplicar_layout_grafico(fig, height=None):
    """Aplica o estilo institucional Medtextil a qualquer figura Plotly."""
    layout_kwargs = dict(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family='Inter, Segoe UI, Roboto, sans-serif', color='#495057', size=12),
        margin=dict(l=10, r=10, t=36, b=10),
        xaxis=dict(showgrid=False, showline=True, linecolor='#E9ECEF', tickfont=dict(size=11)),
        yaxis=dict(showgrid=True, gridcolor='#F0F0F0', showline=False, tickfont=dict(size=11)),
        coloraxis_showscale=False,
        hoverlabel=dict(bgcolor='#1F4788', font_color='white', font_size=12,
                        bordercolor='#1F4788'),
    )
    if height:
        layout_kwargs['height'] = height
    fig.update_layout(**layout_kwargs)
    return fig

@st.cache_data(ttl=3600)
def processar_dados(df):
    """Aplica as regras de negócio nos dados"""
    df['Valor_Real'] = df.apply(
        lambda row: row['TotalProduto'] if row['TipoMov'] == 'NF Venda' else -row['TotalProduto'],
        axis=1
    )
    # Converter DataEmissao com formato brasileiro e normalizar para meia-noite
    df['DataEmissao'] = pd.to_datetime(df['DataEmissao'], errors='coerce', dayfirst=True)
    # Normalizar para meia-noite (remove hora) para cálculos corretos de dias
    df['DataEmissao'] = df['DataEmissao'].dt.normalize()
    
    df['Mes'] = df['DataEmissao'].dt.month
    df['Ano'] = df['DataEmissao'].dt.year
    df['MesAno'] = df['DataEmissao'].dt.to_period('M').astype(str)
    
    # Calcular prazo histórico se a coluna DataVencimento existir
    if 'DataVencimento' in df.columns:
        df['PrazoHistorico'] = df.apply(
            lambda row: calcular_prazo_historico(row['DataEmissao'], row['DataVencimento']),
            axis=1
        )
    else:
        df['PrazoHistorico'] = ''
    
    return df

def obter_notas_unicas(df):
    """Remove duplicatas de Numero_NF mantendo apenas primeira ocorrência"""
    return df.drop_duplicates(subset=['Numero_NF'], keep='first')

def to_excel(df):
    """Converte DataFrame para Excel"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
    return output.getvalue()

def to_excel_pedidos_pendentes(df):
    """Converte DataFrame de pedidos pendentes para Excel com abas por tipo de produto"""
    output = io.BytesIO()
    
    # Função para identificar tipo de produto pela descrição
    def identificar_tipo(descricao):
        if pd.isna(descricao):
            return 'OUTROS'
        
        descricao_upper = str(descricao).upper()
        
        if 'ATADURA' in descricao_upper:
            return 'ATADURAS'
        elif 'CAMPO' in descricao_upper:
            return 'CAMPO'
        elif ('GAZE' in descricao_upper and 'ROLO' in descricao_upper) or ('GAZE' in descricao_upper and 'CIRCULAR' in descricao_upper):
            return 'GAZE EM ROLO'
        elif 'NAO ESTERIL' in descricao_upper or 'NÃO ESTERIL' in descricao_upper or 'NÃO ESTÉRIL' in descricao_upper or 'NAO ESTÉRIL' in descricao_upper:
            return 'NÃO ESTERIL'
        elif 'ESTERIL' in descricao_upper or 'ESTÉRIL' in descricao_upper:
            return 'ESTERIL'
        else:
            return 'OUTROS'
    
    # Função para identificar se é HOSPITALAR ou FARMA (apenas para ATADURAS)
    def identificar_categoria(descricao, tipo):
        if tipo == 'ATADURAS':
            if pd.notna(descricao) and 'HOSP' in str(descricao).upper():
                return 'HOSPITALAR'
            else:
                return 'FARMA'
        return ''
    
    # Adicionar coluna de tipo
    df_export = df.copy()
    df_export['TipoProduto'] = df_export['Descricao'].apply(identificar_tipo)
    df_export['Categoria'] = df_export.apply(lambda row: identificar_categoria(row['Descricao'], row['TipoProduto']), axis=1)
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Definir ordem das abas
        tipos_ordem = ['ATADURAS', 'CAMPO', 'ESTERIL', 'NÃO ESTERIL', 'GAZE EM ROLO', 'OUTROS']
        
        for tipo in tipos_ordem:
            df_tipo = df_export[df_export['TipoProduto'] == tipo].copy()
            
            if len(df_tipo) > 0:
                # Para ATADURAS, ordenar por Categoria (HOSPITALAR primeiro)
                if tipo == 'ATADURAS':
                    df_tipo = df_tipo.sort_values('Categoria')
                
                # Remover colunas auxiliares antes de exportar
                colunas_para_remover = ['TipoProduto']
                # Manter coluna Categoria apenas para ATADURAS
                if tipo != 'ATADURAS':
                    colunas_para_remover.append('Categoria')
                
                df_tipo = df_tipo.drop(columns=[col for col in colunas_para_remover if col in df_tipo.columns])
                
                df_tipo.to_excel(writer, index=False, sheet_name=tipo)
    
    return output.getvalue()

def formatar_moeda(valor):
    """Formata valor para moeda brasileira"""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_dataframe_moeda(df, colunas_moeda):
    """Formata colunas de moeda em um dataframe para exibição"""
    df_formatado = df.copy()
    for col in colunas_moeda:
        if col in df_formatado.columns:
            df_formatado[col] = df_formatado[col].apply(lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00")
    return df_formatado

def formatar_numero_br(valor, casas=0):
    """Formata número no padrão brasileiro (ponto milhar, vírgula decimal), sem prefixo R$"""
    try:
        texto = f"{valor:,.{casas}f}"
    except (ValueError, TypeError):
        return str(valor)
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")

@st.cache_data(ttl=3600)
def processar_inadimplencia(df):
    """Processa dados de inadimplência"""
    # Padronizar nomes das colunas
    # Tentar várias variações de nomes de colunas
    rename_map = {
        'Funcionário': 'Vendedor',
        'Razão Social': 'Cliente',
        'N_Doc': 'NumeroDoc',
        'N Doc': 'NumeroDoc',
        'NDoc': 'NumeroDoc',
        'Numero Doc': 'NumeroDoc',
        'Dt.Vencimento': 'DataVencimento',
        'Dt Vencimento': 'DataVencimento',
        'Data Vencimento': 'DataVencimento',
        'Vr.Líquido': 'ValorLiquido',
        'Vr Líquido': 'ValorLiquido',
        'Valor Líquido': 'ValorLiquido',
        'Valor Liquido': 'ValorLiquido',
        'Conta/Caixa': 'Banco',
        'Conta Caixa': 'Banco',
        'UF': 'Estado'
    }
    
    df = df.rename(columns=rename_map)
    
    # Converter data de vencimento
    df['DataVencimento'] = pd.to_datetime(df['DataVencimento'], errors='coerce')
    
    # Calcular dias de atraso
    hoje = pd.Timestamp.now()
    df['DiasAtraso'] = (hoje - df['DataVencimento']).dt.days
    df['DiasAtraso'] = df['DiasAtraso'].apply(lambda x: max(0, x))  # Não mostrar valores negativos
    
    # Classificar inadimplência
    def classificar_inadimplencia(dias):
        if dias == 0:
            return 'A Vencer'
        elif dias <= 30:
            return '1-30 dias'
        elif dias <= 60:
            return '31-60 dias'
        elif dias <= 90:
            return '61-90 dias'
        else:
            return 'Acima de 90 dias'
    
    df['FaixaAtraso'] = df['DiasAtraso'].apply(classificar_inadimplencia)
    
    return df


# ====================== PROPOSTA PDF (HISTÓRICO DE CLIENTE) ======================
def gerar_proposta_pdf_historico(cliente_info_dict, historico_df, vendas_resumo):
    """
    Gera PDF de Proposta Comercial baseada no histórico de compras do cliente.
    Usa apenas a biblioteca fpdf2 (pip install fpdf2).
    Fallback: se fpdf2 não estiver disponível, usa ReportLab.
    """
    import io, requests
    from datetime import date

    razao    = str(cliente_info_dict.get('RazaoSocial', ''))
    cpf_cnpj = str(cliente_info_dict.get('CPF_CNPJ', ''))
    cidade   = str(cliente_info_dict.get('Cidade', ''))
    estado   = str(cliente_info_dict.get('Estado', ''))
    vendedor = str(cliente_info_dict.get('Vendedor', ''))
    hoje     = date.today().strftime('%d/%m/%Y')

    # ── Tentar fpdf2 primeiro ─────────────────────────────────────────────
    try:
        from fpdf import FPDF

        class PropostaPDF(FPDF):
            def header(self):
                # Logo
                try:
                    resp = requests.get("https://i.imgur.com/gt3rgyL.png", timeout=8)
                    if resp.status_code == 200:
                        tmp = io.BytesIO(resp.content)
                        tmp.seek(0)
                        import tempfile, os
                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f_tmp:
                            f_tmp.write(resp.content)
                            tmp_path = f_tmp.name
                        self.image(tmp_path, x=10, y=8, w=32)
                        os.unlink(tmp_path)
                except Exception:
                    pass
                self.set_xy(46, 10)
                self.set_font('Helvetica', 'B', 13)
                self.set_text_color(31, 71, 136)
                self.cell(0, 6, 'MEDTEXTIL PRODUTOS TEXTIL HOSPITALARES', ln=True)
                self.set_xy(46, 16)
                self.set_font('Helvetica', '', 8)
                self.set_text_color(100, 100, 100)
                self.cell(0, 5, 'CNPJ: 40.357.820/0001-50  |  IE: 16.390.286-0', ln=True)
                self.set_draw_color(31, 71, 136)
                self.set_line_width(0.6)
                self.line(10, 26, 200, 26)
                self.ln(4)

            def footer(self):
                self.set_y(-14)
                self.set_font('Helvetica', 'I', 7)
                self.set_text_color(160, 160, 160)
                self.cell(0, 6,
                    f'Medtextil — Proposta gerada em {hoje}  |  Pág. {self.page_no()}',
                    align='C')

        pdf = PropostaPDF()
        pdf.set_auto_page_break(auto=True, margin=18)
        pdf.add_page()
        pdf.set_margins(10, 10, 10)

        # ── Título ────────────────────────────────────────────────────────
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_text_color(31, 71, 136)
        pdf.cell(0, 8, 'PROPOSTA COMERCIAL', align='C', ln=True)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(130, 130, 130)
        pdf.cell(0, 5, f'Emitida em {hoje}', align='C', ln=True)
        pdf.ln(4)

        # ── Dados do Cliente ──────────────────────────────────────────────
        pdf.set_fill_color(240, 244, 255)
        pdf.set_draw_color(200, 210, 230)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(31, 71, 136)
        pdf.cell(0, 7, ' DADOS DO CLIENTE', fill=True, border=1, ln=True)

        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(50, 50, 50)
        w1, w2 = 38, 152
        for label, valor in [
            ('Razão Social', razao),
            ('CPF / CNPJ',   cpf_cnpj),
            ('Cidade / UF',  f'{cidade} / {estado}'),
            ('Vendedor',     vendedor),
        ]:
            pdf.set_font('Helvetica', 'B', 8)
            pdf.cell(w1, 6, f'  {label}:', border='LB', fill=False)
            pdf.set_font('Helvetica', '', 8)
            pdf.cell(w2, 6, f'  {valor}', border='RB', ln=True)
        pdf.ln(5)

        # ── Resumo Financeiro ─────────────────────────────────────────────
        pdf.set_fill_color(240, 244, 255)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(31, 71, 136)
        pdf.cell(0, 7, ' RESUMO FINANCEIRO (PERÍODO SELECIONADO)', fill=True, border=1, ln=True)

        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(50, 50, 50)
        w3 = 95
        itens_resumo = list(vendas_resumo.items())
        for i in range(0, len(itens_resumo), 2):
            k1, v1 = itens_resumo[i]
            pdf.set_font('Helvetica', 'B', 8)
            pdf.cell(w3, 6, f'  {k1}:', border='LB')
            pdf.set_font('Helvetica', '', 8)
            if i + 1 < len(itens_resumo):
                k2, v2 = itens_resumo[i+1]
                pdf.cell(w3, 6, f'  {v1}', border='B')
                pdf.set_font('Helvetica', 'B', 8)
                pdf.cell(0, 6, f'  {k2}:', border='B')
            else:
                pdf.cell(0, 6, f'  {v1}', border='RB')
            pdf.ln()
        pdf.ln(5)

        # ── Tabela de Produtos ────────────────────────────────────────────
        pdf.set_fill_color(31, 71, 136)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 8)
        cols_w = [18, 72, 18, 22, 26, 26]
        cols_h = ['Código', 'Produto', 'Qtd', 'Prazo', 'Preço Unit.', 'Total']
        for cw, ch in zip(cols_w, cols_h):
            pdf.cell(cw, 7, ch, border=1, fill=True, align='C')
        pdf.ln()

        pdf.set_text_color(50, 50, 50)
        fill_row = False
        vendas_only = historico_df[historico_df['TipoMov'] == 'NF Venda'].copy()
        # Agrupar por produto para resumo
        grp_cols = ['CodigoProduto', 'NomeProduto']
        if 'PrazoHistorico' in vendas_only.columns:
            grp_cols_agg = {
                'Quantidade': 'sum',
                'PrecoUnit': 'mean',
                'TotalProduto': 'sum',
                'PrazoHistorico': 'first'
            }
        else:
            grp_cols_agg = {
                'Quantidade': 'sum',
                'PrecoUnit': 'mean',
                'TotalProduto': 'sum'
            }
        try:
            resumo_prod = vendas_only.groupby(
                ['CodigoProduto', 'NomeProduto'], as_index=False
            ).agg({k: v for k, v in grp_cols_agg.items() if k in vendas_only.columns})
            resumo_prod = resumo_prod.sort_values('TotalProduto', ascending=False)
        except Exception:
            resumo_prod = vendas_only[['CodigoProduto','NomeProduto','Quantidade','PrecoUnit','TotalProduto']].head(30)

        for _, row in resumo_prod.iterrows():
            pdf.set_fill_color(247, 249, 255) if fill_row else pdf.set_fill_color(255, 255, 255)
            pdf.set_font('Helvetica', '', 7)
            cod  = str(row.get('CodigoProduto', ''))[:8]
            nome = str(row.get('NomeProduto', ''))[:38]
            qtd  = f"{formatar_numero_br(row.get('Quantidade', 0), 0)}"
            prazo = str(row.get('PrazoHistorico', '-'))[:10] if 'PrazoHistorico' in row.index else '-'
            preco = f"R$ {formatar_numero_br(row.get('PrecoUnit', 0), 2)}"
            total = f"R$ {formatar_numero_br(row.get('TotalProduto', 0), 2)}"
            row_vals = [cod, nome, qtd, prazo, preco, total]
            aligns   = ['C', 'L', 'C', 'C', 'R', 'R']
            for cw, rv, al in zip(cols_w, row_vals, aligns):
                pdf.cell(cw, 6, rv, border=1, fill=True, align=al)
            pdf.ln()
            fill_row = not fill_row

        # Total geral
        total_geral = vendas_only['TotalProduto'].sum() if len(vendas_only) > 0 else 0
        pdf.set_fill_color(31, 71, 136)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(sum(cols_w[:5]), 7, 'TOTAL GERAL', border=1, fill=True, align='R')
        pdf.cell(cols_w[5], 7, f'R$ {formatar_numero_br(total_geral, 2)}', border=1, fill=True, align='R')
        pdf.ln(8)

        # ── Rodapé da proposta ────────────────────────────────────────────
        pdf.set_font('Helvetica', 'I', 8)
        pdf.set_text_color(130, 130, 130)
        pdf.multi_cell(0, 5,
            'Esta proposta é baseada no histórico de compras do cliente e não representa um pedido confirmado. '
            'Valores sujeitos a alteração. Validade: 15 dias.')

        return pdf.output()

    # ── Fallback: ReportLab ───────────────────────────────────────────────
    except ImportError:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=15*mm, leftMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        azul = colors.HexColor('#1F4788')
        elements = []

        # Título
        p_title = ParagraphStyle('T', parent=styles['Heading1'],
                                 fontSize=16, textColor=azul, alignment=1)
        elements.append(Paragraph('PROPOSTA COMERCIAL', p_title))
        elements.append(Paragraph(f'<font size=9 color="grey">Medtextil — {hoje}</font>', styles['Normal']))
        elements.append(Spacer(1, 6*mm))

        # Dados do cliente
        dados = [['Razão Social', razao], ['CPF/CNPJ', cpf_cnpj],
                 ['Cidade/UF', f'{cidade}/{estado}'], ['Vendedor', vendedor]]
        t_dados = Table(dados, colWidths=[40*mm, 150*mm])
        t_dados.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(0,-1), colors.HexColor('#F0F4FF')),
            ('TEXTCOLOR', (0,0),(0,-1), azul),
            ('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0),(-1,-1), 8),
            ('BOX', (0,0),(-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0),(-1,-1), 0.3, colors.lightgrey),
            ('LEFTPADDING', (0,0),(-1,-1), 5),
        ]))
        elements.append(t_dados)
        elements.append(Spacer(1, 6*mm))

        # Resumo
        for k, v in vendas_resumo.items():
            elements.append(Paragraph(f'<b>{k}:</b> {v}', styles['Normal']))
        elements.append(Spacer(1, 4*mm))

        # Tabela de produtos
        vendas_only = historico_df[historico_df['TipoMov'] == 'NF Venda']
        header = ['Código', 'Produto', 'Qtd', 'Preço Unit.', 'Total']
        rows = [header]
        try:
            grp = vendas_only.groupby(['CodigoProduto','NomeProduto'], as_index=False).agg(
                {'Quantidade':'sum','PrecoUnit':'mean','TotalProduto':'sum'})
            grp = grp.sort_values('TotalProduto', ascending=False)
            for _, r in grp.iterrows():
                rows.append([str(r['CodigoProduto'])[:8], str(r['NomeProduto'])[:40],
                             f"{formatar_numero_br(r['Quantidade'], 0)}", f"R$ {formatar_numero_br(r['PrecoUnit'], 2)}",
                             f"R$ {formatar_numero_br(r['TotalProduto'], 2)}"])
        except Exception:
            pass
        total_g = vendas_only['TotalProduto'].sum() if len(vendas_only) > 0 else 0
        rows.append(['','','','Total Geral', f'R$ {formatar_numero_br(total_g, 2)}'])

        t_prod = Table(rows, colWidths=[20*mm, 80*mm, 18*mm, 28*mm, 28*mm])
        t_prod.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0), azul),
            ('TEXTCOLOR', (0,0),(-1,0), colors.white),
            ('FONTNAME', (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0),(-1,-1), 7),
            ('ROWBACKGROUNDS', (0,1),(-1,-2), [colors.white, colors.HexColor('#F7F9FF')]),
            ('BACKGROUND', (0,-1),(-1,-1), colors.HexColor('#F0F4FF')),
            ('FONTNAME', (0,-1),(-1,-1), 'Helvetica-Bold'),
            ('BOX', (0,0),(-1,-1), 0.5, colors.grey),
            ('INNERGRID', (0,0),(-1,-1), 0.3, colors.lightgrey),
            ('LEFTPADDING', (0,0),(-1,-1), 3),
            ('ALIGN', (2,0),(-1,-1), 'RIGHT'),
        ]))
        elements.append(t_prod)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


# ====================== FILTROS DE DATA LOCAIS ======================
def renderizar_filtros_locais(key_prefix, label="📅 Ajustar Período"):
    """
    Expander compacto com date_inputs lado a lado.
    Retorna (data_inicial, data_final) — None se não preenchido.
    Usa key_prefix para evitar conflito de keys entre módulos.
    """
    data_ini = None
    data_fim = None
    with st.expander(label, expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            data_ini = st.date_input(
                "De", value=None,
                key=f"local_ini_{key_prefix}",
                format="DD/MM/YYYY",
                label_visibility="visible"
            )
        with c2:
            data_fim = st.date_input(
                "Até", value=None,
                key=f"local_fim_{key_prefix}",
                format="DD/MM/YYYY",
                label_visibility="visible"
            )
    return data_ini, data_fim

# ====================== INÍCIO DO APP ======================
if not check_password():
    st.stop()

# Obter informações do usuário logado
usuario = st.session_state.get("usuario", {})
tipo_usuario = usuario.get("tipo", "")
nome_usuario = usuario.get("nome", "Usuário")
modulos_permitidos = usuario.get("modulos", [])

# ── Sidebar: Logo + Usuário + Navegação (clean) ──────────────────────────
with st.sidebar:
    # ── Logo centralizado ──
    st.markdown("""
    <div style="text-align:center;padding:20px 0 14px 0;border-bottom:1px solid #E9ECEF;margin-bottom:14px;">
        <img src="https://i.imgur.com/gt3rgyL.png" height="56"
             style="border-radius:10px;box-shadow:0 2px 8px rgba(31,71,136,0.18);"
             onerror="this.style.display='none'"/>
        <div style="font-size:0.8rem;font-weight:800;color:#4A7BC8;letter-spacing:0.1em;
                    text-transform:uppercase;margin-top:10px;">Medtextil</div>
        <div style="font-size:0.65rem;color:#ADB5BD;letter-spacing:0.06em;margin-top:1px;">
            BI Dashboard 2.0
        </div>
    </div>
    """, unsafe_allow_html=True)
    # ── Badge do usuário ──
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#F0F4FF,#E8EFFD);border:1px solid #C5D5F0;
                border-radius:10px;padding:10px 14px;font-size:0.83rem;color:#4A7BC8;
                font-weight:600;text-align:center;margin-bottom:6px;">
        👤 &nbsp;{nome_usuario}
    </div>
    """, unsafe_allow_html=True)
    # ── Botão sair compacto ──
    if st.button("🚪 Sair", use_container_width=True):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
    st.markdown("<div style='margin-bottom:10px;'></div>", unsafe_allow_html=True)

# ── Cabeçalho principal ───────────────────────────────────────────────────
col_titulo, col_actions = st.columns([4, 1])
with col_titulo:
    st.markdown("# 📊 Dashboard Comercial")
    st.markdown('<p class="page-subtitle">Medtextil Produtos Textil Hospitalares — Análise de Vendas & BI</p>',
                unsafe_allow_html=True)

# ── Carregamento silencioso + Status no sidebar expander ─────────────────
with st.sidebar:
    with st.expander("🛠️ Status das Planilhas", expanded=False):
        with st.spinner("Conectando ao GitHub..."):
            planilhas_disponiveis = listar_planilhas_github()

        if planilhas_disponiveis['vendas']:
            st.success(f"✅ Vendas: {planilhas_disponiveis['vendas']['nome']}")
            url_planilha_vendas = planilhas_disponiveis['vendas']['url']
        else:
            st.error("❌ Planilha de vendas não encontrada")
            st.info("Procurando por arquivo com 'CONSULTA_VENDEDORES' no nome")

        if planilhas_disponiveis['inadimplencia']:
            st.success(f"✅ Inadimplência: {planilhas_disponiveis['inadimplencia']['nome']}")
        else:
            st.warning("⚠️ Inadimplência não encontrada (módulo desabilitado)")

        if planilhas_disponiveis.get('produtos_agrupados'):
            st.success(f"✅ Produtos: {planilhas_disponiveis['produtos_agrupados']['nome']}")

        if planilhas_disponiveis.get('contrato'):
            st.success(f"✅ Contratos: {planilhas_disponiveis['contrato']['nome']}")
        else:
            st.warning("⚠️ Planilha de contratos não encontrada")

        if planilhas_disponiveis.get('pedidos_pendentes'):
            st.success(f"✅ Pedidos Pendentes: {planilhas_disponiveis['pedidos_pendentes']['nome']}")
        else:
            st.warning("⚠️ Planilha de pedidos pendentes não encontrada")

        if st.button("🔄 Recarregar Dados", use_container_width=True, key="btn_reload"):
            st.cache_data.clear()
            st.rerun()

# Validação crítica fora do expander (sem mensagem visual)
if not planilhas_disponiveis.get('vendas'):
    st.error("❌ Planilha de vendas não encontrada no GitHub. Verifique o repositório.")
    st.stop()

with st.spinner(""):
    df = carregar_planilha_github(url_planilha_vendas)

if df is None:
    st.error("❌ Não foi possível carregar os dados de vendas.")
    st.stop()

df = processar_dados(df)

# Carregar planilha de produtos para cálculo de comissão
if planilhas_disponiveis.get('produtos_agrupados'):
    df_ref_preco = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
    if df_ref_preco is not None:
        df_ref_preco.columns = df_ref_preco.columns.str.upper()
        if 'ID_COD' in df_ref_preco.columns and 'PRECO' in df_ref_preco.columns:
            df_ref_preco = df_ref_preco[['ID_COD', 'PRECO']].rename(
                columns={'ID_COD': 'CodigoProduto', 'PRECO': 'PrecoRef'}
            )
            def normalizar_codigo(val):
                try:
                    return str(int(float(str(val).strip())))
                except Exception:
                    return str(val).strip()
            df['CodigoProduto'] = df['CodigoProduto'].apply(normalizar_codigo)
            df_ref_preco['CodigoProduto'] = df_ref_preco['CodigoProduto'].apply(normalizar_codigo)
            df_ref_preco = df_ref_preco.drop_duplicates(subset=['CodigoProduto'], keep='first')
            df = df.merge(df_ref_preco, on='CodigoProduto', how='left')
            df['Comissao'] = df.apply(
                lambda row: calcular_comissao(row['PrecoUnit'], row['PrecoRef']),
                axis=1
            )
        else:
            df['PrecoRef'] = None
            df['Comissao'] = ''
else:
    df['PrecoRef'] = None
    df['Comissao'] = ''

# ── Filtros Globais — dentro de expander único ───────────────────────────
with st.expander("⚙️ Filtros", expanded=False):
    # Linha 1: Datas lado a lado
    fc1, fc2 = st.columns(2)
    with fc1:
        data_inicial = st.date_input("📅 Data Inicial", value=None,
                                     key="data_ini", format="DD/MM/YYYY")
    with fc2:
        data_final = st.date_input("📅 Data Final", value=None,
                                   key="data_fim", format="DD/MM/YYYY")
    # Linha 2: Vendedor e Estado lado a lado
    fc3, fc4 = st.columns(2)
    with fc3:
        vendedores = ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist())
        vendedor_filtro = st.selectbox("👤 Vendedor", vendedores, key="vend_global")
    with fc4:
        estados = ['Todos'] + sorted(df['Estado'].dropna().unique().tolist())
        estado_filtro = st.selectbox("🗺️ Estado", estados, key="est_global")
    # Linha 3: Mês e Ano lado a lado
    fc5, fc6 = st.columns(2)
    with fc5:
        meses_opcoes = ['Todos'] + list(range(1, 13))
        mes_filtro = st.selectbox("📆 Mês", meses_opcoes, key="mes_global")
    with fc6:
        anos_opcoes = ['Todos'] + sorted(df['Ano'].dropna().unique().tolist(), reverse=True)
        ano_filtro = st.selectbox("🗓️ Ano", anos_opcoes, key="ano_global")
df_filtrado = df.copy()

if data_inicial:
    df_filtrado = df_filtrado[df_filtrado['DataEmissao'] >= pd.to_datetime(data_inicial)]
if data_final:
    df_filtrado = df_filtrado[df_filtrado['DataEmissao'] <= pd.to_datetime(data_final)]
if vendedor_filtro != 'Todos':
    df_filtrado = df_filtrado[df_filtrado['Vendedor'] == vendedor_filtro]
if estado_filtro != 'Todos':
    df_filtrado = df_filtrado[df_filtrado['Estado'] == estado_filtro]
if mes_filtro != 'Todos':
    df_filtrado = df_filtrado[df_filtrado['Mes'] == mes_filtro]
if ano_filtro != 'Todos':
    df_filtrado = df_filtrado[df_filtrado['Ano'] == ano_filtro]

notas_unicas = obter_notas_unicas(df_filtrado)

st.sidebar.markdown("---")

# ====================== NAVEGAÇÃO ======================

_DESC = {
    "Dashboard":          "Visão geral de faturamento",
    "Positivação":        "Clientes atendidos no período",
    "Inadimplência":      "Títulos em aberto e atrasos",
    "Clientes sem Compra":"Clientes inativos para reativar",
    "Histórico":          "Consulta por cliente ou vendedor",
    "Pedidos Pendentes":  "Itens aguardando faturamento",
    "Rankings":           "Top vendedores e clientes",
    "Performance de Vendedores": "Painel completo de KPIs por vendedor",
}
_INFO_CARD = {}  # preenchido depois dos dados

# ── Mapeamento de categorias na sidebar ──────────────────────────────────
_CATEGORIAS_NAV = {
    "GESTÃO COMERCIAL": [
        "Dashboard",
        "Performance de Vendedores",
        "Positivação",
        "Clientes sem Compra",
        "Novo Pedido",
    ],
    "PEDIDOS ERP": [
        "Novo Pedido ERP",
        "Meus Pedidos",
        "Fila de Aprovação",
        "Todos os Pedidos",
    ],
    "RELATÓRIOS E ATENÇÃO": [
        "Pedidos Pendentes",
        "Inadimplência",
    ],
    "CONSULTAS RÁPIDAS": [
        "Tabela de Preços",
        "Histórico do Cliente",
    ],
}

# Alias: label exibido → nome real do módulo no session_state
_ALIAS_MODULO = {
    "Novo Pedido":          "__novo_pedido__",
    "Novo Pedido ERP":      "__erp_novo_pedido__",
    "Meus Pedidos":         "__erp_meus_pedidos__",
    "Fila de Aprovação":    "__erp_fila_aprovacao__",
    "Todos os Pedidos":     "__erp_todos_pedidos__",
    "Tabela de Preços":     "Consulta Clientes",
    "Histórico do Cliente": "__historico_cliente__",
}
_ICONES_NAV = {
    "Dashboard":"▦","Positivação":"✓","Inadimplência":"⚠",
    "Clientes sem Compra":"＋","Histórico":"◷",
    "Pedidos Pendentes":"▣","Rankings":"▲","Performance de Vendedores":"★",
    "Novo Pedido":"📝","Tabela de Preços":"＄","Histórico do Cliente":"◷",
    "Novo Pedido ERP":"🆕","Meus Pedidos":"📋",
    "Fila de Aprovação":"⏳","Todos os Pedidos":"🗂️",
}

if 'menu_option' not in st.session_state:
    # Admin vai para home; colaborador vai direto para primeiro módulo
    _tipo_usuario = usuario.get('tipo', 'administrador')
    if _tipo_usuario == 'administrador':
        st.session_state.menu_option = '__home__'
    else:
        _primeiro_modulo = modulos_permitidos[0] if modulos_permitidos else 'Inadimplência'
        st.session_state.menu_option = _primeiro_modulo

modulos_visiveis = modulos_permitidos if modulos_permitidos else [
    "Dashboard","Positivação","Inadimplência","Clientes sem Compra",
    "Histórico","Pedidos Pendentes","Rankings","Performance de Vendedores"
]

# ══════════════════════════════════════════════════════════════════
# CSS GLOBAL — injetado uma única vez no topo do fluxo de página
# ══════════════════════════════════════════════════════════════════
# CSS unificado já injetado acima — bloco secundário removido.
# ══════════════════════════════════════════════════════════════════

# ── Sidebar: ícones por módulo ───────────────────────────────────────────
# _ICONES_NAV já definido acima com todos os módulos (incluindo Novo Pedido, Tabela de Preços, Histórico do Cliente)
_ICONES_CARD = {
    "Dashboard":"▦","Positivação":"✓","Inadimplência":"⚠",
    "Clientes sem Compra":"＋","Histórico":"◷",
    "Pedidos Pendentes":"▣","Rankings":"▲","Performance de Vendedores":"★",
    "Consulta Clientes":"＄","Histórico do Cliente":"◷","Novo Pedido":"📝",
}

with st.sidebar:
    st.markdown("""<div style="font-size:0.62rem;font-weight:700;color:#ADB5BD;
        letter-spacing:0.1em;text-transform:uppercase;
        margin-bottom:5px;padding-left:4px;">Navegação</div>""",
        unsafe_allow_html=True)

    # Botão Início
    if st.button("🏠  Início", key="nav_home", use_container_width=True,
                 type="primary" if st.session_state.menu_option == '__home__' else "secondary"):
        st.session_state.menu_option = '__home__'
        st.rerun()

    st.sidebar.markdown("---")

    # ── Navegação categorizada ──────────────────────────────────────────
    _CAT_ICONS = {
        "GESTÃO COMERCIAL":       "💰",
        "PEDIDOS ERP":            "🛒",
        "RELATÓRIOS E ATENÇÃO":   "📋",
        "CONSULTAS RÁPIDAS":      "🔍",
    }
    # Todos os módulos reais que o usuário tem acesso + aliases sempre visíveis
    _todos_reais = set(modulos_visiveis) | {"Consulta Clientes"}

    # Módulos ERP — só mostrar se Supabase disponível
    _modulos_erp = {
        "__erp_novo_pedido__", "__erp_meus_pedidos__",
        "__erp_fila_aprovacao__", "__erp_todos_pedidos__",
    }
    # Perfis que podem ver Fila de Aprovação e Todos os Pedidos
    _perfis_gestor = {"admin", "administrador", "gestor"}
    _tipo_atual = usuario.get("tipo", "colaborador")

    for _cat_label, _cat_itens in _CATEGORIAS_NAV.items():
        # Ocultar categoria PEDIDOS ERP se Supabase não configurado
        if _cat_label == "PEDIDOS ERP" and not supa_disponivel():
            continue
        _cat_icon = _CAT_ICONS.get(_cat_label, "•")
        st.sidebar.markdown(
            f"""<div style="font-size:0.60rem;font-weight:700;color:#8A96A8;
            letter-spacing:0.09em;text-transform:uppercase;
            margin:10px 0 3px 4px;">{_cat_icon} {_cat_label}</div>""",
            unsafe_allow_html=True
        )
        for _item in _cat_itens:
            # Resolver módulo real
            _modulo_real = _ALIAS_MODULO.get(_item, _item)
            # Controle de visibilidade ERP por perfil
            if _modulo_real in _modulos_erp:
                # Fila de Aprovação e Todos os Pedidos: só gestor/admin
                if _modulo_real in ("__erp_fila_aprovacao__", "__erp_todos_pedidos__"):
                    if _tipo_atual not in _perfis_gestor:
                        continue
            else:
                # Módulos clássicos: verificar permissão normal
                _especial = _modulo_real.startswith("__")
                if not _especial and _modulo_real not in _todos_reais:
                    continue
            _icone = _ICONES_NAV.get(_item, "•")
            _sel = (st.session_state.menu_option == _modulo_real)
            if st.button(f"{_icone}  {_item}", key=f"nav_{_item.replace(' ','_')}",
                         use_container_width=True,
                         type="primary" if _sel else "secondary"):
                st.session_state.menu_option = _modulo_real
                st.rerun()

    # Rankings como item avulso (sem categoria nova)
    if "Rankings" in _todos_reais:
        st.sidebar.markdown(
            """<div style="font-size:0.60rem;font-weight:700;color:#8A96A8;
            letter-spacing:0.09em;text-transform:uppercase;
            margin:10px 0 3px 4px;">📊 OUTROS</div>""",
            unsafe_allow_html=True
        )
        if st.button("▲  Rankings", key="nav_Rankings",
                     use_container_width=True,
                     type="primary" if st.session_state.menu_option == "Rankings" else "secondary"):
            st.session_state.menu_option = "Rankings"
            st.rerun()

    st.sidebar.markdown("---")
    st.sidebar.markdown("""<div style="font-size:0.62rem;font-weight:700;color:#ADB5BD;
        letter-spacing:0.1em;text-transform:uppercase;
        margin-bottom:5px;padding-left:4px;">Relatório Semanal</div>""",
        unsafe_allow_html=True)

    if st.button("📦 Gerar Relatório Semanal", key="btn_semanal", use_container_width=True):
        import zipfile, io as _io
        from datetime import date

        _hoje = pd.Timestamp.now()
        _inicio_mes = _hoje.replace(day=1)

        _zip_buf = _io.BytesIO()

        with st.sidebar.spinner("Gerando relatórios..."):
            try:
                # ── Carregar inadimplência ──
                _df_inad_sem = None
                if planilhas_disponiveis.get('inadimplencia'):
                    _raw_inad = carregar_planilha_github(planilhas_disponiveis['inadimplencia']['url'])
                    if _raw_inad is not None:
                        _df_inad_sem = processar_inadimplencia(_raw_inad)

                # ── Carregar pedidos pendentes ──
                _df_pend_sem = None
                if planilhas_disponiveis.get('pedidos_pendentes'):
                    try:
                        import zipfile as _zf, xml.etree.ElementTree as _ET
                        _resp_p = requests.get(planilhas_disponiveis['pedidos_pendentes']['url'])
                        _ef = _io.BytesIO(_resp_p.content)
                        _data_p = []
                        _cur_cli = None
                        with _zf.ZipFile(_ef) as _z:
                            with _z.open('xl/sharedStrings.xml') as _f:
                                _st = _ET.parse(_f)
                                _ns_s = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                _ss = [si.text if si.text else '' for si in _st.findall('.//ss:t', _ns_s)]
                            with _z.open('xl/worksheets/sheet1.xml') as _f:
                                _sh = _ET.parse(_f)
                                _ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                for _row in _sh.findall('.//ss:row', _ns):
                                    _rd = {}
                                    for _cell in _row.findall('.//ss:c', _ns):
                                        _ref = _cell.get('r', '')
                                        _col = ''.join([c for c in _ref if c.isalpha()])
                                        _ve = _cell.find('.//ss:v', _ns)
                                        if _ve is not None and _ve.text:
                                            if _cell.get('t', 'n') == 's':
                                                _idx = int(_ve.text)
                                                _rd[_col] = _ss[_idx] if _idx < len(_ss) else _ve.text
                                            else:
                                                _rd[_col] = _ve.text
                                    if not _rd:
                                        continue
                                    _ca, _cb = _rd.get('A', ''), _rd.get('B', '')
                                    if _ca and not _cb and 'N° do pedido' not in _ca and 'Valor Total' not in _ca and _ca != 'Subgrupo:':
                                        _cur_cli = _ca
                                    elif 'N° do pedido' in _ca:
                                        _desc = _rd.get('C', '')
                                        if _desc and ' - ' in _desc:
                                            try:
                                                import re as _re2
                                                _obs_m = _re2.search(r'observa[çc][aã]o[:\s]*', _desc, _re2.IGNORECASE)
                                                if _obs_m:
                                                    _obs = _desc[_obs_m.end():].strip()
                                                    _desc = _desc[:_obs_m.start()].strip()
                                                else:
                                                    _obs = ''
                                                _qtdc = float(_rd.get('D', 0))
                                                _vunit = float(_rd.get('E', 0))
                                                _qtde = float(_rd.get('H', 0))
                                                _qtdp = _qtdc - _qtde
                                                _dt_v = _rd.get('G', '')
                                                _dt_em = (pd.Timestamp('1899-12-30') + pd.Timedelta(days=float(_dt_v))) if _dt_v else None
                                                _data_p.append({
                                                    'Cliente': _cur_cli,
                                                    'NumeroPedido': _cb,
                                                    'CodigoProduto': _desc.split(' - ')[0].strip(),
                                                    'Descricao': _desc,
                                                    'Observacoes': _obs,
                                                    'QtdContratada': _qtdc,
                                                    'QtdEntregue': _qtde,
                                                    'QtdPendente': _qtdp,
                                                    'ValorUnit': _vunit,
                                                    'ValorPendente': _qtdp * _vunit,
                                                    'DataEmissao': _dt_em,
                                                    'Vendedor': _rd.get('J', ''),
                                                    'PercEntregue': float(_rd.get('I', 0))
                                                })
                                            except Exception:
                                                continue
                        _df_pend_sem = pd.DataFrame(_data_p)
                        if len(_df_pend_sem) > 0:
                            _df_pend_sem = _df_pend_sem.drop_duplicates(subset=['NumeroPedido', 'CodigoProduto'])
                        # Filtrar: apenas com quantidade pendente (independente do mês)
                        if len(_df_pend_sem) > 0 and 'QtdPendente' in _df_pend_sem.columns:
                            _df_pend_sem = _df_pend_sem[_df_pend_sem['QtdPendente'] > 0]
                    except Exception:
                        _df_pend_sem = None

                # ── Faturados: início do mês vigente até hoje ──
                _df_fat_sem = df[
                    (df['TipoMov'] == 'NF Venda') &
                    (df['DataEmissao'] >= _inicio_mes) &
                    (df['DataEmissao'] <= _hoje)
                ].copy()

                # ── Lista de vendedores por regra independente ──
                # Faturados: apenas quem tem NF Venda do início do mês vigente até hoje
                _vends_fat = set(df[
                    (df['TipoMov'] == 'NF Venda') &
                    (df['DataEmissao'] >= _inicio_mes) &
                    (df['DataEmissao'] <= _hoje)
                ]['Vendedor'].dropna().unique().tolist())
                # Inadimplência: todos com clientes na planilha de inadimplência
                _vends_inad = set(_df_inad_sem['Vendedor'].dropna().unique().tolist()) \
                    if _df_inad_sem is not None and len(_df_inad_sem) > 0 else set()
                # Pendentes: todos com clientes na planilha de pedidos pendentes
                _vends_pend = set(_df_pend_sem['Vendedor'].dropna().unique().tolist()) \
                    if _df_pend_sem is not None and len(_df_pend_sem) > 0 else set()
                _vendedores_ativos = sorted(_vends_fat | _vends_inad | _vends_pend)

                with zipfile.ZipFile(_zip_buf, 'w', zipfile.ZIP_DEFLATED) as _zout:
                    for _vend in _vendedores_ativos:
                        _vend_pasta = _vend.upper().replace(' ', '_')
                        _prefixo = f"RELATORIO SEMANAL/{_vend_pasta}/"

                        # ── 1. FATURADOS ──
                        _df_v_fat = _df_fat_sem[_df_fat_sem['Vendedor'] == _vend].copy()
                        if len(_df_v_fat) > 0:
                            _cols = [c for c in ['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado', 'Vendedor',
                                                  'DataEmissao', 'Numero_NF', 'TipoMov',
                                                  'CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit',
                                                  'TotalProduto', 'Valor_Real'] if c in _df_v_fat.columns]
                            _df_exp_f = _df_v_fat[_cols].copy()
                            _df_exp_f['DataEmissao'] = _df_exp_f['DataEmissao'].dt.strftime('%d/%m/%Y')

                            # Aba FATURAMENTO TOTAL: dedup Numero_NF + soma
                            _cols_oc = ['CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit', 'TotalProduto', 'Valor_Real']
                            _cols_ft = [c for c in _cols if c not in _cols_oc]
                            _df_ft = _df_v_fat.drop_duplicates(subset=['Numero_NF'], keep='first')[_cols_ft + ['TotalProduto']].copy()
                            _df_ft['DataEmissao'] = _df_ft['DataEmissao'].dt.strftime('%d/%m/%Y')
                            _soma_ft = _df_ft['TotalProduto'].sum()
                            _ln_tot = {c: '' for c in _df_ft.columns}
                            _ln_tot['TotalProduto'] = _soma_ft
                            _ln_tot['RazaoSocial'] = 'TOTAL'
                            _df_ft = pd.concat([_df_ft, pd.DataFrame([_ln_tot])], ignore_index=True)

                            _buf_f = _io.BytesIO()
                            with pd.ExcelWriter(_buf_f, engine='xlsxwriter') as _wr:
                                _wb = _wr.book
                                _df_exp_f.to_excel(_wr, index=False, sheet_name='PRODUTOS POR CLIENTE')
                                _ws1 = _wr.sheets['PRODUTOS POR CLIENTE']
                                if len(_df_exp_f) > 0:
                                    _ws1.add_table(0, 0, len(_df_exp_f), len(_df_exp_f.columns)-1, {
                                        'name': f'TblPC_{_vend_pasta[:10]}',
                                        'style': 'Table Style Medium 2',
                                        'columns': [{'header': c} for c in _df_exp_f.columns]
                                    })
                                _df_ft.to_excel(_wr, index=False, sheet_name='FATURAMENTO TOTAL')
                                _ws2 = _wr.sheets['FATURAMENTO TOTAL']
                                _nft = len(_df_ft) - 1
                                if _nft > 0:
                                    _ws2.add_table(0, 0, _nft, len(_df_ft.columns)-1, {
                                        'name': f'TblFT_{_vend_pasta[:10]}',
                                        'style': 'Table Style Medium 2',
                                        'columns': [{'header': c} for c in _df_ft.columns]
                                    })
                                _fmt_b = _wb.add_format({'bold': True, 'num_format': '#,##0.00'})
                                _sc = list(_df_ft.columns).index('TotalProduto')
                                _ws2.write(_nft + 1, _sc, _soma_ft, _fmt_b)
                            _zout.writestr(_prefixo + f"{_vend_pasta}_FATURADOS.xlsx", _buf_f.getvalue())

                        # ── 2. PENDENTES ──
                        if _df_pend_sem is not None and len(_df_pend_sem) > 0:
                            _df_v_pend = _df_pend_sem[_df_pend_sem['Vendedor'] == _vend].copy()
                            if len(_df_v_pend) > 0:
                                _buf_p = _io.BytesIO()
                                with pd.ExcelWriter(_buf_p, engine='xlsxwriter') as _wr:
                                    _df_v_pend.to_excel(_wr, index=False, sheet_name='PENDENTES')
                                    _wsp = _wr.sheets['PENDENTES']
                                    _wsp.add_table(0, 0, len(_df_v_pend), len(_df_v_pend.columns)-1, {
                                        'name': f'TblPend_{_vend_pasta[:10]}',
                                        'style': 'Table Style Medium 2',
                                        'columns': [{'header': c} for c in _df_v_pend.columns]
                                    })
                                _zout.writestr(_prefixo + f"PENDENTES_{_vend_pasta}.xlsx", _buf_p.getvalue())

                        # ── 3. INADIMPLÊNCIA ──
                        if _df_inad_sem is not None and len(_df_inad_sem) > 0:
                            _df_v_inad = _df_inad_sem[_df_inad_sem['Vendedor'] == _vend].copy()
                            if len(_df_v_inad) > 0:
                                _cols_inad = [c for c in ['Vendedor', 'Cliente', 'NumeroDoc', 'DataVencimento',
                                                           'ValorLiquido', 'DiasAtraso', 'FaixaAtraso', 'Banco', 'Estado']
                                              if c in _df_v_inad.columns]
                                _df_v_inad = _df_v_inad[_cols_inad].copy()
                                if 'DataVencimento' in _df_v_inad.columns:
                                    _df_v_inad['DataVencimento'] = _df_v_inad['DataVencimento'].dt.strftime('%d/%m/%Y')
                                _buf_i = _io.BytesIO()
                                with pd.ExcelWriter(_buf_i, engine='xlsxwriter') as _wr:
                                    _df_v_inad.to_excel(_wr, index=False, sheet_name='INADIMPLENCIA')
                                    _wsi = _wr.sheets['INADIMPLENCIA']
                                    _wsi.add_table(0, 0, len(_df_v_inad), len(_df_v_inad.columns)-1, {
                                        'name': f'TblInad_{_vend_pasta[:10]}',
                                        'style': 'Table Style Medium 2',
                                        'columns': [{'header': c} for c in _df_v_inad.columns]
                                    })
                                _zout.writestr(_prefixo + f"INADIMPLENCIA_{_vend_pasta}.xlsx", _buf_i.getvalue())

                st.session_state['_zip_semanal'] = _zip_buf.getvalue()
                st.session_state['_zip_semanal_nome'] = f"RELATORIO_SEMANAL_{_hoje.strftime('%d-%m-%Y')}.zip"
                st.rerun()

            except Exception as _e:
                st.sidebar.error(f"Erro: {_e}")

    if st.session_state.get('_zip_semanal'):
        st.sidebar.download_button(
            "💾 Baixar ZIP Semanal",
            st.session_state['_zip_semanal'],
            st.session_state.get('_zip_semanal_nome', 'RELATORIO_SEMANAL.zip'),
            "application/zip",
            key="download_zip_semanal"
        )


if st.session_state.menu_option == '__home__':
    usuario_info = st.session_state.get("usuario", {})

    # ── Calcular previews dos cards ──────────────────────────────────────
    _now   = pd.Timestamp.now()
    _mes   = _now.month
    _ano   = _now.year
    _6m    = _now - pd.DateOffset(months=6)

    # Dashboard
    try:
        vendas_mes = notas_unicas[
            (notas_unicas['DataEmissao'].dt.month == _mes) &
            (notas_unicas['DataEmissao'].dt.year  == _ano)
        ]['Valor_Real'].sum()
    except Exception:
        vendas_mes = 0

    try:
        _ano_anterior = _ano - 1
        _meses_pt = {1:'jan',2:'fev',3:'mar',4:'abr',5:'mai',6:'jun',
                     7:'jul',8:'ago',9:'set',10:'out',11:'nov',12:'dez'}
        vendas_mes_ano_ant = notas_unicas[
            (notas_unicas['TipoMov'] == 'NF Venda') &
            (notas_unicas['DataEmissao'].dt.month == _mes) &
            (notas_unicas['DataEmissao'].dt.year  == _ano_anterior)
        ]['TotalProduto'].sum()
        _info_dash_ano_ant = f" · R$ {formatar_numero_br(vendas_mes_ano_ant, 0)} em {_meses_pt[_mes]}/{_ano_anterior}"
    except Exception:
        _info_dash_ano_ant = ""

    # Positivação
    try:
        _base_total   = df['CPF_CNPJ'].nunique()
        _posit_mes    = df[
            (df['TipoMov']=='NF Venda') &
            (df['DataEmissao'].dt.month==_mes) &
            (df['DataEmissao'].dt.year==_ano)
        ]['CPF_CNPJ'].nunique()
        _info_posit   = f"{formatar_numero_br(_base_total, 0)} na base · {formatar_numero_br(_posit_mes, 0)} positivados no mês"
    except Exception:
        _info_posit   = "Base de clientes"

    # Inadimplência — carregada separadamente, usar placeholder se não disponível
    try:
        if planilhas_disponiveis.get('inadimplencia'):
            _df_inad = carregar_planilha_github(planilhas_disponiveis['inadimplencia']['url'])
            if _df_inad is not None:
                _df_inad = processar_inadimplencia(_df_inad)
                _val_inad = _df_inad['ValorLiquido'].sum()
                _cli_inad = _df_inad['Cliente'].nunique()
                _info_inad = f"R$ {formatar_numero_br(_val_inad, 0)} · {formatar_numero_br(_cli_inad, 0)} clientes"
            else:
                _info_inad = "Dados não disponíveis"
        else:
            _info_inad = "Planilha não configurada"
    except Exception:
        _info_inad = "Títulos em aberto e atrasos"

    # Clientes sem compra há mais de 6 meses
    try:
        _ultima_compra = df[df['TipoMov']=='NF Venda'].groupby('CPF_CNPJ')['DataEmissao'].max()
        _sem_6m = (_ultima_compra < _6m).sum()
        _info_churn = f"{formatar_numero_br(_sem_6m, 0)} clientes sem compra há +6 meses"
    except Exception:
        _info_churn = "Clientes inativos"

    # Pedidos pendentes — carregado separadamente
    try:
        if planilhas_disponiveis.get('pedidos_pendentes'):
            _resp = requests.get(planilhas_disponiveis['pedidos_pendentes']['url'], timeout=15)
            import zipfile, xml.etree.ElementTree as ET
            from io import BytesIO as _BytesIO
            _ef = _BytesIO(_resp.content)
            with zipfile.ZipFile(_ef) as _z:
                with _z.open('xl/sharedStrings.xml') as _f:
                    _st = ET.parse(_f)
                    _ns = {'ss':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    _ss = [s.text or '' for s in _st.findall('.//ss:t',_ns)]
                with _z.open('xl/worksheets/sheet1.xml') as _f:
                    _sh = ET.parse(_f)
                    _val_pend = 0.0; _cli_pend = set()
                    for _row in _sh.findall('.//ss:row',_ns):
                        _rd = {}
                        for _c in _row.findall('.//ss:c',_ns):
                            _ref=''.join([x for x in _c.get('r','') if x.isalpha()])
                            _v=_c.find('.//ss:v',_ns)
                            if _v is not None and _v.text:
                                _rd[_ref]=_ss[int(_v.text)] if _c.get('t')=='s' else _v.text
                        _ca=_rd.get('A',''); _cb=_rd.get('B','')
                        if _ca and not _cb and 'N° do pedido' not in _ca:
                            _cur_cli=_ca
                        elif 'N° do pedido' in _ca:
                            try:
                                _qc=float(_rd.get('D',0)); _vu=float(_rd.get('E',0))
                                _qe=float(_rd.get('H',0)); _qp=_qc-_qe
                                _val_pend+=_qp*_vu
                                if hasattr(_cur_cli,'__len__'): _cli_pend.add(_cur_cli)
                            except Exception: pass
            _info_pend = f"R$ {formatar_numero_br(_val_pend, 0)} · {formatar_numero_br(len(_cli_pend), 0)} clientes"
        else:
            _info_pend = "Aguardando faturamento"
    except Exception:
        _info_pend = "Aguardando faturamento"

    # Rankings — top 3 vendedores DO MÊS ATUAL
    try:
        # Filtrar apenas vendas do mês atual (mesma lógica do Dashboard)
        vendas_mes_atual = notas_unicas[
            (notas_unicas['TipoMov'] == 'NF Venda') &
            (notas_unicas['DataEmissao'].dt.month == _mes) &
            (notas_unicas['DataEmissao'].dt.year == _ano)
        ]
        _rank = vendas_mes_atual.groupby('Vendedor')['TotalProduto'].sum().nlargest(3)
        _info_rank = " · ".join([f"{v.split()[0]} R${formatar_numero_br(r, 0)}" for v,r in _rank.items()])
    except Exception:
        _info_rank = "Top vendedores e clientes"

    cards_data = [
        # 💰 GESTÃO COMERCIAL
        {'nome':'Dashboard',                'info':f'R$ {formatar_numero_br(vendas_mes, 0)} no mês atual{_info_dash_ano_ant}', 'cat':'💰 Gestão Comercial'},
        {'nome':'Performance de Vendedores','info':'Análise completa por vendedor',              'cat':'💰 Gestão Comercial'},
        {'nome':'Positivação',              'info':_info_posit,                                  'cat':'💰 Gestão Comercial'},
        {'nome':'Clientes sem Compra',      'info':_info_churn,                                  'cat':'💰 Gestão Comercial'},
        # 📋 RELATÓRIOS E ATENÇÃO
        {'nome':'Pedidos Pendentes',        'info':_info_pend,                                   'cat':'📋 Relatórios'},
        {'nome':'Inadimplência',            'info':_info_inad,                                   'cat':'📋 Relatórios'},
        # 🔍 CONSULTAS RÁPIDAS
        {'nome':'Consulta Clientes',        'info':'Busca de produtos e preços por estado',      'cat':'🔍 Consultas'},
        {'nome':'Histórico',                'info':'Por cliente, vendedor ou produto',            'cat':'🔍 Consultas'},
    ]
    cards_visiveis = [c for c in cards_data if c['nome'] in modulos_visiveis or c['nome'] in ('Consulta Clientes',)]

    st.markdown(f"""
    <div style="margin-bottom:20px;padding-bottom:14px;border-bottom:1px solid #E9ECEF;">
        <div style="font-size:1.45rem;font-weight:600;color:#2C5AA0;margin-bottom:3px;">
            Olá, {usuario_info.get('nome','Usuário')}
        </div>
        <div style="color:#8A96A8;font-size:0.87rem;">
            Selecione um módulo abaixo para iniciar a análise.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Grid 4 colunas — técnica de botão overlay que FUNCIONA
    try:
        from streamlit_card import card as st_card
        _USE_CARD_LIB = True
    except ImportError:
        _USE_CARD_LIB = False

    # Grid 2 colunas — funciona em mobile e desktop sem dependências externas
    _n_cols = 2

    for row_start in range(0, len(cards_visiveis), _n_cols):
        row = cards_visiveis[row_start:row_start+_n_cols]
        cols = st.columns(_n_cols)
        for j, c in enumerate(row):
            with cols[j]:
                nome = c['nome']
                desc = _DESC.get(nome, '')
                info = c['info']
                ic   = _ICONES_CARD.get(nome, '•')

                if _USE_CARD_LIB:
                    clicked = st_card(
                        title=f"{ic}  {nome}",
                        text=[desc, info],
                        key=f"hc_{nome}",
                        styles={
                            "card": {
                                "width": "100%",
                                "height": "190px",
                                "background-color": "var(--secondary-background-color)",
                                "border": "1px solid #E4E9F0",
                                "border-radius": "14px",
                                "box-shadow": "0 1px 6px rgba(31,71,136,0.07)",
                                "cursor": "pointer",
                                "padding": "22px 20px 18px 20px",
                                "transition": "box-shadow 0.18s ease, transform 0.18s ease, border-color 0.18s ease",
                                "font-family": "'Inter', 'Segoe UI', Roboto, sans-serif",
                                "margin": "0",
                            },
                            "title": {
                                "font-size": "1rem",
                                "font-weight": "700",
                                "color": "#2C5AA0",
                                "font-family": "'Inter', 'Segoe UI', Roboto, sans-serif",
                                "letter-spacing": "-0.01em",
                                "margin-bottom": "6px",
                                "white-space": "nowrap",
                                "overflow": "hidden",
                                "text-overflow": "ellipsis",
                            },
                            "text": {
                                "font-size": "0.78rem",
                                "color": "#6C757D",
                                "font-family": "'Inter', 'Segoe UI', Roboto, sans-serif",
                                "line-height": "1.45",
                                "font-weight": "400",
                            },
                            "filter": {"background": "rgba(0,0,0,0)"},
                        }
                    )
                    if clicked:
                        st.session_state.menu_option = nome
                        st.rerun()
                else:
                    st.markdown(f"""
                    <div style="background:var(--secondary-background-color);border:1px solid rgba(128,128,128,0.2);
                                border-radius:14px;padding:20px 18px;min-height:148px;
                                box-shadow:0 1px 6px rgba(31,71,136,0.07);
                                font-family:'Inter','Segoe UI',sans-serif;">
                        <div style="font-size:1rem;margin-bottom:10px;">{ic}</div>
                        <div style="font-size:0.95rem;font-weight:700;color:#2C5AA0;
                                    margin-bottom:5px;">{nome}</div>
                        <div style="font-size:0.76rem;color:#6C757D;
                                    margin-bottom:8px;">{desc}</div>
                        <div style="font-size:0.70rem;color:#ADB5BD;
                                    border-top:1px solid #F0F2F5;padding-top:7px;">{info}</div>
                    </div>
                    """, unsafe_allow_html=True)
                    if st.button(f"Abrir", key=f"hc_{nome}",
                                 use_container_width=True):
                        st.session_state.menu_option = nome
                        st.rerun()

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    st.stop()

# ── Módulo ativo ──────────────────────────────────────────────────────────
menu = st.session_state.menu_option

# Resolver alias para nome de exibição no breadcrumb
_ALIAS_DISPLAY = {
    "__novo_pedido__":       "Novo Pedido",
    "__historico_cliente__": "Histórico do Cliente",
    "Consulta Clientes":     "Tabela de Preços",
}
_menu_display = _ALIAS_DISPLAY.get(menu, menu)

# Módulos especiais que não precisam de verificação de permissão
_MENU_ESPECIAIS = {"__novo_pedido__", "__historico_cliente__", "Consulta Clientes", "Rankings"}

st.markdown(f"""
<div style="font-size:0.74rem;color:#ADB5BD;margin-bottom:14px;
            padding-bottom:10px;border-bottom:1px solid #F0F2F5;">
    <span style="color:#6C757D;">Início</span>
    <span style="margin:0 6px;color:#D0D5DE;">›</span>
    <span style="color:#4A7BC8;font-weight:600;">{_menu_display}</span>
</div>
""", unsafe_allow_html=True)


# ── Helpers e dados ERP (definidos antes da cadeia de menu) ─────────
# ╔══════════════════════════════════════════════════════════════════════╗
# ║               ERP — MÓDULO DE PEDIDOS (MVP v1.0)                   ║
# ╚══════════════════════════════════════════════════════════════════════╝

# ── Helpers visuais compartilhados pelos módulos ERP ─────────────────────

def _erp_badge(status):
    """Retorna HTML do badge colorido para cada status."""
    cfg = {
        "rascunho":      ("#F1F5F9", "#64748B", "📝 Rascunho"),
        "devolvido":     ("#FFF7ED", "#C2410C", "↩️ Devolvido"),
        "enviado":       ("#EFF6FF", "#1D4ED8", "📤 Aguard. Aprovação"),
        "aprovado":      ("#F0FDF4", "#15803D", "✅ Aprovado"),
        "em_separacao":  ("#FFF7ED", "#C2410C", "📦 Em Separação"),
        "faturado":      ("#14532D", "#FFFFFF", "🧾 Faturado"),
        "cancelado":     ("#FEF2F2", "#B91C1C", "❌ Cancelado"),
    }
    bg, cor, label = cfg.get(status, ("#F1F5F9", "#64748B", status.title()))
    return (f'<span style="background:{bg};color:{cor};padding:3px 10px;'
            f'border-radius:12px;font-size:0.78rem;font-weight:600;">'
            f'{label}</span>')

def _erp_kpi(col, label, valor, cor="#1F4788"):
    """KPI card compacto para os painéis ERP."""
    with col:
        st.markdown(
            f'<div style="background:#F8FAFC;border-left:3px solid {cor};'
            f'border-radius:8px;padding:10px 14px;margin-bottom:8px;">'
            f'<div style="font-size:0.72rem;color:#6C757D;">{label}</div>'
            f'<div style="font-size:1.15rem;font-weight:700;color:{cor};">{valor}</div>'
            f'</div>', unsafe_allow_html=True
        )

def _erp_aviso_sem_supabase():
    st.warning(
        "⚙️ **Módulo ERP não configurado.**\n\n"
        "Para ativar o controle de pedidos, configure as credenciais do Supabase "
        "em `.streamlit/secrets.toml`:\n\n"
        "```toml\n[supabase]\nurl = \"https://XXXX.supabase.co\"\n"
        "key = \"eyJ...\"\n```\n\n"
        "Consulte a documentação do projeto para criar as tabelas necessárias."
    )

# ── Dados do usuário atual (disponíveis em todos os módulos ERP) ─────────
_erp_usuario    = st.session_state.get("usuario", {})
_erp_user_id    = _erp_usuario.get("id", "")
_erp_user_nome  = _erp_usuario.get("nome", "Usuário")
_erp_user_tipo  = _erp_usuario.get("tipo", "colaborador")
_erp_is_gestor  = _erp_user_tipo in ("admin", "administrador", "gestor")


if menu not in modulos_permitidos and menu not in _MENU_ESPECIAIS:
    st.markdown("""
    <div style="background:#FFF3F3;border:1px solid #F5C6CB;border-radius:10px;
                padding:16px 20px;color:#721C24;font-size:0.9rem;">
        Acesso negado. Você não tem permissão para acessar este módulo.
    </div>""", unsafe_allow_html=True)
    st.stop()
# ====================== DASHBOARD ======================
elif menu == "Dashboard":
    # Filtro local de período (não afeta os demais módulos)
    _dash_ini, _dash_fim = renderizar_filtros_locais("dash", "📅 Ajustar Período")
    _dash_notas = notas_unicas.copy()
    _dash_df_filtrado = df_filtrado.copy()
    if _dash_ini:
        _dash_notas = _dash_notas[_dash_notas['DataEmissao'] >= pd.to_datetime(_dash_ini)]
        _dash_df_filtrado = _dash_df_filtrado[_dash_df_filtrado['DataEmissao'] >= pd.to_datetime(_dash_ini)]
    if _dash_fim:
        _dash_notas = _dash_notas[_dash_notas['DataEmissao'] <= pd.to_datetime(_dash_fim)]
        _dash_df_filtrado = _dash_df_filtrado[_dash_df_filtrado['DataEmissao'] <= pd.to_datetime(_dash_fim)]

    # KPIs principais com cards customizados
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        vendas_brutas = _dash_notas[_dash_notas['TipoMov'] == 'NF Venda']['TotalProduto'].sum()
        render_kpi_card("Faturamento Bruto", f"R$ {formatar_numero_br(vendas_brutas, 0)}", icon="💰", color="#1F4788")
    
    with col2:
        faturamento_liquido = (
            _dash_notas[_dash_notas['TipoMov'] == 'NF Venda']['TotalProduto'].sum() -
            _dash_notas[_dash_notas['TipoMov'] == 'NF Dev.Venda']['TotalProduto'].sum()
        )
        render_kpi_card("Faturamento Líquido", f"R$ {formatar_numero_br(faturamento_liquido, 0)}", icon="💵", color="#10B981")
    
    with col3:
        clientes_unicos = _dash_df_filtrado['CPF_CNPJ'].nunique()
        render_kpi_card("Clientes Únicos", f"{formatar_numero_br(clientes_unicos, 0)}", icon="👥", color="#F59E0B")
    
    with col4:
        total_notas = len(_dash_notas[_dash_notas['TipoMov'] == 'NF Venda'])
        render_kpi_card("Notas de Venda", f"{formatar_numero_br(total_notas, 0)}", icon="📄", color="#EF4444")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Segunda linha de KPIs
    col1b, col2b, col3b, col4b = st.columns(4)
    
    with col1b:
        total_devolucoes = _dash_notas[_dash_notas['TipoMov'] == 'NF Dev.Venda']['TotalProduto'].sum()
        render_kpi_card("Devoluções", f"R$ {formatar_numero_br(total_devolucoes, 0)}", icon="↩️", color="#E5E7EB")
    
    with col2b:
        ticket_medio = vendas_brutas / clientes_unicos if clientes_unicos > 0 else 0
        render_kpi_card("Ticket Médio", f"R$ {formatar_numero_br(ticket_medio, 0)}", icon="🎯", color="#E5E7EB")
    
    with col3b:
        qtd_notas_dev = len(_dash_notas[_dash_notas['TipoMov'] == 'NF Dev.Venda'])
        render_kpi_card("Notas Devolução", f"{formatar_numero_br(qtd_notas_dev, 0)}", icon="📋", color="#E5E7EB")
    
    with col4b:
        taxa_devolucao = (total_devolucoes / vendas_brutas * 100) if vendas_brutas > 0 else 0
        render_kpi_card("Taxa Devolução", f"{taxa_devolucao:.1f}%", icon="📊", color="#E5E7EB")
    
    st.markdown("---")

    # Linha 1: 3 gráficos
    col5, col6, col7 = st.columns(3)

    with col5:
        st.subheader("📈 Evolução de Vendas")
        vendas_tempo = _dash_notas[_dash_notas['TipoMov'] == 'NF Venda'].groupby('MesAno')['TotalProduto'].sum().reset_index().sort_values('MesAno')
        if len(vendas_tempo) > 0:
            fig_linha = px.line(vendas_tempo, x='MesAno', y='TotalProduto',
                labels={'MesAno': 'Período', 'TotalProduto': 'Valor (R$)'})
            fig_linha.update_traces(line_color='#1F4788', line_width=3, mode='lines+markers', marker=dict(size=5, color='#1F4788'))
            fig_linha.update_layout(xaxis_title="Período", yaxis_title="Valor (R$)", hovermode='x unified')
            fig_linha = aplicar_layout_grafico(fig_linha)
            st.plotly_chart(fig_linha, use_container_width=True)
        else:
            st.info("Sem dados para exibir")

    with col6:
        st.subheader("🗺️ Top 10 Estados")
        vendas_estado = _dash_notas[_dash_notas['TipoMov'] == 'NF Venda'].groupby('Estado')['TotalProduto'].sum().reset_index().sort_values('TotalProduto', ascending=False).head(10)
        fig_bar = px.bar(vendas_estado, x='Estado', y='TotalProduto',
            labels={'Estado': 'Estado', 'TotalProduto': 'Valor (R$)'},
            color_discrete_sequence=['#2E86AB'])
        fig_bar = aplicar_layout_grafico(fig_bar)
        st.plotly_chart(fig_bar, use_container_width=True)

    with col7:
        st.subheader("👥 Positivação por Vendedor")
        atendidos = _dash_df_filtrado[_dash_df_filtrado['TipoMov'] == 'NF Venda'].groupby('Vendedor')['CPF_CNPJ'].nunique().reset_index()
        atendidos.columns = ['Vendedor', 'Clientes']
        atendidos = atendidos.sort_values('Clientes', ascending=False).head(10)
        fig_posit = px.bar(atendidos, x='Vendedor', y='Clientes',
            labels={'Vendedor': 'Vendedor', 'Clientes': 'Clientes Atendidos'},
            color_discrete_sequence=['#1F4788'])
        fig_posit = aplicar_layout_grafico(fig_posit)
        st.plotly_chart(fig_posit, use_container_width=True)

    st.markdown("---")

    # Linha 2: 3 gráficos
    col8, col9, col10 = st.columns(3)

    with col8:
        st.subheader("🏆 Top 10 Clientes")
        ranking_clientes = _dash_notas[_dash_notas['TipoMov'] == 'NF Venda'].groupby('RazaoSocial')['TotalProduto'].sum().reset_index().sort_values('TotalProduto', ascending=False).head(10)
        fig_clientes = px.bar(ranking_clientes, x='TotalProduto', y='RazaoSocial', orientation='h',
            labels={'RazaoSocial': 'Cliente', 'TotalProduto': 'Valor (R$)'},
            color_discrete_sequence=['#4A7BC8'])
        fig_clientes = aplicar_layout_grafico(fig_clientes)
        st.plotly_chart(fig_clientes, use_container_width=True)

    with col9:
        st.subheader("⚠️ Clientes sem Compra")
        _com_venda = set(_dash_df_filtrado[_dash_df_filtrado['TipoMov'] == 'NF Venda']['CPF_CNPJ'].unique())
        _todos = df.sort_values('DataEmissao').groupby('CPF_CNPJ').last().reset_index()
        _vhist = df[df['TipoMov'] == 'NF Venda'].groupby('CPF_CNPJ')['TotalProduto'].sum().reset_index()
        _vhist.columns = ['CPF_CNPJ', 'ValorHistorico']
        _todos = pd.merge(_todos, _vhist, on='CPF_CNPJ', how='left').fillna(0)
        _sem = _todos[~_todos['CPF_CNPJ'].isin(_com_venda)].sort_values('ValorHistorico', ascending=False).head(10)
        fig_churn = px.bar(_sem, x='ValorHistorico', y='RazaoSocial', orientation='h',
            labels={'RazaoSocial': 'Cliente', 'ValorHistorico': 'Valor Histórico (R$)'},
            color_discrete_sequence=['#1F4788'])
        fig_churn = aplicar_layout_grafico(fig_churn)
        st.plotly_chart(fig_churn, use_container_width=True)

    with col10:
        st.subheader("📊 Ranking de Vendedores")
        ranking_vendedores = _dash_notas[_dash_notas['TipoMov'] == 'NF Venda'].groupby('Vendedor')['TotalProduto'].sum().reset_index().sort_values('TotalProduto', ascending=False).head(10)
        fig_rank_vend = px.bar(ranking_vendedores, x='TotalProduto', y='Vendedor', orientation='h',
            labels={'Vendedor': 'Vendedor', 'TotalProduto': 'Valor Total (R$)'},
            color_discrete_sequence=['#163561'])
        fig_rank_vend = aplicar_layout_grafico(fig_rank_vend)
        st.plotly_chart(fig_rank_vend, use_container_width=True)

# ====================== POSITIVAÇÃO ======================
elif menu == "Positivação":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Relatório de Positivação</h2>', unsafe_allow_html=True)

    # ── KPIs do mês vigente no topo ───────────────────────────────────────
    _mes_atual = pd.Timestamp.now().month
    _ano_atual = pd.Timestamp.now().year
    _vendas_mes = df_filtrado[
        (df_filtrado['TipoMov'] == 'NF Venda') &
        (df_filtrado['DataEmissao'].dt.month == _mes_atual) &
        (df_filtrado['DataEmissao'].dt.year == _ano_atual)
    ]
    _posit_mes    = _vendas_mes['CPF_CNPJ'].nunique()
    _total_base   = df['CPF_CNPJ'].nunique()
    _perc_posit   = (_posit_mes / _total_base * 100) if _total_base > 0 else 0

    _kp1, _kp2, _kp3 = st.columns(3)
    with _kp1:
        st.metric("Positivados no Mês", f"{formatar_numero_br(_posit_mes, 0)} clientes",
                  help="Clientes com ao menos uma compra no mês vigente")
    with _kp2:
        st.metric("Total da Base", f"{formatar_numero_br(_total_base, 0)} clientes",
                  help="Total de clientes únicos na base")
    with _kp3:
        st.metric("% da Base Positivada", f"{_perc_posit:.1f}%",
                  help="Percentual da base que comprou no mês vigente")

    st.markdown("---")

    tab1, tab2, tab3_fat, tab4_prod = st.tabs(["📊 Por Vendedor", "🗺️ Por Estado", "🧾 Pedidos Faturados", "📦 Faturamento por Produto"])
    
    with tab1:
        base_vendedor = df.groupby('Vendedor')['CPF_CNPJ'].nunique().reset_index()
        base_vendedor.columns = ['Vendedor', 'TotalBase']
        
        vendas_periodo = df_filtrado[df_filtrado['TipoMov'] == 'NF Venda']
        atendidos = vendas_periodo.groupby('Vendedor')['CPF_CNPJ'].nunique().reset_index()
        atendidos.columns = ['Vendedor', 'QtdAtendidos']
        
        valor_vendedor = obter_notas_unicas(vendas_periodo).groupby('Vendedor')['Valor_Real'].sum().reset_index()
        valor_vendedor.columns = ['Vendedor', 'ValorTotal']
        
        relatorio_positivacao = pd.merge(base_vendedor, atendidos, on='Vendedor', how='left')
        relatorio_positivacao = pd.merge(relatorio_positivacao, valor_vendedor, on='Vendedor', how='left')
        relatorio_positivacao['QtdAtendidos'] = relatorio_positivacao['QtdAtendidos'].fillna(0).astype(int)
        relatorio_positivacao['ValorTotal'] = relatorio_positivacao['ValorTotal'].fillna(0)
        relatorio_positivacao['Percentual'] = (relatorio_positivacao['QtdAtendidos'] / relatorio_positivacao['TotalBase'] * 100).round(1)
        relatorio_positivacao = relatorio_positivacao.sort_values('QtdAtendidos', ascending=False)
        
        fig_posit_vend = px.bar(
            relatorio_positivacao.head(15),
            x='Vendedor',
            y='Percentual',
            labels={'Vendedor': 'Vendedor', 'Percentual': 'Positivação (%)'},
            color='Percentual',
            color_discrete_sequence=['#1F4788'],
            title='Top 15 Vendedores - Taxa de Positivação'
        )
        fig_posit_vend = aplicar_layout_grafico(fig_posit_vend)
        st.plotly_chart(fig_posit_vend, use_container_width=True)
        
        # Formatar para exibição
        relatorio_positivacao_display = formatar_dataframe_moeda(relatorio_positivacao, ['ValorTotal'])
        st.dataframe(relatorio_positivacao_display, use_container_width=True)
        
        st.download_button(
            "📥 Exportar Positivação por Vendedor",
            to_excel(relatorio_positivacao),
            "positivacao_vendedor.xlsx",
            "application/vnd.ms-excel",
            key="dl_posit_vendedor"
        )
        
        st.markdown("---")
        
        st.subheader("📋 Detalhamento de Clientes")
        vendedor_selecionado = st.selectbox(
            "Selecione o vendedor",
            relatorio_positivacao['Vendedor'].tolist()
        )
        
        if vendedor_selecionado:
            notas_vendedor = obter_notas_unicas(vendas_periodo[vendas_periodo['Vendedor'] == vendedor_selecionado])
            
            clientes_vendedor = notas_vendedor.groupby(['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado']).agg({
                'Valor_Real': 'sum'
            }).reset_index()
            clientes_vendedor.columns = ['CPF/CNPJ', 'Razão Social', 'Cidade', 'Estado', 'Valor Total']
            clientes_vendedor = clientes_vendedor.sort_values('Valor Total', ascending=False)
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Clientes Atendidos", len(clientes_vendedor))
            with col2:
                st.metric("Valor Total", f"R$ {formatar_numero_br(clientes_vendedor['Valor Total'].sum(), 2)}")
            
            # Formatar para exibição
            clientes_vendedor_display = formatar_dataframe_moeda(clientes_vendedor, ['Valor Total'])
            st.dataframe(clientes_vendedor_display, use_container_width=True)
            
            st.download_button(
                f"📥 Exportar Clientes - {vendedor_selecionado}",
                to_excel(clientes_vendedor),
                f"clientes_{vendedor_selecionado}.xlsx",
                "application/vnd.ms-excel",
                key="dl_posit_cliente_det"
            )
    
    with tab2:
        st.subheader("🗺️ Positivação por Estado")
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            vendedor_estado_filtro = st.selectbox(
                "Filtrar por Vendedor",
                ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist()),
                key="vend_estado"
            )
        with col_f2:
            ano_estado_filtro = st.selectbox(
                "Filtrar por Ano",
                ['Todos'] + sorted(df['Ano'].dropna().unique().tolist(), reverse=True),
                key="ano_estado"
            )
        
        df_estado_filtrado = df_filtrado.copy()
        if vendedor_estado_filtro != 'Todos':
            df_estado_filtrado = df_estado_filtrado[df_estado_filtrado['Vendedor'] == vendedor_estado_filtro]
        if ano_estado_filtro != 'Todos':
            df_estado_filtrado = df_estado_filtrado[df_estado_filtrado['Ano'] == ano_estado_filtro]
        
        base_estado = df.groupby('Estado')['CPF_CNPJ'].nunique().reset_index()
        base_estado.columns = ['Estado', 'TotalBase']
        
        vendas_estado = df_estado_filtrado[df_estado_filtrado['TipoMov'] == 'NF Venda']
        atendidos_estado = vendas_estado.groupby('Estado')['CPF_CNPJ'].nunique().reset_index()
        atendidos_estado.columns = ['Estado', 'QtdAtendidos']
        
        valor_estado = obter_notas_unicas(vendas_estado).groupby('Estado')['Valor_Real'].sum().reset_index()
        valor_estado.columns = ['Estado', 'ValorTotal']
        
        relatorio_estado = pd.merge(base_estado, atendidos_estado, on='Estado', how='left')
        relatorio_estado = pd.merge(relatorio_estado, valor_estado, on='Estado', how='left')
        relatorio_estado['QtdAtendidos'] = relatorio_estado['QtdAtendidos'].fillna(0).astype(int)
        relatorio_estado['ValorTotal'] = relatorio_estado['ValorTotal'].fillna(0)
        relatorio_estado['Percentual'] = (relatorio_estado['QtdAtendidos'] / relatorio_estado['TotalBase'] * 100).round(1)
        relatorio_estado = relatorio_estado.sort_values('Percentual', ascending=False)
        
        fig_posit_estado = px.bar(
            relatorio_estado.head(15),
            x='Estado',
            y='Percentual',
            labels={'Estado': 'Estado', 'Percentual': 'Positivação (%)'},
            color='Percentual',
            color_discrete_sequence=['#2E86AB'],
            title='Top 15 Estados - Taxa de Positivação'
        )
        fig_posit_estado = aplicar_layout_grafico(fig_posit_estado)
        st.plotly_chart(fig_posit_estado, use_container_width=True)
        
        # Formatar para exibição
        relatorio_estado_display = formatar_dataframe_moeda(relatorio_estado, ['ValorTotal'])
        st.dataframe(relatorio_estado_display, use_container_width=True)
        
        st.download_button(
            "📥 Exportar Positivação por Estado",
            to_excel(relatorio_estado),
            "positivacao_estado.xlsx",
            "application/vnd.ms-excel",
            key="dl_posit_estado"
        )

    with tab3_fat:
        st.subheader("🧾 Relatório de Pedidos Faturados")

        # ── Filtros locais ──
        _fc1, _fc2, _fc3 = st.columns(3)
        with _fc1:
            _fat_vendedores = ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist())
            _fat_vend = st.selectbox("👤 Vendedor", _fat_vendedores, key="fat_vend")
        with _fc2:
            _fat_regioes = ['Todos'] + sorted(df['Estado'].dropna().unique().tolist())
            _fat_reg = st.selectbox("🗺️ Estado/Região", _fat_regioes, key="fat_reg")
        with _fc3:
            _fat_col1, _fat_col2 = st.columns(2)
            with _fat_col1:
                _fat_di = st.date_input("📅 De", value=None, key="fat_di", format="DD/MM/YYYY")
            with _fat_col2:
                _fat_df = st.date_input("📅 Até", value=None, key="fat_df", format="DD/MM/YYYY")

        # ── Aplicar filtros ──
        _df_fat = df[df['TipoMov'] == 'NF Venda'].copy()
        if _fat_vend != 'Todos':
            _df_fat = _df_fat[_df_fat['Vendedor'] == _fat_vend]
        if _fat_reg != 'Todos':
            _df_fat = _df_fat[_df_fat['Estado'] == _fat_reg]
        if _fat_di:
            _df_fat = _df_fat[_df_fat['DataEmissao'] >= pd.to_datetime(_fat_di)]
        if _fat_df:
            _df_fat = _df_fat[_df_fat['DataEmissao'] <= pd.to_datetime(_fat_df)]

        if len(_df_fat) == 0:
            st.info("Nenhum registro encontrado com os filtros selecionados.")
        else:
            # ── KPIs ──
            _fk1, _fk2, _fk3 = st.columns(3)
            with _fk1:
                st.metric("Total Faturado", f"R$ {formatar_numero_br(_df_fat['TotalProduto'].sum(), 2)}")
            with _fk2:
                st.metric("Notas Fiscais", obter_notas_unicas(_df_fat)['Numero_NF'].nunique())
            with _fk3:
                st.metric("Clientes", _df_fat['CPF_CNPJ'].nunique())

            # ── Colunas para exibição/exportação ──
            _cols_base = ['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado', 'Vendedor',
                          'DataEmissao', 'Numero_NF', 'TipoMov',
                          'CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit',
                          'TotalProduto', 'Valor_Real']
            _cols_disp = [c for c in _cols_base if c in _df_fat.columns]
            _df_fat_disp = _df_fat[_cols_disp].copy()
            _df_fat_disp['DataEmissao'] = _df_fat_disp['DataEmissao'].dt.strftime('%d/%m/%Y')
            st.dataframe(_df_fat_disp, use_container_width=True, height=400)

            # ── Exportar Excel com duas abas como tabela ──
            def _gerar_excel_faturado(df_src, nome_vend):
                import io
                _cols = [c for c in ['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado', 'Vendedor',
                                     'DataEmissao', 'Numero_NF', 'TipoMov',
                                     'CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit',
                                     'TotalProduto', 'Valor_Real'] if c in df_src.columns]
                _df_exp = df_src[_cols].copy()
                _df_exp['DataEmissao'] = _df_exp['DataEmissao'].dt.strftime('%d/%m/%Y')

                # Aba FATURAMENTO TOTAL: dedup por Numero_NF + soma TotalProduto
                _cols_ocultas = ['CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit', 'TotalProduto', 'Valor_Real']
                _cols_fat_total = [c for c in _cols if c not in _cols_ocultas]
                _df_fat_total = (
                    df_src.drop_duplicates(subset=['Numero_NF'], keep='first')
                    [_cols_fat_total + ['TotalProduto']]
                    .copy()
                )
                _df_fat_total['DataEmissao'] = _df_fat_total['DataEmissao'].dt.strftime('%d/%m/%Y')
                _soma = _df_fat_total['TotalProduto'].sum()
                _linha_total = {c: '' for c in _df_fat_total.columns}
                _linha_total['TotalProduto'] = _soma
                _linha_total['RazaoSocial'] = 'TOTAL'
                _df_fat_total = pd.concat([_df_fat_total, pd.DataFrame([_linha_total])], ignore_index=True)

                import io
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    wb = writer.book

                    # ── Aba 1: PRODUTOS POR CLIENTE ──
                    _df_exp.to_excel(writer, index=False, sheet_name='PRODUTOS POR CLIENTE')
                    ws1 = writer.sheets['PRODUTOS POR CLIENTE']
                    if len(_df_exp) > 0:
                        ws1.add_table(0, 0, len(_df_exp), len(_df_exp.columns) - 1, {
                            'name': 'TblProdutosCliente',
                            'style': 'Table Style Medium 2',
                            'columns': [{'header': c} for c in _df_exp.columns]
                        })

                    # ── Aba 2: FATURAMENTO TOTAL ──
                    _df_fat_total.to_excel(writer, index=False, sheet_name='FATURAMENTO TOTAL')
                    ws2 = writer.sheets['FATURAMENTO TOTAL']
                    _nrows_ft = len(_df_fat_total) - 1  # última linha é total, fora da tabela
                    if _nrows_ft > 0:
                        ws2.add_table(0, 0, _nrows_ft, len(_df_fat_total.columns) - 1, {
                            'name': 'TblFaturamentoTotal',
                            'style': 'Table Style Medium 2',
                            'columns': [{'header': c} for c in _df_fat_total.columns]
                        })
                    # Linha de soma logo após a tabela
                    _fmt_bold = wb.add_format({'bold': True, 'num_format': '#,##0.00'})
                    _soma_row = _nrows_ft + 1
                    _soma_col = list(_df_fat_total.columns).index('TotalProduto')
                    ws2.write(_soma_row, _soma_col, _soma, _fmt_bold)

                return output.getvalue()

            _nome_arquivo_fat = f"{_fat_vend.upper().replace(' ', '_')}.xlsx" if _fat_vend != 'Todos' else "FATURADO_GERAL.xlsx"

            st.download_button(
                "📥 Exportar Pedidos Faturados",
                _gerar_excel_faturado(_df_fat, _fat_vend),
                _nome_arquivo_fat,
                "application/vnd.ms-excel",
                key="download_fat"
            )

    with tab4_prod:
        st.subheader("📦 Faturamento por Produto no Período")

        # Filtros de período — value=None para não pré-selecionar data
        _fp_col_d1, _fp_col_d2 = st.columns(2)
        with _fp_col_d1:
            _fp_dt_ini = st.date_input("Data inicial", value=None, key="fp_dt_ini_posit")
        with _fp_col_d2:
            _fp_dt_fim = st.date_input("Data final",   value=None, key="fp_dt_fim_posit")

        # Filtros de produto e vendedor
        _col_fp1, _col_fp2, _col_fp3, _col_fp4 = st.columns(4)
        with _col_fp1:
            _fp_vend = st.selectbox(
                "Vendedor", ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist()),
                key="fp_vend_posit"
            )
        with _col_fp2:
            _fp_cod = st.text_input("🔍 Código", placeholder="Ex: 85", key="fp_cod_posit")
        with _col_fp3:
            _fp_busca = st.text_input("🔍 Produto", placeholder="Ex: Atadura", key="fp_busca_posit")
        with _col_fp4:
            _fp_ordem = st.selectbox("Ordenar por",
                ["Faturamento (Maior)", "Quantidade (Maior)", "Nome (A-Z)"],
                key="fp_ordem_posit")

        # Base: NF Venda e NF Dev.Venda (devolução é descontada por produto,
        # mesma regra do "Faturamento Líquido" usado nos demais relatórios do sistema)
        _prod_fat = df[df['TipoMov'].isin(['NF Venda', 'NF Dev.Venda'])].copy()
        _prod_fat['DataEmissao'] = pd.to_datetime(_prod_fat['DataEmissao'], errors='coerce').dt.normalize()

        # Aplicar filtros de data
        if _fp_dt_ini:
            _prod_fat = _prod_fat[_prod_fat['DataEmissao'] >= pd.Timestamp(_fp_dt_ini)]
        if _fp_dt_fim:
            _prod_fat = _prod_fat[_prod_fat['DataEmissao'] <= pd.Timestamp(_fp_dt_fim)]
        if _fp_vend != 'Todos':
            _prod_fat = _prod_fat[_prod_fat['Vendedor'] == _fp_vend]
        if _fp_cod:
            _prod_fat = _prod_fat[_prod_fat['CodigoProduto'].astype(str).str.strip() == str(_fp_cod).strip()]
        if _fp_busca and len(_fp_busca) >= 2:
            _prod_fat = _prod_fat[_prod_fat['NomeProduto'].str.contains(_fp_busca, case=False, na=False)]

        if len(_prod_fat) == 0:
            st.info("ℹ️ Nenhum produto encontrado. Ajuste os filtros acima.")
        else:
            # Fonte de dados: exclusivamente CONSULTA_VENDEDORES.xlsx (df).
            # ValorItem = PrecoUnit * Quantidade (TotalProduto é o total da NOTA inteira,
            # repetido em cada linha de produto — somá-lo direto infla o faturamento
            # sempre que a nota tem mais de um produto). Mesmo padrão já usado em
            # Performance de Vendedores > Resultado por Produto.
            # Sinal negativo para NF Dev.Venda, para descontar devoluções do produto
            # (mesma regra de sinal já usada em Valor_Real).
            _sinal = _prod_fat['TipoMov'].apply(lambda t: 1 if t == 'NF Venda' else -1)
            _prod_fat['ValorItem'] = _prod_fat['PrecoUnit'] * _prod_fat['Quantidade'] * _sinal
            _prod_fat['QtdItem'] = _prod_fat['Quantidade'] * _sinal
            _prod_agrup = _prod_fat.groupby(['CodigoProduto', 'NomeProduto']).agg(
                Quantidade=('QtdItem', 'sum'),
                TotalProduto=('ValorItem', 'sum')
            ).reset_index()

            if _fp_ordem == "Faturamento (Maior)":
                _prod_agrup = _prod_agrup.sort_values('TotalProduto', ascending=False)
            elif _fp_ordem == "Quantidade (Maior)":
                _prod_agrup = _prod_agrup.sort_values('Quantidade', ascending=False)
            else:
                _prod_agrup = _prod_agrup.sort_values('NomeProduto')

            # Adicionar Gramatura via lookup da planilha de produtos
            if planilhas_disponiveis.get('produtos_agrupados'):
                try:
                    _fp_gram_df = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
                    if _fp_gram_df is not None:
                        _fp_gram_df.columns = _fp_gram_df.columns.str.upper().str.strip()
                        _fp_kc = next((c for c in _fp_gram_df.columns if any(x in c for x in ['ID_COD','CODIGO','COD'])), None)
                        _fp_gc = next((c for c in _fp_gram_df.columns if 'GRAMATUR' in c), None)
                        if _fp_kc and _fp_gc:
                            def _fp_norm(v):
                                try: return str(int(float(str(v).strip())))
                                except Exception: return str(v).strip()
                            _fp_gram_df['_K'] = _fp_gram_df[_fp_kc].apply(_fp_norm)
                            _fp_gmap = _fp_gram_df.drop_duplicates(subset='_K').set_index('_K')[_fp_gc]
                            _prod_agrup['Gramatura'] = _prod_agrup['CodigoProduto'].apply(_fp_norm).map(_fp_gmap).fillna('')
                except Exception:
                    _prod_agrup['Gramatura'] = ''

            _col_fp_m1, _col_fp_m2, _col_fp_m3 = st.columns(3)
            with _col_fp_m1:
                st.metric("Total Produtos", len(_prod_agrup))
            with _col_fp_m2:
                st.metric("Faturamento Total", f"R$ {formatar_numero_br(_prod_agrup['TotalProduto'].sum(), 2)}")
            with _col_fp_m3:
                st.metric("Qtd Total", f"{formatar_numero_br(_prod_agrup['Quantidade'].sum(), 0)}")

            _prod_display = _prod_agrup.copy()
            _prod_display['TotalProduto'] = _prod_display['TotalProduto'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
            _prod_display['Quantidade']   = _prod_display['Quantidade'].apply(lambda x: f"{formatar_numero_br(x, 0)}")
            _col_order = ['CodigoProduto', 'NomeProduto']
            if 'Gramatura' in _prod_display.columns:
                _col_order.append('Gramatura')
            _col_order += ['Quantidade', 'TotalProduto']
            _prod_display = _prod_display[_col_order].rename(columns={
                'CodigoProduto': 'Código', 'NomeProduto': 'Produto',
                'Gramatura': 'Gramatura', 'Quantidade': 'Qtd',
                'TotalProduto': 'Faturamento'
            })
            st.dataframe(_prod_display, use_container_width=True, height=420, hide_index=True)

            st.download_button(
                "📥 Exportar Faturamento por Produto",
                to_excel(_prod_agrup),
                "faturamento_por_produto.xlsx",
                "application/vnd.ms-excel",
                key="dl_fat_produto_posit"
            )

# ====================== INADIMPLÊNCIA ======================
elif menu == "Inadimplência":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Relatório de Inadimplência</h2>', unsafe_allow_html=True)
    
    # Verificar se a planilha de inadimplência existe
    if not planilhas_disponiveis['inadimplencia']:
        st.error("❌ Planilha de inadimplência não encontrada")
        st.info("💡 Para usar este módulo, adicione no GitHub um arquivo com 'LANCAMENTO A RECEBER' no nome")
        st.info(f"📂 Local: {GITHUB_REPO}/{GITHUB_FOLDER}/")
        st.info("📋 Colunas necessárias: Funcionário, Razão Social, N_Doc, Dt.Vencimento, Vr.Líquido, Conta/Caixa, UF")
        st.stop()
    
    # Carregar dados de inadimplência
    with st.spinner("📥 Carregando dados de inadimplência..."):
        df_inadimplencia = carregar_planilha_github(planilhas_disponiveis['inadimplencia']['url'])
    
    if df_inadimplencia is not None and len(df_inadimplencia) > 0:
        df_inadimplencia = processar_inadimplencia(df_inadimplencia)
        
        
        # ========== FILTROS ==========
        st.subheader("🔍 Filtros")
        col_f1, col_f2 = st.columns(2)
        
        with col_f1:
            vendedores_inad = ['Todos'] + sorted(df_inadimplencia['Vendedor'].dropna().unique().tolist())
            vendedor_inad_filtro = st.selectbox("Vendedor", vendedores_inad, key="vend_inad")
        
        with col_f2:
            estados_inad = ['Todos'] + sorted(df_inadimplencia['Estado'].dropna().unique().tolist())
            estado_inad_filtro = st.selectbox("Estado", estados_inad, key="est_inad")
        
        data_inicial_inad, data_final_inad = renderizar_filtros_locais("inad", "📅 Ajustar Período de Vencimento")
        
        # Aplicar filtros
        df_inad_filtrado = df_inadimplencia.copy()
        
        if vendedor_inad_filtro != 'Todos':
            df_inad_filtrado = df_inad_filtrado[df_inad_filtrado['Vendedor'] == vendedor_inad_filtro]
        if estado_inad_filtro != 'Todos':
            df_inad_filtrado = df_inad_filtrado[df_inad_filtrado['Estado'] == estado_inad_filtro]
        if data_inicial_inad:
            df_inad_filtrado = df_inad_filtrado[df_inad_filtrado['DataVencimento'] >= pd.to_datetime(data_inicial_inad)]
        if data_final_inad:
            df_inad_filtrado = df_inad_filtrado[df_inad_filtrado['DataVencimento'] <= pd.to_datetime(data_final_inad)]
        
        st.markdown("---")
        
        # ========== CARDS DE RESUMO ==========
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_inadimplencia = df_inad_filtrado['ValorLiquido'].sum()
            st.metric("Total em Aberto", f"R$ {formatar_numero_br(total_inadimplencia, 2)}")
        
        with col2:
            qtd_titulos = len(df_inad_filtrado)
            st.metric("Qtd. Títulos", f"{formatar_numero_br(qtd_titulos, 0)}")
        
        with col3:
            clientes_inadimplentes = df_inad_filtrado['Cliente'].nunique()
            st.metric("Clientes Inadimplentes", f"{formatar_numero_br(clientes_inadimplentes, 0)}")
        
        with col4:
            atraso_medio = df_inad_filtrado['DiasAtraso'].mean()
            st.metric("Atraso Médio", f"{atraso_medio:.0f} dias")
        
        st.markdown("---")
        
        # ========== GRÁFICOS — 4 por linha ==========
        col5, col6, col7, col8 = st.columns(4)

        with col5:
            st.markdown("**📊 Por Faixa de Atraso**")
            ordem_faixas = ['A Vencer', '1-30 dias', '31-60 dias', '61-90 dias', 'Acima de 90 dias']
            inad_por_faixa = df_inad_filtrado.groupby('FaixaAtraso')['ValorLiquido'].sum().reset_index()
            inad_por_faixa['FaixaAtraso'] = pd.Categorical(inad_por_faixa['FaixaAtraso'], categories=ordem_faixas, ordered=True)
            inad_por_faixa = inad_por_faixa.sort_values('FaixaAtraso')
            fig_faixa = px.bar(inad_por_faixa, x='FaixaAtraso', y='ValorLiquido',
                labels={'FaixaAtraso': '', 'ValorLiquido': 'R$'},
                color_discrete_sequence=['#1F4788'])
            fig_faixa = aplicar_layout_grafico(fig_faixa, height=280)
            st.plotly_chart(fig_faixa, use_container_width=True)

        with col6:
            st.markdown("**🏦 Por Banco**")
            inad_por_banco = df_inad_filtrado.groupby('Banco')['ValorLiquido'].sum().reset_index()
            inad_por_banco = inad_por_banco.sort_values('ValorLiquido', ascending=False).head(10)
            fig_banco = px.bar(inad_por_banco, x='ValorLiquido', y='Banco', orientation='h',
                labels={'Banco': '', 'ValorLiquido': 'R$'},
                color_discrete_sequence=['#1F4788'])
            fig_banco = aplicar_layout_grafico(fig_banco, height=280)
            st.plotly_chart(fig_banco, use_container_width=True)

        with col7:
            st.markdown("**👤 Top Vendedores**")
            if 'NumeroDoc' not in df_inad_filtrado.columns:
                possiveis_nomes = [col for col in df_inad_filtrado.columns if 'DOC' in col.upper() or 'NUMERO' in col.upper()]
                df_inad_filtrado['NumeroDoc'] = df_inad_filtrado[possiveis_nomes[0]] if possiveis_nomes else 1
            inad_por_vendedor = df_inad_filtrado.groupby('Vendedor').agg(
                {'ValorLiquido': 'sum', 'NumeroDoc': 'count'}).reset_index()
            inad_por_vendedor.columns = ['Vendedor', 'Valor', 'QtdTitulos']
            inad_por_vendedor = inad_por_vendedor.sort_values('Valor', ascending=False).head(10)
            fig_vend_inad = px.bar(inad_por_vendedor, x='Valor', y='Vendedor', orientation='h',
                labels={'Vendedor': '', 'Valor': 'R$'},
                color_discrete_sequence=['#4A7BC8'])
            fig_vend_inad = aplicar_layout_grafico(fig_vend_inad, height=280)
            st.plotly_chart(fig_vend_inad, use_container_width=True)

        with col8:
            st.markdown("**🗺️ Top Estados**")
            inad_por_estado = df_inad_filtrado.groupby('Estado')['ValorLiquido'].sum().reset_index()
            inad_por_estado = inad_por_estado.sort_values('ValorLiquido', ascending=False).head(10)
            fig_est_inad = px.bar(inad_por_estado, x='ValorLiquido', y='Estado', orientation='h',
                labels={'Estado': '', 'ValorLiquido': 'R$'},
                color_discrete_sequence=['#163561'])
            fig_est_inad = aplicar_layout_grafico(fig_est_inad, height=280)
            st.plotly_chart(fig_est_inad, use_container_width=True)
        
        st.markdown("---")
        
        # ========== TABELA DETALHADA ==========
        st.subheader("📋 Detalhamento dos Títulos")
        
        # Preparar dados para exibição
        df_detalhado = df_inad_filtrado[[
            'Vendedor', 'Cliente', 'NumeroDoc', 'DataVencimento', 
            'ValorLiquido', 'DiasAtraso', 'FaixaAtraso', 'Banco', 'Estado'
        ]].copy()
        
        # Formatar data
        df_detalhado['DataVencimento'] = df_detalhado['DataVencimento'].dt.strftime('%d/%m/%Y')
        
        # Formatar valores para exibição
        df_detalhado_display = df_detalhado.copy()
        df_detalhado_display['ValorLiquido'] = df_detalhado_display['ValorLiquido'].apply(
            lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00"
        )
        
        # Renomear colunas
        df_detalhado_display = df_detalhado_display.rename(columns={
            'Vendedor': 'Vendedor',
            'Cliente': 'Cliente',
            'NumeroDoc': 'Nº Documento',
            'DataVencimento': 'Vencimento',
            'ValorLiquido': 'Valor em Aberto',
            'DiasAtraso': 'Dias Atraso',
            'FaixaAtraso': 'Faixa',
            'Banco': 'Banco',
            'Estado': 'UF'
        })
        
        # Ordenar por dias de atraso (maior para menor)
        df_detalhado_display = df_detalhado_display.sort_values('Dias Atraso', ascending=False)
        
        st.dataframe(df_detalhado_display, use_container_width=True, height=400)
        
        # Botão de download
        _nome_inad = (
            f"{vendedor_inad_filtro.upper().replace(' ', '_')}_INADIMPLENCIA.xlsx"
            if vendedor_inad_filtro != 'Todos'
            else "RELATORIO_INADIMPLENCIA.xlsx"
        )
        st.download_button(
            "📥 Exportar Relatório Completo",
            to_excel(df_detalhado),
            _nome_inad,
            "application/vnd.ms-excel",
            key="dl_inad_completo"
        )

# ====================== CLIENTES SEM COMPRA ======================
elif menu == "Clientes sem Compra":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Clientes sem Compra no Período</h2>', unsafe_allow_html=True)

    # ── Filtro de meses sem compra ────────────────────────────────────────
    # Define a janela de positivação: clientes que compraram dentro dessa janela
    # são EXCLUÍDOS do relatório. Os demais (sem compra na janela) aparecem.
    st.markdown("#### 🗓️ Filtro por Tempo sem Compra")
    _opcoes_faixa = (
        ["Selecione..."]
        + [f"{m} mês" if m == 1 else f"{m} meses" for m in range(1, 13)]
        + ["12 a 24 meses"]
    )
    _col_faixa1, _col_faixa2 = st.columns([2, 5])
    with _col_faixa1:
        _faixa_selecionada = st.selectbox(
            "Clientes sem compra há:",
            _opcoes_faixa,
            key="faixa_churn_meses"
        )

    # ── Calcular janela de positivação ────────────────────────────────────
    # A lógica é: "sem compra há N meses" = não fez nenhuma NF Venda
    # nos últimos N meses (janela: de hoje-N_meses até hoje).
    # Para "12 a 24 meses": não comprou nos últimos 12 meses E comprou
    # há pelo menos 12 meses (ou seja, última compra entre 12 e 24 meses atrás).
    _hoje_churn = pd.Timestamp.now().normalize()

    if _faixa_selecionada == "Selecione...":
        # Sem filtro de faixa: usa o período do sidebar global
        _janela_ini = None
        _janela_fim = None
        _label_faixa = "período definido na barra lateral"
    elif _faixa_selecionada == "12 a 24 meses":
        # Excluir quem comprou nos últimos 24 meses E incluir só quem
        # tem última compra entre 12 e 24 meses atrás
        _janela_ini = _hoje_churn - pd.DateOffset(months=24)
        _janela_fim = _hoje_churn
        _ultima_compra_min = _hoje_churn - pd.DateOffset(months=24)
        _ultima_compra_max = _hoje_churn - pd.DateOffset(months=12)
        _label_faixa = "entre 12 e 24 meses sem compra"
    else:
        _n_meses = int(_faixa_selecionada.split()[0])
        # Janela de positivação: últimos N meses (quem comprou aqui é excluído)
        _janela_ini = _hoje_churn - pd.DateOffset(months=_n_meses)
        _janela_fim = _hoje_churn
        _ultima_compra_min = None
        _ultima_compra_max = None
        _label_faixa = f"há {_faixa_selecionada} sem compra"

    st.markdown("---")

    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        vendedor_churn_filtro = st.selectbox(
            "Filtrar por Vendedor",
            ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist()),
            key="vend_churn"
        )
    with col_f2:
        estado_churn_filtro = st.selectbox(
            "Filtrar por Estado",
            ['Todos'] + sorted(df['Estado'].dropna().unique().tolist()),
            key="est_churn"
        )
    with col_f3:
        ordem = st.selectbox(
            "Ordenar por",
            ["Valor Histórico (Maior)", "Valor Histórico (Menor)", "Nome (A-Z)", "Última Compra (Mais Recente)"],
            key="ordem_churn"
        )
    with col_f4:
        busca_cliente_churn = st.text_input(
            "🔍 Buscar Cliente",
            placeholder="Digite o nome...",
            key="busca_churn"
        )

    # ── Base: somente NF Venda no histórico completo ──────────────────────
    _df_nf = df[df['TipoMov'] == 'NF Venda'].copy()
    _df_nf['DataEmissao'] = pd.to_datetime(_df_nf['DataEmissao'], errors='coerce')

    # ── Definir quem POSITIVOU na janela selecionada ──────────────────────
    if _faixa_selecionada == "Selecione...":
        # Fallback: usar o período do sidebar global
        if data_inicial and data_final:
            _label_periodo = f"{data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}"
            _df_janela = _df_nf[
                (_df_nf['DataEmissao'] >= pd.to_datetime(data_inicial)) &
                (_df_nf['DataEmissao'] <= pd.to_datetime(data_final))
            ]
        elif data_inicial:
            _label_periodo = f"A partir de {data_inicial.strftime('%d/%m/%Y')}"
            _df_janela = _df_nf[_df_nf['DataEmissao'] >= pd.to_datetime(data_inicial)]
        elif data_final:
            _label_periodo = f"Até {data_final.strftime('%d/%m/%Y')}"
            _df_janela = _df_nf[_df_nf['DataEmissao'] <= pd.to_datetime(data_final)]
        else:
            _mes_now = _hoje_churn.month
            _ano_now = _hoje_churn.year
            _label_periodo = f"Mês vigente ({_mes_now:02d}/{_ano_now})"
            _df_janela = _df_nf[
                (_df_nf['DataEmissao'].dt.month == _mes_now) &
                (_df_nf['DataEmissao'].dt.year == _ano_now)
            ]
    else:
        _label_periodo = f"{_janela_ini.strftime('%d/%m/%Y')} a {_janela_fim.strftime('%d/%m/%Y')}"
        _df_janela = _df_nf[
            (_df_nf['DataEmissao'] >= _janela_ini) &
            (_df_nf['DataEmissao'] <= _janela_fim)
        ]

    # CPFs que positivaram (compraram) na janela — serão excluídos do relatório
    _cpfs_positivaram = set(_df_janela['CPF_CNPJ'].unique())

    st.info(f"📅 Janela de positivação: **{_label_periodo}** — clientes que NÃO compraram neste período · Faixa: **{_label_faixa}**")

    # ── Montar base cadastral de todos os clientes com histórico de NF Venda ──
    # Dados cadastrais: último registro de cada CPF
    _cadastro = (
        _df_nf.sort_values('DataEmissao')
        .groupby('CPF_CNPJ')
        .last()
        .reset_index()[['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado']]
    )

    # Vendedor do cliente: usa TODOS os vínculos históricos (não apenas o principal).
    # Assim um cliente do Mario que eventualmente teve NF de outro vendedor
    # ainda aparece na carteira do Mario.
    _vendedor_principal = (
        _df_nf.groupby(['CPF_CNPJ', 'Vendedor'])
        .size()
        .reset_index(name='_cnt')
        .sort_values('_cnt', ascending=False)
        .groupby('CPF_CNPJ')
        .first()
        .reset_index()[['CPF_CNPJ', 'Vendedor']]
    )

    # Última compra de cada CPF no histórico completo
    _ultima_compra_hist = (
        _df_nf.groupby('CPF_CNPJ')['DataEmissao']
        .max()
        .reset_index()
        .rename(columns={'DataEmissao': 'UltimaCompra'})
    )

    # Valor histórico total
    _valor_historico = (
        _df_nf.groupby('CPF_CNPJ')['TotalProduto']
        .sum()
        .reset_index()
        .rename(columns={'TotalProduto': 'ValorHistorico'})
    )

    # Montar todos_clientes
    todos_clientes = (
        _cadastro
        .merge(_vendedor_principal,  on='CPF_CNPJ', how='left')
        .merge(_ultima_compra_hist,  on='CPF_CNPJ', how='left')
        .merge(_valor_historico,     on='CPF_CNPJ', how='left')
    )
    todos_clientes['ValorHistorico'] = todos_clientes['ValorHistorico'].fillna(0)
    todos_clientes['UltimaCompra']   = pd.to_datetime(todos_clientes['UltimaCompra'], errors='coerce')

    # ── REGRA PRINCIPAL: excluir quem positivou na janela ────────────────
    clientes_sem_compra = todos_clientes[
        ~todos_clientes['CPF_CNPJ'].isin(_cpfs_positivaram)
    ].copy()

    # ── Filtro adicional de faixa "12 a 24 meses" ────────────────────────
    # Para essa faixa específica também filtramos pela última compra histórica
    if _faixa_selecionada == "12 a 24 meses":
        clientes_sem_compra = clientes_sem_compra[
            (clientes_sem_compra['UltimaCompra'] >= _ultima_compra_min) &
            (clientes_sem_compra['UltimaCompra'] <= _ultima_compra_max)
        ]

    # ── Filtro por vendedor ───────────────────────────────────────────────
    if vendedor_churn_filtro != 'Todos':
        # Todos os CPFs que já tiveram NF Venda com esse vendedor
        _cpfs_do_vendedor = set(_df_nf[_df_nf['Vendedor'] == vendedor_churn_filtro]['CPF_CNPJ'].unique())
        clientes_sem_compra = clientes_sem_compra[
            clientes_sem_compra['CPF_CNPJ'].isin(_cpfs_do_vendedor)
        ]

    if estado_churn_filtro != 'Todos':
        clientes_sem_compra = clientes_sem_compra[
            clientes_sem_compra['Estado'] == estado_churn_filtro
        ]

    if busca_cliente_churn and len(busca_cliente_churn) >= 2:
        clientes_sem_compra = clientes_sem_compra[
            clientes_sem_compra['RazaoSocial'].str.contains(busca_cliente_churn, case=False, na=False)
        ]

    # ── Ordenação ─────────────────────────────────────────────────────────
    if ordem == "Valor Histórico (Maior)":
        clientes_sem_compra = clientes_sem_compra.sort_values('ValorHistorico', ascending=False)
    elif ordem == "Valor Histórico (Menor)":
        clientes_sem_compra = clientes_sem_compra.sort_values('ValorHistorico', ascending=True)
    elif ordem == "Nome (A-Z)":
        clientes_sem_compra = clientes_sem_compra.sort_values('RazaoSocial')
    elif ordem == "Última Compra (Mais Recente)":
        clientes_sem_compra = clientes_sem_compra.sort_values('UltimaCompra', ascending=False)

    # ── Métricas ──────────────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total de Clientes sem Compra", len(clientes_sem_compra))
    with col2:
        st.metric("Valor Potencial Perdido", f"R$ {formatar_numero_br(clientes_sem_compra['ValorHistorico'].sum(), 2)}")
    with col3:
        ticket_medio_churn = clientes_sem_compra['ValorHistorico'].mean() if len(clientes_sem_compra) > 0 else 0
        st.metric("Ticket Médio Histórico", f"R$ {formatar_numero_br(ticket_medio_churn, 2)}")

    if len(clientes_sem_compra) > 0:
        top_churn = clientes_sem_compra.head(15)
        fig_churn = px.bar(
            top_churn,
            x='ValorHistorico',
            y='RazaoSocial',
            orientation='h',
            labels={'RazaoSocial': 'Cliente', 'ValorHistorico': 'Valor Histórico (R$)'},
            color='ValorHistorico',
            color_discrete_sequence=['#1F4788'],
            title='Top 15 Clientes sem Compra por Valor Histórico'
        )
        fig_churn = aplicar_layout_grafico(fig_churn)
        st.plotly_chart(fig_churn, use_container_width=True)

    # ── Tabela de exibição ────────────────────────────────────────────────
    _display_cols = ['RazaoSocial', 'CPF_CNPJ', 'Vendedor', 'Cidade', 'Estado', 'ValorHistorico', 'UltimaCompra']
    clientes_sem_compra_display = clientes_sem_compra[_display_cols].copy()
    clientes_sem_compra_display = formatar_dataframe_moeda(clientes_sem_compra_display, ['ValorHistorico'])
    clientes_sem_compra_display['UltimaCompra'] = pd.to_datetime(
        clientes_sem_compra_display['UltimaCompra']
    ).dt.strftime('%d/%m/%Y')

    clientes_sem_compra_display = clientes_sem_compra_display.rename(columns={
        'RazaoSocial':   'Razão Social',
        'CPF_CNPJ':      'CPF/CNPJ',
        'Vendedor':      'Vendedor',
        'Cidade':        'Cidade',
        'Estado':        'Estado',
        'ValorHistorico':'Valor Histórico',
        'UltimaCompra':  'Última Compra',
    })

    st.dataframe(clientes_sem_compra_display, use_container_width=True, height=400)

    st.download_button(
        "📥 Exportar Clientes sem Compra",
        to_excel(clientes_sem_compra[_display_cols]),
        "clientes_sem_compra.xlsx",
        "application/vnd.ms-excel",
        key="dl_churn_excel"
    )

# ====================== HISTÓRICO ======================
elif menu == "Histórico":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Histórico de Vendas</h2>', unsafe_allow_html=True)
    
    tab1, tab2, tab3, tab4 = st.tabs(["👤 Por Cliente", "🧑‍💼 Por Vendedor", "📝 Pedidos", "📦 Por Produto"])
    
    # ========== ABA: POR CLIENTE ==========
    with tab1:
        st.subheader("Histórico de Vendas por Cliente")
        
        # Buscar cliente por CPF/CNPJ ou Nome
        col_busca1, col_busca2 = st.columns(2)
        
        with col_busca1:
            busca_tipo = st.radio("Buscar por:", ["Nome", "CPF/CNPJ"], horizontal=True, key="busca_tipo_cliente")
        
        with col_busca2:
            if busca_tipo == "Nome":
                busca_texto = st.text_input("Digite o nome do cliente", placeholder="Ex: Nome da Empresa", key="busca_nome_cliente")
            else:
                busca_texto = st.text_input("Digite o CPF/CNPJ", placeholder="Ex: 12345678901234", key="busca_cpf_cliente")
        
        cliente_selecionado = None
        cpf_cnpj = None
        
        if busca_texto and len(busca_texto) >= 3:
            if busca_tipo == "Nome":
                clientes_filtrados = df[df['RazaoSocial'].str.contains(busca_texto, case=False, na=False)][['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado']].drop_duplicates()
            else:
                clientes_filtrados = df[df['CPF_CNPJ'].str.contains(busca_texto, case=False, na=False)][['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado']].drop_duplicates()
            
            if len(clientes_filtrados) > 0:
                clientes_filtrados['Display'] = clientes_filtrados['RazaoSocial'] + " - " + clientes_filtrados['CPF_CNPJ'] + " (" + clientes_filtrados['Cidade'] + "/" + clientes_filtrados['Estado'] + ")"
                
                cliente_selecionado = st.selectbox(
                    f"📋 Clientes encontrados ({len(clientes_filtrados)}):",
                    options=clientes_filtrados['Display'].tolist(),
                    key="cliente_hist"
                )
                
                if cliente_selecionado:
                    cpf_cnpj = cliente_selecionado.split(' - ')[1].split(' (')[0]
            else:
                st.warning("❌ Nenhum cliente encontrado com esse critério")
        
        if cpf_cnpj:
            historico = df[df['CPF_CNPJ'] == cpf_cnpj].sort_values('DataEmissao', ascending=False).copy()
            # Gramatura: buscar na planilha produtos_agrupados pela coluna GRAMATURA/GRAMAT pelo ID_COD
            if planilhas_disponiveis.get('produtos_agrupados'):
                _hg_plan = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
                if _hg_plan is not None:
                    _hg_plan.columns = _hg_plan.columns.str.upper().str.strip()
                    _hg_gcol = next((c for c in _hg_plan.columns if c in ('GRAMATURA','GRAMAT')), None)
                    if _hg_gcol and 'ID_COD' in _hg_plan.columns:
                        _hg_map = (
                            _hg_plan[['ID_COD', _hg_gcol]]
                            .drop_duplicates(subset='ID_COD')
                            .assign(ID_COD=lambda d: d['ID_COD'].apply(
                                lambda v: str(int(float(str(v)))) if str(v).replace('.','',1).isdigit() else str(v)
                            ))
                            .set_index('ID_COD')[_hg_gcol]
                        )
                        historico['Gramatura'] = historico['CodigoProduto'].apply(
                            lambda v: str(int(float(str(v)))) if str(v).replace('.','',1).isdigit() else str(v)
                        ).map(_hg_map).fillna('')
            
            if len(historico) > 0:
                cliente_info = historico.iloc[0]
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Cliente", cliente_info['RazaoSocial'])
                with col2:
                    st.metric("CPF/CNPJ", cliente_info['CPF_CNPJ'])
                with col3:
                    st.metric("Cidade/Estado", f"{cliente_info['Cidade']}/{cliente_info['Estado']}")
                with col4:
                    st.metric("Total de Registros", len(historico))
                
                st.markdown("---")
                
                vendas_cliente = historico[historico['TipoMov'] == 'NF Venda']
                devolucoes_cliente = historico[historico['TipoMov'] == 'NF Dev.Venda']
                
                col5, col6, col7, col8 = st.columns(4)
                with col5:
                    st.metric("Total Vendas", f"R$ {formatar_numero_br(vendas_cliente['TotalProduto'].sum(), 2)}")
                with col6:
                    st.metric("Total Devoluções", f"R$ {formatar_numero_br(devolucoes_cliente['TotalProduto'].sum(), 2)}")
                with col7:
                    st.metric("Qtd Notas Vendas", len(vendas_cliente['Numero_NF'].unique()))
                with col8:
                    st.metric("Qtd Notas Devoluções", len(devolucoes_cliente['Numero_NF'].unique()))
                
                vendas_tempo_cliente = vendas_cliente.groupby('MesAno')['TotalProduto'].sum().reset_index()
                vendas_tempo_cliente = vendas_tempo_cliente.sort_values('MesAno')
                
                if len(vendas_tempo_cliente) > 0:
                    fig_hist = px.line(
                        vendas_tempo_cliente,
                        x='MesAno',
                        y='TotalProduto',
                        labels={'MesAno': 'Período', 'TotalProduto': 'Valor (R$)'},
                        title='Evolução de Compras'
                    )
                    fig_hist.update_traces(line_color='#28A745', line_width=3, mode='lines+markers', marker=dict(size=6, color='#28A745'))
                    fig_hist = aplicar_layout_grafico(fig_hist)
                    st.plotly_chart(fig_hist, use_container_width=True)
                
                st.markdown("---")
                
                st.subheader("📋 Detalhamento de Produtos")
                
                # Verificar se PrazoHistorico e Comissao existem no dataframe
                colunas_display = ['DataEmissao', 'TipoMov', 'Numero_NF', 'CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit', 'TotalProduto']
                if 'Gramatura' in historico.columns:
                    colunas_display.insert(colunas_display.index('NomeProduto') + 1, 'Gramatura')
                if 'PrazoHistorico' in historico.columns:
                    colunas_display.append('PrazoHistorico')
                if 'Comissao' in historico.columns:
                    colunas_display.append('Comissao')
                
                historico_display = historico[colunas_display].copy()
                historico_display['DataEmissao'] = historico_display['DataEmissao'].dt.strftime('%d/%m/%Y')
                
                # Formatar valores monetários
                historico_display['PrecoUnit'] = historico_display['PrecoUnit'].apply(lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00")
                historico_display['TotalProduto'] = historico_display['TotalProduto'].apply(lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00")
                
                # Renomear colunas
                colunas_rename = {
                    'DataEmissao': 'Data',
                    'TipoMov': 'Tipo',
                    'Gramatura': 'Gramatura',
                    'Numero_NF': 'Nota Fiscal',
                    'CodigoProduto': 'Código',
                    'NomeProduto': 'Produto',
                    'Quantidade': 'Qtd',
                    'PrecoUnit': 'Preço Unit.',
                    'TotalProduto': 'Total'
                }
                if 'PrazoHistorico' in historico_display.columns:
                    colunas_rename['PrazoHistorico'] = 'Prazo (dias)'
                if 'Comissao' in historico_display.columns:
                    colunas_rename['Comissao'] = 'Comissão%'
                
                historico_display = historico_display.rename(columns=colunas_rename)
                
                st.dataframe(historico_display, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Exportar Histórico Excel",
                    to_excel(historico),
                    f"historico_{cpf_cnpj}.xlsx",
                    "application/vnd.ms-excel",
                    key="dl_hist_excel"
                )

                st.markdown("---")
                st.markdown("""
                <div style="background:#F0F4FF;border:1px solid #C5D5F0;border-radius:10px;
                            padding:14px 18px;margin-bottom:8px;">
                    <div style="font-size:0.88rem;font-weight:700;color:#4A7BC8;margin-bottom:4px;">
                        Gerar Proposta Comercial PDF
                    </div>
                    <div style="font-size:0.78rem;color:#6C757D;">
                        Exporta os produtos do histórico do cliente em formato de proposta
                        comercial com cabeçalho Medtextil, dados do cliente e tabela de itens.
                    </div>
                </div>
                """, unsafe_allow_html=True)

                _col_pdf1, _col_pdf2 = st.columns([2, 1])
                with _col_pdf1:
                    _vendas_resumo = {
                        'Total de Vendas':   f"R$ {formatar_numero_br(vendas_cliente['TotalProduto'].sum(), 2)}",
                        'Total Devoluções':  f"R$ {formatar_numero_br(devolucoes_cliente['TotalProduto'].sum(), 2)}",
                        'Notas de Venda':    str(len(vendas_cliente['Numero_NF'].unique())),
                        'Clientes (CNPJ)':   cpf_cnpj,
                    }
                    _cliente_info_dict = {
                        'RazaoSocial': cliente_info.get('RazaoSocial',''),
                        'CPF_CNPJ':    cpf_cnpj,
                        'Cidade':      cliente_info.get('Cidade',''),
                        'Estado':      cliente_info.get('Estado',''),
                        'Vendedor':    cliente_info.get('Vendedor',''),
                    }
                    if st.button("Gerar Proposta PDF", key="btn_gerar_proposta",
                                 use_container_width=True, type="primary"):
                        with st.spinner("Gerando proposta..."):
                            try:
                                _pdf_bytes = gerar_proposta_pdf_historico(
                                    _cliente_info_dict, historico, _vendas_resumo
                                )
                                st.session_state['proposta_pdf_bytes'] = _pdf_bytes
                                st.session_state['proposta_pdf_nome'] = (
                                    f"Proposta_{razao_curta}_{hoje_str}.pdf"
                                    if 'razao_curta' in dir() else
                                    f"Proposta_{cpf_cnpj}.pdf"
                                )
                                st.success("Proposta gerada! Clique em Download abaixo.")
                            except Exception as _e:
                                st.error(f"Erro ao gerar proposta: {_e}")

                with _col_pdf2:
                    if st.session_state.get('proposta_pdf_bytes'):
                        import datetime as _dt
                        _nome_pdf = (
                            f"Proposta_Medtextil_{cliente_info.get('RazaoSocial','cliente')[:20].replace(' ','_')}"
                            f"_{_dt.date.today().strftime('%Y%m%d')}.pdf"
                        )
                        st.download_button(
                            "Download PDF",
                            data=st.session_state['proposta_pdf_bytes'],
                            file_name=_nome_pdf,
                            mime="application/pdf",
                            key="dl_proposta_pdf",
                            use_container_width=True
                        )
            else:
                st.warning("Nenhum registro encontrado para este cliente")
        else:
            st.info("👆 Digite pelo menos 3 caracteres para buscar um cliente")
    
    # ========== ABA: POR VENDEDOR ==========
    with tab2:
        st.subheader("Histórico de Vendas por Vendedor")
        
        # Filtros
        col_f1, = st.columns(1)
        
        with col_f1:
            vendedores_hist = ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist())
            vendedor_hist_filtro = st.selectbox("Vendedor", vendedores_hist, key="vend_hist")
        
        data_inicial_hist, data_final_hist = renderizar_filtros_locais("hist_vend", "📅 Ajustar Período")
        
        # Aplicar filtros
        df_hist_vendedor = df[df['TipoMov'] == 'NF Venda'].copy()
        
        if vendedor_hist_filtro != 'Todos':
            df_hist_vendedor = df_hist_vendedor[df_hist_vendedor['Vendedor'] == vendedor_hist_filtro]
        if data_inicial_hist:
            df_hist_vendedor = df_hist_vendedor[df_hist_vendedor['DataEmissao'] >= pd.to_datetime(data_inicial_hist)]
        if data_final_hist:
            df_hist_vendedor = df_hist_vendedor[df_hist_vendedor['DataEmissao'] <= pd.to_datetime(data_final_hist)]
        
        if len(df_hist_vendedor) > 0:
            # Obter notas únicas e agrupar por NF para somar o valor total
            notas_vendedor = obter_notas_unicas(df_hist_vendedor)
            
            # Preparar dados para exibição
            colunas_vendedor = ['DataEmissao', 'RazaoSocial', 'Numero_NF', 'TotalProduto', 'Vendedor']
            if 'PrazoHistorico' in notas_vendedor.columns:
                colunas_vendedor.append('PrazoHistorico')
            
            historico_vendedor = notas_vendedor[colunas_vendedor].copy()
            historico_vendedor = historico_vendedor.sort_values('DataEmissao', ascending=False)
            
            # Métricas resumidas
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total de Vendas", f"R$ {formatar_numero_br(historico_vendedor['TotalProduto'].sum(), 2)}")
            with col2:
                st.metric("Quantidade de Notas", len(historico_vendedor))
            with col3:
                st.metric("Clientes Atendidos", historico_vendedor['RazaoSocial'].nunique())
            with col4:
                ticket_medio_vend = historico_vendedor['TotalProduto'].mean() if len(historico_vendedor) > 0 else 0
                st.metric("Ticket Médio", f"R$ {formatar_numero_br(ticket_medio_vend, 2)}")
            
            st.markdown("---")
            
            # Formatar para exibição
            historico_vendedor_display = historico_vendedor.copy()
            historico_vendedor_display['DataEmissao'] = historico_vendedor_display['DataEmissao'].dt.strftime('%d/%m/%Y')
            historico_vendedor_display['TotalProduto'] = historico_vendedor_display['TotalProduto'].apply(
                lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00"
            )
            
            # Renomear colunas
            colunas_rename_vendedor = {
                'DataEmissao': 'Data',
                'RazaoSocial': 'Cliente',
                'Numero_NF': 'Nota Fiscal',
                'TotalProduto': 'Valor Total',
                'Vendedor': 'Vendedor'
            }
            if 'PrazoHistorico' in historico_vendedor_display.columns:
                colunas_rename_vendedor['PrazoHistorico'] = 'Prazo (dias)'
            
            historico_vendedor_display = historico_vendedor_display.rename(columns=colunas_rename_vendedor)
            
            st.dataframe(historico_vendedor_display, use_container_width=True, height=400)
            
            # Botão de download
            st.download_button(
                "📥 Exportar Histórico de Vendas",
                to_excel(historico_vendedor),
                f"historico_vendedor_{vendedor_hist_filtro if vendedor_hist_filtro != 'Todos' else 'todos'}.xlsx",
                "application/vnd.ms-excel",
                key="download_hist_vend"
            )
        else:
            st.info("Nenhuma venda encontrada com os filtros selecionados")

    
    # ========== ABA: PEDIDOS ==========
    with tab3:
        st.subheader("📝 Gerar Pedido/Proposta")
        
        # Inicializar session_state para os itens do pedido
        if 'itens_pedido' not in st.session_state:
            st.session_state.itens_pedido = []
        
        # Carregar dados de produtos se disponível
        df_produtos_pedido = None
        if planilhas_disponiveis.get('produtos_agrupados'):
            with st.spinner("📥 Carregando catálogo de produtos..."):
                df_produtos_pedido = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
                if df_produtos_pedido is not None:
                    df_produtos_pedido.columns = df_produtos_pedido.columns.str.upper()
        
        # SEÇÃO 1: DADOS DO CLIENTE
        st.markdown("### 👤 Informações do Cliente")
        
        col_cli1, col_cli2 = st.columns(2)
        
        with col_cli1:
            # Buscar cliente
            clientes_lista = sorted(df['RazaoSocial'].dropna().unique().tolist())
            cliente_selecionado = st.selectbox("Selecione o Cliente", [''] + clientes_lista, key="cliente_pedido")
        
        # Buscar dados do cliente
        dados_cliente = {}
        if cliente_selecionado:
            df_cliente = df[df['RazaoSocial'] == cliente_selecionado].iloc[0]
            dados_cliente = {
                'razao_social': df_cliente.get('RazaoSocial', ''),
                'cpf_cnpj': df_cliente.get('CPF_CNPJ', ''),
                'cidade': df_cliente.get('Cidade', ''),
                'estado': df_cliente.get('Estado', ''),
                'vendedor': df_cliente.get('Vendedor', '')
            }
        
        with col_cli2:
            representante = st.text_input("Representante", value=dados_cliente.get('vendedor', ''), key="representante_pedido")
        
        col_cli3, col_cli4, col_cli5 = st.columns(3)
        
        with col_cli3:
            nome_fantasia = st.text_input("Nome Fantasia", value=dados_cliente.get('razao_social', ''), key="fantasia_pedido")
        
        with col_cli4:
            cnpj_pedido = st.text_input("CNPJ", value=dados_cliente.get('cpf_cnpj', ''), key="cnpj_pedido")
        
        with col_cli5:
            insc_estadual = st.text_input("Inscrição Estadual", key="ie_pedido")
        
        col_cli6, col_cli7 = st.columns(2)
        
        with col_cli6:
            telefone_pedido = st.text_input("Telefone", key="tel_pedido")
        
        with col_cli7:
            email_pedido = st.text_input("Email NF-e", key="email_pedido")
        
        endereco_pedido = st.text_input("Endereço", value=f"{dados_cliente.get('cidade', '')}/{dados_cliente.get('estado', '')}" if dados_cliente else "", key="end_pedido")
        
        obs_cliente = st.text_area("Observação (Cliente)", key="obs_cli_pedido", height=80)
        
        st.markdown("---")
        
        # SEÇÃO 2: DADOS DO PEDIDO
        st.markdown("### 📋 Informações do Pedido")
        
        col_ped1, col_ped2, col_ped3, col_ped4 = st.columns(4)
        
        with col_ped1:
            num_pedido = st.text_input("Nº do Pedido", key="num_pedido")
        
        with col_ped2:
            tabela_preco = st.text_input("Tabela de Preço", key="tab_preco")
        
        with col_ped3:
            tipo_frete = st.selectbox("Tipo de Frete", ["CIF", "FOB"], key="tipo_frete")
        
        with col_ped4:
            data_venda = st.date_input("Data da Venda", value=pd.Timestamp.now(), key="data_venda")
        
        condicoes_pagto = st.text_input("Condições de Pagamento", key="cond_pagto")
        
        st.markdown("---")
        
        # SEÇÃO 3: ADICIONAR PRODUTOS
        st.markdown("### 🛒 Adicionar Produtos ao Pedido")
        
        col_prod1, col_prod2, col_prod3, col_prod4 = st.columns([2, 1, 1, 1])
        
        with col_prod1:
            # Buscar por código ou descrição
            tipo_busca_prod = st.radio("Buscar por:", ["Código", "Descrição"], horizontal=True, key="tipo_busca_prod")
            
            if tipo_busca_prod == "Código":
                if df_produtos_pedido is not None:
                    codigos = [''] + sorted(df_produtos_pedido['ID_COD'].dropna().astype(str).unique().tolist())
                    codigo_selecionado = st.selectbox("Código do Produto", codigos, key="cod_prod_pedido")
                else:
                    codigo_selecionado = st.text_input("Código do Produto", key="cod_prod_pedido_txt")
            else:
                busca_desc = st.text_input("Descrição do Produto", key="desc_prod_pedido")
                codigo_selecionado = None
        
        # Buscar informações do produto
        produto_info = {}
        if df_produtos_pedido is not None and codigo_selecionado:
            prod = df_produtos_pedido[df_produtos_pedido['ID_COD'].astype(str) == str(codigo_selecionado)]
            if len(prod) > 0:
                prod = prod.iloc[0]
                # Montar descrição completa
                descricao_completa = f"{prod.get('GRUPO', '')} {prod.get('DESCRIÇÃO', '') or prod.get('DESCRICAO', '')} {prod.get('LINHA', '') or prod.get('LINHAS', '')}".strip()
                
                produto_info = {
                    'codigo': str(prod.get('ID_COD', '')),
                    'descricao': descricao_completa,
                    'peso': prod.get('GRAMATURA', ''),
                    'cx_embarque': prod.get('CX_EMB', ''),
                    'preco_ref': prod.get('PRECO', 0)
                }
                
                # Buscar último preço que o cliente comprou
                if cliente_selecionado:
                    hist_cliente = df[(df['RazaoSocial'] == cliente_selecionado) & 
                                     (df['CodigoProduto'].astype(str) == str(codigo_selecionado))]
                    if len(hist_cliente) > 0:
                        hist_cliente = hist_cliente.sort_values('DataEmissao', ascending=False)
                        produto_info['preco_sugerido'] = hist_cliente.iloc[0]['PrecoUnit']
                        produto_info['preco_historico'] = hist_cliente.iloc[0]['PrecoUnit']  # Para mostrar na tabela
                    else:
                        produto_info['preco_sugerido'] = prod.get('PRECO', 0)
                        produto_info['preco_historico'] = prod.get('PRECO', 0)
                else:
                    produto_info['preco_sugerido'] = prod.get('PRECO', 0)
                    produto_info['preco_historico'] = prod.get('PRECO', 0)
        
        with col_prod2:
            qtde_item = st.number_input("Quantidade", min_value=0, value=0, key="qtde_item_pedido")
        
        with col_prod3:
            valor_item = st.number_input("Valor Unit.", min_value=0.0, value=float(produto_info.get('preco_sugerido', 0)), format="%.2f", key="valor_item_pedido")
        
        with col_prod4:
            st.write("")
            st.write("")
            if st.button("➕ Adicionar Item", use_container_width=True, key="add_item_pedido"):
                if produto_info and qtde_item > 0:
                    # Calcular comissão
                    comissao = calcular_comissao(valor_item, produto_info.get('preco_ref', 0))
                    
                    item = {
                        'codigo': produto_info['codigo'],
                        'descricao': produto_info['descricao'],
                        'peso': produto_info.get('peso', ''),
                        'cx_embarque': produto_info.get('cx_embarque', ''),
                        'quantidade': qtde_item,
                        'valor_unit': valor_item,
                        'preco_historico': produto_info.get('preco_historico', 0),
                        'total': qtde_item * valor_item,
                        'comissao': comissao
                    }
                    st.session_state.itens_pedido.append(item)
                    st.success(f"✅ Item adicionado: {produto_info['descricao']}")
                    st.rerun()
        
        # PREVIEW EM TEMPO REAL - Mostrar antes de adicionar
        if produto_info and qtde_item > 0 and valor_item > 0:
            st.markdown("---")
            st.markdown("### 👁️ Preview do Item")
            
            # Calcular valores preview
            total_preview = qtde_item * valor_item
            comissao_preview = calcular_comissao(valor_item, produto_info.get('preco_ref', 0))
            preco_hist_preview = produto_info.get('preco_historico', 0)
            
            # Mostrar em tabela estilizada
            preview_data = {
                'Código': [produto_info['codigo']],
                'Produto': [produto_info['descricao'][:50]],
                'Peso': [produto_info.get('peso', '')],
                'Cx Embarque': [produto_info.get('cx_embarque', '')],
                'Qtde': [f"{formatar_numero_br(qtde_item, 0)}"],
                'Preço Histórico': [f"R$ {formatar_numero_br(preco_hist_preview, 2)}"],
                'Valor Unit.': [f"R$ {formatar_numero_br(valor_item, 2)}"],
                'Total': [f"R$ {formatar_numero_br(total_preview, 2)}"],
                'Comissão%': [comissao_preview]
            }
            
            df_preview = pd.DataFrame(preview_data)
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            
            # Comparação com preço histórico
            if preco_hist_preview > 0:
                variacao = ((valor_item - preco_hist_preview) / preco_hist_preview) * 100
                if variacao > 0:
                    st.info(f"📈 Valor {variacao:.1f}% **acima** do histórico (R$ {formatar_numero_br(preco_hist_preview, 2)})")
                elif variacao < 0:
                    st.warning(f"📉 Valor {abs(variacao):.1f}% **abaixo** do histórico (R$ {formatar_numero_br(preco_hist_preview, 2)})")
                else:
                    st.success(f"✅ Valor **igual** ao histórico (R$ {formatar_numero_br(preco_hist_preview, 2)})")
            
            st.markdown("---")
        
        # Mostrar produtos adicionados
        if st.session_state.itens_pedido:
            st.markdown("---")
            st.markdown("### 📦 Itens do Pedido")
            
            # Criar DataFrame dos itens
            df_itens = pd.DataFrame(st.session_state.itens_pedido)
            
            # Formatar para exibição
            df_itens_display = df_itens.copy()
            df_itens_display['preco_historico'] = df_itens_display['preco_historico'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
            df_itens_display['valor_unit'] = df_itens_display['valor_unit'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
            df_itens_display['total'] = df_itens_display['total'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
            
            df_itens_display = df_itens_display.rename(columns={
                'codigo': 'COD.',
                'descricao': 'PRODUTO',
                'peso': 'PESO',
                'cx_embarque': 'CAIXA EMBARQUE',
                'quantidade': 'QTDE',
                'preco_historico': 'PREÇO HISTÓRICO',
                'valor_unit': 'VALOR',
                'total': 'TOTAL',
                'comissao': 'COMISSÃO%'
            })
            
            st.dataframe(df_itens_display, use_container_width=True, height=300)
            
            # Métricas do pedido
            col_met1, col_met2, col_met3 = st.columns(3)
            
            with col_met1:
                total_itens = df_itens['quantidade'].sum()
                st.metric("Qtde Total de Itens", f"{formatar_numero_br(total_itens, 0)}")
            
            with col_met2:
                st.metric("Frete", tipo_frete)
            
            with col_met3:
                total_pedido = df_itens['total'].sum()
                st.metric("Total Final", f"R$ {formatar_numero_br(total_pedido, 2)}")
            
            # Observação final
            obs_pedido = st.text_area("Observação (Pedido)", key="obs_pedido", height=100)
            
            col_btn1, col_btn2 = st.columns(2)
            
            with col_btn1:
                if st.button("🗑️ Limpar Pedido", use_container_width=True, key="limpar_pedido"):
                    st.session_state.itens_pedido = []
                    st.rerun()
            
            with col_btn2:
                if st.button("📄 Gerar PDF do Pedido", use_container_width=True, key="gerar_pdf_pedido", type="primary"):
                    try:
                        # Preparar dados para o PDF
                        dados_cliente_pdf = {
                            'representante': representante,
                            'razao_social': cliente_selecionado,
                            'nome_fantasia': nome_fantasia,
                            'cnpj': cnpj_pedido,
                            'ie': insc_estadual,
                            'telefone': telefone_pedido,
                            'email': email_pedido,
                            'endereco': endereco_pedido,
                            'obs_cliente': obs_cliente
                        }
                        
                        dados_pedido_pdf = {
                            'numero': num_pedido,
                            'tabela_preco': tabela_preco,
                            'tipo_frete': tipo_frete,
                            'data_venda': data_venda.strftime('%d/%m/%Y'),
                            'condicoes_pagto': condicoes_pagto
                        }
                        
                        # Gerar PDF
                        pdf_bytes = gerar_pdf_pedido(dados_cliente_pdf, dados_pedido_pdf, st.session_state.itens_pedido, obs_pedido)
                        
                        # Botão de download
                        st.download_button(
                            label="📥 Baixar PDF do Pedido",
                            data=pdf_bytes,
                            file_name=f"Pedido_{num_pedido or 'SN'}_{cliente_selecionado.replace(' ', '_')}.pdf",
                            mime="application/pdf",
                            key="download_pdf_pedido"
                        )
                        
                        st.success("✅ PDF gerado com sucesso!")
                        
                    except Exception as e:
                        st.error(f"❌ Erro ao gerar PDF: {str(e)}")
                        st.info("💡 Certifique-se de que a biblioteca ReportLab está instalada")
        else:
            st.info("ℹ️ Nenhum item adicionado ao pedido ainda. Use o formulário acima para adicionar produtos.")

    with tab4:
        st.markdown("#### 📦 Vendas por Produto")

        # ── Filtros ───────────────────────────────────────────────────────
        _tp1, _tp2 = st.columns(2)
        with _tp1:
            _tp_di = st.date_input("📅 Data Inicial", value=None,
                                   key="tp_data_ini", format="DD/MM/YYYY")
        with _tp2:
            _tp_df = st.date_input("📅 Data Final", value=None,
                                   key="tp_data_fim", format="DD/MM/YYYY")

        _tp3, _tp4 = st.columns(2)
        with _tp3:
            # Multiselect de produtos
            _prods_disponiveis = sorted(
                df[df['NomeProduto'].notna()]['NomeProduto'].unique().tolist()
            )
            _tp_prods = st.multiselect(
                "🔍 Filtrar por Produto(s)",
                options=_prods_disponiveis,
                placeholder="Todos os produtos",
                key="tp_produtos"
            )
        with _tp4:
            _codigos_disponiveis = sorted(
                df[df['CodigoProduto'].notna()]['CodigoProduto'].astype(str).unique().tolist()
            )
            _tp_cods = st.multiselect(
                "🔍 Filtrar por Código(s)",
                options=_codigos_disponiveis,
                placeholder="Todos os códigos",
                key="tp_codigos"
            )

        # ── Aplicar filtros ───────────────────────────────────────────────
        _df_tp = df[df['TipoMov'] == 'NF Venda'].copy()

        if _tp_di:
            _df_tp = _df_tp[_df_tp['DataEmissao'] >= pd.to_datetime(_tp_di)]
        if _tp_df:
            _df_tp = _df_tp[_df_tp['DataEmissao'] <= pd.to_datetime(_tp_df)]
        if _tp_prods:
            _df_tp = _df_tp[_df_tp['NomeProduto'].isin(_tp_prods)]
        if _tp_cods:
            _df_tp = _df_tp[_df_tp['CodigoProduto'].astype(str).isin(_tp_cods)]

        if len(_df_tp) == 0:
            st.info("Nenhuma venda encontrada com os filtros aplicados.")
        else:
            # ── KPIs ──────────────────────────────────────────────────────
            _m1, _m2, _m3, _m4 = st.columns(4)
            with _m1:
                st.metric("Total Faturado", f"R$ {formatar_numero_br(_df_tp['TotalProduto'].sum(), 2)}")
            with _m2:
                st.metric("Qtd Vendida", f"{formatar_numero_br(_df_tp['Quantidade'].sum(), 0)}")
            with _m3:
                _pm = (_df_tp['TotalProduto'].sum() / _df_tp['Quantidade'].sum()
                       if _df_tp['Quantidade'].sum() > 0 else 0)
                st.metric("Preço Médio", f"R$ {formatar_numero_br(_pm, 2)}")
            with _m4:
                st.metric("Produtos Únicos", f"{formatar_numero_br(_df_tp['CodigoProduto'].nunique(), 0)}")

            st.markdown("---")

            # ── Gráficos ──────────────────────────────────────────────────
            _g1, _g2 = st.columns(2)

            with _g1:
                st.markdown("**Top 10 por Faturamento**")
                _top_fat = (
                    _df_tp.groupby('NomeProduto')['TotalProduto']
                    .sum().reset_index()
                    .sort_values('TotalProduto', ascending=False).head(10)
                )
                _fig1 = px.bar(_top_fat, x='TotalProduto', y='NomeProduto',
                               orientation='h',
                               labels={'NomeProduto': '', 'TotalProduto': 'R$'},
                               color_discrete_sequence=['#1F4788'])
                _fig1 = aplicar_layout_grafico(_fig1, height=320)
                st.plotly_chart(_fig1, use_container_width=True)

            with _g2:
                st.markdown("**Top 10 por Quantidade**")
                _top_qtd = (
                    _df_tp.groupby('NomeProduto')['Quantidade']
                    .sum().reset_index()
                    .sort_values('Quantidade', ascending=False).head(10)
                )
                _fig2 = px.bar(_top_qtd, x='Quantidade', y='NomeProduto',
                               orientation='h',
                               labels={'NomeProduto': '', 'Quantidade': 'Unidades'},
                               color_discrete_sequence=['#2E86AB'])
                _fig2 = aplicar_layout_grafico(_fig2, height=320)
                st.plotly_chart(_fig2, use_container_width=True)

            # ── Evolução temporal ─────────────────────────────────────────
            st.markdown("**Evolução Mensal**")
            _prods_evo = ['Todos'] + sorted(_df_tp['NomeProduto'].dropna().unique().tolist())
            _prod_sel  = st.selectbox("Produto para evolução:", _prods_evo, key="tp_evo_prod")
            _df_evo    = _df_tp if _prod_sel == 'Todos' else _df_tp[_df_tp['NomeProduto'] == _prod_sel]

            if 'MesAno' in _df_evo.columns and len(_df_evo) > 0:
                _evo = _df_evo.groupby('MesAno').agg(
                    Faturamento=('TotalProduto', 'sum'),
                    Quantidade=('Quantidade', 'sum')
                ).reset_index().sort_values('MesAno')

                _e1, _e2 = st.columns(2)
                with _e1:
                    _fe = px.line(_evo, x='MesAno', y='Faturamento',
                                  labels={'MesAno': 'Período', 'Faturamento': 'R$'},
                                  color_discrete_sequence=['#1F4788'])
                    _fe.update_traces(line_width=2, mode='lines+markers', marker=dict(size=5))
                    _fe = aplicar_layout_grafico(_fe, height=240)
                    st.plotly_chart(_fe, use_container_width=True)
                with _e2:
                    _qe = px.line(_evo, x='MesAno', y='Quantidade',
                                  labels={'MesAno': 'Período', 'Quantidade': 'Unidades'},
                                  color_discrete_sequence=['#2E86AB'])
                    _qe.update_traces(line_width=2, mode='lines+markers', marker=dict(size=5))
                    _qe = aplicar_layout_grafico(_qe, height=240)
                    st.plotly_chart(_qe, use_container_width=True)

            # ── Tabela detalhada ──────────────────────────────────────────
            st.markdown("**Detalhamento por Produto**")
            _df_tab = (
                _df_tp.groupby(['CodigoProduto', 'NomeProduto'])
                .agg(QtdVendida=('Quantidade', 'sum'),
                     PrecoMedio=('PrecoUnit', 'mean'),
                     TotalFaturado=('TotalProduto', 'sum'))
                .reset_index()
                .sort_values('TotalFaturado', ascending=False)
            )
            _df_tab['PrecoMedio']    = _df_tab['PrecoMedio'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
            _df_tab['TotalFaturado'] = _df_tab['TotalFaturado'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
            _df_tab['QtdVendida']    = _df_tab['QtdVendida'].apply(lambda x: f"{formatar_numero_br(x, 0)}")
            _df_tab = _df_tab.rename(columns={
                'CodigoProduto': 'Código',
                'NomeProduto':   'Produto',
                'QtdVendida':    'Qtd Vendida',
                'PrecoMedio':    'Preço Médio',
                'TotalFaturado': 'Total (R$)',
            })
            st.dataframe(_df_tab, use_container_width=True, height=350)

            st.download_button(
                "📥 Exportar Vendas por Produto",
                to_excel(_df_tp),
                "historico_por_produto.xlsx",
                "application/vnd.ms-excel",
                key="dl_hist_produto"
            )


# ====================== NOVO PEDIDO (alias de Histórico > tab Pedidos) ======================
elif menu == "__novo_pedido__":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">📝 Novo Pedido & Simulador</h2>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6C757D;font-size:0.88rem;margin-bottom:20px;">Criação de pedidos e simulador de comissão</p>', unsafe_allow_html=True)

    if 'itens_pedido' not in st.session_state:
        st.session_state.itens_pedido = []

    df_produtos_pedido = None
    if planilhas_disponiveis.get('produtos_agrupados'):
        with st.spinner("📥 Carregando catálogo de produtos..."):
            df_produtos_pedido = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
            if df_produtos_pedido is not None:
                df_produtos_pedido.columns = df_produtos_pedido.columns.str.upper()

    st.markdown("### 👤 Informações do Cliente")
    col_cli1, col_cli2 = st.columns(2)
    with col_cli1:
        clientes_lista = sorted(df['RazaoSocial'].dropna().unique().tolist())
        cliente_selecionado = st.selectbox("Selecione o Cliente", [''] + clientes_lista, key="cliente_pedido_np")
    dados_cliente = {}
    if cliente_selecionado:
        df_cliente_row = df[df['RazaoSocial'] == cliente_selecionado].iloc[0]
        dados_cliente = {
            'razao_social': df_cliente_row.get('RazaoSocial', ''),
            'cpf_cnpj':     df_cliente_row.get('CPF_CNPJ', ''),
            'cidade':       df_cliente_row.get('Cidade', ''),
            'estado':       df_cliente_row.get('Estado', ''),
            'vendedor':     df_cliente_row.get('Vendedor', '')
        }
    with col_cli2:
        representante = st.text_input("Representante", value=dados_cliente.get('vendedor', ''), key="representante_pedido_np")
    col_cli3, col_cli4, col_cli5 = st.columns(3)
    with col_cli3:
        nome_fantasia = st.text_input("Nome Fantasia", value=dados_cliente.get('razao_social', ''), key="fantasia_pedido_np")
    with col_cli4:
        cnpj_pedido = st.text_input("CNPJ", value=dados_cliente.get('cpf_cnpj', ''), key="cnpj_pedido_np")
    with col_cli5:
        insc_estadual = st.text_input("Inscrição Estadual", key="ie_pedido_np")
    col_cli6, col_cli7 = st.columns(2)
    with col_cli6:
        telefone_pedido = st.text_input("Telefone", key="tel_pedido_np")
    with col_cli7:
        email_pedido = st.text_input("Email NF-e", key="email_pedido_np")
    _end_default = f"{dados_cliente.get('cidade','')}/{dados_cliente.get('estado','')}" if dados_cliente else ""
    endereco_pedido = st.text_input("Endereço", value=_end_default, key="end_pedido_np")
    obs_cliente = st.text_area("Observação (Cliente)", key="obs_cli_pedido_np", height=80)
    st.markdown("---")

    st.markdown("### 📋 Informações do Pedido")
    col_ped1, col_ped2, col_ped3, col_ped4 = st.columns(4)
    with col_ped1:
        num_pedido = st.text_input("Nº do Pedido", key="num_pedido_np")
    with col_ped2:
        tabela_preco = st.text_input("Tabela de Preço", key="tab_preco_np")
    with col_ped3:
        tipo_frete = st.selectbox("Tipo de Frete", ["CIF", "FOB"], key="tipo_frete_np")
    with col_ped4:
        data_venda = st.date_input("Data da Venda", value=pd.Timestamp.now(), key="data_venda_np")
    condicoes_pagto = st.text_input("Condições de Pagamento", key="cond_pagto_np")
    st.markdown("---")

    st.markdown("### 🛒 Adicionar Produtos ao Pedido")
    col_prod1, col_prod2, col_prod3, col_prod4 = st.columns([2, 1, 1, 1])
    produto_info = {}
    with col_prod1:
        tipo_busca_prod = st.radio("Buscar por:", ["Código", "Descrição"], horizontal=True, key="tipo_busca_prod_np")
        if tipo_busca_prod == "Código":
            if df_produtos_pedido is not None:
                codigos_lista = [''] + sorted(df_produtos_pedido['ID_COD'].dropna().astype(str).unique().tolist())
                codigo_selecionado = st.selectbox("Código do Produto", codigos_lista, key="cod_prod_pedido_np")
            else:
                codigo_selecionado = st.text_input("Código do Produto", key="cod_prod_pedido_np_txt")
        else:
            busca_desc = st.text_input("Descrição do Produto", key="desc_prod_pedido_np")
            codigo_selecionado = None

    if df_produtos_pedido is not None and codigo_selecionado:
        prod_row = df_produtos_pedido[df_produtos_pedido['ID_COD'].astype(str) == str(codigo_selecionado)]
        if len(prod_row) > 0:
            prod_row = prod_row.iloc[0]
            _desc_parts = [str(prod_row.get(c, '') or '') for c in ['GRUPO','DESCRIÇÃO','DESCRICAO','LINHA','LINHAS'] if c in df_produtos_pedido.columns]
            descricao_completa = ' '.join(p.strip() for p in _desc_parts if p.strip())
            produto_info = {
                'codigo':      str(prod_row.get('ID_COD', '')),
                'descricao':   descricao_completa or str(codigo_selecionado),
                'peso':        prod_row.get('GRAMATURA', ''),
                'cx_embarque': prod_row.get('CX_EMB', ''),
                'preco_ref':   prod_row.get('PRECO', 0) or 0
            }
            if cliente_selecionado:
                hist_cli = df[(df['RazaoSocial'] == cliente_selecionado) &
                              (df['CodigoProduto'].astype(str) == str(codigo_selecionado))]
                if len(hist_cli) > 0:
                    hist_cli = hist_cli.sort_values('DataEmissao', ascending=False)
                    produto_info['preco_sugerido'] = hist_cli.iloc[0]['PrecoUnit']
                else:
                    produto_info['preco_sugerido'] = produto_info['preco_ref']
            else:
                produto_info['preco_sugerido'] = produto_info['preco_ref']

    with col_prod2:
        qtde_item = st.number_input("Quantidade", min_value=0, value=0, key="qtde_item_pedido_np")
    with col_prod3:
        valor_item = st.number_input("Valor Unit.", min_value=0.0,
                                     value=float(produto_info.get('preco_sugerido', 0)),
                                     format="%.2f", key="valor_item_pedido_np")
    with col_prod4:
        st.write(""); st.write("")
        if st.button("➕ Adicionar Item", use_container_width=True, key="add_item_pedido_np"):
            if produto_info and qtde_item > 0:
                comissao = calcular_comissao(valor_item, produto_info.get('preco_ref', 0))
                st.session_state.itens_pedido.append({
                    'codigo':          produto_info['codigo'],
                    'descricao':       produto_info['descricao'],
                    'peso':            produto_info.get('peso', ''),
                    'cx_embarque':     produto_info.get('cx_embarque', ''),
                    'quantidade':      qtde_item,
                    'valor_unit':      valor_item,
                    'preco_historico': produto_info.get('preco_sugerido', 0),
                    'total':           qtde_item * valor_item,
                    'comissao':        comissao
                })
                st.success(f"✅ Item adicionado: {produto_info['descricao']}")
                st.rerun()

    # Preview em tempo real
    if produto_info and qtde_item > 0 and valor_item > 0:
        st.markdown("---")
        st.markdown("### 👁️ Preview do Item")
        _total_prev = qtde_item * valor_item
        _preco_hist  = produto_info.get('preco_sugerido', 0)
        _comissao_p  = calcular_comissao(valor_item, produto_info.get('preco_ref', 0))
        st.dataframe(pd.DataFrame([{
            'Código': produto_info['codigo'],
            'Produto': produto_info['descricao'][:50],
            'Peso': produto_info.get('peso', ''),
            'Cx Embarque': produto_info.get('cx_embarque', ''),
            'Qtde': f"{formatar_numero_br(qtde_item, 0)}",
            'Preço Histórico': f"R$ {formatar_numero_br(_preco_hist, 2)}",
            'Valor Unit.': f"R$ {formatar_numero_br(valor_item, 2)}",
            'Total': f"R$ {formatar_numero_br(_total_prev, 2)}",
            'Comissão%': _comissao_p
        }]), use_container_width=True, hide_index=True)
        if _preco_hist > 0:
            _var = ((valor_item - _preco_hist) / _preco_hist) * 100
            if _var > 0:
                st.info(f"📈 Valor {_var:.1f}% **acima** do histórico (R$ {formatar_numero_br(_preco_hist, 2)})")
            elif _var < 0:
                st.warning(f"📉 Valor {abs(_var):.1f}% **abaixo** do histórico (R$ {formatar_numero_br(_preco_hist, 2)})")
            else:
                st.success(f"✅ Valor **igual** ao histórico")
        st.markdown("---")

    # Itens do pedido
    if st.session_state.itens_pedido:
        st.markdown("---")
        st.markdown("### 📦 Itens do Pedido")
        df_itens = pd.DataFrame(st.session_state.itens_pedido)
        df_itens_display = df_itens.copy()
        df_itens_display['preco_historico'] = df_itens_display['preco_historico'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
        df_itens_display['valor_unit']      = df_itens_display['valor_unit'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
        df_itens_display['total']           = df_itens_display['total'].apply(lambda x: f"R$ {formatar_numero_br(x, 2)}")
        df_itens_display = df_itens_display.rename(columns={
            'codigo':'COD.','descricao':'PRODUTO','peso':'PESO',
            'cx_embarque':'CX EMBARQUE','quantidade':'QTDE',
            'preco_historico':'PREÇO HISTÓRICO','valor_unit':'VALOR',
            'total':'TOTAL','comissao':'COMISSÃO%'
        })
        st.dataframe(df_itens_display, use_container_width=True, height=300)
        col_met1, col_met2, col_met3 = st.columns(3)
        with col_met1:
            st.metric("Qtde Total", f"{formatar_numero_br(df_itens['quantidade'].sum(), 0)}")
        with col_met2:
            st.metric("Frete", tipo_frete)
        with col_met3:
            st.metric("Total Final", f"R$ {formatar_numero_br(df_itens['total'].sum(), 2)}")
        obs_pedido = st.text_area("Observação (Pedido)", key="obs_pedido_np", height=100)
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🗑️ Limpar Pedido", use_container_width=True, key="limpar_pedido_np"):
                st.session_state.itens_pedido = []
                st.rerun()
        with col_btn2:
            if st.button("📄 Gerar PDF do Pedido", use_container_width=True, key="gerar_pdf_pedido_np", type="primary"):
                try:
                    dados_cliente_pdf = {
                        'representante': representante, 'razao_social': cliente_selecionado,
                        'nome_fantasia': nome_fantasia, 'cnpj': cnpj_pedido,
                        'ie': insc_estadual, 'telefone': telefone_pedido,
                        'email': email_pedido, 'endereco': endereco_pedido,
                        'obs_cliente': obs_cliente
                    }
                    dados_pedido_pdf = {
                        'numero': num_pedido, 'tabela_preco': tabela_preco,
                        'tipo_frete': tipo_frete,
                        'data_venda': data_venda.strftime('%d/%m/%Y'),
                        'condicoes_pagto': condicoes_pagto
                    }
                    pdf_bytes = gerar_pdf_pedido(dados_cliente_pdf, dados_pedido_pdf,
                                                  st.session_state.itens_pedido, obs_pedido)
                    st.download_button(
                        label="📥 Baixar PDF do Pedido",
                        data=pdf_bytes,
                        file_name=f"Pedido_{num_pedido or 'SN'}_{(cliente_selecionado or 'cliente').replace(' ','_')}.pdf",
                        mime="application/pdf",
                        key="download_pdf_pedido_np"
                    )
                    st.success("✅ PDF gerado com sucesso!")
                except Exception as _e_pdf:
                    st.error(f"❌ Erro ao gerar PDF: {str(_e_pdf)}")
    else:
        st.info("ℹ️ Nenhum item adicionado. Use o formulário acima para adicionar produtos.")



# ====================== HISTÓRICO DO CLIENTE (alias de Histórico > tab Por Cliente) ======================
elif menu == "__historico_cliente__":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">🔍 Histórico do Cliente</h2>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6C757D;font-size:0.88rem;margin-bottom:20px;">Consulta passiva de compras anteriores por CNPJ ou Nome</p>', unsafe_allow_html=True)

    col_busca1, col_busca2 = st.columns(2)
    with col_busca1:
        busca_tipo = st.radio("Buscar por:", ["Nome", "CPF/CNPJ"], horizontal=True, key="busca_tipo_hc")
    with col_busca2:
        if busca_tipo == "Nome":
            termo_busca = st.text_input("Digite o nome do cliente (mín. 3 caracteres)", key="busca_nome_hc", placeholder="Ex: Farmácia...")
        else:
            termo_busca = st.text_input("Digite o CPF/CNPJ", key="busca_cnpj_hc", placeholder="Ex: 12.345.678/0001-90")

    if termo_busca and len(termo_busca) >= 3:
        if busca_tipo == "Nome":
            clientes_encontrados = df[df['RazaoSocial'].str.contains(termo_busca, case=False, na=False)]['RazaoSocial'].unique()
        else:
            clientes_encontrados = df[df['CPF_CNPJ'].str.contains(termo_busca, na=False)]['RazaoSocial'].unique()

        if len(clientes_encontrados) > 0:
            cliente_sel = st.selectbox("Selecione o cliente:", sorted(clientes_encontrados), key="cli_sel_hc")
            historico_cli = df[df['RazaoSocial'] == cliente_sel].copy()
            cliente_info = historico_cli.iloc[0].to_dict() if len(historico_cli) > 0 else {}
            # Gramatura: buscar na planilha produtos_agrupados
            _hg_url = None
            if planilhas_disponiveis.get('produtos_agrupados'):
                _hg_url = planilhas_disponiveis['produtos_agrupados']['url']
            else:
                _hg_url = next((p['url'] for p in planilhas_disponiveis.get('todas', []) if 'PRODUTO' in p['nome'].upper()), None)
            if _hg_url:
                try:
                    _hg_plan = carregar_planilha_github(_hg_url)
                    if _hg_plan is not None:
                        _hg_plan.columns = _hg_plan.columns.str.upper().str.strip()
                        _hg_kcol = next((c for c in _hg_plan.columns if any(x in c for x in ['ID_COD','CODIGO','COD'])), None)
                        _hg_gcol = next((c for c in _hg_plan.columns if 'GRAMATUR' in c), None)
                        if _hg_kcol and _hg_gcol:
                            def _hg_norm2(v):
                                try: return str(int(float(str(v).strip())))
                                except Exception: return str(v).strip()
                            _hg_plan['_K'] = _hg_plan[_hg_kcol].apply(_hg_norm2)
                            _hg_map = _hg_plan.drop_duplicates(subset='_K').set_index('_K')[_hg_gcol]
                            historico_cli['Gramatura'] = historico_cli['CodigoProduto'].apply(_hg_norm2).map(_hg_map).fillna('')
                except Exception:
                    pass

            _hc1, _hc2, _hc3, _hc4 = st.columns(4)
            with _hc1:
                st.metric("CNPJ", cliente_info.get('CPF_CNPJ', ''))
            with _hc2:
                st.metric("Cidade/UF", f"{cliente_info.get('Cidade','')} / {cliente_info.get('Estado','')}")
            with _hc3:
                st.metric("Vendedor", cliente_info.get('Vendedor', ''))
            with _hc4:
                _notas_venda_cli = obter_notas_unicas(historico_cli[historico_cli['TipoMov']=='NF Venda'])
                st.metric("Total Comprado", formatar_moeda(_notas_venda_cli['TotalProduto'].sum()))

            st.markdown("---")
            vendas_cli = historico_cli[historico_cli['TipoMov'] == 'NF Venda']
            devolucoes_cli = historico_cli[historico_cli['TipoMov'] == 'NF Dev.Venda']
            colunas_display_hc = ['DataEmissao', 'TipoMov', 'Numero_NF', 'CodigoProduto', 'NomeProduto', 'Quantidade', 'PrecoUnit', 'TotalProduto']
            if 'Gramatura' in historico_cli.columns:
                colunas_display_hc.insert(colunas_display_hc.index('CodigoProduto') + 1, 'Gramatura')
            if 'PrazoHistorico' in historico_cli.columns:
                colunas_display_hc.append('PrazoHistorico')
            if 'Comissao' in historico_cli.columns:
                colunas_display_hc.append('Comissao')
            _colunas_disp = [c for c in colunas_display_hc if c in historico_cli.columns]

            _ht1, _ht2 = st.tabs(["🛒 Vendas", "↩️ Devoluções"])
            with _ht1:
                if len(vendas_cli) > 0:
                    _vd = vendas_cli[_colunas_disp].copy()
                    _vd['DataEmissao'] = _vd['DataEmissao'].dt.strftime('%d/%m/%Y')
                    _vd = formatar_dataframe_moeda(_vd, ['PrecoUnit', 'TotalProduto'])
                    st.dataframe(_vd.sort_values('DataEmissao', ascending=False), use_container_width=True, height=350)
                else:
                    st.info("Sem vendas registradas")
            with _ht2:
                if len(devolucoes_cli) > 0:
                    _dd = devolucoes_cli[_colunas_disp].copy()
                    _dd['DataEmissao'] = _dd['DataEmissao'].dt.strftime('%d/%m/%Y')
                    _dd = formatar_dataframe_moeda(_dd, ['PrecoUnit', 'TotalProduto'])
                    st.dataframe(_dd.sort_values('DataEmissao', ascending=False), use_container_width=True, height=350)
                else:
                    st.info("Sem devoluções registradas")
        else:
            st.warning("Nenhum cliente encontrado com esse critério.")
    else:
        st.info("👆 Digite pelo menos 3 caracteres para buscar um cliente")

elif menu == "Pedidos Pendentes":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Pedidos Pendentes de Faturamento</h2>', unsafe_allow_html=True)
    
    # Verificar se a planilha existe
    if not planilhas_disponiveis.get('pedidos_pendentes'):
        st.error("❌ Planilha 'PEDIDOSPENDENTES.xlsx' não encontrada")
        st.info("💡 Adicione no GitHub um arquivo com 'PEDIDOSPENDENTES' no nome")
        st.info(f"📂 Local: {GITHUB_REPO}/{GITHUB_FOLDER}/")
        st.stop()
    
    # Carregar planilha
    with st.spinner("📥 Carregando pedidos pendentes..."):
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            from io import BytesIO
            
            # Baixar arquivo
            response = requests.get(planilhas_disponiveis['pedidos_pendentes']['url'])
            excel_file = BytesIO(response.content)
            
            # Extrair shared strings
            with zipfile.ZipFile(excel_file) as z:
                with z.open('xl/sharedStrings.xml') as f:
                    strings_tree = ET.parse(f)
                    ns_str = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    shared_strings = [si.text if si.text else '' for si in strings_tree.findall('.//ss:t', ns_str)]
                
                # Extrair sheet
                with z.open('xl/worksheets/sheet1.xml') as f:
                    sheet_tree = ET.parse(f)
                    ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    # Parsear dados
                    data = []
                    current_client = None
                    current_pedido = None
                    
                    for row in sheet_tree.findall('.//ss:row', ns):
                        row_data = {}
                        for cell in row.findall('.//ss:c', ns):
                            ref = cell.get('r', '')
                            col = ''.join([c for c in ref if c.isalpha()])
                            v_elem = cell.find('.//ss:v', ns)
                            if v_elem is not None and v_elem.text:
                                val_type = cell.get('t', 'n')
                                if val_type == 's':
                                    idx = int(v_elem.text)
                                    value = shared_strings[idx] if idx < len(shared_strings) else v_elem.text
                                else:
                                    value = v_elem.text
                                row_data[col] = value
                        
                        if not row_data:
                            continue
                        
                        # Detectar tipo de linha
                        col_a = row_data.get('A', '')
                        col_b = row_data.get('B', '')
                        
                        # Linha de cliente (apenas coluna A preenchida com nome)
                        if col_a and not col_b and 'N° do pedido' not in col_a and 'Valor Total' not in col_a and col_a != 'Subgrupo:':
                            current_client = col_a
                        
                        # Linha de pedido (tem "N° do pedido:")
                        elif 'N° do pedido' in col_a:
                            current_pedido = col_b
                            
                            # Extrair dados do produto
                            descricao = row_data.get('C', '')
                            if descricao and ' - ' in descricao:
                                # Separar observação da descrição (tudo após a palavra "observa")
                                import re as _re
                                _obs_match = _re.search(r'observa[çc][aã]o[:\s]*', descricao, _re.IGNORECASE)
                                if _obs_match:
                                    observacao = descricao[_obs_match.end():].strip()
                                    descricao = descricao[:_obs_match.start()].strip()
                                else:
                                    observacao = ''

                                # Extrair código do produto (ex: "476 - ATADURA...")
                                codigo_produto = descricao.split(' - ')[0].strip()
                                
                                try:
                                    qtd_contratada = float(row_data.get('D', 0))
                                    valor_unit = float(row_data.get('E', 0))  # Corrigido: E é o valor unitário
                                    qtd_entregue = float(row_data.get('H', 0))  # Corrigido: H é qtd entregue
                                    qtd_pendente = qtd_contratada - qtd_entregue
                                    valor_pendente = qtd_pendente * valor_unit
                                    
                                    # Converter data de emissão (coluna G)
                                    dt_emissao_val = row_data.get('G', '')
                                    if dt_emissao_val:
                                        try:
                                            # Data vem como número (days since 1900)
                                            dt_emissao = pd.Timestamp('1899-12-30') + pd.Timedelta(days=float(dt_emissao_val))
                                        except Exception:
                                            dt_emissao = None
                                    else:
                                        dt_emissao = None
                                    
                                    data.append({
                                        'Cliente': current_client,
                                        'NumeroPedido': current_pedido,
                                        'CodigoProduto': codigo_produto,
                                        'Descricao': descricao,
                                        'Observacoes': observacao,
                                        'QtdContratada': qtd_contratada,
                                        'QtdEntregue': qtd_entregue,
                                        'QtdPendente': qtd_pendente,
                                        'ValorUnit': valor_unit,
                                        'ValorPendente': valor_pendente,
                                        'DataEmissao': dt_emissao,
                                        'Vendedor': row_data.get('J', ''),  # Corrigido: J é o vendedor
                                        'PercEntregue': float(row_data.get('I', 0))  # Corrigido: I é % entregue
                                    })
                                except Exception:
                                    continue
            
            df_pendentes = pd.DataFrame(data)
            if len(df_pendentes) > 0:
                df_pendentes = df_pendentes.drop_duplicates(subset=['NumeroPedido', 'CodigoProduto'])
            
            if len(df_pendentes) == 0:
                st.warning("⚠️ Nenhum pedido pendente encontrado na planilha")
                st.stop()
            
            
        except Exception as e:
            st.error(f"❌ Erro ao processar planilha: {str(e)}")
            st.stop()
    
    st.markdown("---")
    
    # Filtros
    st.subheader("🔍 Filtros")
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    
    with col_f1:
        clientes_pend = ['Todos'] + sorted(df_pendentes['Cliente'].dropna().unique().tolist())
        cliente_pend_filtro = st.selectbox("Cliente", clientes_pend, key="cli_pend")
    
    with col_f2:
        vendedores_pend = ['Todos'] + sorted(df_pendentes['Vendedor'].dropna().unique().tolist())
        vendedor_pend_filtro = st.selectbox("Vendedor", vendedores_pend, key="vend_pend")
    
    with col_f3:
        busca_produto = st.text_input("🔍 Buscar Produto", placeholder="Digite código ou descrição", key="busca_prod_pend")
    
    with col_f4:
        apenas_pendentes = st.checkbox("Apenas com pendência", value=True, key="apenas_pend")
    
    # Aplicar filtros
    df_pend_filtrado = df_pendentes.copy()
    
    if cliente_pend_filtro != 'Todos':
        df_pend_filtrado = df_pend_filtrado[df_pend_filtrado['Cliente'] == cliente_pend_filtro]
    if vendedor_pend_filtro != 'Todos':
        df_pend_filtrado = df_pend_filtrado[df_pend_filtrado['Vendedor'] == vendedor_pend_filtro]
    if busca_produto:
        df_pend_filtrado = df_pend_filtrado[
            df_pend_filtrado['CodigoProduto'].str.contains(busca_produto, case=False, na=False) |
            df_pend_filtrado['Descricao'].str.contains(busca_produto, case=False, na=False)
        ]
    if apenas_pendentes:
        df_pend_filtrado = df_pend_filtrado[df_pend_filtrado['QtdPendente'] > 0]
    
    st.markdown("---")
    
    # Métricas
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_pendente = df_pend_filtrado['ValorPendente'].sum()
        st.metric("Valor Total Pendente", f"R$ {formatar_numero_br(total_pendente, 2)}")
    
    with col2:
        qtd_pendente = df_pend_filtrado['QtdPendente'].sum()
        st.metric("Qtd. Total Pendente", f"{formatar_numero_br(qtd_pendente, 0)}")
    
    with col3:
        pedidos_unicos = df_pend_filtrado['NumeroPedido'].nunique()
        st.metric("Pedidos Únicos", f"{formatar_numero_br(pedidos_unicos, 0)}")
    
    with col4:
        perc_medio = df_pend_filtrado['PercEntregue'].mean() if len(df_pend_filtrado) > 0 else 0
        st.metric("% Médio Entregue", f"{perc_medio:.1f}%")
    
    st.markdown("---")
    
    # Gráficos
    col5, col6 = st.columns(2)
    
    with col5:
        st.subheader("🏢 Top 10 Clientes - Valor Pendente")
        top_clientes = df_pend_filtrado.groupby('Cliente')['ValorPendente'].sum().reset_index()
        top_clientes = top_clientes.sort_values('ValorPendente', ascending=False).head(10)
        
        fig_cli = px.bar(
            top_clientes,
            x='ValorPendente',
            y='Cliente',
            orientation='h',
            labels={'Cliente': 'Cliente', 'ValorPendente': 'Valor Pendente (R$)'},
            color='ValorPendente',
            color_discrete_sequence=['#1F4788']
        )
        fig_cli = aplicar_layout_grafico(fig_cli)
        st.plotly_chart(fig_cli, use_container_width=True)
    
    with col6:
        st.subheader("👤 Top 10 Vendedores - Valor Pendente")
        top_vend = df_pend_filtrado.groupby('Vendedor')['ValorPendente'].sum().reset_index()
        top_vend = top_vend.sort_values('ValorPendente', ascending=False).head(10)
        
        fig_vend = px.bar(
            top_vend,
            x='ValorPendente',
            y='Vendedor',
            orientation='h',
            labels={'Vendedor': 'Vendedor', 'ValorPendente': 'Valor Pendente (R$)'},
            color='ValorPendente',
            color_discrete_sequence=['#4A7BC8']
        )
        fig_vend = aplicar_layout_grafico(fig_vend)
        st.plotly_chart(fig_vend, use_container_width=True)
    
    st.markdown("---")
    
    # Tabela detalhada
    st.subheader("📋 Detalhamento de Pedidos Pendentes")
    
    # Preparar dados para exibição
    df_pend_display = df_pend_filtrado[[
        'Cliente', 'NumeroPedido', 'CodigoProduto', 'Descricao', 
        'QtdContratada', 'QtdEntregue', 'QtdPendente',
        'ValorUnit', 'ValorPendente', 'PercEntregue', 'DataEmissao', 'Vendedor'
    ]].copy()
    
    # Formatar valores
    df_pend_display['ValorUnit'] = df_pend_display['ValorUnit'].apply(
        lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00"
    )
    df_pend_display['ValorPendente'] = df_pend_display['ValorPendente'].apply(
        lambda x: formatar_moeda(x) if pd.notnull(x) else "R$ 0,00"
    )
    df_pend_display['DataEmissao'] = df_pend_display['DataEmissao'].apply(
        lambda x: x.strftime('%d/%m/%Y') if pd.notnull(x) else ''
    )
    df_pend_display['PercEntregue'] = df_pend_display['PercEntregue'].apply(
        lambda x: f"{x:.1f}%" if pd.notnull(x) else "0%"
    )
    
    # Renomear colunas
    df_pend_display = df_pend_display.rename(columns={
        'Cliente': 'Cliente',
        'NumeroPedido': 'N° Pedido',
        'CodigoProduto': 'Código',
        'Descricao': 'Descrição',
        'QtdContratada': 'Qtd Contratada',
        'QtdEntregue': 'Qtd Entregue',
        'QtdPendente': 'Qtd Pendente',
        'ValorUnit': 'Valor Unit.',
        'ValorPendente': 'Valor Pendente',
        'PercEntregue': '% Entregue',
        'DataEmissao': 'Data Emissão',
        'Vendedor': 'Vendedor'
    })
    
    st.dataframe(df_pend_display, use_container_width=True, height=400)
    
    # Botão de download — nome do arquivo reflete o vendedor filtrado
    _nome_arquivo_pend = (
        f"{vendedor_pend_filtro.upper().replace(' ', '_')}_PENDENTES.xlsx"
        if vendedor_pend_filtro != 'Todos'
        else "PEDIDOS_PENDENTES.xlsx"
    )
    st.download_button(
        "📥 Exportar Pedidos Pendentes (Separado por Tipo)",
        to_excel_pedidos_pendentes(df_pend_filtrado),
        _nome_arquivo_pend,
        "application/vnd.ms-excel",
        key="download_pendentes"
    )

    # ===== PREVISÃO DE FATURAMENTO POR CAPACIDADE PRODUTIVA =====
    st.markdown("---")
    st.markdown("### 🏭 Previsão de Faturamento por Capacidade Produtiva")
    st.caption("Converte unidades pendentes em caixas e estima datas de conclusão com base na capacidade diária de cada produto.")

    import math
    import re as _re_prod
    import unicodedata as _unicodedata_prod
    from datetime import date, timedelta

    # ====================== CAPACIDADE PRODUTIVA — VARIÁVEIS EDITÁVEIS ======================
    # Taxa unitária de cada linha é regra de negócio fixa (validada com a diretoria).
    # Só o número de pessoas/máquinas (e o modo da Gaze Estéril) é editável e fica salvo.
    TAXAS_PRODUCAO = {
        'campo_45x50':        {'nome': 'Campo 45x50',                             'recurso': 'pessoas',  'taxa_unit': 25.0,   'default': 7, 'unidade_cap': 'cx'},
        'campo_25x28_c5':     {'nome': 'Campo 25x28 c/5',                          'recurso': 'pessoas',  'taxa_unit': 37 / 7, 'default': 7, 'unidade_cap': 'cx'},
        'campo_25x28_c2':     {'nome': 'Campo 25x28 c/2',                          'recurso': 'pessoas',  'taxa_unit': 3.5,    'default': 2, 'unidade_cap': 'cx'},
        'atadura_farma':      {'nome': 'Atadura Farma',                            'recurso': 'maquinas', 'taxa_unit': 15.0,   'default': 2, 'unidade_cap': 'fd'},
        'atadura_hospitalar': {'nome': 'Atadura Hospitalar',                       'recurso': 'pessoas',  'taxa_unit': 10.0,   'default': 2, 'unidade_cap': 'cx'},
        'gaze_pacote_geral':  {'nome': 'Gaze não estéril pacote',                  'recurso': 'pessoas',  'taxa_unit': 10.0,   'default': 3, 'unidade_cap': 'cx'},
        'gaze_pacote_105gr':  {'nome': 'Gaze não estéril pacote 105gr (09 fios)',  'recurso': 'pessoas',  'taxa_unit': 8.0,    'default': 3, 'unidade_cap': 'cx'},
        'gaze_esteril_pct10': {'nome': 'Gaze estéril pct 10',                      'recurso': 'modo',
                                'modos': {'1 máq. grande + 1 pequena': 100.0, '2 máq. pequenas': 65.0},
                                'default': '1 máq. grande + 1 pequena', 'unidade_cap': 'cx'},
        'gaze_rolo_queijo':   {'nome': 'Gaze em rolo (queijo/circular)',           'recurso': 'pessoas',  'taxa_unit': 29.0,   'default': 1, 'unidade_cap': 'cx'},
    }
    FARDO_PARA_CAIXA = 2  # 1 fardo de Atadura Farma = 2 caixas (para comparar com CAIXAS_NECESSARIAS)

    def carregar_config_producao():
        """Lê pessoas/máquinas/modo salvos no Supabase; usa o default de TAXAS_PRODUCAO se não houver registro."""
        cfg = {k: v['default'] for k, v in TAXAS_PRODUCAO.items()}
        if supa_disponivel():
            for reg in supa_select("producao_capacidade"):
                linha = reg.get('linha')
                if linha in cfg:
                    val = reg.get('valor')
                    if TAXAS_PRODUCAO[linha]['recurso'] != 'modo':
                        try:
                            val = int(float(val))
                        except Exception:
                            continue
                    cfg[linha] = val
        return cfg

    def salvar_config_producao(linha, valor, usuario_nome=None):
        """Salva (upsert) uma variável de capacidade no Supabase."""
        if not supa_disponivel():
            return False
        existentes = supa_select("producao_capacidade", filtros={"linha": linha})
        dados = {
            "linha": linha, "valor": str(valor),
            "atualizado_em": datetime.now().isoformat(),
            "atualizado_por": usuario_nome or "",
        }
        if existentes:
            return supa_update("producao_capacidade", linha, dados, id_col="linha")
        return supa_insert("producao_capacidade", dados) is not None

    def capacidade_dia(linha_key, cfg):
        """Capacidade diária na unidade nativa da linha (cx, exceto Atadura Farma que é fd)."""
        info = TAXAS_PRODUCAO[linha_key]
        if info['recurso'] == 'modo':
            return info['modos'].get(cfg.get(linha_key, info['default']), 0.0)
        return float(cfg.get(linha_key, info['default'])) * info['taxa_unit']

    def identificar_linha_producao(descricao, gramatura=None):
        """
        Identifica a linha de produção (Módulo A) a partir da descrição do produto.
        Retorna a chave de TAXAS_PRODUCAO, ou None se o produto está fora do escopo
        definido (ex.: Gaze Estéril 11 Fios / 50x91).
        """
        d = _unicodedata_prod.normalize('NFKD', str(descricao or '')).encode('ascii', 'ignore').decode('ascii').upper()

        if 'ATADURA' in d:
            return 'atadura_hospitalar' if 'HOSPITALAR' in d else 'atadura_farma'

        if any(x in d for x in ['CAMPO OPERATORIO', 'CAMPO OP']):
            if '45X50' in d or '45 X 50' in d:
                return 'campo_45x50'
            if '25X28' in d or '25 X 28' in d:
                if 'PCT 2' in d or _re_prod.search(r'\bC\s*2\b', d):
                    return 'campo_25x28_c2'
                return 'campo_25x28_c5'
            return None

        if 'CIRCULAR' in d or 'QUEIJO' in d:
            return 'gaze_rolo_queijo'

        if 'NAO ESTERIL' in d or 'PACOTE' in d:
            gram = str(gramatura or '').replace(',', '.').strip()
            fios_09 = bool(_re_prod.search(r'\b0?9\s*F', d))
            try:
                is_105 = abs(float(gram) - 105) < 1
            except Exception:
                is_105 = False
            return 'gaze_pacote_105gr' if (is_105 and fios_09) else 'gaze_pacote_geral'

        if 'ESTERIL' in d:
            if 'PCT 10' in d or _re_prod.search(r'\b13\s*F', d):
                return 'gaze_esteril_pct10'
            return None  # ex.: 11 Fios / 50x91 — fora do escopo definido

        return None

    def adicionar_dias_uteis(data_inicio, dias):
        """Avança N dias úteis (seg–sáb), ignorando domingo."""
        atual = data_inicio
        contados = 0
        while contados < dias:
            atual += timedelta(days=1)
            if atual.weekday() != 6:  # 6 = domingo
                contados += 1
        return atual

    # ── Painel de variáveis de capacidade (pessoas/máquinas/modo) ──────────
    st.markdown("**⚙️ Variáveis de Capacidade Produtiva**")
    st.caption("Ajuste pessoas ou máquinas por linha — a previsão abaixo recalcula na hora.")

    _cfg_prod = carregar_config_producao()
    _cfg_editado = {}
    _linhas_prod = list(TAXAS_PRODUCAO.keys())
    _cols_prod = st.columns(3)
    for _i, _linha_key in enumerate(_linhas_prod):
        _info = TAXAS_PRODUCAO[_linha_key]
        with _cols_prod[_i % 3]:
            if _info['recurso'] == 'modo':
                _opcoes = list(_info['modos'].keys())
                _idx_atual = _opcoes.index(_cfg_prod[_linha_key]) if _cfg_prod[_linha_key] in _opcoes else 0
                _cfg_editado[_linha_key] = st.selectbox(
                    _info['nome'], _opcoes, index=_idx_atual, key=f"cfg_prod_{_linha_key}"
                )
            else:
                _cfg_editado[_linha_key] = st.number_input(
                    f"{_info['nome']} ({_info['recurso']})",
                    min_value=0, value=int(_cfg_prod[_linha_key]), step=1, key=f"cfg_prod_{_linha_key}"
                )

    if st.button("💾 Salvar variáveis de produção", key="salvar_cfg_prod"):
        _usuario_nome_cfg = st.session_state.get('usuario_nome', '')
        _ok_salvar = all(
            salvar_config_producao(k, v, _usuario_nome_cfg) for k, v in _cfg_editado.items()
        )
        if _ok_salvar:
            st.success("✅ Variáveis de produção salvas.")
        else:
            st.warning("⚠️ Não foi possível salvar no Supabase (verifique a conexão). Os valores acima continuam valendo só para esta sessão.")

    st.markdown("---")

    # Carregar produtos_agrupados para obter CX_EMB e PRECO via ID_COD
    _df_prod_prev = None
    if planilhas_disponiveis.get('produtos_agrupados'):
        with st.spinner("Carregando dados de produtos para previsão..."):
            _df_prod_prev = carregar_planilha_github(
                planilhas_disponiveis['produtos_agrupados']['url']
            )
        if _df_prod_prev is not None:
            _df_prod_prev.columns = _df_prod_prev.columns.str.upper().str.strip()

    if _df_prod_prev is None:
        st.warning("⚠️ Planilha Produtos_Agrupados não disponível. Previsão desabilitada.")
    else:
        # Normalizar ID_COD
        def _norm_cod(v):
            try:
                return str(int(float(str(v).strip())))
            except Exception:
                return str(v).strip()

        _df_prod_prev['ID_COD_N'] = _df_prod_prev['ID_COD'].apply(_norm_cod)

        # Colunas necessárias
        _cx_col    = next((c for c in _df_prod_prev.columns if 'CX_EMB' in c), None)
        _preco_col = next((c for c in _df_prod_prev.columns if 'PRECO' in c or 'PREÇO' in c), None)
        _desc_col  = next((c for c in _df_prod_prev.columns if 'DESCRI' in c or 'GRUPO' in c), None)
        _gram_col  = next((c for c in _df_prod_prev.columns if 'GRAMATUR' in c), None)

        if not _cx_col or not _preco_col:
            st.warning(f"⚠️ Colunas CX_EMB ou PRECO não encontradas. Colunas disponíveis: {_df_prod_prev.columns.tolist()}")
        else:
            # Preparar base de pendentes com ID_COD normalizado
            _df_base = df_pend_filtrado.copy()
            _df_base['COD_N'] = _df_base['CodigoProduto'].apply(_norm_cod)

            # Merge com produtos
            _cols_merge = ['ID_COD_N', _cx_col, _preco_col] + ([_desc_col] if _desc_col else []) + ([_gram_col] if _gram_col else [])
            _df_prod_merge = _df_prod_prev[_cols_merge].drop_duplicates(subset=['ID_COD_N'])
            _df_merge = _df_base.merge(
                _df_prod_merge,
                left_on='COD_N',
                right_on='ID_COD_N',
                how='left'
            )

            # Converter unidades → caixas (ceil, evitar div/0)
            def _calc_caixas(row):
                try:
                    cx = float(row[_cx_col])
                    if cx <= 0 or pd.isna(cx):
                        return None
                    return math.ceil(float(row['QtdPendente']) / cx)
                except Exception:
                    return None

            _df_merge['CAIXAS_NECESSARIAS'] = _df_merge.apply(_calc_caixas, axis=1)

            # Identificar linha de produção e capacidade diária dinâmica (pessoas/máquinas configurados acima)
            def _linha_e_capacidade(row):
                desc_ref = row.get(_desc_col, '') if _desc_col else ''
                gram_ref = row.get(_gram_col, '') if _gram_col else ''
                linha_key = identificar_linha_producao(desc_ref if desc_ref else row.get('Descricao', ''), gram_ref)
                if linha_key is None:
                    return pd.Series([None, None, 'SEM CAPACIDADE'])
                cap = capacidade_dia(linha_key, _cfg_editado)
                if TAXAS_PRODUCAO[linha_key]['unidade_cap'] == 'fd':
                    cap = cap * FARDO_PARA_CAIXA  # converte fardos/dia → caixas/dia para comparar com CAIXAS_NECESSARIAS
                return pd.Series([linha_key, cap if cap > 0 else None, TAXAS_PRODUCAO[linha_key]['nome']])

            _df_merge[['LINHA_KEY', 'CAPACIDADE_DIA', 'GRUPO_PROD']] = _df_merge.apply(_linha_e_capacidade, axis=1)
            _df_merge['PEDIDO_ITEM_ID'] = _df_merge['NumeroPedido'].astype(str) + '||' + _df_merge['CodigoProduto'].astype(str)

            # ── Fila de priorização por linha de produção ────────────────
            def carregar_prioridades():
                """Lê a ordem manual salva por linha. Retorna {linha_key: [pedido_item_id, ...]}."""
                prioridades = {}
                if supa_disponivel():
                    for reg in supa_select("producao_prioridade"):
                        try:
                            prioridades[reg['linha']] = json.loads(reg.get('ordem') or '[]')
                        except Exception:
                            prioridades[reg['linha']] = []
                return prioridades

            def salvar_prioridade(linha_key, ordem_lista, usuario_nome=None):
                if not supa_disponivel():
                    return False
                existentes = supa_select("producao_prioridade", filtros={"linha": linha_key})
                dados = {
                    "linha": linha_key, "ordem": json.dumps(ordem_lista),
                    "atualizado_em": datetime.now().isoformat(),
                    "atualizado_por": usuario_nome or "",
                }
                if existentes:
                    return supa_update("producao_prioridade", linha_key, dados, id_col="linha")
                return supa_insert("producao_prioridade", dados) is not None

            def _ordenar_fila(df_linha, ordem_manual):
                """Prioridade manual salva primeiro (na ordem salva); resto por data de emissão."""
                df_linha = df_linha.copy()
                df_linha['_DATA_ORD'] = pd.to_datetime(df_linha['DataEmissao'], errors='coerce')
                df_linha = df_linha.sort_values('_DATA_ORD', na_position='last')
                if not ordem_manual:
                    return df_linha
                ordem_pos = {pid: i for i, pid in enumerate(ordem_manual)}
                df_com_ordem = df_linha[df_linha['PEDIDO_ITEM_ID'].isin(ordem_pos)].copy()
                df_com_ordem['_ORD'] = df_com_ordem['PEDIDO_ITEM_ID'].map(ordem_pos)
                df_com_ordem = df_com_ordem.sort_values('_ORD')
                df_sem_ordem = df_linha[~df_linha['PEDIDO_ITEM_ID'].isin(ordem_pos)]
                return pd.concat([df_com_ordem, df_sem_ordem], ignore_index=True)

            _prioridades_salvas = carregar_prioridades()
            _hoje = date.today()
            _partes_calculadas = []
            _filas_por_linha = {}

            for _linha_key_iter in [k for k in _df_merge['LINHA_KEY'].dropna().unique()]:
                _df_linha = _df_merge[_df_merge['LINHA_KEY'] == _linha_key_iter].copy()
                _cap_linha = _df_linha['CAPACIDADE_DIA'].iloc[0]
                _ordem_salva = _prioridades_salvas.get(_linha_key_iter, [])
                _df_linha = _ordenar_fila(_df_linha, _ordem_salva)

                _acumulado, _dias_lista, _data_lista = 0.0, [], []
                for _, _linha_row in _df_linha.iterrows():
                    _acumulado += (_linha_row['CAIXAS_NECESSARIAS'] or 0)
                    _dias = math.ceil(_acumulado / _cap_linha) if _cap_linha else None
                    _dias_lista.append(_dias)
                    _data_lista.append(adicionar_dias_uteis(_hoje, _dias) if _dias is not None else None)
                _df_linha['DIAS_PRODUCAO'] = _dias_lista
                _df_linha['DATA_PREVISTA'] = _data_lista
                _filas_por_linha[_linha_key_iter] = _df_linha
                _partes_calculadas.append(_df_linha)

            _df_sem_linha = _df_merge[_df_merge['LINHA_KEY'].isna()].copy()
            _df_sem_linha['DIAS_PRODUCAO'] = None
            _df_sem_linha['DATA_PREVISTA'] = None
            _partes_calculadas.append(_df_sem_linha)

            _df_merge = pd.concat(_partes_calculadas, ignore_index=True) if _partes_calculadas else _df_merge
            _df_merge['PREVISAO_FORMATADA'] = _df_merge.apply(
                lambda r: f"{r['DATA_PREVISTA'].strftime('%d/%m/%Y')} ({int(r['DIAS_PRODUCAO'])} dias)"
                if r['DATA_PREVISTA'] is not None else "SEM CAPACIDADE",
                axis=1
            )

            # Valor total por linha
            def _calc_valor(row):
                try:
                    return float(row['QtdPendente']) * float(row[_preco_col])
                except Exception:
                    return 0.0

            _df_merge['VALOR_TOTAL'] = _df_merge.apply(_calc_valor, axis=1)

            # ── KPIs ────────────────────────────────────────────────────
            _kp1, _kp2, _kp3 = st.columns(3)
            _com_prev  = _df_merge[_df_merge['DATA_PREVISTA'].notna()]
            _sem_prev  = _df_merge[_df_merge['DATA_PREVISTA'].isna()]
            _dias_pond = (
                (_com_prev['DIAS_PRODUCAO'] * _com_prev['VALOR_TOTAL']).sum()
                / _com_prev['VALOR_TOTAL'].sum()
            ) if _com_prev['VALOR_TOTAL'].sum() > 0 else 0

            with _kp1:
                st.metric("Valor Total Previsto", f"R$ {formatar_numero_br(_com_prev['VALOR_TOTAL'].sum(), 2)}")
            with _kp2:
                st.metric("Média Ponderada de Dias", f"{_dias_pond:.1f} dias")
            with _kp3:
                st.metric("Itens sem Capacidade", f"{formatar_numero_br(len(_sem_prev), 0)}")

            st.markdown("---")

            # ── Faturamento agrupado por data ────────────────────────────
            st.markdown("**Faturamento Previsto por Data de Conclusão**")
            _fat_data = (
                _com_prev.groupby('PREVISAO_FORMATADA')['VALOR_TOTAL']
                .sum()
                .reset_index()
                .sort_values('PREVISAO_FORMATADA')
            )
            _fat_data.columns = ['Data Prevista', 'Faturamento (R$)']

            _fig_fat = px.bar(
                _fat_data,
                x='Data Prevista',
                y='Faturamento (R$)',
                labels={'Data Prevista': 'Data', 'Faturamento (R$)': 'R$'},
                color_discrete_sequence=['#1F4788']
            )
            _fig_fat = aplicar_layout_grafico(_fig_fat, height=300)
            st.plotly_chart(_fig_fat, use_container_width=True)

            # ── Tabela de previsão ───────────────────────────────────────
            st.markdown("**Relatório Detalhado de Previsão**")
            _cols_show = ['Cliente', 'CodigoProduto', 'Descricao',
                          'QtdPendente', 'CAIXAS_NECESSARIAS',
                          'CAPACIDADE_DIA', 'DIAS_PRODUCAO',
                          'PREVISAO_FORMATADA', 'VALOR_TOTAL']
            _cols_show = [c for c in _cols_show if c in _df_merge.columns]
            _df_show   = _df_merge[_cols_show].copy()
            _df_show['VALOR_TOTAL'] = _df_show['VALOR_TOTAL'].apply(
                lambda x: f"R$ {formatar_numero_br(x, 2)}" if pd.notnull(x) else "R$ 0,00"
            )
            _df_show = _df_show.rename(columns={
                'CodigoProduto':      'Código',
                'Descricao':          'Produto',
                'QtdPendente':        'Qtd (un)',
                'CAIXAS_NECESSARIAS': 'Caixas',
                'CAPACIDADE_DIA':     'Cap/Dia',
                'DIAS_PRODUCAO':      'Dias',
                'PREVISAO_FORMATADA': 'Previsão',
                'VALOR_TOTAL':        'Faturamento',
            })
            st.dataframe(_df_show, use_container_width=True, height=380)

            # ── Fila de Priorização por Produto ────────────────────────
            st.markdown("---")
            st.markdown("**🔀 Fila de Priorização por Produto**")
            st.caption("Use as setas para mudar a ordem de produção. A previsão de todos os pedidos abaixo recalcula na hora.")

            _nome_usuario_fila = st.session_state.get('usuario_nome', '')

            for _linha_key_ui, _df_linha_ui in _filas_por_linha.items():
                _nome_linha_ui = TAXAS_PRODUCAO[_linha_key_ui]['nome']
                with st.expander(f"{_nome_linha_ui} — {len(_df_linha_ui)} pedido(s) na fila"):
                    _ids_ordem = _df_linha_ui['PEDIDO_ITEM_ID'].tolist()
                    _cab1, _cab2, _cab3, _cab4, _cab5, _cab6 = st.columns([0.5, 0.5, 3, 1.2, 1, 1.5])
                    _cab3.markdown("**Cliente / Pedido**")
                    _cab4.markdown("**Caixas**")
                    _cab5.markdown("**Dias**")
                    _cab6.markdown("**Previsão**")
                    for _pos, (_, _item) in enumerate(_df_linha_ui.iterrows()):
                        _c1, _c2, _c3, _c4, _c5, _c6 = st.columns([0.5, 0.5, 3, 1.2, 1, 1.5])
                        _item_id = _item['PEDIDO_ITEM_ID']
                        with _c1:
                            if st.button("↑", key=f"fila_up_{_linha_key_ui}_{_item_id}", disabled=(_pos == 0)):
                                _nova_ordem = _ids_ordem.copy()
                                _nova_ordem[_pos - 1], _nova_ordem[_pos] = _nova_ordem[_pos], _nova_ordem[_pos - 1]
                                salvar_prioridade(_linha_key_ui, _nova_ordem, _nome_usuario_fila)
                                st.rerun()
                        with _c2:
                            if st.button("↓", key=f"fila_down_{_linha_key_ui}_{_item_id}", disabled=(_pos == len(_ids_ordem) - 1)):
                                _nova_ordem = _ids_ordem.copy()
                                _nova_ordem[_pos + 1], _nova_ordem[_pos] = _nova_ordem[_pos], _nova_ordem[_pos + 1]
                                salvar_prioridade(_linha_key_ui, _nova_ordem, _nome_usuario_fila)
                                st.rerun()
                        with _c3:
                            st.write(f"**{_pos + 1}.** {_item.get('Cliente', '')} — Pedido {_item.get('NumeroPedido', '')}")
                        with _c4:
                            st.write(f"{_item.get('CAIXAS_NECESSARIAS') or '—'}")
                        with _c5:
                            st.write(f"{_item.get('DIAS_PRODUCAO') or '—'}")
                        with _c6:
                            _dp_ui = _item.get('DATA_PREVISTA')
                            st.write(_dp_ui.strftime('%d/%m/%Y') if _dp_ui is not None else '—')

            # ── Downloads ───────────────────────────────────────────────
            def _gerar_relatorio_previsao(df_merge, df_prod_prev, cx_col, preco_col, desc_col):
                """
                Gera Excel com abas separadas por grupo de produto.
                Colunas: Cliente, Código, Volumes(cx), Descrição, Contratado,
                         Entregue, Pendente, Valor Unit, Valor Pendente,
                         Data Emissão, Dias Pendentes, Vendedor, %Entregue,
                         Previsão(branco), Categoria, Observações(branco)
                """
                import math
                from datetime import date
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                hoje = date.today()

                # Mapeamento de grupos → abas
                # Grupos: palavras-chave em ordem de prioridade
                # ORDEM CRÍTICA:
                #   1. Campo  — sempre primeiro
                #   2. Tipo Queijo — CIRCULAR deve ser verificado ANTES de NAO ESTERIL
                #      porque "GAZE CIRCULAR NAO ESTERIL" contém ambos os termos.
                #      A presença de CIRCULAR indica Tipo Queijo, independentemente
                #      de conter também NAO ESTERIL.
                #   3. Pacote — Gaze não estéril em pacote (sem CIRCULAR)
                #   4. Esteril — após Pacote para não capturar NAO ESTERIL
                GRUPOS_REGRAS = [
                    ('Campo',       ['CAMPO OPERATORIO', 'CAMPO OPERATÓRIO', 'CAMPO OP']),
                    ('Tipo Queijo', ['GAZE CIRCULAR', 'QUEIJO', 'TIPO QUEIJO', 'CIRCULAR']),
                    ('Pacote',      ['NAO ESTERIL', 'NÃO ESTERIL', 'PACOTE']),
                    ('Esteril',     ['GAZE ESTERIL', 'ESTERIL']),
                ]

                def _norm(s):
                    """Remove acentos e normaliza para comparação segura."""
                    import unicodedata
                    return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii').upper()

                def identificar_aba(descricao):
                    if not descricao:
                        return 'Outros'
                    d = _norm(descricao)

                    # Atadura — sempre primeiro
                    if 'ATADURA' in d:
                        return 'Atadura Hospitalar' if 'HOSPITALAR' in d else 'Atadura Farma'

                    # Campo — antes de qualquer gaze
                    if any(x in d for x in ['CAMPO OPERATORIO', 'CAMPO OPERATORIO', 'CAMPO OP']):
                        return 'Campo'

                    # Tipo Queijo — CIRCULAR tem prioridade sobre NAO ESTERIL.
                    # "GAZE CIRCULAR NAO ESTERIL" deve ir para Tipo Queijo, não Pacote.
                    if 'CIRCULAR' in d or 'QUEIJO' in d or 'TIPO QUEIJO' in d:
                        return 'Tipo Queijo'

                    # Pacote — Gaze não estéril sem CIRCULAR
                    if 'NAO ESTERIL' in d or 'PACOTE' in d:
                        return 'Pacote'

                    # Estéril — após todos os não estéril/circular
                    if 'ESTERIL' in d:
                        return 'Esteril'

                    return 'Outros'

                def identificar_categoria(descricao, aba):
                    if 'Atadura' not in aba:
                        return ''
                    d = str(descricao).upper()
                    if 'HOSPITALAR' in d:
                        return 'Hospitalar'
                    if 'FARMA' in d:
                        return 'Farma'
                    return 'Farma'  # sem indicação → Farma

                def extrair_descricao_pura(descricao, codigo):
                    """Remove código e nome genérico — retorna só descrição adicional"""
                    if not descricao:
                        return ''
                    d = str(descricao).strip()
                    # Remover código se presente no início
                    cod_str = str(codigo).strip()
                    if d.startswith(cod_str):
                        d = d[len(cod_str):].strip(' -|')
                    return d

                # Calcular CX_EMB lookup
                cx_lookup = {}
                if df_prod_prev is not None and cx_col:
                    for _, row in df_prod_prev.iterrows():
                        try:
                            k = str(int(float(str(row['ID_COD_N'])))).strip()
                            v = float(row[cx_col])
                            if v > 0:
                                cx_lookup[k] = v
                        except Exception:
                            pass

                # Gramatura lookup (mesma lógica do módulo consulta tabela)
                gram_col_g = next((c for c in df_prod_prev.columns if 'GRAMATUR' in c), None) if df_prod_prev is not None else None
                gram_lookup = {}
                if df_prod_prev is not None and gram_col_g:
                    for _, row in df_prod_prev.iterrows():
                        try:
                            k = str(int(float(str(row['ID_COD_N'])))).strip()
                            gv = str(row.get(gram_col_g, '')).strip()
                            if gv and gv.lower() not in ('nan', '0', '0.0', ''):
                                gram_lookup[k] = gv
                        except Exception:
                            pass

                COLUNAS_BASE = [
                    'N° Pedido', 'Cliente', 'Código', 'Gramatura', 'Volumes (cx)', 'Descrição',
                    'Contratado', 'Entregue', 'Pendente',
                    'Valor Unitário', 'Valor Pendente',
                    'Data Emissão', 'Dias Pendentes', 'Vendedor',
                    '% Entregue', 'Previsão', 'Categoria', 'Observações'
                ]

                # Abas que exibem Gramatura
                ABAS_COM_GRAM  = {'Tipo Queijo', 'Pacote'}  # Gaze não estéril e Gaze circular
                # Abas com agrupamento por fios
                ABAS_COM_FIOS  = {'Esteril', 'Tipo Queijo', 'Pacote'}
                ORDEM_FIOS_REL = ['09', '11', '13', 'Outros']

                IDX_GRAM_B  = COLUNAS_BASE.index('Gramatura')
                IDX_CONT_B  = COLUNAS_BASE.index('Contratado')
                IDX_ENT_B   = COLUNAS_BASE.index('Entregue')
                IDX_PEND_B  = COLUNAS_BASE.index('Pendente')
                IDX_VUNIT_B = COLUNAS_BASE.index('Valor Unitário')
                IDX_VPEND_B = COLUNAS_BASE.index('Valor Pendente')
                IDX_DESC_B  = COLUNAS_BASE.index('Descrição')

                import re as _re_fios
                def _fios(desc):
                    d = str(desc).upper()
                    for f in ['13', '11', '09', '9']:
                        if _re_fios.search(r'\b' + f + r'\s*F', d) or                            _re_fios.search(r'(^|[\s\-_])' + f + r'(\s|$)', d):
                            return '09' if f == '9' else f
                    return 'Outros'

                # Agrupar linhas por aba
                abas_data = {}
                for _, row in df_merge.iterrows():
                    desc_raw = str(row.get('Descricao', '') or '')
                    # Preferir descrição da planilha base para classificar a aba
                    # pois é a fonte de verdade do produto (evita classificação errada
                    # quando a descrição do PEDIDOSPENDENTES contém termos genéricos)
                    _desc_base = str(row.get(desc_col, '') or '') if desc_col else ''
                    aba = identificar_aba(_desc_base if _desc_base else desc_raw)
                    if aba not in abas_data:
                        abas_data[aba] = []

                    cod = str(row.get('CodigoProduto', '')).strip()
                    try:
                        cod_n = str(int(float(cod)))
                    except Exception:
                        cod_n = cod

                    cx = cx_lookup.get(cod_n, 0)
                    qtd_cont = float(row.get('QtdContratada', 0) or 0)
                    qtd_ent  = float(row.get('QtdEntregue', 0) or 0)
                    qtd_pend = float(row.get('QtdPendente', 0) or 0)
                    val_unit = float(row.get('ValorUnit', 0) or 0)
                    val_pend = val_unit * qtd_pend

                    # Volumes em caixas
                    volumes_cx = math.ceil(qtd_pend / cx) if cx > 0 else ''

                    # Dias pendentes
                    try:
                        _dt_raw2 = row.get('DataEmissao', '')
                        if _dt_raw2 not in (None, '', 'None', 'nan'):
                            _dt_p2 = pd.to_datetime(_dt_raw2, dayfirst=True, errors='coerce')
                            dias_pend = (hoje - _dt_p2.date()).days if pd.notna(_dt_p2) else ''
                        else:
                            dias_pend = ''
                    except Exception:
                        dias_pend = ''

                    # % entregue
                    perc_ent = round((qtd_ent / qtd_cont * 100), 1) if qtd_cont > 0 else 0

                    desc_pura = extrair_descricao_pura(desc_raw, cod)
                    categoria = identificar_categoria(desc_raw, aba)

                    dt_em_fmt = ''
                    try:
                        _dt_raw = row.get('DataEmissao', '')
                        if _dt_raw not in (None, '', 'None', 'nan'):
                            _dt_parsed = pd.to_datetime(_dt_raw, dayfirst=True, errors='coerce')
                            if pd.notna(_dt_parsed):
                                dt_em_fmt = _dt_parsed.strftime('%d/%m/%y')
                    except Exception:
                        pass

                    # Gramatura pelo código
                    gram_val = gram_lookup.get(cod_n, '')

                    abas_data[aba].append([
                        # N° Pedido — remover .0 se vier como float
                        (lambda v: str(int(float(v))) if str(v).replace('.','',1).isdigit() else str(v))(row.get('NumeroPedido', '')),
                        row.get('Cliente', ''),
                        cod,
                        gram_val,      # Gramatura
                        volumes_cx,
                        desc_pura,
                        qtd_cont,
                        qtd_ent,
                        qtd_pend,
                        val_unit,
                        val_pend,
                        dt_em_fmt,
                        dias_pend,
                        row.get('Vendedor', ''),
                        f"{perc_ent:.1f}%",
                        row.get('Previsao', ''),    # Previsão — vem da conciliação ou branco
                        categoria,
                        row.get('Observacoes', ''), # Observações — vem da conciliação ou branco
                    ])

                # Criar workbook
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

                # Estilos
                HDR_FILL  = PatternFill("solid", fgColor="1F4788")
                HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
                HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
                BORDER    = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ALT_FILL  = PatternFill("solid", fgColor="EEF3FC")

                ORDEM_ABAS = ['Atadura Farma', 'Atadura Hospitalar', 'Campo', 'Tipo Queijo', 'Esteril', 'Pacote', 'Outros']
                SEP_FILL   = PatternFill("solid", fgColor="1F4788")
                SEP_FONT   = Font(bold=True, color="FFFFFF", size=10)

                # Subgrupos da aba Campo
                CAMPO_GRUPOS = [
                    ('CAMPO 45X50',   lambda d: '45X50' in d.upper() or '45 X 50' in d.upper()),
                    ('CAMPO 25X28 C5', lambda d: ('25X28' in d.upper() or '25 X 28' in d.upper()) and ('PCT 5' in d.upper() or 'PCT5' in d.upper() or 'C5' in d.upper() or ' 5' in d.upper())),
                    ('CAMPO 25X28 C2', lambda d: ('25X28' in d.upper() or '25 X 28' in d.upper()) and ('PCT 2' in d.upper() or 'PCT2' in d.upper() or 'C2' in d.upper() or ' 2' in d.upper())),
                ]

                def _campo_grupo(desc):
                    d = str(desc)
                    for label, fn in CAMPO_GRUPOS:
                        if fn(d):
                            return label
                    return 'Outros Campo'

                ABAS_COM_CAMPO_SUBGRUPOS = {'Campo'}

                for nome_aba in ORDEM_ABAS:
                    linhas_base = abas_data.get(nome_aba, [])
                    ws = wb.create_sheet(title=nome_aba)

                    # Gramatura: apenas Tipo Queijo (Gaze circular) e Pacote (Gaze não estéril)
                    sem_gram = nome_aba not in ABAS_COM_GRAM
                    COLUNAS  = [c for c in COLUNAS_BASE if not (sem_gram and c == 'Gramatura')]

                    # col_map: índice em COLUNAS_BASE → posição 1-based em COLUNAS
                    col_map = {i_b: i_a for i_a, i_b in
                               enumerate([j for j, c in enumerate(COLUNAS_BASE)
                                          if not (sem_gram and c == 'Gramatura')], 1)}

                    def _linha_aba(lb):
                        return [v for i, v in enumerate(lb) if not (sem_gram and i == IDX_GRAM_B)]

                    # Cabeçalho
                    ws.append(COLUNAS)
                    for col_idx in range(1, len(COLUNAS) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.fill = HDR_FILL; cell.font = HDR_FONT
                        cell.alignment = HDR_ALIGN; cell.border = BORDER
                    ws.row_dimensions[1].height = 30

                    # Índices de coluna para formatos (1-based em COLUNAS)
                    ci_cont  = col_map.get(IDX_CONT_B,  0)
                    ci_ent   = col_map.get(IDX_ENT_B,   0)
                    ci_pend  = col_map.get(IDX_PEND_B,  0)
                    ci_vunit = col_map.get(IDX_VUNIT_B, 0)
                    ci_vpend = col_map.get(IDX_VPEND_B, 0)

                    def _estilizar(r_idx, fill):
                        for ci in range(1, len(COLUNAS) + 1):
                            c = ws.cell(r_idx, ci)
                            c.border = BORDER
                            c.alignment = Alignment(vertical="center")
                            if fill and fill.fill_type: c.fill = fill
                            if ci in (ci_vunit, ci_vpend): c.number_format = 'R$ #,##0.00'
                            if ci in (ci_cont, ci_ent, ci_pend): c.number_format = '#,##0'

                    r_idx = 2

                    if nome_aba in ABAS_COM_CAMPO_SUBGRUPOS:
                        # Agrupar por subgrupo Campo
                        _campo_ordem = [g[0] for g in CAMPO_GRUPOS] + ['Outros Campo']
                        _campo_grps  = {g: [] for g in _campo_ordem}
                        for lb in linhas_base:
                            _campo_grps[_campo_grupo(lb[IDX_DESC_B])].append(lb)

                        dados_escritos = []
                        for _cg_label in _campo_ordem:
                            _cg_linhas = _campo_grps[_cg_label]
                            if not _cg_linhas:
                                continue
                            # Linha separadora
                            ws.cell(r_idx, 1, _cg_label)
                            ws.merge_cells(start_row=r_idx, start_column=1,
                                           end_row=r_idx, end_column=len(COLUNAS))
                            for ci in range(1, len(COLUNAS) + 1):
                                c = ws.cell(r_idx, ci)
                                c.fill = SEP_FILL; c.font = SEP_FONT
                                c.alignment = Alignment(horizontal="center", vertical="center")
                                c.border = BORDER
                            ws.row_dimensions[r_idx].height = 18
                            r_idx += 1
                            for lb in _cg_linhas:
                                ws.append(_linha_aba(lb))
                                _estilizar(r_idx, ALT_FILL if r_idx % 2 == 0 else PatternFill())
                                dados_escritos.append(lb)
                                r_idx += 1
                            # Subtotal por subgrupo Campo
                            _sub_label = f"TOTAL {_cg_label.upper()}"
                            _sub_fill  = PatternFill("solid", fgColor="D9E1F2")
                            _sub_font  = Font(bold=True, size=10)
                            ws.cell(r_idx, 1, _sub_label).font = _sub_font
                            ws.cell(r_idx, 1).fill = _sub_fill
                            for i_b, ci in col_map.items():
                                if i_b in (IDX_CONT_B, IDX_ENT_B, IDX_PEND_B, IDX_VPEND_B):
                                    if i_b == IDX_VPEND_B:
                                        _stot = sum(
                                            float(lb[IDX_PEND_B]) * float(lb[IDX_VUNIT_B])
                                            if isinstance(lb[IDX_PEND_B], (int,float)) and isinstance(lb[IDX_VUNIT_B], (int,float)) else 0
                                            for lb in _cg_linhas
                                        )
                                    else:
                                        _stot = sum(float(lb[i_b]) if isinstance(lb[i_b], (int,float)) else 0 for lb in _cg_linhas)
                                    _sc = ws.cell(r_idx, ci, _stot)
                                    _sc.font = _sub_font
                                    _sc.fill = _sub_fill
                                    _sc.number_format = 'R$ #,##0.00' if i_b == IDX_VPEND_B else '#,##0'
                            r_idx += 1

                        if dados_escritos:
                            ws.cell(r_idx, 1, 'TOTAL').font = Font(bold=True)
                            for i_b, ci in col_map.items():
                                if i_b in (IDX_CONT_B, IDX_ENT_B, IDX_PEND_B, IDX_VPEND_B):
                                    if i_b == IDX_VPEND_B:
                                        tot = sum(
                                            float(lb[IDX_PEND_B]) * float(lb[IDX_VUNIT_B])
                                            if isinstance(lb[IDX_PEND_B], (int,float)) and isinstance(lb[IDX_VUNIT_B], (int,float)) else 0
                                            for lb in dados_escritos
                                        )
                                    else:
                                        tot = sum(float(lb[i_b]) if isinstance(lb[i_b], (int,float)) else 0 for lb in dados_escritos)
                                    c = ws.cell(r_idx, ci, tot)
                                    c.font = Font(bold=True)
                                    c.number_format = 'R$ #,##0.00' if i_b == IDX_VPEND_B else '#,##0'

                    elif nome_aba in ABAS_COM_FIOS:
                        # Agrupar por fios
                        grupos_f = {f: [] for f in ORDEM_FIOS_REL}
                        for lb in linhas_base:
                            grupos_f[_fios(lb[IDX_DESC_B])].append(lb)

                        dados_escritos = []
                        for fio in ORDEM_FIOS_REL:
                            grp = grupos_f[fio]
                            if not grp:
                                continue
                            # Linha separadora
                            label = f"{fio} Fios" if fio != 'Outros' else "Outros"
                            ws.cell(r_idx, 1, label)
                            ws.merge_cells(start_row=r_idx, start_column=1,
                                           end_row=r_idx, end_column=len(COLUNAS))
                            for ci in range(1, len(COLUNAS) + 1):
                                c = ws.cell(r_idx, ci)
                                c.fill = SEP_FILL; c.font = SEP_FONT
                                c.alignment = Alignment(horizontal="center", vertical="center")
                                c.border = BORDER
                            ws.row_dimensions[r_idx].height = 18
                            r_idx += 1
                            for lb in grp:
                                ws.append(_linha_aba(lb))
                                _estilizar(r_idx, ALT_FILL if r_idx % 2 == 0 else PatternFill())
                                dados_escritos.append(lb)
                                r_idx += 1
                            # Subtotal por grupo de fio
                            _sub_label = f"TOTAL {label.upper()}"
                            _sub_fill  = PatternFill("solid", fgColor="D9E1F2")
                            _sub_font  = Font(bold=True, size=10)
                            ws.cell(r_idx, 1, _sub_label).font = _sub_font
                            ws.cell(r_idx, 1).fill = _sub_fill
                            for i_b, ci in col_map.items():
                                if i_b in (IDX_CONT_B, IDX_ENT_B, IDX_PEND_B, IDX_VPEND_B):
                                    if i_b == IDX_VPEND_B:
                                        _stot = sum(
                                            float(lb[IDX_PEND_B]) * float(lb[IDX_VUNIT_B])
                                            if isinstance(lb[IDX_PEND_B], (int,float)) and isinstance(lb[IDX_VUNIT_B], (int,float)) else 0
                                            for lb in grp
                                        )
                                    else:
                                        _stot = sum(float(lb[i_b]) if isinstance(lb[i_b], (int,float)) else 0 for lb in grp)
                                    _sc = ws.cell(r_idx, ci, _stot)
                                    _sc.font = _sub_font
                                    _sc.fill = _sub_fill
                                    _sc.number_format = 'R$ #,##0.00' if i_b == IDX_VPEND_B else '#,##0'
                            r_idx += 1

                        # Total — r_idx aponta para linha após último dado
                        if dados_escritos:
                            ws.cell(r_idx, 1, 'TOTAL').font = Font(bold=True)
                            for i_b, ci in col_map.items():
                                if i_b in (IDX_CONT_B, IDX_ENT_B, IDX_PEND_B, IDX_VPEND_B):
                                    # Valor Pendente = Pendente × Valor Unitário (recalculado)
                                    if i_b == IDX_VPEND_B:
                                        tot = sum(
                                            float(lb[IDX_PEND_B]) * float(lb[IDX_VUNIT_B])
                                            if isinstance(lb[IDX_PEND_B], (int,float)) and isinstance(lb[IDX_VUNIT_B], (int,float)) else 0
                                            for lb in dados_escritos
                                        )
                                    else:
                                        tot = sum(float(lb[i_b]) if isinstance(lb[i_b], (int,float)) else 0 for lb in dados_escritos)
                                    c = ws.cell(r_idx, ci, tot)
                                    c.font = Font(bold=True)
                                    c.number_format = 'R$ #,##0.00' if i_b == IDX_VPEND_B else '#,##0'
                    else:
                        for lb in linhas_base:
                            ws.append(_linha_aba(lb))
                            _estilizar(r_idx, ALT_FILL if r_idx % 2 == 0 else PatternFill())
                            r_idx += 1

                        # Total
                        if linhas_base:
                            ws.cell(r_idx, 1, 'TOTAL').font = Font(bold=True)
                            for i_b, ci in col_map.items():
                                if i_b in (IDX_CONT_B, IDX_ENT_B, IDX_PEND_B, IDX_VPEND_B):
                                    if i_b == IDX_VPEND_B:
                                        tot = sum(
                                            float(lb[IDX_PEND_B]) * float(lb[IDX_VUNIT_B])
                                            if isinstance(lb[IDX_PEND_B], (int,float)) and isinstance(lb[IDX_VUNIT_B], (int,float)) else 0
                                            for lb in linhas_base
                                        )
                                    else:
                                        tot = sum(float(lb[i_b]) if isinstance(lb[i_b], (int,float)) else 0 for lb in linhas_base)
                                    c = ws.cell(r_idx, ci, tot)
                                    c.font = Font(bold=True)
                                    c.number_format = 'R$ #,##0.00' if i_b == IDX_VPEND_B else '#,##0'

                    # Larguras
                    if sem_gram:
                        larguras = [14, 30, 10, 10, 35, 12, 12, 12, 14, 14, 14, 12, 20, 10, 14, 12, 20]
                    else:
                        larguras = [14, 30, 10, 12, 10, 35, 12, 12, 12, 14, 14, 14, 12, 20, 10, 14, 12, 20]
                    for i, larg in enumerate(larguras[:len(COLUNAS)], 1):
                        ws.column_dimensions[get_column_letter(i)].width = larg

                output = BytesIO()
                wb.save(output)
                return output.getvalue()

            _dc1, _dc2 = st.columns(2)
            with _dc1:
                st.download_button(
                    "📥 Relatório Final com Previsão",
                    _gerar_relatorio_previsao(
                        _df_merge, _df_prod_prev, _cx_col, _preco_col, _desc_col
                    ),
                    "RELATORIO_FINAL_COM_PREVISAO.xlsx",
                    "application/vnd.ms-excel",
                    key="dl_previsao_final"
                )
            with _dc2:
                st.download_button(
                    "📥 Faturamento por Data",
                    to_excel(_fat_data),
                    "FATURAMENTO_POR_DATA.xlsx",
                    "application/vnd.ms-excel",
                    key="dl_fat_data"
                )


    # =====================================================================
    # CONCILIAÇÃO: Relatório Atual + Relatório Anterior com Observações
    # =====================================================================
    st.markdown("---")
    st.markdown("### 🔀 Conciliar Relatórios")
    st.caption(
        "Carregue o **Relatório Atual** (recém gerado, sem observações) e o "
        "**Relatório Anterior** onde você preencheu Previsão e Observações. "
        "O sistema transfere as colunas preenchidas para o arquivo atualizado, "
        "vinculando pelo N° Pedido."
    )

    _cc1, _cc2 = st.columns(2)
    with _cc1:
        st.markdown("**1. Relatório Atual** (gerado agora, sem observações)")
        _f_atual = st.file_uploader(
            "Relatório Atual", type=["xlsx"], key="conc_atual",
            label_visibility="collapsed"
        )
    with _cc2:
        st.markdown("**2. Relatório Anterior** (com Previsão e Observações preenchidas)")
        _f_anterior = st.file_uploader(
            "Relatório Anterior", type=["xlsx"], key="conc_anterior",
            label_visibility="collapsed"
        )

    # Salvar bytes em session_state para sobreviver reruns
    if _f_atual:
        st.session_state['_conc_bytes_atual']    = _f_atual.read()
    if _f_anterior:
        st.session_state['_conc_bytes_anterior'] = _f_anterior.read()

    _bytes_atual    = st.session_state.get('_conc_bytes_atual')
    _bytes_anterior = st.session_state.get('_conc_bytes_anterior')

    if _bytes_atual and _bytes_anterior:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO as _BIO

            # ── PASSO 1: extrair mapa {N°Pedido -> {previsao, obs}} do arquivo ANTERIOR ──
            _obs_map = {}
            _wb_ant  = openpyxl.load_workbook(_BIO(_bytes_anterior), data_only=True)

            for _ws_ant in _wb_ant.worksheets:
                _all_rows = list(_ws_ant.iter_rows(values_only=True))
                if len(_all_rows) < 2:
                    continue

                # cabeçalho: normalizar para maiúsculo sem espaços extras
                _hdr = [str(c).strip() if c is not None else '' for c in _all_rows[0]]

                # localizar colunas pelo nome exato (maiúsculo)
                _i_num = _i_prev = _i_obs = None
                for _i, _h in enumerate(_hdr):
                    _hu = _h.upper()
                    if _i_num  is None and any(x in _hu for x in ['N° PEDIDO','N PEDIDO','NUMERO','NUM']):
                        _i_num = _i
                    if _i_prev is None and 'PREV' in _hu:
                        _i_prev = _i
                    if _i_obs  is None and ('OBSERV' in _hu or _hu == 'OBS'):
                        _i_obs = _i

                # detectar também coluna Código e Cliente para chave composta (fallback)
                _i_cod = _i_cli = None
                for _i, _h in enumerate(_hdr):
                    _hu = _h.upper()
                    if _i_cod is None and any(x in _hu for x in ['CÓDIGO','CODIGO']) and 'N°' not in _hu:
                        _i_cod = _i
                    if _i_cli is None and 'CLIENTE' in _hu:
                        _i_cli = _i

                # se não tem N° Pedido E não tem nenhuma coluna útil, pular
                if _i_num is None and _i_cod is None:
                    continue

                for _row in _all_rows[1:]:
                    # chave primária: N° Pedido; fallback: Código|Cliente
                    if _i_num is not None:
                        _raw_num = _row[_i_num] if _i_num < len(_row) else None
                        _k = str(_raw_num).strip() if _raw_num not in (None, '', 'None') else ''
                    else:
                        _cod_v = str(_row[_i_cod]).strip() if _i_cod < len(_row) and _row[_i_cod] not in (None,'','None') else ''
                        _cli_v = str(_row[_i_cli]).strip() if _i_cli is not None and _i_cli < len(_row) and _row[_i_cli] not in (None,'','None') else ''
                        _k = f"{_cod_v}|{_cli_v}" if _cod_v else ''

                    if not _k or _k.upper().startswith('TOTAL'):
                        continue

                    _prev_v = ''
                    _obs_v  = ''
                    if _i_prev is not None and _i_prev < len(_row):
                        _rv = _row[_i_prev]
                        if _rv not in (None, '', 'None'):
                            _prev_v = str(_rv).strip()
                    if _i_obs is not None and _i_obs < len(_row):
                        _rv = _row[_i_obs]
                        if _rv not in (None, '', 'None'):
                            _obs_v = str(_rv).strip()

                    # não sobrescrever com vazio se já tem valor de outra aba
                    _ex = _obs_map.get(_k, {})
                    _obs_map[_k] = {
                        'previsao': _prev_v or _ex.get('previsao', ''),
                        'obs':      _obs_v  or _ex.get('obs', '')
                    }

            # ── PASSO 2: gramatura via tabela de produtos do GitHub ──────────
            _gram_map = {}
            if planilhas_disponiveis.get('produtos_agrupados'):
                _df_gram = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
                if _df_gram is not None:
                    _df_gram.columns = _df_gram.columns.str.upper().str.strip()
                    _gc = next((c for c in _df_gram.columns if any(x in c for x in ['ID_COD','CODIGO','COD'])), None)
                    _gg = next((c for c in _df_gram.columns if 'GRAMATUR' in c), None)
                    if _gc and _gg:
                        for _, _gr in _df_gram.iterrows():
                            try:    _gk = str(int(float(str(_gr[_gc]).strip())))
                            except Exception: _gk = str(_gr[_gc]).strip()
                            _gv = str(_gr[_gg]).strip()
                            if _gv and _gv.lower() not in ('nan','0','0.0',''):
                                _gram_map[_gk] = _gv

            # ── PASSO 3: copiar arquivo ATUAL aba a aba injetando os valores ──
            # ── PASSO 3: reutilizar _gerar_relatorio_previsao com dados conciliados ──
            # Reconstruir df_merge a partir do arquivo atual, injetando Previsão/Obs do anterior

            _wb_at2 = openpyxl.load_workbook(_BIO(_bytes_atual))
            _total_aplicados = 0

            # Reconstruir linhas no formato que _gerar_relatorio_previsao espera (df_merge)
            _rows_conc = []
            for _ws2 in _wb_at2.worksheets:
                _src2 = list(_ws2.iter_rows(values_only=True))
                if len(_src2) < 2:
                    continue
                _hdr2 = [str(c).strip() if c is not None else '' for c in _src2[0]]

                # Mapear cabeçalho → índice
                def _ci2(keywords):
                    for i, h in enumerate(_hdr2):
                        hu = h.upper()
                        if all(k.upper() in hu for k in (keywords if isinstance(keywords, list) else [keywords])):
                            return i
                    return None

                # Usar função que retorna None sem falsy issue com índice 0
                def _find2(*kwlist):
                    for kw in kwlist:
                        r = _ci2(kw if isinstance(kw, list) else [kw])
                        if r is not None:
                            return r
                    return None

                _i_num2  = _find2('N° PEDIDO', 'N°PEDIDO', 'NUMERO PEDIDO', 'N° ', 'PEDIDO', 'NUM')
                _i_cli2  = _find2('CLIENTE')
                _i_cod2  = _find2('CÓDIGO', 'CODIGO')
                _i_desc2 = _ci2(['DESCRIÇÃO']) or _ci2(['DESCRICAO']) or _ci2(['DESCRI'])
                _i_cont2 = _ci2(['CONTRAT'])
                _i_ent2  = _ci2(['ENTREGUE'])
                _i_pend2 = _ci2(['PENDENTE'])
                _i_vunt2 = _ci2(['VALOR UNIT']) or _ci2(['VLUNIT']) or _ci2(['UNIT'])
                _i_vped2 = _ci2(['VALOR PEND']) or _ci2(['VLPEND'])
                _i_data2 = _ci2(['DATA'])
                _i_dias2 = _ci2(['DIAS'])
                _i_vend2 = _ci2(['VENDEDOR'])
                _i_perc2 = _ci2(['%'])
                _i_cat2  = _ci2(['CATEG'])
                _i_obs2  = _ci2(['OBSERV']) or _ci2(['OBS'])

                for _row2 in _src2[1:]:
                    _first2 = str(_row2[0]).strip().upper() if _row2[0] is not None else ''
                    # Pular TOTAL, cabeçalhos e separadores de fios/campo
                    if _first2 in ('', ) or _first2.startswith('TOTAL') or any(f in _first2 for f in ['FIOS', 'OUTROS']):
                        continue

                    def _gv(idx):
                        if idx is not None and idx < len(_row2):
                            v = _row2[idx]
                            return '' if v is None else v
                        return ''

                    _num2 = str(_gv(_i_num2)).strip()

                    # Buscar previsão/obs
                    _lookup2 = _num2 if _num2 and _num2.upper() != 'TOTAL' else ''
                    if not _lookup2:
                        _cod2v = str(_gv(_i_cod2)).strip()
                        _cli2v = str(_gv(_i_cli2)).strip()
                        _lookup2 = f"{_cod2v}|{_cli2v}"

                    _prev2 = _obs2 = ''
                    if _lookup2:
                        _ent2 = _obs_map.get(_lookup2, {})
                        _prev2 = _ent2.get('previsao', '')
                        _obs2  = _ent2.get('obs', '')
                        if _prev2 or _obs2:
                            _total_aplicados += 1

                    # Observação do arquivo atual (extraída da planilha de origem)
                    _obs2_atual = str(_gv(_i_obs2)).strip() if _i_obs2 is not None else ''
                    # Prioridade: arquivo anterior (preenchido manualmente) > arquivo atual (extraído)
                    _obs2 = _obs2 or _obs2_atual

                    # Montar linha como dict compatível com df_merge
                    # Normalizar data: converter datetime do Excel para string dd/mm/yy
                    _dt_conc = _gv(_i_data2)
                    if _dt_conc not in (None, '', 'None', 'nan'):
                        try:
                            _dt_conc_p = pd.to_datetime(_dt_conc, dayfirst=True, errors='coerce')
                            if pd.notna(_dt_conc_p):
                                _dt_conc = _dt_conc_p.strftime('%d/%m/%y')
                        except Exception:
                            pass

                    _rows_conc.append({
                        'NumeroPedido':  _gv(_i_num2),
                        'Cliente':       _gv(_i_cli2),
                        'CodigoProduto': _gv(_i_cod2),
                        'Descricao':     _gv(_i_desc2),
                        'QtdContratada': _gv(_i_cont2),
                        'QtdEntregue':   _gv(_i_ent2),
                        'QtdPendente':   _gv(_i_pend2),
                        'ValorUnit':     _gv(_i_vunt2),
                        'ValorPendente': _gv(_i_vped2),
                        'DataEmissao':   _dt_conc,
                        'Vendedor':      _gv(_i_vend2),
                        'Previsao':      _prev2,
                        'Observacoes':   _obs2,
                    })

            _df_merge_conc = pd.DataFrame(_rows_conc)

            # Gerar Excel usando a mesma função com todas as regras de formatação
            _buf = _BIO(_gerar_relatorio_previsao(_df_merge_conc, _df_prod_prev, _cx_col, _preco_col, _desc_col))

            _n_prev = sum(1 for v in _obs_map.values() if v.get('previsao'))
            _n_obs  = sum(1 for v in _obs_map.values() if v.get('obs'))
            st.success(
                f"✅ Conciliação concluída — "
                f"{len(_obs_map)} pedidos lidos do arquivo anterior | "
                f"{_n_prev} com Previsão | {_n_obs} com Observações | "
                f"{_total_aplicados} linhas preenchidas no arquivo final"
            )

            # debug: mostrar amostra do mapa lido
            if _obs_map:
                with st.expander("🔍 Ver pedidos lidos do arquivo anterior"):
                    _sample = {k: v for k, v in list(_obs_map.items())[:20] if v.get('previsao') or v.get('obs')}
                    if _sample:
                        st.dataframe(
                            pd.DataFrame([{'N° Pedido': k, 'Previsão': v['previsao'], 'Observações': v['obs']} for k, v in _sample.items()])
                        )
                    else:
                        st.warning("Nenhum pedido com Previsão ou Observação encontrado no arquivo anterior.")

            st.download_button(
                "📥 Baixar Relatório Conciliado",
                _buf.getvalue(),
                "RELATORIO_CONCILIADO.xlsx",
                "application/vnd.ms-excel",
                key="dl_conciliado"
            )

        except Exception as _e:
            import traceback
            st.error(f"❌ Erro na conciliação: {_e}")
            st.code(traceback.format_exc())

    elif _bytes_atual or _bytes_anterior:
        st.info("⬆️ Carregue os dois arquivos para habilitar a conciliação.")

# ====================== PERFORMANCE DE VENDEDORES ======================
elif menu == "Performance de Vendedores":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">📈 Performance de Vendedores</h2>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6C757D;font-size:0.88rem;margin-bottom:20px;">Painel gerencial completo — análise individual e comparativa por vendedor</p>', unsafe_allow_html=True)

    # ── Filtros locais do módulo ──────────────────────────────────────────────
    with st.expander("⚙️ Filtros do Módulo", expanded=True):
        _pv_c1, _pv_c2, _pv_c3 = st.columns(3)
        with _pv_c1:
            _pv_vendedores_lista = ['Todos'] + sorted(df['Vendedor'].dropna().unique().tolist())
            _pv_vendedor = st.selectbox("👤 Vendedor", _pv_vendedores_lista, key="pv_vendedor")
        with _pv_c2:
            _pv_regioes_lista = ['Todas'] + sorted(df['Estado'].dropna().unique().tolist())
            _pv_regiao = st.selectbox("🗺️ Região (Estado)", _pv_regioes_lista, key="pv_regiao")
        with _pv_c3:
            _pv_periodo = st.selectbox(
                "📅 Período",
                ["Filtro Global", "Mês Atual", "Últimos 3 Meses", "Últimos 6 Meses", "Ano Atual", "Personalizado"],
                key="pv_periodo"
            )
        if _pv_periodo == "Personalizado":
            _pv_d1, _pv_d2 = st.columns(2)
            with _pv_d1:
                _pv_data_ini = st.date_input("De", value=None, key="pv_data_ini", format="DD/MM/YYYY")
            with _pv_d2:
                _pv_data_fim = st.date_input("Até", value=None, key="pv_data_fim", format="DD/MM/YYYY")
        else:
            _pv_data_ini = None
            _pv_data_fim = None

    # ── Aplicar filtros ───────────────────────────────────────────────────────
    _pv_now = pd.Timestamp.now()
    _pv_df = df.copy()

    # Filtro período
    if _pv_periodo == "Filtro Global":
        _pv_df = df_filtrado.copy()
    elif _pv_periodo == "Mês Atual":
        _pv_df = _pv_df[
            (_pv_df['DataEmissao'].dt.month == _pv_now.month) &
            (_pv_df['DataEmissao'].dt.year == _pv_now.year)
        ]
    elif _pv_periodo == "Últimos 3 Meses":
        _pv_df = _pv_df[_pv_df['DataEmissao'] >= (_pv_now - pd.DateOffset(months=3))]
    elif _pv_periodo == "Últimos 6 Meses":
        _pv_df = _pv_df[_pv_df['DataEmissao'] >= (_pv_now - pd.DateOffset(months=6))]
    elif _pv_periodo == "Ano Atual":
        _pv_df = _pv_df[_pv_df['DataEmissao'].dt.year == _pv_now.year]
    elif _pv_periodo == "Personalizado":
        if _pv_data_ini:
            _pv_df = _pv_df[_pv_df['DataEmissao'] >= pd.to_datetime(_pv_data_ini)]
        if _pv_data_fim:
            _pv_df = _pv_df[_pv_df['DataEmissao'] <= pd.to_datetime(_pv_data_fim)]

    # Filtro região
    if _pv_regiao != 'Todas':
        _pv_df = _pv_df[_pv_df['Estado'] == _pv_regiao]

    # Filtro vendedor
    if _pv_vendedor != 'Todos':
        _pv_df = _pv_df[_pv_df['Vendedor'] == _pv_vendedor]

    # Base apenas vendas e devoluções
    _pv_vendas = _pv_df[_pv_df['TipoMov'] == 'NF Venda'].copy()
    _pv_devol  = _pv_df[_pv_df['TipoMov'] == 'NF Dev.Venda'].copy()
    _pv_notas  = obter_notas_unicas(_pv_df)
    _pv_notas_v = _pv_notas[_pv_notas['TipoMov'] == 'NF Venda']
    _pv_notas_d = _pv_notas[_pv_notas['TipoMov'] == 'NF Dev.Venda']

    # ── KPIs Consolidados ─────────────────────────────────────────────────────
    _pv_fat_bruto   = _pv_notas_v['TotalProduto'].sum()
    _pv_fat_devol   = _pv_notas_d['TotalProduto'].sum()
    _pv_fat_liq     = _pv_fat_bruto - _pv_fat_devol
    _pv_clientes    = _pv_vendas['CPF_CNPJ'].nunique()
    _pv_qtd_notas   = len(_pv_notas_v)
    _pv_ticket      = _pv_fat_bruto / _pv_clientes if _pv_clientes > 0 else 0
    _pv_vol_total   = _pv_vendas['Quantidade'].sum() if 'Quantidade' in _pv_vendas.columns else 0

    # Prazo médio
    def _pv_prazo_medio(df_v):
        try:
            if 'PrazoHistorico' not in df_v.columns:
                return 0
            prazos = []
            for val in df_v['PrazoHistorico'].dropna():
                for p in str(val).split('/'):
                    try:
                        prazos.append(int(p))
                    except Exception:
                        pass
            return sum(prazos) / len(prazos) if prazos else 0
        except Exception:
            return 0

    _pv_prazo = _pv_prazo_medio(_pv_vendas)

    # Comissão média
    def _pv_comissao_media(df_v):
        try:
            if 'Comissao' not in df_v.columns:
                return "N/D"
            mapa = {'4%': 4.0, '3%': 3.0, '2,5%': 2.5, '2%': 2.0}
            vals = df_v['Comissao'].map(mapa).dropna()
            if len(vals) == 0:
                return "N/D"
            return f"{vals.mean():.2f}%"
        except Exception:
            return "N/D"

    _pv_comissao = _pv_comissao_media(_pv_vendas)

    # ── Exibir KPI Cards ─────────────────────────────────────────────────────
    _pv_k1, _pv_k2, _pv_k3, _pv_k4 = st.columns(4)
    with _pv_k1:
        render_kpi_card("Faturamento Líquido", f"R$ {formatar_numero_br(_pv_fat_liq, 0)}", icon="💰", color="#1F4788")
    with _pv_k2:
        render_kpi_card("Faturamento Bruto", f"R$ {formatar_numero_br(_pv_fat_bruto, 0)}", icon="💵", color="#2E86AB")
    with _pv_k3:
        render_kpi_card("Devoluções", f"R$ {formatar_numero_br(_pv_fat_devol, 0)}", icon="↩️", color="#EF4444")
    with _pv_k4:
        render_kpi_card("Clientes Positivados", f"{formatar_numero_br(_pv_clientes, 0)}", icon="👥", color="#28A745")

    st.markdown("<br>", unsafe_allow_html=True)

    _pv_k5, _pv_k6, _pv_k7, _pv_k8 = st.columns(4)
    with _pv_k5:
        render_kpi_card("Ticket Médio", f"R$ {formatar_numero_br(_pv_ticket, 0)}", icon="🎯", color="#F4A261")
    with _pv_k6:
        render_kpi_card("Volume Vendido", f"{formatar_numero_br(_pv_vol_total, 0)} un", icon="📦", color="#6C757D")
    with _pv_k7:
        render_kpi_card("Prazo Médio", f"{_pv_prazo:.0f} dias", icon="📅", color="#163561")
    with _pv_k8:
        render_kpi_card("Comissão Média", _pv_comissao, icon="💎", color="#1B5E8A")

    st.markdown("---")

    # ── Inadimplência por Vendedor ─────────────────────────────────────────
    _pv_df_inad = None
    _pv_inad_vendedor = 0
    _pv_inad_total = 0
    _pv_perc_inad = 0.0
    if planilhas_disponiveis.get('inadimplencia'):
        try:
            _pv_raw_inad = carregar_planilha_github(planilhas_disponiveis['inadimplencia']['url'])
            if _pv_raw_inad is not None:
                _pv_df_inad = processar_inadimplencia(_pv_raw_inad)
                if _pv_vendedor != 'Todos' and 'Vendedor' in _pv_df_inad.columns:
                    _pv_inad_vend_df = _pv_df_inad[_pv_df_inad['Vendedor'] == _pv_vendedor]
                else:
                    _pv_inad_vend_df = _pv_df_inad.copy()
                if _pv_regiao != 'Todas' and 'Estado' in _pv_df_inad.columns:
                    _pv_inad_vend_df = _pv_inad_vend_df[_pv_inad_vend_df['Estado'] == _pv_regiao]
                _pv_inad_vendedor = _pv_inad_vend_df['ValorLiquido'].sum() if 'ValorLiquido' in _pv_inad_vend_df.columns else 0
                _pv_inad_total    = _pv_df_inad['ValorLiquido'].sum() if 'ValorLiquido' in _pv_df_inad.columns else 0
                _pv_perc_inad     = (_pv_inad_vendedor / _pv_fat_bruto * 100) if _pv_fat_bruto > 0 else 0
        except Exception:
            pass

    # ── Contratos por Vendedor (Realizado x Contratado) ─────────────────────
    _pv_df_contrato = None
    _pv_contrato_valor = 0
    _pv_contrato_total = 0
    _pv_perc_realizacao = 0.0
    _pv_contrato_por_vendedor = None
    _pv_col_data_contrato = None
    _pv_ctr_filtrado = None
    if planilhas_disponiveis.get('contrato'):
        try:
            _pv_raw_contrato = carregar_planilha_github(planilhas_disponiveis['contrato']['url'])
            if _pv_raw_contrato is not None:
                _pv_df_contrato = _pv_raw_contrato.copy()
                _pv_df_contrato.columns = [str(c).strip() for c in _pv_df_contrato.columns]

                # Normalizar nome do vendedor (coluna "Funcionário") para casar com "Vendedor"
                if 'Funcionário' in _pv_df_contrato.columns:
                    _pv_df_contrato['_FuncNorm'] = _pv_df_contrato['Funcionário'].astype(str).str.strip().str.upper()
                else:
                    _pv_df_contrato['_FuncNorm'] = ''

                if 'Total Contrato (R$)' in _pv_df_contrato.columns:
                    _pv_df_contrato['_ValorContrato'] = pd.to_numeric(
                        _pv_df_contrato['Total Contrato (R$)'], errors='coerce'
                    ).fillna(0)
                else:
                    _pv_df_contrato['_ValorContrato'] = 0

                # Filtro de período — coluna de data é "Dt.Emissão"
                if 'Dt.Emissão' in _pv_df_contrato.columns:
                    _pv_col_data_contrato = 'Dt.Emissão'
                else:
                    _pv_col_data_contrato = next(
                        (c for c in _pv_df_contrato.columns if 'data' in c.lower() or 'emiss' in c.lower()), None
                    )
                _pv_ctr_filtrado = _pv_df_contrato.copy()
                if _pv_col_data_contrato:
                    _pv_ctr_filtrado[_pv_col_data_contrato] = pd.to_datetime(
                        _pv_ctr_filtrado[_pv_col_data_contrato], errors='coerce'
                    )
                    if _pv_periodo == "Mês Atual":
                        _pv_ctr_filtrado = _pv_ctr_filtrado[
                            (_pv_ctr_filtrado[_pv_col_data_contrato].dt.month == _pv_now.month) &
                            (_pv_ctr_filtrado[_pv_col_data_contrato].dt.year == _pv_now.year)
                        ]
                    elif _pv_periodo == "Últimos 3 Meses":
                        _pv_ctr_filtrado = _pv_ctr_filtrado[_pv_ctr_filtrado[_pv_col_data_contrato] >= (_pv_now - pd.DateOffset(months=3))]
                    elif _pv_periodo == "Últimos 6 Meses":
                        _pv_ctr_filtrado = _pv_ctr_filtrado[_pv_ctr_filtrado[_pv_col_data_contrato] >= (_pv_now - pd.DateOffset(months=6))]
                    elif _pv_periodo == "Ano Atual":
                        _pv_ctr_filtrado = _pv_ctr_filtrado[_pv_ctr_filtrado[_pv_col_data_contrato].dt.year == _pv_now.year]
                    elif _pv_periodo == "Personalizado":
                        if _pv_data_ini:
                            _pv_ctr_filtrado = _pv_ctr_filtrado[_pv_ctr_filtrado[_pv_col_data_contrato] >= pd.to_datetime(_pv_data_ini)]
                        if _pv_data_fim:
                            _pv_ctr_filtrado = _pv_ctr_filtrado[_pv_ctr_filtrado[_pv_col_data_contrato] <= pd.to_datetime(_pv_data_fim)]

                _pv_contrato_total = _pv_ctr_filtrado['_ValorContrato'].sum()

                if _pv_vendedor != 'Todos':
                    _pv_contrato_valor = _pv_ctr_filtrado[
                        _pv_ctr_filtrado['_FuncNorm'] == str(_pv_vendedor).strip().upper()
                    ]['_ValorContrato'].sum()
                else:
                    _pv_contrato_valor = _pv_contrato_total

                _pv_perc_realizacao = (_pv_fat_bruto / _pv_contrato_valor * 100) if _pv_contrato_valor > 0 else 0

                _pv_contrato_por_vendedor = _pv_ctr_filtrado.groupby('_FuncNorm')['_ValorContrato'].sum().reset_index()
                _pv_contrato_por_vendedor.columns = ['_FuncNorm', 'ValorContratado']
        except Exception:
            pass

    # Card de inadimplência
    _pv_ki1, _pv_ki2 = st.columns(2)
    with _pv_ki1:
        render_kpi_card(
            "Índice de Inadimplência (R$)",
            f"R$ {formatar_numero_br(_pv_inad_vendedor, 0)}",
            icon="⚠️",
            color="#EF4444" if _pv_inad_vendedor > 0 else "#28A745"
        )
    with _pv_ki2:
        render_kpi_card(
            "Inadimplência sobre Faturamento",
            f"{_pv_perc_inad:.1f}%",
            icon="📊",
            color="#EF4444" if _pv_perc_inad > 5 else "#F4A261"
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # Card de Contrato x Faturado
    _pv_kc1, _pv_kc2 = st.columns(2)
    with _pv_kc1:
        render_kpi_card(
            "Valor Contratado",
            f"R$ {formatar_numero_br(_pv_contrato_valor, 0)}",
            icon="📄",
            color="#1F4788"
        )
    with _pv_kc2:
        render_kpi_card(
            "% Realização (Faturado / Contratado)",
            f"{_pv_perc_realizacao:.1f}%",
            icon="🎯",
            color="#28A745" if _pv_perc_realizacao >= 100 else "#F4A261"
        )

    st.markdown("---")

    # ── Tabs de análise ───────────────────────────────────────────────────────
    _pv_tab1, _pv_tab2, _pv_tab3, _pv_tab4 = st.tabs([
        "📊 Comparativo", "📈 Evolução Temporal", "🌐 Capilaridade", "🛒 Mix de Produtos"
    ])

    # ─── Tab 1: Comparativo de Vendedores ────────────────────────────────────
    with _pv_tab1:
        st.markdown("#### Comparativo de Desempenho por Vendedor")

        _pv_comp = _pv_notas_v.groupby('Vendedor').agg(
            FaturamentoBruto=('TotalProduto', 'sum'),
            QtdNotas=('Numero_NF', 'count'),
            ClientesAtendidos=('CPF_CNPJ', 'nunique'),
        ).reset_index()

        # Ticket médio por vendedor
        _pv_comp['TicketMedio'] = _pv_comp['FaturamentoBruto'] / _pv_comp['ClientesAtendidos'].replace(0, 1)

        # Volume por vendedor
        if 'Quantidade' in _pv_vendas.columns:
            _pv_vol_vend = _pv_vendas.groupby('Vendedor')['Quantidade'].sum().reset_index()
            _pv_vol_vend.columns = ['Vendedor', 'VolumeTotal']
            _pv_comp = _pv_comp.merge(_pv_vol_vend, on='Vendedor', how='left')
        else:
            _pv_comp['VolumeTotal'] = 0

        # Comissão média por vendedor
        if 'Comissao' in _pv_vendas.columns:
            _mapa_com = {'4%': 4.0, '3%': 3.0, '2,5%': 2.5, '2%': 2.0}
            _pv_vendas_c = _pv_vendas.copy()
            _pv_vendas_c['ComissaoNum'] = _pv_vendas_c['Comissao'].map(_mapa_com)
            _pv_com_vend = _pv_vendas_c.groupby('Vendedor')['ComissaoNum'].mean().reset_index()
            _pv_com_vend.columns = ['Vendedor', 'ComissaoMedia']
            _pv_comp = _pv_comp.merge(_pv_com_vend, on='Vendedor', how='left')
        else:
            _pv_comp['ComissaoMedia'] = None

        # Prazo médio por vendedor
        if 'PrazoHistorico' in _pv_vendas.columns:
            def _prazo_med_vend(series):
                prazos = []
                for val in series.dropna():
                    for p in str(val).split('/'):
                        try:
                            prazos.append(int(p))
                        except Exception:
                            pass
                return sum(prazos) / len(prazos) if prazos else 0
            _pv_prazo_vend = _pv_vendas.groupby('Vendedor')['PrazoHistorico'].apply(_prazo_med_vend).reset_index()
            _pv_prazo_vend.columns = ['Vendedor', 'PrazoMedio']
            _pv_comp = _pv_comp.merge(_pv_prazo_vend, on='Vendedor', how='left')
        else:
            _pv_comp['PrazoMedio'] = 0

        # Valor Contratado e % Realização por vendedor
        if _pv_contrato_por_vendedor is not None:
            _pv_comp['_VendNorm'] = _pv_comp['Vendedor'].astype(str).str.strip().str.upper()
            _pv_comp = _pv_comp.merge(
                _pv_contrato_por_vendedor, left_on='_VendNorm', right_on='_FuncNorm', how='left'
            ).drop(columns=['_FuncNorm', '_VendNorm'])
            _pv_comp['ValorContratado'] = _pv_comp['ValorContratado'].fillna(0)
        else:
            _pv_comp['ValorContratado'] = 0

        _pv_comp['PercRealizacao'] = _pv_comp.apply(
            lambda r: (r['FaturamentoBruto'] / r['ValorContratado'] * 100) if r['ValorContratado'] > 0 else 0,
            axis=1
        )

        _pv_comp = _pv_comp.sort_values('FaturamentoBruto', ascending=False)

        # Gráfico comparativo faturamento
        _pv_cv1, _pv_cv2 = st.columns(2)
        with _pv_cv1:
            _fig_fat = px.bar(
                _pv_comp.head(15),
                x='Vendedor', y='FaturamentoBruto',
                title='Faturamento Bruto por Vendedor',
                labels={'FaturamentoBruto': 'R$', 'Vendedor': ''},
                color='FaturamentoBruto',
                color_continuous_scale=['#A8C4E8', '#1F4788']
            )
            _fig_fat = aplicar_layout_grafico(_fig_fat, height=340)
            _fig_fat.update_traces(
                hovertemplate='<b>%{x}</b><br>R$ %{formatar_numero_br(y, 2)}<extra></extra>'
            )
            st.plotly_chart(_fig_fat, use_container_width=True)

        with _pv_cv2:
            _fig_cli = px.bar(
                _pv_comp.head(15),
                x='Vendedor', y='ClientesAtendidos',
                title='Clientes Atendidos por Vendedor',
                labels={'ClientesAtendidos': 'Clientes', 'Vendedor': ''},
                color='ClientesAtendidos',
                color_continuous_scale=['#B8E0C8', '#28A745']
            )
            _fig_cli = aplicar_layout_grafico(_fig_cli, height=340)
            _fig_cli.update_traces(
                hovertemplate='<b>%{x}</b><br>%{y} clientes<extra></extra>'
            )
            st.plotly_chart(_fig_cli, use_container_width=True)

        _pv_cv3, _pv_cv4 = st.columns(2)
        with _pv_cv3:
            _fig_ticket = px.bar(
                _pv_comp.head(15),
                x='Vendedor', y='TicketMedio',
                title='Ticket Médio por Vendedor',
                labels={'TicketMedio': 'R$', 'Vendedor': ''},
                color='TicketMedio',
                color_continuous_scale=['#F8D9B8', '#F4A261']
            )
            _fig_ticket = aplicar_layout_grafico(_fig_ticket, height=340)
            _fig_ticket.update_traces(
                hovertemplate='<b>%{x}</b><br>Ticket: R$ %{formatar_numero_br(y, 2)}<extra></extra>'
            )
            st.plotly_chart(_fig_ticket, use_container_width=True)

        with _pv_cv4:
            if _pv_comp['ComissaoMedia'].notna().any():
                _fig_com = px.bar(
                    _pv_comp[_pv_comp['ComissaoMedia'].notna()].head(15),
                    x='Vendedor', y='ComissaoMedia',
                    title='Comissão Média por Vendedor (%)',
                    labels={'ComissaoMedia': 'Comissão (%)', 'Vendedor': ''},
                    color='ComissaoMedia',
                    color_continuous_scale=['#C5D5F0', '#163561']
                )
                _fig_com = aplicar_layout_grafico(_fig_com, height=340)
                _fig_com.update_traces(
                    hovertemplate='<b>%{x}</b><br>Comissão: %{y:.2f}%<extra></extra>'
                )
                st.plotly_chart(_fig_com, use_container_width=True)
            else:
                st.info("Dados de comissão não disponíveis. Verifique se a planilha de produtos está carregada.")

        # Tabela consolidada
        st.markdown("#### Tabela Consolidada de Performance")
        _pv_comp_disp = _pv_comp.copy()
        _pv_comp_disp['FaturamentoBruto'] = _pv_comp_disp['FaturamentoBruto'].apply(formatar_moeda)
        _pv_comp_disp['TicketMedio']      = _pv_comp_disp['TicketMedio'].apply(formatar_moeda)
        _pv_comp_disp['VolumeTotal']      = _pv_comp_disp['VolumeTotal'].apply(lambda x: f"{formatar_numero_br(x, 0)} un")
        _pv_comp_disp['ComissaoMedia']    = _pv_comp_disp['ComissaoMedia'].apply(
            lambda x: f"{x:.2f}%" if pd.notnull(x) else "N/D"
        )
        _pv_comp_disp['PrazoMedio']       = _pv_comp_disp['PrazoMedio'].apply(lambda x: f"{x:.0f} dias")
        _pv_comp_disp['ValorContratado']  = _pv_comp_disp['ValorContratado'].apply(formatar_moeda)
        _pv_comp_disp['PercRealizacao']   = _pv_comp_disp['PercRealizacao'].apply(lambda x: f"{x:.1f}%")
        _pv_comp_disp.insert(0, 'Posição', range(1, len(_pv_comp_disp) + 1))
        _pv_comp_disp = _pv_comp_disp.rename(columns={
            'FaturamentoBruto': 'Faturamento',
            'QtdNotas':         'Nº Notas',
            'ClientesAtendidos':'Clientes',
            'TicketMedio':      'Ticket Médio',
            'VolumeTotal':      'Volume',
            'ComissaoMedia':    'Comissão Média',
            'PrazoMedio':       'Prazo Médio',
            'ValorContratado':  'Valor Contratado',
            'PercRealizacao':   '% Realização',
        })
        st.dataframe(_pv_comp_disp, use_container_width=True)

        # ── Geração do Excel com 3 abas ──────────────────────────────────────
        def _gerar_excel_performance(
            _vendas_periodo=None,
            _comp_data=None,
            _vendedor_sel=None,
            _regiao_sel=None,
            _periodo_sel=None,
            _now_ts=None,
            _df_full=None,
            _ctr_data=None,
            _ctr_col_data=None
        ):
            # Usar dados passados explicitamente para evitar problema de closure/cache
            _pv_vendas   = _vendas_periodo
            _pv_comp     = _comp_data
            _pv_vendedor = _vendedor_sel
            _pv_regiao   = _regiao_sel
            _pv_periodo  = _periodo_sel
            _pv_now      = _now_ts
            df           = _df_full
            _pv_ctr      = _ctr_data
            _pv_ctr_col  = _ctr_col_data
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                wb = writer.book

                # ── Formatos ──────────────────────────────────────────────
                fmt_header = wb.add_format({
                    'bold': True, 'bg_color': '#1F4788', 'font_color': '#FFFFFF',
                    'border': 1, 'align': 'center', 'valign': 'vcenter',
                    'font_name': 'Calibri', 'font_size': 10
                })
                fmt_moeda = wb.add_format({
                    'num_format': 'R$ #,##0.00', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9
                })
                fmt_perc = wb.add_format({
                    'num_format': '0.00%', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9
                })
                fmt_text = wb.add_format({
                    'border': 1, 'font_name': 'Calibri', 'font_size': 9
                })
                fmt_text_bold = wb.add_format({
                    'bold': True, 'border': 1, 'font_name': 'Calibri', 'font_size': 9,
                    'bg_color': '#EEF3FC'
                })
                fmt_moeda_bold = wb.add_format({
                    'bold': True, 'num_format': 'R$ #,##0.00', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9, 'bg_color': '#EEF3FC'
                })
                fmt_perc_bold = wb.add_format({
                    'bold': True, 'num_format': '0.00%', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9, 'bg_color': '#EEF3FC'
                })
                fmt_mes_header = wb.add_format({
                    'bold': True, 'bg_color': '#2D5AA0', 'font_color': '#FFFFFF',
                    'border': 1, 'align': 'center', 'valign': 'vcenter',
                    'font_name': 'Calibri', 'font_size': 9
                })
                fmt_mes_atual_header = wb.add_format({
                    'bold': True, 'bg_color': '#E8650A', 'font_color': '#FFFFFF',
                    'border': 1, 'align': 'center', 'valign': 'vcenter',
                    'font_name': 'Calibri', 'font_size': 9
                })
                fmt_num = wb.add_format({
                    'num_format': '#,##0', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9
                })

                # ══════════════════════════════════════════════════════════
                # ABA 1 — Comparativo (existente, melhorada)
                # ══════════════════════════════════════════════════════════
                ws1 = wb.add_worksheet('Comparativo')
                writer.sheets['Comparativo'] = ws1

                _comp_export = _pv_comp.copy().sort_values('FaturamentoBruto', ascending=False)
                _comp_export.insert(0, 'Posição', range(1, len(_comp_export) + 1))

                cols1 = ['Posição', 'Vendedor', 'Faturamento Bruto (R$)', 'Nº Notas',
                         'Clientes Atendidos', 'Ticket Médio (R$)', 'Volume Total (un)',
                         'Comissão Média (%)', 'Prazo Médio (dias)',
                         'Valor Contratado (R$)', '% Realização']

                ws1.set_row(0, 22)
                ws1.write(0, 0, 'COMPARATIVO DE PERFORMANCE DE VENDEDORES', wb.add_format({
                    'bold': True, 'font_color': '#1F4788', 'font_size': 13,
                    'font_name': 'Calibri'
                }))
                ws1.merge_range(0, 0, 0, len(cols1)-1, 'COMPARATIVO DE PERFORMANCE DE VENDEDORES',
                    wb.add_format({'bold': True, 'font_color': '#1F4788', 'font_size': 13,
                                   'font_name': 'Calibri', 'align': 'center'}))
                ws1.write(1, 0, f'Vendedor: {_pv_vendedor}  |  Região: {_pv_regiao}  |  Gerado em: {_pv_now.strftime("%d/%m/%Y %H:%M")}',
                    wb.add_format({'italic': True, 'font_color': '#6C757D', 'font_size': 9, 'font_name': 'Calibri'}))

                for c_idx, col in enumerate(cols1):
                    ws1.write(3, c_idx, col, fmt_header)

                col_widths1 = [8, 28, 20, 10, 18, 18, 16, 16, 16, 20, 14]
                for i, w in enumerate(col_widths1):
                    ws1.set_column(i, i, w)

                for r_idx, row in _comp_export.iterrows():
                    row_num = list(_comp_export.index).index(r_idx) + 4
                    ws1.write(row_num, 0, row['Posição'], fmt_text)
                    ws1.write(row_num, 1, str(row['Vendedor']), fmt_text)
                    ws1.write(row_num, 2, row['FaturamentoBruto'], fmt_moeda)
                    ws1.write(row_num, 3, row['QtdNotas'], fmt_num)
                    ws1.write(row_num, 4, row['ClientesAtendidos'], fmt_num)
                    ws1.write(row_num, 5, row['TicketMedio'], fmt_moeda)
                    ws1.write(row_num, 6, row.get('VolumeTotal', 0), fmt_num)
                    _com_val = row.get('ComissaoMedia')
                    ws1.write(row_num, 7, (_com_val/100) if pd.notnull(_com_val) else '', fmt_perc)
                    ws1.write(row_num, 8, row.get('PrazoMedio', 0), fmt_num)
                    ws1.write(row_num, 9, row.get('ValorContratado', 0), fmt_moeda)
                    _real_val = row.get('PercRealizacao')
                    ws1.write(row_num, 10, (_real_val/100) if pd.notnull(_real_val) else '', fmt_perc)

                # Linha de total
                _tot_row = len(_comp_export) + 4
                ws1.write(_tot_row, 0, '', fmt_text_bold)
                ws1.write(_tot_row, 1, 'TOTAL', fmt_text_bold)
                ws1.write(_tot_row, 2, _comp_export['FaturamentoBruto'].sum(), fmt_moeda_bold)
                ws1.write(_tot_row, 3, _comp_export['QtdNotas'].sum(), fmt_text_bold)
                ws1.write(_tot_row, 4, '', fmt_text_bold)
                ws1.write(_tot_row, 5, '', fmt_text_bold)
                ws1.write(_tot_row, 6, _comp_export['VolumeTotal'].sum(), fmt_text_bold)
                ws1.write(_tot_row, 7, '', fmt_text_bold)
                ws1.write(_tot_row, 8, '', fmt_text_bold)
                _tot_contratado = _comp_export.get('ValorContratado', pd.Series(dtype=float)).sum()
                ws1.write(_tot_row, 9, _tot_contratado, fmt_moeda_bold)
                _tot_fat_geral = _comp_export['FaturamentoBruto'].sum()
                _perc_real_geral = (_tot_fat_geral / _tot_contratado) if _tot_contratado > 0 else ''
                ws1.write(_tot_row, 10, _perc_real_geral, fmt_perc_bold)

                # ══════════════════════════════════════════════════════════
                # ABA 2 — Mês a Mês
                # ══════════════════════════════════════════════════════════
                ws2 = wb.add_worksheet('Mês a Mês')
                writer.sheets['Mês a Mês'] = ws2

                # Usar df completo (sem filtro de período), mas com filtro de vendedor/região
                _mm_df_base = df.copy()
                if _pv_regiao != 'Todas':
                    _mm_df_base = _mm_df_base[_mm_df_base['Estado'] == _pv_regiao]

                _mm_vendas = _mm_df_base[_mm_df_base['TipoMov'] == 'NF Venda'].copy()
                _mm_notas  = obter_notas_unicas(_mm_df_base)
                _mm_notas_v = _mm_notas[_mm_notas['TipoMov'] == 'NF Venda'].copy()

                # Total empresa por mês (sem filtro de vendedor)
                _mm_total_empresa = _mm_notas_v.groupby('MesAno')['TotalProduto'].sum().to_dict()

                # Vendedores a exibir
                if _pv_vendedor != 'Todos':
                    _mm_vendedores = [_pv_vendedor]
                else:
                    _mm_vendedores = sorted(_mm_notas_v['Vendedor'].dropna().unique().tolist())

                # Meses disponíveis ordenados
                _mm_meses = sorted(_mm_notas_v['MesAno'].dropna().unique().tolist())
                _pv_now_mesano = _pv_now.to_period('M').strftime('%Y-%m')

                # Título e filtros
                ws2.merge_range(0, 0, 0, 5 + len(_mm_meses)*2,
                    'PERFORMANCE MENSAL — MÊS A MÊS',
                    wb.add_format({'bold': True, 'font_color': '#1F4788', 'font_size': 13,
                                   'font_name': 'Calibri', 'align': 'center'}))
                ws2.write(1, 0, f'Região: {_pv_regiao}  |  Vendedor: {_pv_vendedor}  |  Todos os períodos históricos  |  Gerado em: {_pv_now.strftime("%d/%m/%Y %H:%M")}',
                    wb.add_format({'italic': True, 'font_color': '#6C757D', 'font_size': 9, 'font_name': 'Calibri'}))

                # Cabeçalhos fixos
                _mm_fixed_cols = ['Vendedor', 'Região']
                _mm_col_offset = len(_mm_fixed_cols)

                ws2.set_row(3, 30)
                ws2.write(3, 0, 'Vendedor', fmt_header)
                ws2.write(3, 1, 'Região', fmt_header)
                ws2.set_column(0, 0, 28)
                ws2.set_column(1, 1, 14)

                # Cabeçalhos de meses (par: valor + %)
                for m_idx, mesano in enumerate(_mm_meses):
                    col_v = _mm_col_offset + m_idx * 2
                    col_p = col_v + 1
                    _is_atual = (mesano == _pv_now_mesano)
                    _hfmt = fmt_mes_atual_header if _is_atual else fmt_mes_header
                    _label = f"{mesano}{'*' if _is_atual else ''}"
                    ws2.write(3, col_v, _label, _hfmt)
                    ws2.write(3, col_p, f'% Emp.', _hfmt)
                    ws2.set_column(col_v, col_v, 16)
                    ws2.set_column(col_p, col_p, 10)

                # Colunas de total
                _tot_col_v = _mm_col_offset + len(_mm_meses) * 2
                _tot_col_p = _tot_col_v + 1
                ws2.write(3, _tot_col_v, 'Total Geral', fmt_header)
                ws2.write(3, _tot_col_p, '% Média Emp.', fmt_header)
                ws2.set_column(_tot_col_v, _tot_col_v, 18)
                ws2.set_column(_tot_col_p, _tot_col_p, 14)

                # Dados por vendedor
                _mm_row = 4
                for vend in _mm_vendedores:
                    _vend_df = _mm_notas_v[_mm_notas_v['Vendedor'] == vend]
                    _vend_por_mes = _vend_df.groupby('MesAno')['TotalProduto'].sum().to_dict()
                    _regiao_vend  = _vend_df['Estado'].mode()[0] if len(_vend_df) > 0 and 'Estado' in _vend_df.columns else _pv_regiao

                    ws2.write(_mm_row, 0, vend, fmt_text)
                    ws2.write(_mm_row, 1, _regiao_vend if _pv_regiao == 'Todas' else _pv_regiao, fmt_text)

                    _total_vend = 0
                    _perc_list  = []
                    for m_idx, mesano in enumerate(_mm_meses):
                        col_v = _mm_col_offset + m_idx * 2
                        col_p = col_v + 1
                        _val  = _vend_por_mes.get(mesano, 0)
                        _emp  = _mm_total_empresa.get(mesano, 0)
                        _perc = (_val / _emp) if _emp > 0 else 0
                        ws2.write(_mm_row, col_v, _val, fmt_moeda)
                        ws2.write(_mm_row, col_p, _perc, fmt_perc)
                        _total_vend += _val
                        if _val > 0:
                            _perc_list.append(_perc)

                    _perc_media = sum(_perc_list) / len(_perc_list) if _perc_list else 0
                    ws2.write(_mm_row, _tot_col_v, _total_vend, fmt_moeda_bold)
                    ws2.write(_mm_row, _tot_col_p, _perc_media, fmt_perc_bold)
                    _mm_row += 1

                # Linha de total empresa por mês
                ws2.write(_mm_row, 0, 'TOTAL EMPRESA', fmt_text_bold)
                ws2.write(_mm_row, 1, '', fmt_text_bold)
                _grand_total = 0
                for m_idx, mesano in enumerate(_mm_meses):
                    col_v = _mm_col_offset + m_idx * 2
                    col_p = col_v + 1
                    _emp_val = _mm_total_empresa.get(mesano, 0)
                    ws2.write(_mm_row, col_v, _emp_val, fmt_moeda_bold)
                    ws2.write(_mm_row, col_p, '', fmt_text_bold)
                    _grand_total += _emp_val
                ws2.write(_mm_row, _tot_col_v, _grand_total, fmt_moeda_bold)
                ws2.write(_mm_row, _tot_col_p, '', fmt_text_bold)

                # Nota de rodapé
                _mm_row += 2
                ws2.write(_mm_row, 0, f'* Mês atual: acumulado até {_pv_now.strftime("%d/%m/%Y")}  |  % Emp. = participação do vendedor no total faturado pela empresa no mês',
                    wb.add_format({'italic': True, 'font_color': '#6C757D', 'font_size': 8, 'font_name': 'Calibri'}))

                # ══════════════════════════════════════════════════════════
                # ABA 3 — Resultado por Produto (separado por vendedor)
                # ══════════════════════════════════════════════════════════
                ws3 = wb.add_worksheet('Resultado por Produto')
                writer.sheets['Resultado por Produto'] = ws3

                # ── Classificador de grupo ────────────────────────────────
                def _classificar_grupo(nome):
                    if pd.isna(nome):
                        return 'OUTROS'
                    n = str(nome).upper()
                    if 'ATADURA' in n:
                        return 'ATADURA'
                    elif 'CAMPO' in n:
                        return 'CAMPO OPERATÓRIO'
                    elif 'GAZE' in n and ('CIRCULAR' in n or 'ROLO' in n):
                        return 'GAZE CIRCULAR'
                    elif ('ESTERIL' in n or 'ESTÉRIL' in n) and 'NAO' not in n and 'NÃO' not in n:
                        return 'ESTÉRIL'
                    elif 'NAO ESTERIL' in n or 'NÃO ESTERIL' in n or 'NÃO ESTÉRIL' in n or 'NAO ESTÉRIL' in n:
                        return 'NÃO ESTÉRIL (PACOTE)'
                    else:
                        return 'OUTROS'

                # ── Base de dados com grupo ───────────────────────────────
                _rp_df = _pv_vendas.copy()
                _rp_df['Grupo'] = _rp_df['NomeProduto'].apply(_classificar_grupo)
                # ValorItem = PrecoUnit * Quantidade (TotalProduto é o total da nota, repetido em cada item)
                _rp_df['ValorItem'] = _rp_df['PrecoUnit'] * _rp_df['Quantidade']

                # Total geral de todos os vendedores no período (para % relativo)
                _rp_total_empresa = _rp_df['ValorItem'].sum()

                # Lista de vendedores a exibir
                if _pv_vendedor != 'Todos':
                    _rp_vendedores_lista = [_pv_vendedor]
                else:
                    _rp_vendedores_lista = sorted(_rp_df['Vendedor'].dropna().unique().tolist())

                # ── Formatos específicos da aba ───────────────────────────
                fmt_vend_header = wb.add_format({
                    'bold': True, 'bg_color': '#163561', 'font_color': '#FFFFFF',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 11,
                    'valign': 'vcenter'
                })
                fmt_grupo_sep = wb.add_format({
                    'bold': True, 'bg_color': '#D0E4F7', 'font_color': '#1F4788',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 9
                })
                fmt_grupo_perc = wb.add_format({
                    'bold': True, 'bg_color': '#D0E4F7', 'font_color': '#1F4788',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 9,
                    'num_format': '0.00%'
                })
                fmt_subtotal = wb.add_format({
                    'bold': True, 'bg_color': '#F0F4FA', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9
                })
                fmt_subtotal_moeda = wb.add_format({
                    'bold': True, 'bg_color': '#F0F4FA', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9, 'num_format': 'R$ #,##0.00'
                })
                fmt_subtotal_perc = wb.add_format({
                    'bold': True, 'bg_color': '#F0F4FA', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9, 'num_format': '0.00%'
                })
                fmt_subtotal_num = wb.add_format({
                    'bold': True, 'bg_color': '#F0F4FA', 'border': 1,
                    'font_name': 'Calibri', 'font_size': 9, 'num_format': '#,##0'
                })

                # ── Configurar larguras de colunas ────────────────────────
                _rp_col_widths = [12, 42, 22, 12, 20, 10, 18, 22]
                for i, w in enumerate(_rp_col_widths):
                    ws3.set_column(i, i, w)

                # ── Título da aba ─────────────────────────────────────────
                fmt_titulo = wb.add_format({
                    'bold': True, 'font_color': '#1F4788', 'font_size': 13,
                    'font_name': 'Calibri', 'align': 'center'
                })
                fmt_subtitulo = wb.add_format({
                    'italic': True, 'font_color': '#6C757D', 'font_size': 9,
                    'font_name': 'Calibri'
                })
                ws3.merge_range(0, 0, 0, 7, 'RESULTADO POR PRODUTO — POR VENDEDOR', fmt_titulo)
                ws3.write(1, 0,
                    f'Região: {_pv_regiao}  |  Período: {_pv_periodo}  |  '
                    f'Total Empresa no Período: R$ {formatar_numero_br(_rp_total_empresa, 2)}  |  '
                    f'Gerado em: {_pv_now.strftime("%d/%m/%Y %H:%M")}',
                    fmt_subtitulo)

                _rp_row = 3
                _grupos_ordem = ['ATADURA', 'CAMPO OPERATÓRIO', 'GAZE CIRCULAR',
                                 'ESTÉRIL', 'NÃO ESTÉRIL (PACOTE)', 'OUTROS']

                # ── Iterar por vendedor ───────────────────────────────────
                for _vend_nome in _rp_vendedores_lista:
                    _vend_df = _rp_df[_rp_df['Vendedor'] == _vend_nome].copy()
                    if len(_vend_df) == 0:
                        continue

                    # Total do vendedor no período
                    _vend_total = _vend_df['ValorItem'].sum()
                    _vend_perc_empresa = (_vend_total / _rp_total_empresa) if _rp_total_empresa > 0 else 0

                    # ── Cabeçalho do vendedor ─────────────────────────────
                    ws3.set_row(_rp_row, 22)
                    ws3.merge_range(_rp_row, 0, _rp_row, 7,
                        f'👤  {_vend_nome}   —   Total: R$ {formatar_numero_br(_vend_total, 2)}   '
                        f'({_vend_perc_empresa:.2%} do total da empresa no período)',
                        fmt_vend_header)
                    _rp_row += 1

                    # ── Cabeçalho das colunas ─────────────────────────────
                    _rp_cols = ['Código', 'Produto', 'Grupo', 'Quantidade',
                                'Faturamento (R$)', 'Clientes',
                                '% no Total do Vendedor', '% do Grupo no Total']
                    ws3.set_row(_rp_row, 20)
                    for c_idx, col in enumerate(_rp_cols):
                        ws3.write(_rp_row, c_idx, col, fmt_header)
                    _rp_row += 1

                    # ── Agregar por produto deste vendedor ────────────────
                    _vp = _vend_df.groupby(['NomeProduto', 'CodigoProduto', 'Grupo']).agg(
                        Quantidade=('Quantidade', 'sum'),
                        Faturamento=('ValorItem', 'sum'),
                        Clientes=('CPF_CNPJ', 'nunique')
                    ).reset_index().sort_values('Faturamento', ascending=False)

                    _vp['Perc_Total'] = _vp['Faturamento'] / _vend_total if _vend_total > 0 else 0
                    _vp_grupo_total = _vend_df.groupby('Grupo')['ValorItem'].sum().to_dict()
                    _vp['Perc_Grupo'] = _vp['Grupo'].map(
                        lambda g: (_vp_grupo_total.get(g, 0) / _vend_total) if _vend_total > 0 else 0
                    )

                    # Ordenar por grupo e faturamento
                    _vp_sorted_parts = [
                        _vp[_vp['Grupo'] == g] for g in _grupos_ordem if g in _vp['Grupo'].values
                    ]
                    _vp_outros = _vp[~_vp['Grupo'].isin(_grupos_ordem)]
                    _vp_sorted = pd.concat(_vp_sorted_parts + ([_vp_outros] if len(_vp_outros) > 0 else []))

                    _ultimo_grupo = None
                    for _, row in _vp_sorted.iterrows():
                        # ── Separador de grupo ────────────────────────────
                        if row['Grupo'] != _ultimo_grupo:
                            _g = row['Grupo']
                            _g_fat = _vp_grupo_total.get(_g, 0)
                            _g_perc = (_g_fat / _vend_total) if _vend_total > 0 else 0
                            ws3.write(_rp_row, 0, '', fmt_grupo_sep)
                            ws3.merge_range(_rp_row, 1, _rp_row, 5,
                                f'▶  {_g}  —  Total: R$ {formatar_numero_br(_g_fat, 2)}  ({_g_perc:.2%} do total do vendedor)',
                                fmt_grupo_sep)
                            ws3.write(_rp_row, 6, _g_perc, fmt_grupo_perc)
                            ws3.write(_rp_row, 7, '', fmt_grupo_sep)
                            _rp_row += 1
                            _ultimo_grupo = _g

                        # ── Linha do produto ──────────────────────────────
                        ws3.write(_rp_row, 0, str(row.get('CodigoProduto', '')), fmt_text)
                        ws3.write(_rp_row, 1, str(row['NomeProduto']), fmt_text)
                        ws3.write(_rp_row, 2, str(row['Grupo']), fmt_text)
                        ws3.write(_rp_row, 3, row['Quantidade'], fmt_num)
                        ws3.write(_rp_row, 4, row['Faturamento'], fmt_moeda)
                        ws3.write(_rp_row, 5, row['Clientes'], fmt_num)
                        ws3.write(_rp_row, 6, row['Perc_Total'], fmt_perc)
                        ws3.write(_rp_row, 7, row['Perc_Grupo'], fmt_perc)
                        _rp_row += 1

                    # ── Subtotal do vendedor ──────────────────────────────
                    ws3.write(_rp_row, 0, '', fmt_subtotal)
                    ws3.merge_range(_rp_row, 1, _rp_row, 2, f'SUBTOTAL — {_vend_nome}', fmt_subtotal)
                    ws3.write(_rp_row, 3, _vp['Quantidade'].sum(), fmt_subtotal_num)
                    ws3.write(_rp_row, 4, _vend_total, fmt_subtotal_moeda)
                    ws3.write(_rp_row, 5, _vp['Clientes'].sum(), fmt_subtotal_num)
                    ws3.write(_rp_row, 6, 1.0 if _vend_total > 0 else 0, fmt_subtotal_perc)
                    ws3.write(_rp_row, 7, _vend_perc_empresa, fmt_subtotal_perc)
                    _rp_row += 1

                    # ── Resumo por grupo do vendedor ──────────────────────
                    _rp_row += 1
                    ws3.write(_rp_row, 0, f'Resumo por Grupo — {_vend_nome}', wb.add_format({
                        'bold': True, 'font_color': '#163561', 'font_size': 9,
                        'font_name': 'Calibri', 'italic': True
                    }))
                    _rp_row += 1
                    _rg_cols_h = ['Grupo', 'Faturamento (R$)', '% no Total do Vendedor', 'Qtd Produtos Distintos']
                    for c_idx, col in enumerate(_rg_cols_h):
                        ws3.write(_rp_row, c_idx, col, fmt_mes_header)
                    _rp_row += 1

                    _vp_grupo_agg = _vp.groupby('Grupo').agg(
                        Faturamento=('Faturamento', 'sum'),
                        QtdProdutos=('NomeProduto', 'count')
                    ).reset_index()
                    _vp_grupo_agg['Perc'] = _vp_grupo_agg['Faturamento'] / _vend_total if _vend_total > 0 else 0
                    _vp_grupo_agg = _vp_grupo_agg.sort_values('Faturamento', ascending=False)

                    for _, grow in _vp_grupo_agg.iterrows():
                        ws3.write(_rp_row, 0, grow['Grupo'], fmt_text)
                        ws3.write(_rp_row, 1, grow['Faturamento'], fmt_moeda)
                        ws3.write(_rp_row, 2, grow['Perc'], fmt_perc)
                        ws3.write(_rp_row, 3, grow['QtdProdutos'], fmt_num)
                        _rp_row += 1

                    # Espaço entre vendedores
                    _rp_row += 2

                # ── Totalizador geral ao final ────────────────────────────
                fmt_grand = wb.add_format({
                    'bold': True, 'bg_color': '#1F4788', 'font_color': '#FFFFFF',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 10
                })
                fmt_grand_moeda = wb.add_format({
                    'bold': True, 'bg_color': '#1F4788', 'font_color': '#FFFFFF',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 10,
                    'num_format': 'R$ #,##0.00'
                })
                fmt_grand_num = wb.add_format({
                    'bold': True, 'bg_color': '#1F4788', 'font_color': '#FFFFFF',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 10,
                    'num_format': '#,##0'
                })
                ws3.set_row(_rp_row, 20)
                ws3.write(_rp_row, 0, '', fmt_grand)
                ws3.merge_range(_rp_row, 1, _rp_row, 2, 'TOTAL GERAL — TODOS OS VENDEDORES', fmt_grand)
                ws3.write(_rp_row, 3, _rp_df['Quantidade'].sum(), fmt_grand_num)
                ws3.write(_rp_row, 4, _rp_total_empresa, fmt_grand_moeda)
                ws3.write(_rp_row, 5, _rp_df['CPF_CNPJ'].nunique(), fmt_grand_num)
                ws3.write(_rp_row, 6, 1.0, wb.add_format({
                    'bold': True, 'bg_color': '#1F4788', 'font_color': '#FFFFFF',
                    'border': 1, 'font_name': 'Calibri', 'font_size': 10, 'num_format': '0.00%'
                }))
                ws3.write(_rp_row, 7, '', fmt_grand)

                # ══════════════════════════════════════════════════════════
                # ABA 4 — Clientes sem Compra (últimos 3 meses)
                # ══════════════════════════════════════════════════════════
                ws4 = wb.add_worksheet('Clientes sem Compra')
                writer.sheets['Clientes sem Compra'] = ws4

                # Construir base de clientes sem compra nos últimos 3 meses
                _cs_hoje = pd.Timestamp.now().normalize()
                _cs_janela_ini = _cs_hoje - pd.DateOffset(months=3)

                _cs_nf = df[df['TipoMov'] == 'NF Venda'].copy()
                _cs_nf['DataEmissao'] = pd.to_datetime(_cs_nf['DataEmissao'], errors='coerce')

                # CPFs que compraram nos últimos 3 meses (serão excluídos)
                _cs_positivaram = set(
                    _cs_nf[(_cs_nf['DataEmissao'] >= _cs_janela_ini) &
                            (_cs_nf['DataEmissao'] <= _cs_hoje)]['CPF_CNPJ'].unique()
                )

                # Cadastro: último registro de cada CPF
                _cs_cadastro = (
                    _cs_nf.sort_values('DataEmissao')
                    .groupby('CPF_CNPJ').last().reset_index()
                    [['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado']]
                )

                # Vendedor principal
                _cs_vend_principal = (
                    _cs_nf.groupby(['CPF_CNPJ', 'Vendedor']).size()
                    .reset_index(name='_cnt').sort_values('_cnt', ascending=False)
                    .groupby('CPF_CNPJ').first().reset_index()[['CPF_CNPJ', 'Vendedor']]
                )

                # Última compra e valor histórico
                _cs_ultima = (_cs_nf.groupby('CPF_CNPJ')['DataEmissao'].max()
                              .reset_index().rename(columns={'DataEmissao': 'UltimaCompra'}))
                _cs_valor  = (_cs_nf.groupby('CPF_CNPJ')['TotalProduto'].sum()
                              .reset_index().rename(columns={'TotalProduto': 'ValorHistorico'}))

                _cs_todos = (_cs_cadastro
                             .merge(_cs_vend_principal, on='CPF_CNPJ', how='left')
                             .merge(_cs_ultima, on='CPF_CNPJ', how='left')
                             .merge(_cs_valor,  on='CPF_CNPJ', how='left'))
                _cs_todos['ValorHistorico'] = _cs_todos['ValorHistorico'].fillna(0)
                _cs_todos['UltimaCompra']   = pd.to_datetime(_cs_todos['UltimaCompra'], errors='coerce')

                # Excluir quem comprou na janela
                _cs_result = _cs_todos[~_cs_todos['CPF_CNPJ'].isin(_cs_positivaram)].copy()

                # Aplicar filtros do módulo
                if _pv_vendedor != 'Todos':
                    _cpfs_vend = set(_cs_nf[_cs_nf['Vendedor'] == _pv_vendedor]['CPF_CNPJ'].unique())
                    _cs_result = _cs_result[_cs_result['CPF_CNPJ'].isin(_cpfs_vend)]
                if _pv_regiao != 'Todas':
                    _cs_result = _cs_result[_cs_result['Estado'] == _pv_regiao]

                _cs_result = _cs_result.sort_values('ValorHistorico', ascending=False)

                # Título
                ws4.merge_range(0, 0, 0, 6,
                    'CLIENTES SEM COMPRA — ÚLTIMOS 3 MESES',
                    wb.add_format({'bold': True, 'font_color': '#1F4788', 'font_size': 13,
                                   'font_name': 'Calibri', 'align': 'center'}))
                ws4.write(1, 0,
                    f'Vendedor: {_pv_vendedor}  |  Região: {_pv_regiao}  |  '
                    f'Período: {_cs_janela_ini.strftime("%d/%m/%Y")} a {_cs_hoje.strftime("%d/%m/%Y")}  |  '
                    f'Total: {len(_cs_result)} clientes  |  Gerado em: {_pv_now.strftime("%d/%m/%Y %H:%M")}',
                    wb.add_format({'italic': True, 'font_color': '#6C757D', 'font_size': 9, 'font_name': 'Calibri'}))

                # Cabeçalho
                _cs_cols = ['Razão Social', 'CPF/CNPJ', 'Vendedor', 'Cidade', 'Estado',
                            'Valor Histórico (R$)', 'Última Compra']
                _cs_widths = [38, 20, 28, 22, 10, 22, 16]
                ws4.set_row(3, 20)
                for c_idx, (col, w) in enumerate(zip(_cs_cols, _cs_widths)):
                    ws4.write(3, c_idx, col, fmt_header)
                    ws4.set_column(c_idx, c_idx, w)

                # Dados
                fmt_data = wb.add_format({'border': 1, 'font_name': 'Calibri', 'font_size': 9,
                                          'num_format': 'DD/MM/YYYY'})
                for r_idx, row in enumerate(_cs_result.itertuples(), start=4):
                    ws4.write(r_idx, 0, str(row.RazaoSocial) if pd.notnull(row.RazaoSocial) else '', fmt_text)
                    ws4.write(r_idx, 1, str(row.CPF_CNPJ)    if pd.notnull(row.CPF_CNPJ)    else '', fmt_text)
                    ws4.write(r_idx, 2, str(row.Vendedor)    if pd.notnull(row.Vendedor)     else '', fmt_text)
                    ws4.write(r_idx, 3, str(row.Cidade)      if pd.notnull(row.Cidade)       else '', fmt_text)
                    ws4.write(r_idx, 4, str(row.Estado)      if pd.notnull(row.Estado)       else '', fmt_text)
                    ws4.write(r_idx, 5, row.ValorHistorico, fmt_moeda)
                    ws4.write(r_idx, 6,
                        row.UltimaCompra.to_pydatetime() if pd.notnull(row.UltimaCompra) else '',
                        fmt_data)

                # Linha de total
                _cs_tot_row = len(_cs_result) + 4
                ws4.write(_cs_tot_row, 0, '', fmt_text_bold)
                ws4.merge_range(_cs_tot_row, 1, _cs_tot_row, 4, f'TOTAL — {len(_cs_result)} clientes', fmt_text_bold)
                ws4.write(_cs_tot_row, 5, _cs_result['ValorHistorico'].sum(), fmt_moeda_bold)
                ws4.write(_cs_tot_row, 6, '', fmt_text_bold)

                # ══════════════════════════════════════════════════════════
                # ABA 5 — Vendas / Contratos (detalhamento)
                # ══════════════════════════════════════════════════════════
                ws5 = wb.add_worksheet('Vendas - Contratos')
                writer.sheets['Vendas - Contratos'] = ws5

                if _pv_ctr is not None and len(_pv_ctr) > 0:
                    _vc_df = _pv_ctr.copy()

                    # Aplicar filtro de vendedor do módulo
                    if _pv_vendedor != 'Todos':
                        _vc_df = _vc_df[_vc_df['_FuncNorm'] == str(_pv_vendedor).strip().upper()]

                    # Aplicar filtro de região do módulo, se a planilha tiver essa informação
                    _vc_col_regiao = next(
                        (c for c in _vc_df.columns if c.strip().lower() in
                         ('estado', 'uf', 'regiao', 'região')), None
                    )
                    if _pv_regiao != 'Todas' and _vc_col_regiao:
                        _vc_df = _vc_df[_vc_df[_vc_col_regiao] == _pv_regiao]

                    # Ordenar pela data de emissão (mais recente primeiro), se disponível
                    if _pv_ctr_col and _pv_ctr_col in _vc_df.columns:
                        _vc_df = _vc_df.sort_values(_pv_ctr_col, ascending=False)

                    # Montar colunas de saída: prioridade para Cliente / Data / Valor / Vendedor,
                    # seguidas de todas as demais colunas originais da planilha (produtos, etc.)
                    _vc_col_cliente = next(
                        (c for c in _vc_df.columns if c.strip().lower() in
                         ('nome cliente', 'cliente', 'razão social', 'razao social')), None
                    )
                    _vc_prioritarias = [c for c in [_vc_col_cliente, 'Funcionário', _pv_ctr_col,
                                                     'Total Contrato (R$)'] if c and c in _vc_df.columns]
                    _vc_demais = [c for c in _vc_df.columns
                                  if c not in _vc_prioritarias and not c.startswith('_')]
                    _vc_cols_final = _vc_prioritarias + _vc_demais

                    _vc_export = _vc_df[_vc_cols_final].copy()

                    _vc_rename = {}
                    if _vc_col_cliente:
                        _vc_rename[_vc_col_cliente] = 'Cliente'
                    if 'Funcionário' in _vc_export.columns:
                        _vc_rename['Funcionário'] = 'Vendedor'
                    if _pv_ctr_col and _pv_ctr_col in _vc_export.columns:
                        _vc_rename[_pv_ctr_col] = 'Data'
                    if 'Total Contrato (R$)' in _vc_export.columns:
                        _vc_rename['Total Contrato (R$)'] = 'Valor Contrato (R$)'
                    _vc_export = _vc_export.rename(columns=_vc_rename)

                    # Título e cabeçalho informativo
                    _vc_ncols = len(_vc_export.columns)
                    ws5.merge_range(0, 0, 0, max(_vc_ncols - 1, 1),
                        'VENDAS / CONTRATOS — DETALHAMENTO',
                        wb.add_format({'bold': True, 'font_color': '#1F4788', 'font_size': 13,
                                       'font_name': 'Calibri', 'align': 'center'}))
                    ws5.write(1, 0,
                        f'Vendedor: {_pv_vendedor}  |  Região: {_pv_regiao}  |  Período: {_pv_periodo}  |  '
                        f'Total: {len(_vc_export)} contratos  |  Gerado em: {_pv_now.strftime("%d/%m/%Y %H:%M")}',
                        wb.add_format({'italic': True, 'font_color': '#6C757D', 'font_size': 9, 'font_name': 'Calibri'}))

                    # Cabeçalho das colunas
                    ws5.set_row(3, 20)
                    for c_idx, col in enumerate(_vc_export.columns):
                        ws5.write(3, c_idx, str(col), fmt_header)
                        _w = 30 if col in ('Cliente',) else (18 if col in ('Data', 'Vendedor', 'Valor Contrato (R$)') else 20)
                        ws5.set_column(c_idx, c_idx, _w)

                    # Dados — detecta tipo de cada coluna para formatar adequadamente
                    fmt_vc_data = wb.add_format({'border': 1, 'font_name': 'Calibri', 'font_size': 9,
                                                  'num_format': 'DD/MM/YYYY'})
                    for r_idx, (_, row) in enumerate(_vc_export.iterrows(), start=4):
                        for c_idx, col in enumerate(_vc_export.columns):
                            _val = row[col]
                            if col == 'Data':
                                _dt = pd.to_datetime(_val, errors='coerce')
                                ws5.write(r_idx, c_idx, _dt.to_pydatetime() if pd.notnull(_dt) else '', fmt_vc_data)
                            elif col == 'Valor Contrato (R$)':
                                ws5.write(r_idx, c_idx, float(_val) if pd.notnull(_val) else 0, fmt_moeda)
                            elif isinstance(_val, (int, float)) and pd.notnull(_val):
                                ws5.write(r_idx, c_idx, _val, fmt_num)
                            else:
                                ws5.write(r_idx, c_idx, str(_val) if pd.notnull(_val) else '', fmt_text)

                    # Linha de total
                    _vc_tot_row = len(_vc_export) + 4
                    _vc_val_col_idx = list(_vc_export.columns).index('Valor Contrato (R$)') \
                        if 'Valor Contrato (R$)' in _vc_export.columns else None
                    ws5.write(_vc_tot_row, 0, '', fmt_text_bold)
                    if _vc_val_col_idx is not None and _vc_val_col_idx > 0:
                        ws5.merge_range(_vc_tot_row, 1, _vc_tot_row, _vc_val_col_idx - 1,
                                         f'TOTAL — {len(_vc_export)} contratos', fmt_text_bold)
                        ws5.write(_vc_tot_row, _vc_val_col_idx,
                                  _vc_export['Valor Contrato (R$)'].sum(), fmt_moeda_bold)
                    else:
                        ws5.merge_range(_vc_tot_row, 0, _vc_tot_row, max(_vc_ncols - 1, 1),
                                         f'TOTAL — {len(_vc_export)} contratos', fmt_text_bold)
                else:
                    ws5.write(0, 0, 'Nenhum dado de contrato disponível para o período/filtro selecionado.',
                               wb.add_format({'italic': True, 'font_color': '#6C757D', 'font_size': 10,
                                              'font_name': 'Calibri'}))

            output.seek(0)
            return output.getvalue()

        _excel_bytes = _gerar_excel_performance(
            _vendas_periodo=_pv_vendas,
            _comp_data=_pv_comp,
            _vendedor_sel=_pv_vendedor,
            _regiao_sel=_pv_regiao,
            _periodo_sel=_pv_periodo,
            _now_ts=_pv_now,
            _df_full=df,
            _ctr_data=_pv_ctr_filtrado,
            _ctr_col_data=_pv_col_data_contrato
        )
        st.download_button(
            "📥 Exportar Comparativo (Excel) — 5 abas",
            _excel_bytes,
            f"performance_vendedores_{_pv_now.strftime('%Y%m%d_%H%M')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="pv_dl_comp"
        )

    # ─── Tab 2: Evolução Temporal ─────────────────────────────────────────────
    with _pv_tab2:
        st.markdown("#### Evolução de Vendas ao Longo do Tempo")

        _pv_evol = _pv_notas_v.groupby(['MesAno', 'Vendedor'])['TotalProduto'].sum().reset_index()
        _pv_evol = _pv_evol.sort_values('MesAno')

        if _pv_vendedor != 'Todos':
            # Exibir linha única do vendedor selecionado
            _pv_evol_filt = _pv_evol[_pv_evol['Vendedor'] == _pv_vendedor]
            if len(_pv_evol_filt) > 0:
                _fig_evol = px.line(
                    _pv_evol_filt, x='MesAno', y='TotalProduto',
                    title=f'Evolução Mensal — {_pv_vendedor}',
                    labels={'MesAno': 'Período', 'TotalProduto': 'R$'},
                    markers=True
                )
                _fig_evol.update_traces(line_color='#1F4788', line_width=3, marker=dict(size=7, color='#1F4788'))
                _fig_evol = aplicar_layout_grafico(_fig_evol, height=380)
                st.plotly_chart(_fig_evol, use_container_width=True)
            else:
                st.info("Nenhuma venda encontrada para este vendedor no período.")
        else:
            # Multi-linha: top 8 vendedores
            _top_vend = _pv_notas_v.groupby('Vendedor')['TotalProduto'].sum().nlargest(8).index.tolist()
            _pv_evol_top = _pv_evol[_pv_evol['Vendedor'].isin(_top_vend)]
            _fig_evol = px.line(
                _pv_evol_top, x='MesAno', y='TotalProduto',
                color='Vendedor',
                title='Evolução Mensal — Top 8 Vendedores',
                labels={'MesAno': 'Período', 'TotalProduto': 'R$', 'Vendedor': 'Vendedor'},
                markers=True,
                color_discrete_sequence=CORES_INST
            )
            _fig_evol.update_traces(line_width=2)
            _fig_evol = aplicar_layout_grafico(_fig_evol, height=420)
            st.plotly_chart(_fig_evol, use_container_width=True)

        # Evolução do ticket médio
        st.markdown("#### Evolução do Ticket Médio")
        _pv_tick_evol = _pv_notas_v.groupby(['MesAno', 'Vendedor']).agg(
            Fat=('TotalProduto', 'sum'),
            Cli=('CPF_CNPJ', 'nunique')
        ).reset_index()
        _pv_tick_evol['TicketMedio'] = _pv_tick_evol['Fat'] / _pv_tick_evol['Cli'].replace(0, 1)
        _pv_tick_evol = _pv_tick_evol.sort_values('MesAno')

        if _pv_vendedor != 'Todos':
            _pv_tick_filt = _pv_tick_evol[_pv_tick_evol['Vendedor'] == _pv_vendedor]
            _fig_tick = px.area(
                _pv_tick_filt, x='MesAno', y='TicketMedio',
                title=f'Ticket Médio Mensal — {_pv_vendedor}',
                labels={'MesAno': 'Período', 'TicketMedio': 'R$'},
                color_discrete_sequence=['#2E86AB']
            )
        else:
            _pv_tick_top = _pv_tick_evol[_pv_tick_evol['Vendedor'].isin(_top_vend[:5])]
            _fig_tick = px.line(
                _pv_tick_top, x='MesAno', y='TicketMedio',
                color='Vendedor',
                title='Ticket Médio Mensal — Top 5 Vendedores',
                labels={'MesAno': 'Período', 'TicketMedio': 'R$'},
                color_discrete_sequence=CORES_INST
            )
        _fig_tick = aplicar_layout_grafico(_fig_tick, height=340)
        st.plotly_chart(_fig_tick, use_container_width=True)

    # ─── Tab 3: Capilaridade ─────────────────────────────────────────────────
    with _pv_tab3:
        st.markdown("#### Análise de Capilaridade — Clientes Atendidos por Vendedor")

        _pv_cap = _pv_vendas.groupby(['Vendedor', 'Estado'])['CPF_CNPJ'].nunique().reset_index()
        _pv_cap.columns = ['Vendedor', 'Estado', 'Clientes']

        if _pv_vendedor != 'Todos':
            _pv_cap_filt = _pv_cap[_pv_cap['Vendedor'] == _pv_vendedor]
            _fig_cap = px.bar(
                _pv_cap_filt.sort_values('Clientes', ascending=False),
                x='Estado', y='Clientes',
                title=f'Clientes por Estado — {_pv_vendedor}',
                labels={'Estado': 'Estado', 'Clientes': 'Nº de Clientes'},
                color='Clientes',
                color_continuous_scale=['#B8E0C8', '#1E7B34']
            )
            _fig_cap = aplicar_layout_grafico(_fig_cap, height=380)
            st.plotly_chart(_fig_cap, use_container_width=True)

            # Mapa de calor vendedor × estado
            st.markdown("#### Heatmap de Faturamento por Estado")
            _pv_heat_v = _pv_notas_v[_pv_notas_v['Vendedor'] == _pv_vendedor].groupby('Estado')['TotalProduto'].sum().reset_index()
            _fig_heat = px.bar(
                _pv_heat_v.sort_values('TotalProduto', ascending=True).tail(15),
                x='TotalProduto', y='Estado', orientation='h',
                title=f'Faturamento por Estado — {_pv_vendedor}',
                labels={'TotalProduto': 'R$', 'Estado': ''},
                color='TotalProduto',
                color_continuous_scale=['#C5D5F0', '#1F4788']
            )
            _fig_heat = aplicar_layout_grafico(_fig_heat, height=380)
            st.plotly_chart(_fig_heat, use_container_width=True)

        else:
            # Heatmap vendedor × estado
            _pv_heat = _pv_vendas.groupby(['Vendedor', 'Estado'])['CPF_CNPJ'].nunique().reset_index()
            _pv_heat_pivot = _pv_heat.pivot(index='Vendedor', columns='Estado', values='CPF_CNPJ').fillna(0)

            _fig_heat = go.Figure(data=go.Heatmap(
                z=_pv_heat_pivot.values,
                x=_pv_heat_pivot.columns.tolist(),
                y=_pv_heat_pivot.index.tolist(),
                colorscale=[[0, '#FFFFFF'], [0.5, '#A8C4E8'], [1, '#1F4788']],
                hovertemplate='Vendedor: %{y}<br>Estado: %{x}<br>Clientes: %{z}<extra></extra>'
            ))
            _fig_heat.update_layout(
                title='Capilaridade — Clientes por Vendedor × Estado',
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(family='Inter, Segoe UI, sans-serif', size=11),
                margin=dict(l=10, r=10, t=40, b=10),
                height=480
            )
            st.plotly_chart(_fig_heat, use_container_width=True)

            # Bubble chart: faturamento vs clientes
            _pv_bubble = _pv_notas_v.groupby('Vendedor').agg(
                Fat=('TotalProduto', 'sum'),
                Cli=('CPF_CNPJ', 'nunique'),
                Notas=('Numero_NF', 'count')
            ).reset_index()
            _fig_bub = px.scatter(
                _pv_bubble,
                x='Cli', y='Fat',
                size='Notas', color='Vendedor',
                hover_name='Vendedor',
                title='Faturamento × Clientes Atendidos (tamanho = Qtd Notas)',
                labels={'Cli': 'Clientes Atendidos', 'Fat': 'Faturamento (R$)'},
                color_discrete_sequence=CORES_INST
            )
            _fig_bub = aplicar_layout_grafico(_fig_bub, height=420)
            st.plotly_chart(_fig_bub, use_container_width=True)

    # ─── Tab 4: Mix de Produtos ───────────────────────────────────────────────
    with _pv_tab4:
        st.markdown("#### Concentração de Vendas por Produto")

        _pv_mix = _pv_vendas.groupby('NomeProduto').agg(
            Total=('TotalProduto', 'sum'),
            Volume=('Quantidade', 'sum') if 'Quantidade' in _pv_vendas.columns else ('TotalProduto', 'count'),
            Clientes=('CPF_CNPJ', 'nunique')
        ).reset_index().sort_values('Total', ascending=False)

        _pv_tot_mix = _pv_mix['Total'].sum()
        _pv_mix['Participacao'] = (_pv_mix['Total'] / _pv_tot_mix * 100).round(2)
        _pv_mix['CumulativaPerc'] = _pv_mix['Participacao'].cumsum()

        # Top 15 pizza
        _pv_top15 = _pv_mix.head(15)
        _fig_pie = px.pie(
            _pv_top15, names='NomeProduto', values='Total',
            title='Top 15 Produtos por Faturamento',
            color_discrete_sequence=CORES_INST,
            hole=0.45
        )
        _fig_pie.update_traces(textposition='inside', textinfo='percent+label',
                               hovertemplate='<b>%{label}</b><br>R$ %{formatar_numero_br(value, 2)}<br>%{percent}<extra></extra>')
        _fig_pie.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Inter, Segoe UI, sans-serif', size=11),
            margin=dict(l=10, r=10, t=40, b=10),
            height=420,
            showlegend=False
        )
        st.plotly_chart(_fig_pie, use_container_width=True)

        # Curva ABC
        st.markdown("#### Curva ABC de Produtos")
        _fig_abc = go.Figure()
        _fig_abc.add_trace(go.Bar(
            x=_pv_top15['NomeProduto'], y=_pv_top15['Total'],
            name='Faturamento', marker_color='#1F4788',
            hovertemplate='<b>%{x}</b><br>R$ %{formatar_numero_br(y, 2)}<extra></extra>'
        ))
        _fig_abc.add_trace(go.Scatter(
            x=_pv_top15['NomeProduto'], y=_pv_top15['CumulativaPerc'].head(15),
            name='% Acumulado', yaxis='y2',
            line=dict(color='#EF4444', width=2, dash='dash'),
            marker=dict(size=5, color='#EF4444'),
            hovertemplate='%{x}<br>Acumulado: %{y:.1f}%<extra></extra>'
        ))
        _fig_abc.update_layout(
            title='Curva ABC — Top 15 Produtos',
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Inter, Segoe UI, sans-serif', size=11),
            margin=dict(l=10, r=10, t=40, b=10),
            height=380,
            yaxis=dict(title='Faturamento (R$)', showgrid=True, gridcolor='#F0F0F0'),
            yaxis2=dict(title='% Acumulado', overlaying='y', side='right', range=[0, 110]),
            legend=dict(orientation='h', y=1.08),
            hoverlabel=dict(bgcolor='#1F4788', font_color='white')
        )
        st.plotly_chart(_fig_abc, use_container_width=True)

        # Tabela de mix
        _pv_mix_disp = _pv_mix.copy()
        _pv_mix_disp['Total']        = _pv_mix_disp['Total'].apply(formatar_moeda)
        _pv_mix_disp['Participacao'] = _pv_mix_disp['Participacao'].apply(lambda x: f"{x:.2f}%")
        _pv_mix_disp['CumulativaPerc'] = _pv_mix_disp['CumulativaPerc'].apply(lambda x: f"{x:.2f}%")
        _pv_mix_disp = _pv_mix_disp.rename(columns={
            'NomeProduto':    'Produto',
            'Total':          'Faturamento',
            'Participacao':   '% Part.',
            'CumulativaPerc': '% Acum.',
            'Clientes':       'Clientes'
        })
        st.dataframe(_pv_mix_disp, use_container_width=True)

        st.download_button(
            "📥 Exportar Mix de Produtos (Excel)",
            to_excel(_pv_mix),
            "mix_produtos.xlsx",
            "application/vnd.ms-excel",
            key="pv_dl_mix"
        )

    # ── Geração de PDF ────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📄 Exportar Relatório PDF")
    st.markdown('<p style="color:#6C757D;font-size:0.84rem;">Gera um relatório executivo em PDF com base nos filtros aplicados.</p>', unsafe_allow_html=True)

    if st.button("🖨️ Gerar Relatório PDF", type="primary", key="pv_gerar_pdf"):
        try:
            from fpdf import FPDF
            import math

            class PerformancePDF(FPDF):
                def header(self):
                    # Logo
                    try:
                        import tempfile, os
                        _resp_logo = requests.get("https://i.imgur.com/gt3rgyL.png", timeout=8)
                        if _resp_logo.status_code == 200:
                            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as _ftmp:
                                _ftmp.write(_resp_logo.content)
                                _tmp_path = _ftmp.name
                            self.image(_tmp_path, x=10, y=8, w=28)
                            os.unlink(_tmp_path)
                    except Exception:
                        pass
                    self.set_xy(42, 10)
                    self.set_font('Helvetica', 'B', 13)
                    self.set_text_color(31, 71, 136)
                    self.cell(0, 6, 'MEDTEXTIL PRODUTOS TEXTIL HOSPITALARES', ln=True)
                    self.set_xy(42, 17)
                    self.set_font('Helvetica', '', 8)
                    self.set_text_color(108, 117, 125)
                    self.cell(0, 5, 'CNPJ: 40.357.820/0001-50  |  Dashboard Comercial BI 2.0', ln=True)
                    self.ln(4)
                    self.set_draw_color(31, 71, 136)
                    self.set_line_width(0.8)
                    self.line(10, self.get_y(), 200, self.get_y())
                    self.ln(3)

                def footer(self):
                    self.set_y(-14)
                    self.set_font('Helvetica', 'I', 7)
                    self.set_text_color(173, 181, 189)
                    self.cell(0, 5, f'Performance de Vendedores  ·  Gerado em {_pv_now.strftime("%d/%m/%Y %H:%M")}  ·  Pág. {self.page_no()}', align='C')

            pdf = PerformancePDF()
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.add_page()
            pdf.set_margins(12, 12, 12)

            # ── Título do relatório ───────────────────────────────────────
            pdf.set_font('Helvetica', 'B', 16)
            pdf.set_text_color(31, 71, 136)
            pdf.cell(0, 10, 'RELATÓRIO DE PERFORMANCE DE VENDEDORES', ln=True, align='C')
            pdf.set_font('Helvetica', '', 9)
            pdf.set_text_color(108, 117, 125)
            _pv_label_periodo = f"Período: {_pv_periodo}"
            if _pv_data_ini and _pv_periodo == "Personalizado":
                _pv_label_periodo = f"Período: {_pv_data_ini.strftime('%d/%m/%Y') if hasattr(_pv_data_ini,'strftime') else str(_pv_data_ini)} a {_pv_data_fim.strftime('%d/%m/%Y') if _pv_data_fim and hasattr(_pv_data_fim,'strftime') else '—'}"
            pdf.cell(0, 6, f"Vendedor: {_pv_vendedor}  |  Região: {_pv_regiao}  |  {_pv_label_periodo}", ln=True, align='C')
            pdf.ln(5)

            # ── Resumo Executivo ──────────────────────────────────────────
            pdf.set_fill_color(31, 71, 136)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(0, 7, '  RESUMO EXECUTIVO', fill=True, border=0, ln=True)
            pdf.set_text_color(50, 50, 50)
            pdf.set_font('Helvetica', '', 8)

            _pv_kpis_pdf = [
                ('Faturamento Líquido',       f"R$ {formatar_numero_br(_pv_fat_liq, 2)}"),
                ('Faturamento Bruto',          f"R$ {formatar_numero_br(_pv_fat_bruto, 2)}"),
                ('Devoluções',                 f"R$ {formatar_numero_br(_pv_fat_devol, 2)}"),
                ('Clientes Positivados',       f"{formatar_numero_br(_pv_clientes, 0)}"),
                ('Ticket Médio',               f"R$ {formatar_numero_br(_pv_ticket, 2)}"),
                ('Volume Vendido',             f"{formatar_numero_br(_pv_vol_total, 0)} un"),
                ('Prazo Médio de Venda',       f"{_pv_prazo:.0f} dias"),
                ('Comissão Média',             _pv_comissao),
                ('Índice de Inadimplência',    f"R$ {formatar_numero_br(_pv_inad_vendedor, 2)}"),
                ('Inadimplência / Fat. Bruto', f"{_pv_perc_inad:.1f}%"),
            ]
            w1, w2 = 85, 95
            fill_kpi = False
            for k, v in _pv_kpis_pdf:
                pdf.set_fill_color(240, 244, 255) if fill_kpi else pdf.set_fill_color(255, 255, 255)
                pdf.set_font('Helvetica', 'B', 8)
                pdf.cell(w1, 6, f'  {k}:', border='LB', fill=True)
                pdf.set_font('Helvetica', '', 8)
                pdf.cell(w2, 6, f'  {v}', border='RB', fill=True, ln=True)
                fill_kpi = not fill_kpi
            pdf.ln(6)

            # ── Tabela Comparativa de Vendedores ─────────────────────────
            pdf.set_fill_color(31, 71, 136)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(0, 7, '  COMPARATIVO DE VENDEDORES', fill=True, border=0, ln=True)

            _pv_cols_pdf    = ['Vendedor', 'Faturamento (R$)', 'Clientes', 'Ticket Médio', 'Vol.', 'Prazo (d)', 'Comissão']
            _pv_widths_pdf  = [50, 32, 16, 28, 15, 20, 22]

            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 7)
            for col, w in zip(_pv_cols_pdf, _pv_widths_pdf):
                pdf.cell(w, 7, col, border=1, fill=True, align='C')
            pdf.ln()

            _pv_comp_sorted = _pv_comp.sort_values('FaturamentoBruto', ascending=False).head(20)
            fill_row_pdf = False
            for _, row in _pv_comp_sorted.iterrows():
                pdf.set_fill_color(240, 244, 255) if fill_row_pdf else pdf.set_fill_color(255, 255, 255)
                pdf.set_text_color(50, 50, 50)
                pdf.set_font('Helvetica', '', 7)
                _vend_str = str(row['Vendedor'])[:22]
                _fat_str  = f"R$ {formatar_numero_br(row['FaturamentoBruto'], 2)}"
                _cli_str  = str(int(row['ClientesAtendidos']))
                _tick_str = f"R$ {formatar_numero_br(row['TicketMedio'], 2)}"
                _vol_str  = f"{formatar_numero_br(row.get('VolumeTotal', 0), 0)}"
                _prz_str  = f"{row.get('PrazoMedio', 0):.0f}"
                _com_str  = f"{row['ComissaoMedia']:.2f}%" if pd.notnull(row.get('ComissaoMedia')) else "N/D"
                _row_vals = [_vend_str, _fat_str, _cli_str, _tick_str, _vol_str, _prz_str, _com_str]
                _aligns   = ['L', 'R', 'C', 'R', 'C', 'C', 'C']
                for val, w, align in zip(_row_vals, _pv_widths_pdf, _aligns):
                    pdf.cell(w, 6, val, border=1, fill=True, align=align)
                pdf.ln()
                fill_row_pdf = not fill_row_pdf
            pdf.ln(6)

            # ── Top Produtos ──────────────────────────────────────────────
            pdf.set_fill_color(31, 71, 136)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(0, 7, '  MIX DE PRODUTOS — TOP 15', fill=True, border=0, ln=True)

            _pv_mix_pdf_cols = ['Produto', 'Faturamento (R$)', '% Part.', '% Acum.', 'Clientes']
            _pv_mix_pdf_w    = [70, 32, 18, 18, 18]
            pdf.set_font('Helvetica', 'B', 7)
            for col, w in zip(_pv_mix_pdf_cols, _pv_mix_pdf_w):
                pdf.cell(w, 7, col, border=1, fill=True, align='C')
            pdf.ln()

            fill_mix = False
            for _, row in _pv_mix.head(15).iterrows():
                pdf.set_fill_color(240, 244, 255) if fill_mix else pdf.set_fill_color(255, 255, 255)
                pdf.set_text_color(50, 50, 50)
                pdf.set_font('Helvetica', '', 7)
                _prod_str = str(row['NomeProduto'])[:38]
                _fat_str  = f"R$ {formatar_numero_br(row['Total'], 2)}"
                _part_str = f"{row['Participacao']:.2f}%"
                _acum_str = f"{row['CumulativaPerc']:.2f}%"
                _cli_str  = str(int(row['Clientes']))
                _mix_vals = [_prod_str, _fat_str, _part_str, _acum_str, _cli_str]
                _mix_al   = ['L', 'R', 'C', 'C', 'C']
                for val, w, al in zip(_mix_vals, _pv_mix_pdf_w, _mix_al):
                    pdf.cell(w, 6, val, border=1, fill=True, align=al)
                pdf.ln()
                fill_mix = not fill_mix
            pdf.ln(6)

            # ── Inadimplência resumida ────────────────────────────────────
            if _pv_df_inad is not None and _pv_inad_vendedor > 0:
                pdf.set_fill_color(239, 68, 68)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font('Helvetica', 'B', 9)
                pdf.cell(0, 7, '  INADIMPLÊNCIA', fill=True, border=0, ln=True)
                pdf.set_text_color(50, 50, 50)
                pdf.set_font('Helvetica', '', 8)
                pdf.set_fill_color(255, 240, 240)
                pdf.cell(95, 6, f'  Valor em Aberto: R$ {formatar_numero_br(_pv_inad_vendedor, 2)}', border='LB', fill=True)
                pdf.cell(90, 6, f'  % sobre Fat. Bruto: {_pv_perc_inad:.1f}%', border='RB', fill=True, ln=True)
                pdf.ln(4)

            # ── Rodapé do relatório ───────────────────────────────────────
            pdf.set_font('Helvetica', 'I', 7)
            pdf.set_text_color(173, 181, 189)
            pdf.multi_cell(0, 5,
                'Este relatório é gerado automaticamente com base nos dados filtrados do sistema BI Medtextil. '
                'As informações refletem o período e os filtros selecionados no momento da geração.')

            _pv_pdf_bytes = pdf.output()
            st.download_button(
                label="⬇️ Baixar PDF",
                data=bytes(_pv_pdf_bytes),
                file_name=f"performance_vendedores_{_pv_now.strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                key="pv_dl_pdf_btn"
            )
            st.success("✅ PDF gerado com sucesso! Clique em 'Baixar PDF' para salvar.")

        except ImportError:
            # Fallback ReportLab
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT

                _pv_buf = io.BytesIO()
                _pv_doc = SimpleDocTemplate(_pv_buf, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=20*mm, bottomMargin=15*mm)
                _pv_styles = getSampleStyleSheet()
                _azul = colors.HexColor('#1F4788')
                _elements_rl = []

                _st_titulo = ParagraphStyle('T', parent=_pv_styles['Heading1'], fontSize=14, textColor=_azul, alignment=TA_CENTER)
                _st_normal = ParagraphStyle('N', parent=_pv_styles['Normal'], fontSize=8)

                _elements_rl.append(Paragraph('RELATÓRIO DE PERFORMANCE DE VENDEDORES', _st_titulo))
                _elements_rl.append(Paragraph(f"Vendedor: {_pv_vendedor}  |  Região: {_pv_regiao}  |  Período: {_pv_periodo}", _st_normal))
                _elements_rl.append(Spacer(1, 5*mm))

                # KPI Table
                _kpi_data = [['Indicador', 'Valor']]
                for k, v in _pv_kpis_pdf:
                    _kpi_data.append([k, v])
                _kpi_table = Table(_kpi_data, colWidths=[100*mm, 80*mm])
                _kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0),(-1,0), _azul),
                    ('TEXTCOLOR', (0,0),(-1,0), colors.white),
                    ('FONTNAME', (0,0),(-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0),(-1,-1), 8),
                    ('ROWBACKGROUNDS', (0,1),(-1,-1), [colors.white, colors.HexColor('#F0F4FF')]),
                    ('BOX', (0,0),(-1,-1), 0.5, colors.grey),
                    ('INNERGRID', (0,0),(-1,-1), 0.3, colors.lightgrey),
                    ('LEFTPADDING', (0,0),(-1,-1), 5),
                ]))
                _elements_rl.append(_kpi_table)
                _pv_doc.build(_elements_rl)
                _pv_buf.seek(0)

                st.download_button(
                    "⬇️ Baixar PDF",
                    data=_pv_buf.getvalue(),
                    file_name=f"performance_vendedores_{_pv_now.strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf",
                    key="pv_dl_pdf_rl"
                )
                st.success("✅ PDF gerado com sucesso!")
            except Exception as _e_rl:
                st.error(f"❌ Erro ao gerar PDF: {_e_rl}")
        except Exception as _e_pdf:
            st.error(f"❌ Erro ao gerar PDF: {_e_pdf}")
            st.info("💡 Verifique se a biblioteca fpdf2 está instalada: pip install fpdf2")



    # ── Cards de Resultado por Vendedor ──────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🏅 Resultado por Vendedor")

    _meses_pt = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
                 7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

    # Filtros: período e vendedor
    _cd_c1, _cd_c2, _cd_c3 = st.columns(3)
    _pv_now2 = pd.Timestamp.now()
    with _cd_c1:
        _cd_mes = st.selectbox("Mês de referência",
                               options=list(range(1,13)),
                               index=(_pv_now2.month - 2) % 12,
                               format_func=lambda m: _meses_pt[m],
                               key="cd_mes_sel")
    with _cd_c2:
        _cd_ano = st.number_input("Ano", min_value=2020,
                                  max_value=_pv_now2.year,
                                  value=_pv_now2.year if _pv_now2.month > 1 else _pv_now2.year - 1,
                                  step=1, key="cd_ano_sel")
    with _cd_c3:
        _cd_vend_opts = ['Todos'] + sorted(notas_unicas['Vendedor'].dropna().unique().tolist())
        _cd_vend_sel  = st.selectbox("Vendedor", _cd_vend_opts, key="cd_vend_sel")

    _mes_card = int(_cd_mes)
    _ano_card = int(_cd_ano)
    _mes_meta = _mes_card
    _ano_meta = _ano_card - 1
    _label_mes_card = f"{_meses_pt[_mes_card]}/{_ano_card}"


    # Base para faturamento: notas_unicas com Valor_Real (igual ao dashboard)
    _nu_hist = notas_unicas.copy()
    _nu_hist["DataEmissao"] = pd.to_datetime(_nu_hist["DataEmissao"], errors="coerce")

    # Base para positivação/clientes: df completo (tem CPF_CNPJ por item), só NF Venda
    _df_nf_hist_raw = df[df["TipoMov"] == "NF Venda"].copy()
    _df_nf_hist_raw["DataEmissao"] = pd.to_datetime(_df_nf_hist_raw["DataEmissao"], errors="coerce")
    _df_nf_hist = _df_nf_hist_raw

    # Notas do mês de referência (todas as NF para Valor_Real)
    _nu_mes = _nu_hist[
        (_nu_hist["DataEmissao"].dt.month == _mes_card) &
        (_nu_hist["DataEmissao"].dt.year  == _ano_card)
    ]

    # _df_mes_card: NF Venda do mês no df completo (para positivação e reativados)
    _df_mes_card = _df_nf_hist_raw[
        (_df_nf_hist_raw["DataEmissao"].dt.month == _mes_card) &
        (_df_nf_hist_raw["DataEmissao"].dt.year  == _ano_card)
    ]

    # Vendedores ativos = quem tem NF Venda no mês de referência
    _vends_ativos = sorted(_df_mes_card["Vendedor"].dropna().unique().tolist())

    # Filtro do card (substituí _pv_vendedor pelo seletor próprio dos cards)
    if _cd_vend_sel != "Todos":
        _vends_ativos = [v for v in _vends_ativos if v == _cd_vend_sel]

    if not _vends_ativos:
        st.info(f"Nenhum vendedor com vendas em {_label_mes_card}.")
    else:
        # Faturamento líquido via Valor_Real (notas_unicas) — igual ao dashboard
        _fat_card = _nu_mes.groupby("Vendedor")["Valor_Real"].sum()

        # Positivação: clientes únicos com NF Venda no mês (df completo, tem CPF_CNPJ)
        _posit_card = _df_mes_card.groupby("Vendedor")["CPF_CNPJ"].nunique()
        _base_hist  = _df_nf_hist_raw.groupby("Vendedor")["CPF_CNPJ"].nunique()

        # Cálculo de meta por vendedor — usa NFs deduplicadas para soma correta
        def _meta_card(vendedor):
            # Usar notas_unicas com Valor_Real — igual ao dashboard
            _fat_ano_ant = float(_nu_hist[
                (_nu_hist["Vendedor"] == vendedor) &
                (_nu_hist["DataEmissao"].dt.month == _mes_meta) &
                (_nu_hist["DataEmissao"].dt.year  == _ano_meta)
            ]["Valor_Real"].sum())

            if _fat_ano_ant > 0:
                return _fat_ano_ant * 1.15, f"{_meses_pt[_mes_meta][:3]}/{_ano_meta} +15%", _fat_ano_ant

            _ref_ts = pd.Timestamp(year=_ano_card, month=_mes_card, day=1)
            _3m_ini = _ref_ts - pd.DateOffset(months=3)
            _ult3 = _nu_hist[
                (_nu_hist["Vendedor"] == vendedor) &
                (_nu_hist["DataEmissao"] >= _3m_ini) &
                (_nu_hist["DataEmissao"] <  _ref_ts)
            ]
            if len(_ult3) > 0:
                _fat_3m = _ult3.groupby(
                    [_ult3["DataEmissao"].dt.year, _ult3["DataEmissao"].dt.month]
                )["Valor_Real"].sum()
                return _fat_3m.mean() * 1.15, "Média 3m +15%", 0
            return 0, "Sem histórico", 0

        # ── Pré-calcular indicadores extras para todos os vendedores ativos ──

        _mes_ant_c = _mes_card - 1 if _mes_card > 1 else 12
        _ano_ant_c = _ano_card if _mes_card > 1 else _ano_card - 1

        # Faturamento mês anterior via Valor_Real (igual dashboard)
        _fat_mes_ant_c = _nu_hist[
            (_nu_hist["DataEmissao"].dt.month == _mes_ant_c) &
            (_nu_hist["DataEmissao"].dt.year  == _ano_ant_c)
        ].groupby("Vendedor")["Valor_Real"].sum()

        # Faturamento mesmo mês ano anterior via Valor_Real
        _fat_ano_ant_c = _nu_hist[
            (_nu_hist["DataEmissao"].dt.month == _mes_card) &
            (_nu_hist["DataEmissao"].dt.year  == _ano_card - 1)
        ].groupby("Vendedor")["Valor_Real"].sum()

        # Clientes reativados: compraram no mês de referência MAS não compraram
        # nos 3 meses anteriores (= inativos por 3+ meses que voltaram)
        _ref_ts_c   = pd.Timestamp(year=_ano_card, month=_mes_card, day=1)
        _3m_antes_c = _ref_ts_c - pd.DateOffset(months=3)

        # CPFs que compraram no mês de referência por vendedor
        _cpfs_mes_ref = (
            _df_nf_hist_raw[
                (_df_nf_hist_raw["DataEmissao"].dt.month == _mes_card) &
                (_df_nf_hist_raw["DataEmissao"].dt.year  == _ano_card)
            ].groupby("Vendedor")["CPF_CNPJ"].apply(set)
        )

        # CPFs que compraram nos 3 meses anteriores por vendedor
        _cpfs_3m = (
            _df_nf_hist_raw[
                (_df_nf_hist_raw["DataEmissao"] >= _3m_antes_c) &
                (_df_nf_hist_raw["DataEmissao"] <  _ref_ts_c)
            ].groupby("Vendedor")["CPF_CNPJ"].apply(set)
        )

        def _reativados_vend(vendedor):
            _ativos_mes = _cpfs_mes_ref.get(vendedor, set())
            _ativos_3m  = _cpfs_3m.get(vendedor, set())
            return len(_ativos_mes - _ativos_3m)

        # Renderizar cards — 3 por linha
        _n_cols   = min(3, len(_vends_ativos))
        _rows_v   = [_vends_ativos[i:i+_n_cols] for i in range(0, len(_vends_ativos), _n_cols)]
        st.caption(f"Referência: **{_label_mes_card}** — vendedores que positivaram neste mês")

        for _rv in _rows_v:
            _ccols = st.columns(len(_rv))
            for _ci2, _vend in zip(_ccols, _rv):
                _fat_r   = float(_fat_card.get(_vend, 0))
                _meta_v, _meta_lbl, _base_meta = _meta_card(_vend)
                _perc_m  = (_fat_r / _meta_v * 100) if _meta_v > 0 else 0
                _posit_v = int(_posit_card.get(_vend, 0))
                _base_v  = int(_base_hist.get(_vend, 0))
                _posit_p = (_posit_v / _base_v * 100) if _base_v > 0 else 0

                # Crescimento mensal vs mês anterior
                _fat_ant_m  = float(_fat_mes_ant_c.get(_vend, 0))
                _cresc_m    = ((_fat_r - _fat_ant_m) / _fat_ant_m * 100) if _fat_ant_m > 0 else None
                _cresc_m_str = f"{_cresc_m:+.1f}%" if _cresc_m is not None else "—"
                _cresc_m_cor = "#28A745" if (_cresc_m or 0) >= 0 else "#EF4444"
                _cresc_m_ico = "📈" if (_cresc_m or 0) >= 0 else "📉"
                _mes_ant_label = f"{_meses_pt[_mes_ant_c][:3]}/{_ano_ant_c}"

                # Crescimento anual vs mesmo mês ano anterior
                _fat_ant_a  = float(_fat_ano_ant_c.get(_vend, 0))
                _cresc_a    = ((_fat_r - _fat_ant_a) / _fat_ant_a * 100) if _fat_ant_a > 0 else None
                _cresc_a_str = f"{_cresc_a:+.1f}%" if _cresc_a is not None else "—"
                _cresc_a_cor = "#28A745" if (_cresc_a or 0) >= 0 else "#EF4444"
                _cresc_a_ico = "📈" if (_cresc_a or 0) >= 0 else "📉"
                _mesmo_mes_ant = f"{_meses_pt[_mes_card][:3]}/{_ano_card - 1}"

                # Clientes reativados
                _reat_v = _reativados_vend(_vend)

                # Contratado x Faturado (mês de referência)
                _contrato_v2 = 0.0
                if _pv_df_contrato is not None:
                    try:
                        _ctr_mes_v2 = _pv_df_contrato.copy()
                        if _pv_col_data_contrato and _pv_col_data_contrato in _ctr_mes_v2.columns:
                            _ctr_mes_v2[_pv_col_data_contrato] = pd.to_datetime(
                                _ctr_mes_v2[_pv_col_data_contrato], errors='coerce'
                            )
                            _ctr_mes_v2 = _ctr_mes_v2[
                                (_ctr_mes_v2[_pv_col_data_contrato].dt.month == _mes_card) &
                                (_ctr_mes_v2[_pv_col_data_contrato].dt.year == _ano_card)
                            ]
                        _contrato_v2 = _ctr_mes_v2[
                            _ctr_mes_v2['_FuncNorm'] == str(_vend).strip().upper()
                        ]['_ValorContrato'].sum()
                    except Exception:
                        _contrato_v2 = 0.0
                _perc_real_ctr2 = (_fat_r / _contrato_v2 * 100) if _contrato_v2 > 0 else 0
                _real_cor2 = "#28A745" if _perc_real_ctr2 >= 100 else ("#F4A261" if _perc_real_ctr2 >= 70 else "#EF4444")

                _cor      = "#28A745" if _perc_m >= 100 else ("#F4A261" if _perc_m >= 70 else "#EF4444")
                _barra    = min(int(_perc_m), 100)
                _sinal    = "✅" if _perc_m >= 100 else ("⚠️" if _perc_m >= 70 else "🔴")
                _posit_cor = "#28A745" if _posit_p >= 60 else ("#F4A261" if _posit_p >= 40 else "#EF4444")

                _html = (
                    f'<div style="background:#fff;border:1.5px solid #E2E8F0;border-radius:14px;'
                    f'padding:16px 14px 12px;box-shadow:0 2px 10px rgba(31,71,136,0.09);margin-bottom:14px;">'

                    # Header
                    f'<div style="font-size:0.95rem;font-weight:700;color:#1F4788;border-bottom:2px solid #EEF3FC;'
                    f'padding-bottom:7px;margin-bottom:10px;">👤 {_vend}</div>'
                    f'<div style="font-size:0.7rem;color:#6C757D;margin-bottom:8px;text-align:center;">{_label_mes_card}</div>'

                    # Faturamento
                    f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px;">'
                    f'<span style="color:#6C757D;font-size:0.78rem;">💰 Faturamento</span>'
                    f'<span style="font-weight:700;color:#163561;font-size:0.95rem;">R$ {formatar_numero_br(_fat_r, 0)}</span></div>'

                    # Meta
                    f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:3px;">'
                    f'<span style="color:#6C757D;font-size:0.78rem;">🎯 Meta ({_meta_lbl})</span>'
                    f'<span style="font-weight:600;color:#4A7BC8;font-size:0.85rem;">R$ {formatar_numero_br(_meta_v, 0)}</span></div>'

                    # Barra progresso meta
                    f'<div style="background:#F1F5F9;border-radius:6px;height:7px;margin:7px 0 3px;">'
                    f'<div style="background:{_cor};width:{_barra}%;height:7px;border-radius:6px;"></div></div>'
                    f'<div style="text-align:right;font-size:0.8rem;font-weight:700;color:{_cor};margin-bottom:10px;">'
                    f'{_sinal} {_perc_m:.1f}% da meta</div>'

                    # Crescimentos
                    f'<div style="display:flex;gap:6px;margin-bottom:10px;">'

                    f'<div style="flex:1;background:#F8FAFF;border-radius:8px;padding:6px 4px;border:1px solid #EEF3FC;text-align:center;">'
                    f'<div style="font-size:0.65rem;color:#6C757D;margin-bottom:3px;font-weight:600;">vs {_mes_ant_label}</div>'
                    f'<div style="font-size:0.88rem;font-weight:700;color:{_cresc_m_cor};">{_cresc_m_ico} {_cresc_m_str}</div>'
                    f'<div style="font-size:0.62rem;color:#6C757D;">Cresc. mensal</div></div>'

                    f'<div style="flex:1;background:#F8FAFF;border-radius:8px;padding:6px 4px;border:1px solid #EEF3FC;text-align:center;">'
                    f'<div style="font-size:0.65rem;color:#6C757D;margin-bottom:3px;font-weight:600;">vs {_mesmo_mes_ant}</div>'
                    f'<div style="font-size:0.88rem;font-weight:700;color:{_cresc_a_cor};">{_cresc_a_ico} {_cresc_a_str}</div>'
                    f'<div style="font-size:0.62rem;color:#6C757D;">Cresc. anual</div></div>'

                    f'<div style="flex:1;background:#F8FAFF;border-radius:8px;padding:6px 4px;border:1px solid #EEF3FC;text-align:center;">'
                    f'<div style="font-size:0.65rem;color:#6C757D;margin-bottom:3px;font-weight:600;">Reativados</div>'
                    f'<div style="font-size:0.88rem;font-weight:700;color:#9B59B6;">🔄 {_reat_v}</div>'
                    f'<div style="font-size:0.62rem;color:#6C757D;">3m+ sem compra</div></div>'

                    f'</div>'

                    # Positivação
                    f'<div style="background:#F8FAFF;border-radius:8px;padding:8px 6px;border:1px solid #EEF3FC;">'
                    f'<div style="font-size:0.72rem;color:#6C757D;text-align:center;margin-bottom:6px;font-weight:600;">POSITIVAÇÃO {_label_mes_card}</div>'
                    f'<div style="display:flex;justify-content:space-around;align-items:center;">'

                    f'<div style="text-align:center;">'
                    f'<div style="font-size:1.15rem;font-weight:700;color:#28A745;">{_posit_v}</div>'
                    f'<div style="font-size:0.68rem;color:#6C757D;">Positivados</div></div>'

                    f'<div style="font-size:1.2rem;color:#CDD4E0;">|</div>'

                    f'<div style="text-align:center;">'
                    f'<div style="font-size:1.15rem;font-weight:700;color:#1F4788;">{_base_v}</div>'
                    f'<div style="font-size:0.68rem;color:#6C757D;">Base total</div></div>'

                    f'<div style="font-size:1.2rem;color:#CDD4E0;">|</div>'

                    f'<div style="text-align:center;">'
                    f'<div style="font-size:1.15rem;font-weight:700;color:{_posit_cor};">{_posit_p:.0f}%</div>'
                    f'<div style="font-size:0.68rem;color:#6C757D;">% posit.</div></div>'

                    f'</div></div>'

                    # Contratado x Faturado
                    f'<div style="background:#F8FAFF;border-radius:8px;padding:8px 6px;border:1px solid #EEF3FC;margin-top:8px;">'
                    f'<div style="font-size:0.72rem;color:#6C757D;text-align:center;margin-bottom:6px;font-weight:600;">CONTRATO x FATURADO {_label_mes_card}</div>'
                    f'<div style="display:flex;justify-content:space-around;align-items:center;">'

                    f'<div style="text-align:center;">'
                    f'<div style="font-size:1.0rem;font-weight:700;color:#1F4788;">R$ {formatar_numero_br(_contrato_v2, 0)}</div>'
                    f'<div style="font-size:0.68rem;color:#6C757D;">Contratado</div></div>'

                    f'<div style="font-size:1.2rem;color:#CDD4E0;">|</div>'

                    f'<div style="text-align:center;">'
                    f'<div style="font-size:1.15rem;font-weight:700;color:{_real_cor2};">{_perc_real_ctr2:.1f}%</div>'
                    f'<div style="font-size:0.68rem;color:#6C757D;">% Realização</div></div>'

                    f'</div></div>'
                    f'</div>'
                )
                with _ci2:
                    st.markdown(_html, unsafe_allow_html=True)

    st.markdown("---")

    # ── Download relatorio detalhado por vendedor ─────────────────────────────
    st.markdown("### 📥 Relatório Detalhado por Vendedor")
    _mes_nome_det = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
                     7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

    # ── Geração de imagem PNG por vendedor (estilo post Medtextil) ────────────
    _col_btn1, _col_btn2 = st.columns(2)
    with _col_btn1:
        _vend_img_opts = ["Todos"] + _vends_ativos
        _vend_img_sel  = st.selectbox("Vendedor para imagem:", _vend_img_opts, key="sel_vend_img")
    with _col_btn2:
        _foco1 = st.text_input("Foco 01", placeholder="Ex: Reativação de clientes", key="foco1_img")
        _foco2 = st.text_input("Foco 02", placeholder="Ex: Ampliar mix hospitalar", key="foco2_img")
        _foco3 = st.text_input("Foco 03", placeholder="Ex: Recuperar orçamentos", key="foco3_img")

    if st.button("🖼️ Gerar Card(s) como Imagem PNG", key="btn_gerar_png_card", type="primary"):
        try:
            from PIL import Image as _PilImg, ImageDraw as _PilDraw, ImageFont as _PilFont
            import requests as _req_img
            import io as _io_img

            # Baixar logo do GitHub
            _logo_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{GITHUB_FOLDER}/logo.png"
            try:
                _logo_resp = _req_img.get(_logo_url, timeout=5)
                _logo_pil  = _PilImg.open(_io_img.BytesIO(_logo_resp.content)).convert("RGBA")
            except Exception:
                _logo_pil = None

            # Paleta Medtextil
            _C_NAVY  = (13,  35,  75)   # #0D234B
            _C_GREEN = (39, 174,  96)   # #27AE60
            _C_WHITE = (255,255,255)
            _C_BG    = (240,246,255)    # fundo azul claro
            _C_LGRAY = (220,230,245)
            _C_DARK  = (10, 25, 60)

            # Tamanho: 1080×1080 (quadrado Instagram/WhatsApp)
            _W, _H = 1080, 1080

            def _focos_automaticos(vendedor, cresc_m, cresc_a, posit_p):
                """Gera 3 focos automáticos baseados nos pontos fracos do vendedor."""
                _candidatos = []
                # Crescimento mensal fraco → reativação
                if cresc_m < 5:
                    _candidatos.append(("cresc_m", cresc_m, "Reativacao de clientes inativos"))
                # Positivação fraca → novos clientes
                if posit_p < 50:
                    _candidatos.append(("posit", posit_p, "Ampliar base de clientes ativos"))
                # Crescimento anual fraco → mix de produtos
                if cresc_a < 10:
                    _candidatos.append(("cresc_a", cresc_a, "Ampliar mix de produtos"))
                # Sempre adicionar como fallback
                _candidatos += [
                    ("pad1", 99, "Manter ritmo de positivacao"),
                    ("pad2", 99, "Recuperar orcamentos pendentes"),
                    ("pad3", 99, "Fortalecer relacionamento com clientes"),
                ]
                # Ordenar pelos mais fracos primeiro
                _candidatos.sort(key=lambda x: x[1])
                return [c[2] for c in _candidatos[:3]]

            def _meta_proximo_mes(vendedor):
                """Meta do mês SEGUINTE ao de referência (mesmo algoritmo dos cards)."""
                _mes_prox = _mes_card + 1 if _mes_card < 12 else 1
                _ano_prox = _ano_card if _mes_card < 12 else _ano_card + 1
                _ano_prox_ant = _ano_prox - 1
                # Historico do mesmo mes do ano anterior ao proximo
                _hist = float(_df_nf_hist[
                    (_df_nf_hist["Vendedor"] == vendedor) &
                    (_df_nf_hist["DataEmissao"].dt.month == _mes_prox) &
                    (_df_nf_hist["DataEmissao"].dt.year  == _ano_prox_ant)
                ]["TotalProduto"].sum())
                if _hist > 0:
                    return _hist * 1.15, f"{_meses_pt[_mes_prox][:3]}/{_ano_prox_ant} +15%"
                # Sem historico: media dos 3 meses anteriores ao proximo mes
                _ref2   = pd.Timestamp(year=_ano_prox, month=_mes_prox, day=1)
                _3m2    = _ref2 - pd.DateOffset(months=3)
                _ult3b  = _df_nf_hist[
                    (_df_nf_hist["Vendedor"] == vendedor) &
                    (_df_nf_hist["DataEmissao"] >= _3m2) &
                    (_df_nf_hist["DataEmissao"] <  _ref2)
                ]
                if len(_ult3b) > 0:
                    _fat3b = _ult3b.groupby(
                        [_ult3b["DataEmissao"].dt.year, _ult3b["DataEmissao"].dt.month]
                    )["TotalProduto"].sum()
                    return _fat3b.mean() * 1.15, "Media 3m +15%"
                return 0, "Sem historico"

            def _draw_card_img(vendedor):
                _fat_r   = float(_fat_card.get(vendedor, 0))
                _meta_v, _meta_lbl, _bm = _meta_card(vendedor)
                _perc_m  = (_fat_r / _meta_v * 100) if _meta_v > 0 else 0
                _posit_v = int(_posit_card.get(vendedor, 0))
                _base_v  = int(_base_hist.get(vendedor, 0))
                _posit_p = (_posit_v / _base_v * 100) if _base_v > 0 else 0
                _fat_am  = float(_fat_mes_ant_c.get(vendedor, 0))
                _fat_aa  = float(_fat_ano_ant_c.get(vendedor, 0))
                _cresc_m = ((_fat_r - _fat_am) / _fat_am * 100) if _fat_am > 0 else 0
                _cresc_a = ((_fat_r - _fat_aa) / _fat_aa * 100) if _fat_aa > 0 else 0
                _reat_v  = _reativados_vend(vendedor)

                # Contratado x Faturado (mês de referência)
                _contrato_v = 0.0
                if _pv_df_contrato is not None:
                    try:
                        _ctr_mes_v = _pv_df_contrato.copy()
                        if _pv_col_data_contrato and _pv_col_data_contrato in _ctr_mes_v.columns:
                            _ctr_mes_v[_pv_col_data_contrato] = pd.to_datetime(
                                _ctr_mes_v[_pv_col_data_contrato], errors='coerce'
                            )
                            _ctr_mes_v = _ctr_mes_v[
                                (_ctr_mes_v[_pv_col_data_contrato].dt.month == _mes_card) &
                                (_ctr_mes_v[_pv_col_data_contrato].dt.year == _ano_card)
                            ]
                        _contrato_v = _ctr_mes_v[
                            _ctr_mes_v['_FuncNorm'] == str(vendedor).strip().upper()
                        ]['_ValorContrato'].sum()
                    except Exception:
                        _contrato_v = 0.0
                _perc_real_ctr = (_fat_r / _contrato_v * 100) if _contrato_v > 0 else 0

                # Meta próximo mês
                _meta_prox, _meta_prox_lbl = _meta_proximo_mes(vendedor)
                _mes_prox_nm = _meses_pt[(_mes_card % 12) + 1]
                _ano_prox_nm = _ano_card if _mes_card < 12 else _ano_card + 1

                # Focos automáticos
                _focos_auto = _focos_automaticos(vendedor, _cresc_m, _cresc_a, _posit_p)
                _f1, _f2, _f3 = _focos_auto[0], _focos_auto[1], _focos_auto[2]

                # Cores crescimento
                def _cc(v): return "#27AE60" if v >= 0 else "#E74C3C"
                def _cs(v): return f"{v:+.1f}%"

                # Cor barra meta
                _meta_cor = "#27AE60" if _perc_m >= 100 else ("#F39C12" if _perc_m >= 70 else "#E74C3C")
                _barra_w  = min(int(_perc_m), 100)

                # Logo base64
                _logo_b64 = ""
                try:
                    import base64 as _b64
                    import requests as _req2
                    _lr2 = _req2.get(f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{GITHUB_FOLDER}/logo.png", timeout=5)
                    _logo_b64 = _b64.b64encode(_lr2.content).decode()
                    _logo_tag = f'<img src="data:image/png;base64,{_logo_b64}" style="height:90px;margin-bottom:4px;">'
                except Exception:
                    _logo_tag = '<div style="font-size:48px;font-weight:900;color:#0D234B;">MEDTEXTIL</div>'

                _html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;900&display=swap');
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{
    width:1080px; background:#EAF1FB;
    font-family: 'Montserrat', 'DejaVu Sans', Arial, sans-serif;
  }}
  .wrap {{ width:1080px; background:#EAF1FB; }}

  /* HEADER */
  .header {{
    background:#fff; text-align:center;
    padding:28px 40px 18px; border-bottom:5px solid #27AE60;
  }}
  .header .titulo {{ font-size:96px; font-weight:900; color:#0D234B; line-height:1.0; letter-spacing:2px; }}
  .header .mensal  {{ font-size:96px; font-weight:900; color:#27AE60; line-height:1.0; letter-spacing:2px; }}
  .header .mesref  {{ font-size:32px; color:#6C8EBD; margin-top:8px; font-weight:600; }}

  /* ROW VENDEDOR/MES */
  .row-info {{ display:flex; gap:16px; padding:20px 40px 0; }}
  .info-box {{
    flex:1; background:#0D234B; border-radius:18px;
    display:flex; align-items:center; gap:16px; padding:18px 22px;
  }}
  .info-icon {{
    width:64px; height:64px; border-radius:50%; background:#27AE60;
    display:flex; align-items:center; justify-content:center;
    font-size:32px; color:#fff; font-weight:900; flex-shrink:0;
  }}
  .info-label {{ font-size:22px; color:#27AE60; font-weight:700; }}
  .info-val   {{ font-size:30px; color:#fff; font-weight:800; margin-top:2px; }}

  /* GRID DE CARDS */
  .grid {{ display:flex; gap:16px; padding:16px 40px 0; }}
  .card {{
    flex:1; background:#fff; border-radius:22px;
    overflow:hidden; border:2px solid #D0DFF5;
  }}
  .card-top {{
    background:#0D234B; padding:18px 16px;
    display:flex; align-items:center; gap:14px;
  }}
  .card-icon {{
    width:60px; height:60px; border-radius:50%; background:#27AE60;
    display:flex; align-items:center; justify-content:center;
    font-size:30px; color:#fff; font-weight:900; flex-shrink:0;
  }}
  .card-label {{ font-size:26px; font-weight:800; color:#fff; line-height:1.2; }}
  .card-body  {{ padding:20px 20px 14px; }}
  .card-val   {{ font-size:56px; font-weight:900; color:#0D234B; line-height:1.0; }}
  .card-sub   {{ font-size:22px; color:#7A90B0; margin-top:6px; font-weight:600; }}
  .card-dot   {{
    width:16px; height:16px; border-radius:50%; background:#27AE60;
    margin: 8px 0 4px 20px;
  }}
  .green  {{ color:#27AE60 !important; }}
  .red    {{ color:#E74C3C !important; }}
  .orange {{ color:#F39C12 !important; }}

  /* BARRA META */
  .meta-bar-wrap {{ padding:4px 20px 0; }}
  .meta-bar-bg {{ background:#E8EFF8; border-radius:6px; height:10px; }}
  .meta-bar-fg {{ background:{_meta_cor}; border-radius:6px; height:10px; width:{_barra_w}%; }}

  /* META PROXIMO MES */
  .prox-box {{
    margin:16px 40px 0;
    background:#0D234B; border-radius:18px; padding:22px 28px;
    display:flex; align-items:center; gap:24px;
  }}
  .prox-label {{ font-size:26px; color:#27AE60; font-weight:800; }}
  .prox-val   {{ font-size:44px; color:#fff; font-weight:900; }}
  .prox-sub   {{ font-size:20px; color:#7AAED6; margin-top:2px; }}

  /* FOCO */
  .foco-box {{
    margin:16px 40px 0;
    background:#0B1C45; border-radius:18px; padding:20px 24px;
    display:flex; align-items:center; gap:20px;
  }}
  .foco-title-wrap {{ min-width:170px; }}
  .foco-title {{ font-size:22px; color:#27AE60; font-weight:800; line-height:1.2; }}
  .foco-big   {{ font-size:32px; color:#fff; font-weight:900; line-height:1.2; }}
  .foco-sep   {{ width:3px; background:#27AE60; align-self:stretch; border-radius:2px; }}
  .foco-items {{ display:flex; gap:18px; flex:1; }}
  .foco-item  {{ display:flex; align-items:center; gap:10px; flex:1; }}
  .foco-num   {{
    width:44px; height:44px; border-radius:50%; background:#27AE60;
    display:flex; align-items:center; justify-content:center;
    font-size:24px; font-weight:900; color:#fff; flex-shrink:0;
  }}
  .foco-txt   {{ font-size:21px; color:#fff; font-weight:600; line-height:1.2; }}

  /* RODAPE */
  .rodape {{
    margin:16px 0 0; background:#0B1C45;
    padding:20px 40px; display:flex; align-items:center; justify-content:space-between;
  }}
  .rodape-txt {{ font-size:24px; color:#fff; font-weight:600; }}
  .rodape-txt span {{ color:#27AE60; font-weight:800; }}
</style>
</head><body><div class="wrap">

  <!-- HEADER -->
  <div class="header">
    {_logo_tag}
    <div class="titulo">RESULTADO</div>
    <div class="mensal">MENSAL</div>
    <div class="mesref">{_label_mes_card.upper()}</div>
  </div>

  <!-- VENDEDOR / MÊS -->
  <div class="row-info">
    <div class="info-box">
      <div class="info-icon">V</div>
      <div>
        <div class="info-label">VENDEDOR</div>
        <div class="info-val">{vendedor}</div>
      </div>
    </div>
    <div class="info-box">
      <div class="info-icon">M</div>
      <div>
        <div class="info-label">MÊS DE REFERÊNCIA</div>
        <div class="info-val">{_label_mes_card.upper()}</div>
      </div>
    </div>
  </div>

  <!-- ROW 1: Faturamento | Meta | Cresc. Mensal -->
  <div class="grid">
    <div class="card">
      <div class="card-top">
        <div class="card-icon">$</div>
        <div class="card-label">FATURAMENTO</div>
      </div>
      <div class="card-body">
        <div class="card-val">R$ {formatar_numero_br(_fat_r, 0)}</div>
        <div class="card-sub">Meta: R$ {formatar_numero_br(_meta_v, 0)}</div>
        <div class="meta-bar-wrap">
          <div class="meta-bar-bg"><div class="meta-bar-fg"></div></div>
        </div>
      </div>
      <div class="card-dot"></div>
    </div>
    <div class="card">
      <div class="card-top">
        <div class="card-icon">%</div>
        <div class="card-label">META ATINGIDA</div>
      </div>
      <div class="card-body">
        <div class="card-val {'green' if _perc_m>=100 else ('orange' if _perc_m>=70 else 'red')}">{_perc_m:.1f}%</div>
        <div class="card-sub">{_meta_lbl}</div>
      </div>
      <div class="card-dot"></div>
    </div>
    <div class="card">
      <div class="card-top">
        <div class="card-icon">{'↑' if _cresc_m>=0 else '↓'}</div>
        <div class="card-label">CRESC. MENSAL</div>
      </div>
      <div class="card-body">
        <div class="card-val {'green' if _cresc_m>=0 else 'red'}">{_cs(_cresc_m)}</div>
        <div class="card-sub">vs {_meses_pt[_mes_ant_c][:3]}/{_ano_ant_c}</div>
      </div>
      <div class="card-dot"></div>
    </div>
  </div>

  <!-- ROW 2: Cresc. Anual | Positivados | Reativados -->
  <div class="grid">
    <div class="card">
      <div class="card-top">
        <div class="card-icon">A</div>
        <div class="card-label">CRESC. ANUAL</div>
      </div>
      <div class="card-body">
        <div class="card-val {'green' if _cresc_a>=0 else 'red'}">{_cs(_cresc_a)}</div>
        <div class="card-sub">vs {_meses_pt[_mes_card][:3]}/{_ano_card-1}</div>
      </div>
      <div class="card-dot"></div>
    </div>
    <div class="card">
      <div class="card-top">
        <div class="card-icon">C</div>
        <div class="card-label">CLI. POSITIVADOS</div>
      </div>
      <div class="card-body">
        <div class="card-val">{_posit_v}<span style="font-size:32px;color:#7A90B0;">/{_base_v}</span></div>
        <div class="card-sub">{_posit_p:.0f}% da base ativa</div>
      </div>
      <div class="card-dot"></div>
    </div>
    <div class="card">
      <div class="card-top">
        <div class="card-icon">R</div>
        <div class="card-label">CLI. REATIVADOS</div>
      </div>
      <div class="card-body">
        <div class="card-val">{_reat_v}</div>
        <div class="card-sub">3+ meses sem compra</div>
      </div>
      <div class="card-dot"></div>
    </div>
  </div>

  <!-- ROW 3: Contratado | % Realização -->
  <div class="grid">
    <div class="card">
      <div class="card-top">
        <div class="card-icon">$</div>
        <div class="card-label">VALOR CONTRATADO</div>
      </div>
      <div class="card-body">
        <div class="card-val">R$ {formatar_numero_br(_contrato_v, 0)}</div>
        <div class="card-sub">no mês de referência</div>
      </div>
      <div class="card-dot"></div>
    </div>
    <div class="card">
      <div class="card-top">
        <div class="card-icon">%</div>
        <div class="card-label">% REALIZAÇÃO</div>
      </div>
      <div class="card-body">
        <div class="card-val {'green' if _perc_real_ctr>=100 else ('orange' if _perc_real_ctr>=70 else 'red')}">{_perc_real_ctr:.1f}%</div>
        <div class="card-sub">Faturado / Contratado</div>
      </div>
      <div class="card-dot"></div>
    </div>
  </div>

  <!-- META PRÓXIMO MÊS -->
  <div class="prox-box">
    <div class="card-icon" style="width:70px;height:70px;font-size:34px;">M</div>
    <div>
      <div class="prox-label">META {_mes_prox_nm.upper()}/{_ano_prox_nm}</div>
      <div class="prox-val">R$ {formatar_numero_br(_meta_prox, 0)}</div>
      <div class="prox-sub">{_meta_prox_lbl}</div>
    </div>
  </div>

  <!-- FOCO DO PRÓXIMO MÊS -->
  <div class="foco-box">
    <div class="foco-title-wrap">
      <div class="foco-title">FOCO DO</div>
      <div class="foco-big">PRÓXIMO MÊS</div>
    </div>
    <div class="foco-sep"></div>
    <div class="foco-items">
      <div class="foco-item">
        <div class="foco-num">1</div>
        <div class="foco-txt">{_f1}</div>
      </div>
      <div class="foco-item">
        <div class="foco-num">2</div>
        <div class="foco-txt">{_f2}</div>
      </div>
      <div class="foco-item">
        <div class="foco-num">3</div>
        <div class="foco-txt">{_f3}</div>
      </div>
    </div>
  </div>

  <!-- RODAPÉ -->
  <div class="rodape">
    <div class="rodape-txt">QUALIDADE QUE <span>PROTEGE</span>,<br>CONFIANÇA QUE <span>TRANSFORMA</span>.</div>
    <div style="font-size:40px;">🛡️</div>
  </div>

</div></body></html>"""

                # Gerar PNG via wkhtmltoimage
                import subprocess as _sp, tempfile as _tf, os as _os2
                import io as _io_img

                with _tf.NamedTemporaryFile(suffix=".html", delete=False, mode="w", encoding="utf-8") as _hf:
                    _hf.write(_html)
                    _hf_path = _hf.name

                _png_path = _hf_path.replace(".html", ".png")
                _sp.run([
                    "wkhtmltoimage",
                    "--width", "1080",
                    "--quality", "95",
                    "--zoom", "1.0",
                    "--disable-smart-width",
                    "--quiet",
                    _hf_path, _png_path
                ], capture_output=True)

                _os2.unlink(_hf_path)

                with open(_png_path, "rb") as _pf:
                    _png_data = _pf.read()
                _os2.unlink(_png_path)

                return _png_data


            _vends_gerar = _vends_ativos if _vend_img_sel == "Todos" else [_vend_img_sel]

            if len(_vends_gerar) == 1:
                _png_bytes = _draw_card_img(_vends_gerar[0])
                st.image(_png_bytes, caption=f"Card — {_vends_gerar[0]}")
                st.download_button(
                    "⬇️ Baixar imagem PNG",
                    data=_png_bytes,
                    file_name=f"resultado_{_vends_gerar[0].replace(' ','_')}_{_mes_card:02d}_{_ano_card}.png",
                    mime="image/png",
                    key="dl_png_single"
                )
            else:
                import zipfile as _zf
                _zip_buf = _io_img.BytesIO()
                with _zf.ZipFile(_zip_buf, "w") as _zobj:
                    for _vg in _vends_gerar:
                        _pb = _draw_card_img(_vg)
                        _fname = f"resultado_{_vg.replace(' ','_')}_{_mes_card:02d}_{_ano_card}.png"
                        _zobj.writestr(_fname, _pb)
                _zip_buf.seek(0)
                st.download_button(
                    f"⬇️ Baixar todos os cards ({len(_vends_gerar)} imagens .zip)",
                    data=_zip_buf.getvalue(),
                    file_name=f"cards_resultado_{_mes_card:02d}_{_ano_card}.zip",
                    mime="application/zip",
                    key="dl_png_zip"
                )
            st.success("✅ Imagem(ns) gerada(s)!")
        except Exception as _e_png:
            st.error(f"Erro ao gerar imagem: {_e_png}")
            st.info("Instale Pillow: pip install Pillow")

# ====================== RANKINGS ======================
elif menu == "Rankings":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Rankings</h2>', unsafe_allow_html=True)

    _rank_ini, _rank_fim = renderizar_filtros_locais("rank", "📅 Ajustar Período")
    _rank_notas = notas_unicas.copy()
    if _rank_ini:
        _rank_notas = _rank_notas[_rank_notas['DataEmissao'] >= pd.to_datetime(_rank_ini)]
    if _rank_fim:
        _rank_notas = _rank_notas[_rank_notas['DataEmissao'] <= pd.to_datetime(_rank_fim)]

    tab1, tab2 = st.tabs(["📊 Vendedores", "👥 Clientes"])
    
    with tab1:
        st.subheader("Ranking de Vendedores por Valor")
        
        ranking_vendedores = _rank_notas.groupby('Vendedor').agg({
            'Valor_Real': 'sum',
            'Numero_NF': 'count',
            'CPF_CNPJ': 'nunique'
        }).reset_index()
        ranking_vendedores.columns = ['Vendedor', 'Valor Total', 'Qtd Notas', 'Qtd Clientes']
        ranking_vendedores = ranking_vendedores.sort_values('Valor Total', ascending=False)
        ranking_vendedores.insert(0, 'Posição', range(1, len(ranking_vendedores) + 1))
        
        fig_rank_vend = px.bar(
            ranking_vendedores.head(15),
            x='Vendedor',
            y='Valor Total',
            labels={'Vendedor': 'Vendedor', 'Valor Total': 'Valor Total (R$)'},
            color='Valor Total',
            color_discrete_sequence=['#163561'],
            title='Top 15 Vendedores por Valor'
        )
        fig_rank_vend = aplicar_layout_grafico(fig_rank_vend)
        st.plotly_chart(fig_rank_vend, use_container_width=True)
        
        # Formatar para exibição
        ranking_vendedores_display = formatar_dataframe_moeda(ranking_vendedores, ['Valor Total'])
        st.dataframe(ranking_vendedores_display, use_container_width=True)
        
        st.download_button(
            "📥 Exportar Ranking Vendedores",
            to_excel(ranking_vendedores),
            "ranking_vendedores.xlsx",
            "application/vnd.ms-excel",
            key="dl_rank_vendedores"
        )
    
    with tab2:
        st.subheader("Ranking de Clientes por Valor")
        
        top_n = st.selectbox("Exibir Top:", [10, 20, 50, 100], key="top_clientes")
        
        ranking_clientes = _rank_notas.groupby(['CPF_CNPJ', 'RazaoSocial', 'Cidade', 'Estado']).agg({
            'Valor_Real': 'sum',
            'Numero_NF': 'count'
        }).reset_index()
        ranking_clientes.columns = ['CPF/CNPJ', 'Razão Social', 'Cidade', 'Estado', 'Valor Total', 'Qtd Notas']
        ranking_clientes = ranking_clientes.sort_values('Valor Total', ascending=False).head(top_n)
        ranking_clientes.insert(0, 'Posição', range(1, len(ranking_clientes) + 1))
        
        fig_rank_cli = px.bar(
            ranking_clientes.head(15),
            x='Valor Total',
            y='Razão Social',
            orientation='h',
            labels={'Razão Social': 'Cliente', 'Valor Total': 'Valor Total (R$)'},
            color='Valor Total',
            color_discrete_sequence=['#4A7BC8'],
            title=f'Top 15 Clientes por Valor'
        )
        fig_rank_cli = aplicar_layout_grafico(fig_rank_cli)
        st.plotly_chart(fig_rank_cli, use_container_width=True)
        
        # Formatar para exibição
        ranking_clientes_display = formatar_dataframe_moeda(ranking_clientes, ['Valor Total'])
        st.dataframe(ranking_clientes_display, use_container_width=True)
        
        st.download_button(
            "📥 Exportar Ranking Clientes",
            to_excel(ranking_clientes),
            f"ranking_top{top_n}_clientes.xlsx",
            "application/vnd.ms-excel",
            key="dl_rank_clientes"
        )


# ====================== CONSULTA CLIENTES ======================
elif menu == "Consulta Clientes":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;font-size:1.35rem;">Consulta de Preços por Cliente</h2>', unsafe_allow_html=True)

    # ── Percentuais adicionais por estado ────────────────────────────────
    _PERC_ESTADO = {
        'AC': 6, 'RR': 6, 'RO': 6, 'AP': 6,
        'DF': 5, 'GO': 5,
        'MT': 5, 'MS': 5, 'TO': 5, 'AM': 5,
        'PA': 8,
        'RJ': 6, 'SP': 6, 'PR': 6,
        'RONDONIA': 20,
        'PR_DIRETA': 35,
    }
    _ESTADOS_OPCOES = [
        'Selecione o Estado',
        'AC (6%)', 'RR (6%)', 'RO (6%)', 'AP (6%)',
        'DF (5%)', 'GO (5%)',
        'MT (5%)', 'MS (5%)', 'TO (5%)', 'AM (5%)',
        'PA (8%)',
        'RJ (6%)', 'SP (6%)', 'PR (6%)',
        'RONDONIA (20%)',
        'PR - Venda Direta (35%)',
    ]
    _ESTADO_KEY_MAP = {
        'AC (6%)': ('AC', 6), 'RR (6%)': ('RR', 6), 'RO (6%)': ('RO', 6), 'AP (6%)': ('AP', 6),
        'DF (5%)': ('DF', 5), 'GO (5%)': ('GO', 5),
        'MT (5%)': ('MT', 5), 'MS (5%)': ('MS', 5), 'TO (5%)': ('TO', 5), 'AM (5%)': ('AM', 5),
        'PA (8%)': ('PA', 8),
        'RJ (6%)': ('RJ', 6), 'SP (6%)': ('SP', 6), 'PR (6%)': ('PR', 6),
        'RONDONIA (20%)': ('RONDONIA', 20),
        'PR - Venda Direta (35%)': ('PR_DIRETA', 35),
    }

    # ── Carregar tabela de preços ─────────────────────────────────────────
    _df_tabela = None
    
    # Tentar carregar produtos_agrupados primeiro (mais confiável)
    if planilhas_disponiveis.get('produtos_agrupados'):
        with st.spinner("Carregando catálogo de produtos..."):
            _df_tabela = carregar_planilha_github(planilhas_disponiveis['produtos_agrupados']['url'])
            if _df_tabela is not None:
                _df_tabela.columns = _df_tabela.columns.str.upper().str.strip()
                st.success("✅ Usando: Produtos Agrupados")
    
    # Se não conseguiu, tentar tabela_ne
    if _df_tabela is None and planilhas_disponiveis.get('tabela_ne'):
        with st.spinner("Carregando tabela NE..."):
            try:
                response = requests.get(planilhas_disponiveis['tabela_ne']['url'], timeout=15)
                content = io.BytesIO(response.content)
                
                # Tentar diferentes skiprows para encontrar o cabeçalho correto
                for skip in range(0, 10):
                    try:
                        df_test = pd.read_excel(content, skiprows=skip, nrows=5)
                        content.seek(0)  # Reset para próxima tentativa
                        
                        # Verificar se as colunas parecem ser cabeçalhos válidos
                        cols_str = [str(c).upper() for c in df_test.columns]
                        
                        # Se não tem UNNAMED, provavelmente achou o cabeçalho
                        if not any('UNNAMED' in c for c in cols_str):
                            # Verificar se tem colunas relevantes
                            has_code = any(x in ' '.join(cols_str) for x in ['COD', 'CODIGO', 'CÓDIGO'])
                            has_price = any(x in ' '.join(cols_str) for x in ['PRECO', 'PREÇO', 'VALOR', 'PRICE'])
                            
                            if has_code or has_price or len(cols_str) > 3:
                                # Parece ser o cabeçalho correto!
                                content.seek(0)
                                _df_tabela = pd.read_excel(content, skiprows=skip)
                                _df_tabela.columns = _df_tabela.columns.str.upper().str.strip()
                                _df_tabela = _df_tabela.dropna(how='all')
                                st.success(f"✅ Usando: Tabela NE (skiprows={skip})")
                                break
                    except Exception:
                        continue
            except Exception as e:
                st.warning(f"Erro ao carregar tabela NE: {e}")

    if _df_tabela is None or len(_df_tabela) == 0:
        st.error("❌ Tabela de preços não encontrada")
        st.info("💡 Adicione 'Produtos_Agrupados_Completos_conciliados.xlsx' ou 'TABELA_NE_2026_CRM.xlsx' no GitHub")
        st.stop()

    # Verificar colunas disponíveis
    _cols = _df_tabela.columns.tolist()

    # Identificar coluna de código e preço (busca mais flexível)
    _cod_col   = next((c for c in _cols if any(x in c for x in ['ID_COD', 'CODIGO', 'CÓDIGO', 'COD', 'CÓD'])), None)
    _preco_col = next((c for c in _cols if any(x in c for x in ['PRECO', 'PREÇO', 'PRICE', 'VALOR', 'VLR'])), None)
    _desc_col  = next((c for c in _cols if any(x in c for x in ['DESCRI', 'DESCRIÇÃO', 'NOME', 'PRODUTO', 'GRUPO'])), None)

    if not _cod_col or not _preco_col:
        st.error(f"❌ Colunas necessárias não encontradas")
        st.info(f"📋 Colunas disponíveis: {_cols}")
        st.info(f"🔍 Procurando: Código (COD, CODIGO) e Preço (PRECO, PREÇO, VALOR)")
        with st.expander("🔍 Debug: Ver primeiras linhas da tabela"):
            st.dataframe(_df_tabela.head(10))
        st.stop()

    # ── Seleção de Estado ─────────────────────────────────────────────────
    st.markdown("#### Selecione o Estado")
    _estado_sel = st.selectbox(
        "Estado / Tabela",
        _ESTADOS_OPCOES,
        key="cc_estado",
        label_visibility="collapsed"
    )

    _perc_adicional = 0
    _estado_sigla   = ""
    if _estado_sel != 'Selecione o Estado':
        _estado_sigla, _perc_adicional = _ESTADO_KEY_MAP.get(_estado_sel, ('', 0))
        st.markdown(f"""
        <div style="background:#EEF3FC;border-radius:8px;padding:10px 14px;
                    margin-bottom:12px;font-size:0.85rem;color:#2C5AA0;">
            <b>Estado:</b> {_estado_sigla} &nbsp;·&nbsp;
            <b>Adicional:</b> {_perc_adicional}% &nbsp;·&nbsp;
            <b>Tabela base + {_perc_adicional}% = Tabela 3% comissão</b>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### Consulta de Produto")

    # ── Campo de código do produto ────────────────────────────────────────
    _codigos_lista = [''] + sorted(_df_tabela[_cod_col].dropna().astype(str).unique().tolist())
    _cc1, _cc2, _cc3 = st.columns([1, 2, 1])
    with _cc1:
        _cod_sel = st.selectbox("Código do Produto", _codigos_lista,
                                key="cc_codigo", label_visibility="visible")
        _val_neg = st.number_input("Valor Negociado (R$)",
                                   min_value=0.0,
                                   value=0.0,
                                   format="%.2f",
                                   key="cc_val_neg")

    # ── Buscar produto e calcular preços ──────────────────────────────────
    _prod_row = None
    if _cod_sel:
        _match = _df_tabela[_df_tabela[_cod_col].astype(str) == str(_cod_sel)]
        if len(_match) > 0:
            _prod_row = _match.iloc[0]

    if _prod_row is not None:
        # Descrição
        _descricao = ""
        if _desc_col:
            _parts = []
            for _dc in [c for c in _cols if any(x in c for x in ['GRUPO','DESCRI','LINHA'])]:
                _v = str(_prod_row.get(_dc, '')).strip()
                if _v and _v.lower() not in ('nan', ''):
                    _parts.append(_v)
            _descricao = ' '.join(_parts) if _parts else str(_prod_row.get(_desc_col, ''))

        # Gramatura
        _gram_col = next((c for c in _cols if 'GRAMATUR' in c), None)
        _gramatura = ''
        if _gram_col:
            _gv = str(_prod_row.get(_gram_col, '')).strip()
            if _gv and _gv.lower() not in ('nan', '0', '0.0', ''):
                _gramatura = _gv

        with _cc2:
            st.text_input("Descrição", value=_descricao, disabled=True,
                          key=f"cc_desc_{_cod_sel}")

        with _cc3:
            st.text_input("Gramatura", value=_gramatura, disabled=True,
                          key=f"cc_gram_{_cod_sel}")

        # Preço base da tabela
        try:
            _preco_base = float(_prod_row.get(_preco_col, 0))
        except Exception:
            _preco_base = 0.0

        # Tabela 3% comissão = preco_base * (1 + perc_adicional/100)
        _tab_3pct = _preco_base * (1 + _perc_adicional / 100) if _estado_sel != 'Selecione o Estado' else _preco_base
        # Tabela 4% comissão = tab_3pct * 1.06
        _tab_4pct = _tab_3pct * 1.06

        # Exibir preços calculados
        _pc1, _pc2, _pc3 = st.columns(3)
        with _pc1:
            st.metric("Tabela Base", f"R$ {formatar_numero_br(_preco_base, 2)}",
                      help="Preço da tabela padrão sem adicional de estado")
        with _pc2:
            st.metric(f"Tabela 3% ({_perc_adicional}% estado)",
                      f"R$ {formatar_numero_br(_tab_3pct, 2)}",
                      help="Tabela base + percentual do estado = tabela comissão 3%")
        with _pc3:
            st.metric("Tabela 4%", f"R$ {formatar_numero_br(_tab_4pct, 2)}",
                      help="Tabela 3% + 6% = tabela comissão 4%")

    # ── Calcular comissão sobre o valor negociado ─────────────────────
        # SOLUÇÃO FINAL: Comparação direta com margem de tolerância de 1 centavo
        if '_estado_sel' in locals() and _estado_sel:
            if _val_neg > 0 and _tab_3pct > 0:
                
                # Calculamos o valor exato que seria a tabela de 4% (3% + 6% de margem)
                # Aplicamos round para limpar o 7,77987 para 7,78
                _tabela_4_objetivo = round(_tab_3pct * 1.06, 2)
                _valor_digitado = round(_val_neg, 2)

                # FORÇAR A REGRA: Se o valor digitado for maior ou igual ao objetivo (com margem de 0.001)
                if _valor_digitado >= (_tabela_4_objetivo - 0.001):
                    _comissao_calc = '4%'
                    _variacao = round(((_valor_digitado - _tab_3pct) / _tab_3pct) * 100, 2)
                    _cor = "#10B981"; _msg = f"Comissão **4%** — objetivo de R$ {formatar_numero_br(_tabela_4_objetivo, 2)} atingido"
                else:
                    # Se não atingiu 4%, rodamos a função padrão para as outras faixas
                    _comissao_calc = calcular_comissao(_val_neg, _tab_3pct)
                    _variacao = round(((_val_neg - _tab_3pct) / _tab_3pct) * 100, 2)

                    if _comissao_calc == '3%':
                        _cor = "#2C5AA0"; _msg = "Comissão **3%** — valor igual ou acima da tabela do estado"
                    elif _comissao_calc == '2,5%':
                        _cor = "#F59E0B"; _msg = f"Comissão **2,5%** — valor {abs(_variacao):.1f}% abaixo (até 3%)"
                    elif _comissao_calc == '2%':
                        _cor = "#EF4444"; _msg = f"Comissão **2%** — valor {abs(_variacao):.1f}% abaixo (acima de 3%)"
                    else:
                        _cor = "#6B7280"; _msg = "Comissão não calculada"

                st.markdown(f"""
                <div style="background:{_cor}15;border-left:4px solid {_cor};
                            border-radius:8px;padding:12px 16px;margin-top:8px;">
                    <div style="font-size:1.1rem;font-weight:700;color:{_cor};">
                        Comissão: {_comissao_calc}
                    </div>
                    <div style="font-size:0.82rem;color:#6C757D;margin-top:3px;">{_msg}</div>
                    <div style="font-size:0.78rem;color:#ADB5BD;margin-top:4px;">
                        Valor negociado: R$ {formatar_numero_br(_val_neg, 2)} &nbsp;·&nbsp;
                        Tabela Estado (3%): R$ {formatar_numero_br(round(_tab_3pct, 2), 2)} &nbsp;·&nbsp;
                        Meta para 4%: R$ {formatar_numero_br(_tabela_4_objetivo, 2)}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.info("Insira o valor negociado para calcular a comissão.")
        else:
            st.warning("Selecione um Estado para habilitar o cálculo de comissão.")
    else:
        if _cod_sel:
            st.warning(f"Produto {_cod_sel} não encontrado na tabela.")
        else:
            st.info("Selecione um código de produto para consultar os preços.")

# ══════════════════════════════════════════════════════════════════════════
# MÓDULO ERP 1 — NOVO PEDIDO ERP
# ══════════════════════════════════════════════════════════════════════════
elif menu == "__erp_novo_pedido__":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;'
                'font-size:1.35rem;">🆕 Novo Pedido</h2>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6C757D;font-size:0.88rem;margin-bottom:16px;">'
                'Crie um pedido persistente no ERP.</p>', unsafe_allow_html=True)

    if not supa_disponivel():
        _erp_aviso_sem_supabase()
    else:
        # ── Inicializar estado do pedido ──────────────────────────────────
        if "erp_itens" not in st.session_state:
            st.session_state.erp_itens = []
        if "erp_pedido_id" not in st.session_state:
            st.session_state.erp_pedido_id = None
        if "erp_numero" not in st.session_state:
            st.session_state.erp_numero = ""

        # Mostrar número se já salvo
        if st.session_state.erp_numero:
            st.info(f"📋 Pedido **{st.session_state.erp_numero}** — Rascunho em edição")

        # ── Seção A: Dados do Cliente ─────────────────────────────────────
        st.markdown("### 👤 Cliente")
        _c1, _c2 = st.columns(2)
        with _c1:
            _clientes_lista = sorted(df["RazaoSocial"].dropna().unique().tolist())
            _cli_sel = st.selectbox("Cliente", [""] + _clientes_lista,
                                    key="erp_cliente_sel")
        _dc = {}
        if _cli_sel:
            _row = df[df["RazaoSocial"] == _cli_sel].iloc[0]
            _dc = {
                "razao_social": _row.get("RazaoSocial", ""),
                "cpf_cnpj":     _row.get("CPF_CNPJ", ""),
                "cidade":       _row.get("Cidade", ""),
                "estado":       _row.get("Estado", ""),
                "vendedor":     _row.get("Vendedor", ""),
            }
            # Verificar inadimplência
            if planilhas_disponiveis.get("inadimplencia"):
                _df_inad_check = carregar_planilha_github(
                    planilhas_disponiveis["inadimplencia"]["url"])
                if _df_inad_check is not None:
                    _df_inad_check.columns = _df_inad_check.columns.str.upper()
                    _cnpj_col = next(
                        (c for c in _df_inad_check.columns
                         if "CPF" in c or "CNPJ" in c or "DOCUM" in c), None)
                    if _cnpj_col and _dc.get("cpf_cnpj"):
                        _inad_cli = _df_inad_check[
                            _df_inad_check[_cnpj_col].astype(str).str.strip()
                            == str(_dc["cpf_cnpj"]).strip()
                        ]
                        if not _inad_cli.empty:
                            st.warning(
                                f"⚠️ **Cliente com títulos em aberto** — "
                                f"{len(_inad_cli)} título(s) encontrado(s). "
                                "Verifique antes de enviar o pedido."
                            )
        with _c2:
            _repr = st.text_input("Representante",
                                  value=_dc.get("vendedor", _erp_user_nome),
                                  key="erp_repr")
        _c3, _c4, _c5 = st.columns(3)
        with _c3:
            _cnpj  = st.text_input("CNPJ", value=_dc.get("cpf_cnpj", ""),
                                   key="erp_cnpj")
        with _c4:
            _ie    = st.text_input("Inscrição Estadual", key="erp_ie")
        with _c5:
            _fone  = st.text_input("Telefone", key="erp_fone")
        _c6, _c7 = st.columns(2)
        with _c6:
            _email = st.text_input("E-mail NF-e", key="erp_email")
        with _c7:
            _end   = st.text_input(
                "Endereço",
                value=f"{_dc.get('cidade','')}/{_dc.get('estado','')}" if _dc else "",
                key="erp_end")
        _obs_cli = st.text_area("Observação do cliente", key="erp_obs_cli",
                                height=70)
        st.markdown("---")

        # ── Seção B: Dados do Pedido ──────────────────────────────────────
        st.markdown("### 📋 Pedido")
        _p1, _p2, _p3, _p4 = st.columns(4)
        with _p1:
            _tab_preco = st.text_input("Tabela de Preço", key="erp_tab_preco")
        with _p2:
            _frete = st.selectbox("Tipo de Frete", ["CIF", "FOB"],
                                  key="erp_frete")
        with _p3:
            _data_venda = st.date_input("Data da Venda",
                                        value=pd.Timestamp.now(),
                                        key="erp_data_venda")
        with _p4:
            _cond_pag = st.text_input("Condições de Pagamento",
                                      key="erp_cond_pag")
        st.markdown("---")

        # ── Seção C: Itens ────────────────────────────────────────────────
        st.markdown("### 🛒 Adicionar Produto")
        _df_prod = None
        if planilhas_disponiveis.get("produtos_agrupados"):
            _df_prod = carregar_planilha_github(
                planilhas_disponiveis["produtos_agrupados"]["url"])
            if _df_prod is not None:
                _df_prod.columns = _df_prod.columns.str.upper()

        _pa, _pb, _pc, _pd = st.columns([2, 1, 1, 1])
        with _pa:
            _busca_tipo = st.radio("Buscar por:", ["Código", "Descrição"],
                                   horizontal=True, key="erp_busca_tipo")
            if _busca_tipo == "Código" and _df_prod is not None:
                _cods = [""] + sorted(
                    _df_prod["ID_COD"].dropna().astype(str).unique().tolist())
                _cod_sel_erp = st.selectbox("Código", _cods,
                                            key="erp_cod_sel")
            elif _busca_tipo == "Código":
                _cod_sel_erp = st.text_input("Código", key="erp_cod_txt")
            else:
                _desc_busca = st.text_input("Descrição", key="erp_desc_busca")
                _cod_sel_erp = ""

        _prod_info = {}
        if _df_prod is not None:
            if _busca_tipo == "Código" and _cod_sel_erp:
                _mask = _df_prod["ID_COD"].astype(str) == str(_cod_sel_erp)
                if _mask.any():
                    _pr = _df_prod[_mask].iloc[0]
                    _prod_info = {
                        "codigo":    str(_pr.get("ID_COD", "")),
                        "descricao": str(_pr.get("NOME_PRODUTO",
                                        _pr.get("DESCRICAO",
                                        _pr.get("PRODUTO", "")))),
                        "peso":      str(_pr.get("GRAMATURA", "")),
                        "cx_embarque": str(_pr.get("CX_EMB", "")),
                        "preco_ref": float(_pr.get("PRECO", 0) or 0),
                    }
            elif _busca_tipo == "Descrição" and _desc_busca:
                _mask = _df_prod.apply(
                    lambda r: _desc_busca.upper() in str(r).upper(), axis=1)
                if _mask.any():
                    _pr = _df_prod[_mask].iloc[0]
                    _prod_info = {
                        "codigo":    str(_pr.get("ID_COD", "")),
                        "descricao": str(_pr.get("NOME_PRODUTO",
                                        _pr.get("DESCRICAO",
                                        _pr.get("PRODUTO", "")))),
                        "peso":      str(_pr.get("GRAMATURA", "")),
                        "cx_embarque": str(_pr.get("CX_EMB", "")),
                        "preco_ref": float(_pr.get("PRECO", 0) or 0),
                    }

        with _pb:
            _qtd_erp = st.number_input("Quantidade", min_value=1, value=1,
                                       step=1, key="erp_qtd")
        with _pc:
            _preco_ref_erp = _prod_info.get("preco_ref", 0.0)
            # Sugerir último preço praticado com o cliente
            _preco_hist = 0.0
            if _cli_sel and _prod_info.get("codigo"):
                _mask_hist = (
                    (df["RazaoSocial"] == _cli_sel) &
                    (df["CodigoProduto"].astype(str) == str(_prod_info["codigo"]))
                )
                if _mask_hist.any():
                    _preco_hist = float(
                        df[_mask_hist]["PrecoUnit"].dropna().iloc[-1])
            _val_sug = _preco_hist if _preco_hist > 0 else _preco_ref_erp
            _vunit_erp = st.number_input("Valor Unit. (R$)",
                                         min_value=0.0,
                                         value=round(_val_sug, 2),
                                         step=0.01, key="erp_vunit",
                                         format="%.2f")
            # Alerta de preço abaixo da referência
            if _preco_ref_erp > 0 and _vunit_erp < _preco_ref_erp:
                _perc_desc = ((_preco_ref_erp - _vunit_erp) / _preco_ref_erp) * 100
                st.caption(f"⚠️ {_perc_desc:.1f}% abaixo da tabela")
        with _pd:
            _comiss_erp = calcular_comissao(_vunit_erp, _preco_ref_erp) \
                          if _preco_ref_erp > 0 else "—"
            st.markdown(f"**Comissão**<br>{_comiss_erp}",
                        unsafe_allow_html=True)

        if st.button("➕ Adicionar Item", key="erp_add_item",
                     use_container_width=True):
            if not _prod_info:
                st.error("Selecione um produto antes de adicionar.")
            elif _vunit_erp <= 0:
                st.error("Valor unitário deve ser maior que zero.")
            else:
                _alerta_preco = (
                    _preco_ref_erp > 0 and _vunit_erp < _preco_ref_erp)
                st.session_state.erp_itens.append({
                    "codigo":           _prod_info.get("codigo", ""),
                    "descricao":        _prod_info.get("descricao", ""),
                    "peso":             _prod_info.get("peso", ""),
                    "cx_embarque":      _prod_info.get("cx_embarque", ""),
                    "quantidade":       int(_qtd_erp),
                    "valor_unit":       round(float(_vunit_erp), 2),
                    "preco_ref":        round(_preco_ref_erp, 2),
                    "preco_historico":  round(_preco_hist, 2),
                    "comissao":         _comiss_erp,
                    "total":            round(float(_vunit_erp) * int(_qtd_erp), 2),
                    "alerta_preco_baixo": _alerta_preco,
                })
                st.rerun()

        # ── Tabela de itens adicionados ───────────────────────────────────
        if st.session_state.erp_itens:
            st.markdown("#### Itens do Pedido")
            _df_itens_erp = pd.DataFrame(st.session_state.erp_itens)
            _df_show = _df_itens_erp[[
                "codigo", "descricao", "peso", "quantidade",
                "valor_unit", "preco_ref", "comissao", "total"
            ]].copy()
            _df_show.columns = [
                "Código", "Produto", "Gramatura", "Qtde",
                "Valor Unit.", "Preço Ref.", "Comissão", "Total"
            ]
            st.dataframe(_df_show, use_container_width=True, hide_index=True)

            _tot_erp = _df_itens_erp["total"].sum()
            _m1, _m2, _m3 = st.columns(3)
            with _m1:
                st.metric("Itens", len(st.session_state.erp_itens))
            with _m2:
                st.metric("Total do Pedido", f"R$ {formatar_numero_br(_tot_erp, 2)}")
            with _m3:
                _n_alertas = sum(
                    1 for i in st.session_state.erp_itens
                    if i.get("alerta_preco_baixo"))
                if _n_alertas:
                    st.metric("⚠️ Itens c/ desconto", _n_alertas)

            _obs_ped = st.text_area("Observação do Pedido",
                                    key="erp_obs_ped", height=80)

            _btn1, _btn2, _btn3 = st.columns(3)

            # Montar dicts para salvar
            def _montar_dados_cliente_erp():
                return {
                    "razao_social": _cli_sel,
                    "cpf_cnpj":     _cnpj,
                    "ie":           _ie,
                    "cidade":       _dc.get("cidade", ""),
                    "estado":       _dc.get("estado", ""),
                    "telefone":     _fone,
                    "email":        _email,
                    "endereco":     _end,
                    "representante":_repr,
                    "obs_cliente":  _obs_cli,
                }

            def _montar_dados_pedido_erp():
                return {
                    "numero":         st.session_state.erp_numero,
                    "tabela_preco":   _tab_preco,
                    "tipo_frete":     _frete,
                    "data_venda":     str(_data_venda),
                    "cond_pagto":     _cond_pag,
                    "estado_comissao": _dc.get("estado", ""),
                }

            with _btn1:
                if st.button("💾 Salvar Rascunho", use_container_width=True,
                             key="erp_salvar_rascunho"):
                    if not _cli_sel:
                        st.error("Selecione um cliente.")
                    else:
                        with st.spinner("Salvando..."):
                            _pid, _num = salvar_pedido(
                                _montar_dados_cliente_erp(),
                                _montar_dados_pedido_erp(),
                                st.session_state.erp_itens,
                                _obs_ped,
                                status="rascunho",
                                usuario_id=_erp_user_id,
                                usuario_nome=_erp_user_nome,
                                pedido_id=st.session_state.erp_pedido_id,
                            )
                        if _pid:
                            st.session_state.erp_pedido_id = _pid
                            st.session_state.erp_numero    = _num
                            st.success(
                                f"✅ Rascunho salvo — **{_num}**")
                        else:
                            st.error(
                                "❌ Erro ao salvar. Verifique a conexão "
                                "com o Supabase.")

            with _btn2:
                if st.button("📤 Enviar para Aprovação",
                             use_container_width=True,
                             type="primary",
                             key="erp_enviar_aprovacao"):
                    if not _cli_sel:
                        st.error("Selecione um cliente.")
                    elif not st.session_state.erp_itens:
                        st.error("Adicione ao menos um item.")
                    else:
                        _confirmar = True
                        if _n_alertas > 0:
                            st.warning(
                                f"⚠️ {_n_alertas} item(s) com preço abaixo "
                                "da tabela. Confirme o envio abaixo.")
                            _confirmar = st.checkbox(
                                "Confirmo o envio com preços abaixo da tabela",
                                key="erp_confirma_envio")
                        if _confirmar:
                            with st.spinner("Enviando..."):
                                _pid, _num = salvar_pedido(
                                    _montar_dados_cliente_erp(),
                                    _montar_dados_pedido_erp(),
                                    st.session_state.erp_itens,
                                    _obs_ped,
                                    status="enviado",
                                    usuario_id=_erp_user_id,
                                    usuario_nome=_erp_user_nome,
                                    pedido_id=st.session_state.erp_pedido_id,
                                )
                            if _pid:
                                st.session_state.erp_pedido_id = None
                                st.session_state.erp_numero    = ""
                                st.session_state.erp_itens     = []
                                st.success(
                                    f"✅ Pedido **{_num}** enviado para "
                                    "aprovação!")
                                st.rerun()
                            else:
                                st.error("❌ Erro ao enviar o pedido.")

            with _btn3:
                if st.button("🗑️ Limpar", use_container_width=True,
                             key="erp_limpar"):
                    st.session_state.erp_itens     = []
                    st.session_state.erp_pedido_id = None
                    st.session_state.erp_numero    = ""
                    st.rerun()
        else:
            st.info("Nenhum item adicionado ainda.")

# ══════════════════════════════════════════════════════════════════════════
# MÓDULO ERP 2 — MEUS PEDIDOS
# ══════════════════════════════════════════════════════════════════════════
elif menu == "__erp_meus_pedidos__":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;'
                'font-size:1.35rem;">📋 Meus Pedidos</h2>', unsafe_allow_html=True)

    if not supa_disponivel():
        _erp_aviso_sem_supabase()
    else:
        # Filtros
        _fc1, _fc2, _fc3 = st.columns(3)
        with _fc1:
            _status_filtro_mp = st.selectbox(
                "Status", ["Todos", "rascunho", "enviado", "aprovado",
                           "em_separacao", "faturado", "cancelado"],
                key="erp_mp_status")
        with _fc2:
            _cli_filtro_mp = st.text_input("Cliente (busca)", key="erp_mp_cli")
        with _fc3:
            _periodo_mp = st.selectbox(
                "Período", ["Todos", "Últimos 7 dias", "Últimos 30 dias",
                            "Últimos 90 dias"],
                key="erp_mp_periodo")

        # Buscar pedidos
        _filtros_mp = {}
        if not _erp_is_gestor:
            _filtros_mp["criado_por_id"] = _erp_user_id
        if _status_filtro_mp != "Todos":
            _filtros_mp["status"] = _status_filtro_mp

        _pedidos_mp = supa_select("pedidos", filtros=_filtros_mp,
                                  ordem="criado_em.desc", limite=200)

        # Filtro cliente (client-side)
        if _cli_filtro_mp:
            _pedidos_mp = [
                p for p in _pedidos_mp
                if _cli_filtro_mp.lower() in
                p.get("cliente_razao_social", "").lower()
            ]

        # Filtro período
        if _periodo_mp != "Todos":
            _dias_mp = {"Últimos 7 dias": 7,
                        "Últimos 30 dias": 30,
                        "Últimos 90 dias": 90}[_periodo_mp]
            _corte_mp = pd.Timestamp.now() - pd.Timedelta(days=_dias_mp)
            _pedidos_mp = [
                p for p in _pedidos_mp
                if pd.to_datetime(p.get("criado_em", "")) >= _corte_mp
            ]

        # KPIs
        _total_mp  = len(_pedidos_mp)
        _valor_mp  = sum(float(p.get("valor_total", 0) or 0)
                         for p in _pedidos_mp)
        _abertos_mp = sum(1 for p in _pedidos_mp
                          if p.get("status") not in ("faturado", "cancelado"))
        _k1, _k2, _k3 = st.columns(3)
        _erp_kpi(_k1, "Total de Pedidos", str(_total_mp))
        _erp_kpi(_k2, "Valor Total", f"R$ {formatar_numero_br(_valor_mp, 2)}", "#15803D")
        _erp_kpi(_k3, "Em Aberto", str(_abertos_mp), "#C2410C")

        # Listagem
        if not _pedidos_mp:
            st.info("Nenhum pedido encontrado com os filtros selecionados.")
        else:
            for _p in _pedidos_mp:
                _status_p  = _p.get("status", "rascunho")
                _num_p     = _p.get("numero", "—")
                _cli_p     = _p.get("cliente_razao_social", "—")
                _val_p     = float(_p.get("valor_total", 0) or 0)
                _criado_p  = _p.get("criado_em", "")[:10]
                _id_p      = _p.get("id", "")

                with st.expander(
                    f"{_num_p}  ·  {_cli_p}  ·  R$ {formatar_numero_br(_val_p, 2)}  ·  "
                    f"{_criado_p}", expanded=False
                ):
                    st.markdown(_erp_badge(_status_p), unsafe_allow_html=True)
                    st.markdown(f"**Representante:** {_p.get('representante','—')}")
                    st.markdown(f"**Frete:** {_p.get('tipo_frete','—')}  "
                                f"**Pagamento:** {_p.get('cond_pagto','—')}")

                    # Ações por status
                    _ba, _bb, _bc = st.columns(3)
                    with _ba:
                        if _status_p == "rascunho":
                            if st.button("✏️ Editar", key=f"erp_edit_{_id_p}",
                                         use_container_width=True):
                                st.session_state.erp_pedido_id = _id_p
                                st.session_state.erp_numero    = _num_p
                                st.session_state.erp_itens     = supa_select(
                                    "itens_pedido",
                                    filtros={"pedido_id": _id_p})
                                st.session_state.menu_option = \
                                    "__erp_novo_pedido__"
                                st.rerun()
                    with _bb:
                        if _status_p == "rascunho":
                            if st.button("📤 Enviar", key=f"erp_env_{_id_p}",
                                         use_container_width=True):
                                mudar_status_pedido(
                                    _id_p, "enviado",
                                    _erp_user_id, _erp_user_nome,
                                    status_anterior="rascunho")
                                st.rerun()
                    with _bc:
                        if _status_p in ("rascunho", "enviado"):
                            if st.button("❌ Cancelar",
                                         key=f"erp_canc_{_id_p}",
                                         use_container_width=True):
                                _mot = st.text_input(
                                    "Motivo do cancelamento",
                                    key=f"erp_mot_{_id_p}")
                                if _mot:
                                    mudar_status_pedido(
                                        _id_p, "cancelado",
                                        _erp_user_id, _erp_user_nome,
                                        observacao=_mot,
                                        status_anterior=_status_p)
                                    st.rerun()

                    # Itens do pedido
                    _itens_p = supa_select("itens_pedido",
                                           filtros={"pedido_id": _id_p})
                    if _itens_p:
                        _df_itens_p = pd.DataFrame(_itens_p)[[
                            "codigo_produto", "descricao", "quantidade",
                            "valor_unit", "comissao_perc", "total"
                        ]]
                        _df_itens_p.columns = [
                            "Código", "Produto", "Qtde",
                            "Valor Unit.", "Comissão", "Total"
                        ]
                        st.dataframe(_df_itens_p, use_container_width=True,
                                     hide_index=True)

                    # Histórico de status
                    _hist_p = supa_select("historico_status",
                                          filtros={"pedido_id": _id_p},
                                          ordem="criado_em.asc")
                    if _hist_p:
                        st.markdown("**Histórico:**")
                        for _h in _hist_p:
                            _ts = _h.get("criado_em", "")[:16].replace("T", " ")
                            _obs_h = f" — {_h['observacao']}" \
                                     if _h.get("observacao") else ""
                            st.caption(
                                f"🕐 {_ts}  |  "
                                f"{_h.get('status_anterior','—')} → "
                                f"{_h.get('status_novo','—')}  |  "
                                f"{_h.get('usuario_nome','?')}{_obs_h}"
                            )

                    # PDF
                    if st.button("📄 Gerar PDF", key=f"erp_pdf_{_id_p}",
                                 use_container_width=True):
                        try:
                            _dados_cli_pdf = {
                                "representante": _p.get("representante",""),
                                "razao_social":  _p.get("cliente_razao_social",""),
                                "cnpj":          _p.get("cliente_cpf_cnpj",""),
                                "ie":            _p.get("cliente_ie",""),
                                "telefone":      _p.get("cliente_telefone",""),
                                "email":         _p.get("cliente_email_nfe",""),
                                "endereco":      _p.get("cliente_endereco",""),
                                "obs_cliente":   _p.get("obs_cliente",""),
                            }
                            _dados_ped_pdf = {
                                "numero":        _num_p,
                                "tabela_preco":  _p.get("tabela_preco",""),
                                "tipo_frete":    _p.get("tipo_frete",""),
                                "data_venda":    str(_p.get("data_venda",""))[:10],
                                "condicoes_pagto": _p.get("cond_pagto",""),
                            }
                            _itens_pdf = [
                                {
                                    "codigo":     i.get("codigo_produto",""),
                                    "descricao":  i.get("descricao",""),
                                    "peso":       i.get("gramatura",""),
                                    "cx_embarque":i.get("cx_embarque",""),
                                    "quantidade": i.get("quantidade",0),
                                    "valor_unit": i.get("valor_unit",0),
                                    "total":      i.get("total",0),
                                    "comissao":   i.get("comissao_perc",""),
                                }
                                for i in _itens_p
                            ]
                            _pdf_bytes = gerar_pdf_pedido(
                                _dados_cli_pdf, _dados_ped_pdf,
                                _itens_pdf, _p.get("obs_pedido",""))
                            st.download_button(
                                "📥 Baixar PDF",
                                data=_pdf_bytes,
                                file_name=f"Pedido_{_num_p}.pdf",
                                mime="application/pdf",
                                key=f"erp_dl_pdf_{_id_p}",
                            )
                        except Exception as _e_pdf2:
                            st.error(f"Erro ao gerar PDF: {_e_pdf2}")

# ══════════════════════════════════════════════════════════════════════════
# MÓDULO ERP 3 — FILA DE APROVAÇÃO (somente gestor/admin)
# ══════════════════════════════════════════════════════════════════════════
elif menu == "__erp_fila_aprovacao__":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;'
                'font-size:1.35rem;">⏳ Fila de Aprovação</h2>',
                unsafe_allow_html=True)

    if not supa_disponivel():
        _erp_aviso_sem_supabase()
    elif not _erp_is_gestor:
        st.error("Acesso restrito a gestores e administradores.")
    else:
        _enviados = supa_select("pedidos", filtros={"status": "enviado"},
                                ordem="criado_em.asc", limite=500)

        # KPIs
        _tot_fila = len(_enviados)
        _val_fila = sum(float(p.get("valor_total", 0) or 0)
                        for p in _enviados)
        _urgentes = 0
        for _pf in _enviados:
            try:
                _dias_fila = (
                    pd.Timestamp.now() -
                    pd.to_datetime(_pf.get("criado_em",""))
                ).days
                if _dias_fila >= 2:
                    _urgentes += 1
            except Exception:
                pass

        _kf1, _kf2, _kf3 = st.columns(3)
        _erp_kpi(_kf1, "Aguardando Aprovação", str(_tot_fila))
        _erp_kpi(_kf2, "Valor Total na Fila",
                 f"R$ {formatar_numero_br(_val_fila, 2)}", "#15803D")
        _erp_kpi(_kf3, "⚠️ Urgentes (+48h)", str(_urgentes), "#C2410C")

        if not _enviados:
            st.success("✅ Nenhum pedido aguardando aprovação.")
        else:
            for _pf in _enviados:
                _id_f   = _pf.get("id", "")
                _num_f  = _pf.get("numero", "—")
                _cli_f  = _pf.get("cliente_razao_social", "—")
                _vend_f = _pf.get("criado_por_nome", "—")
                _val_f  = float(_pf.get("valor_total", 0) or 0)
                try:
                    _dias_f = (
                        pd.Timestamp.now() -
                        pd.to_datetime(_pf.get("criado_em",""))
                    ).days
                except Exception:
                    _dias_f = 0

                _urgente_f = _dias_f >= 2
                _titulo_f = (
                    f"{'🔴 ' if _urgente_f else ''}{_num_f}  ·  "
                    f"{_cli_f}  ·  {_vend_f}  ·  "
                    f"R$ {formatar_numero_br(_val_f, 2)}  ·  {_dias_f}d na fila"
                )

                with st.expander(_titulo_f, expanded=_urgente_f):
                    _itens_f = supa_select("itens_pedido",
                                           filtros={"pedido_id": _id_f})
                    if _itens_f:
                        _df_f = pd.DataFrame(_itens_f)[[
                            "codigo_produto", "descricao", "quantidade",
                            "valor_unit", "preco_ref", "comissao_perc", "total"
                        ]]
                        _df_f.columns = [
                            "Código","Produto","Qtde","Valor Unit.",
                            "Preço Ref.","Comissão","Total"
                        ]
                        # Destacar itens com alerta
                        st.dataframe(_df_f, use_container_width=True,
                                     hide_index=True)

                    _obs_fila = _pf.get("obs_pedido","") or \
                                _pf.get("obs_cliente","")
                    if _obs_fila:
                        st.info(f"📝 Obs: {_obs_fila}")

                    _af1, _af2, _af3 = st.columns(3)
                    with _af1:
                        if st.button("✅ Aprovar", key=f"erp_apr_{_id_f}",
                                     use_container_width=True, type="primary"):
                            mudar_status_pedido(
                                _id_f, "aprovado",
                                _erp_user_id, _erp_user_nome,
                                observacao="Aprovado pelo gestor",
                                status_anterior="enviado")
                            st.success(f"Pedido {_num_f} aprovado!")
                            st.rerun()
                    with _af2:
                        _mot_dev = st.text_input(
                            "Motivo (devolução)", key=f"erp_mot_dev_{_id_f}",
                            placeholder="Ex: Preço abaixo do mínimo...")
                        if st.button("↩️ Devolver", key=f"erp_dev_{_id_f}",
                                     use_container_width=True):
                            if not _mot_dev:
                                st.error("Informe o motivo da devolução.")
                            else:
                                mudar_status_pedido(
                                    _id_f, "rascunho",
                                    _erp_user_id, _erp_user_nome,
                                    observacao=_mot_dev,
                                    status_anterior="enviado")
                                st.warning(f"Pedido {_num_f} devolvido.")
                                st.rerun()
                    with _af3:
                        _mot_canc_f = st.text_input(
                            "Motivo (cancelamento)", key=f"erp_mot_cf_{_id_f}")
                        if st.button("❌ Cancelar", key=f"erp_cf_{_id_f}",
                                     use_container_width=True):
                            if not _mot_canc_f:
                                st.error("Informe o motivo.")
                            else:
                                mudar_status_pedido(
                                    _id_f, "cancelado",
                                    _erp_user_id, _erp_user_nome,
                                    observacao=_mot_canc_f,
                                    status_anterior="enviado")
                                st.rerun()

# ══════════════════════════════════════════════════════════════════════════
# MÓDULO ERP 4 — TODOS OS PEDIDOS (somente gestor/admin)
# ══════════════════════════════════════════════════════════════════════════
elif menu == "__erp_todos_pedidos__":
    st.markdown('<h2 style="color:#4A7BC8;font-weight:700;margin-bottom:4px;'
                'font-size:1.35rem;">🗂️ Todos os Pedidos</h2>',
                unsafe_allow_html=True)

    if not supa_disponivel():
        _erp_aviso_sem_supabase()
    elif not _erp_is_gestor:
        st.error("Acesso restrito a gestores e administradores.")
    else:
        # Filtros
        _ft1, _ft2, _ft3, _ft4 = st.columns(4)
        with _ft1:
            _st_todos = st.selectbox(
                "Status",
                ["Todos", "rascunho", "enviado", "aprovado",
                 "em_separacao", "faturado", "cancelado"],
                key="erp_tp_status")
        with _ft2:
            _vend_todos = st.text_input("Vendedor", key="erp_tp_vend")
        with _ft3:
            _cli_todos  = st.text_input("Cliente",  key="erp_tp_cli")
        with _ft4:
            _per_todos  = st.selectbox(
                "Período",
                ["Todos", "Últimos 7 dias", "Últimos 30 dias",
                 "Últimos 90 dias"],
                key="erp_tp_periodo")

        _filtros_tp = {}
        if _st_todos != "Todos":
            _filtros_tp["status"] = _st_todos

        _todos_ped = supa_select("pedidos", filtros=_filtros_tp,
                                 ordem="criado_em.desc", limite=500)

        if _vend_todos:
            _todos_ped = [
                p for p in _todos_ped
                if _vend_todos.lower() in
                p.get("criado_por_nome", "").lower()
            ]
        if _cli_todos:
            _todos_ped = [
                p for p in _todos_ped
                if _cli_todos.lower() in
                p.get("cliente_razao_social", "").lower()
            ]
        if _per_todos != "Todos":
            _dias_tp = {"Últimos 7 dias": 7,
                        "Últimos 30 dias": 30,
                        "Últimos 90 dias": 90}[_per_todos]
            _corte_tp = pd.Timestamp.now() - pd.Timedelta(days=_dias_tp)
            _todos_ped = [
                p for p in _todos_ped
                if pd.to_datetime(p.get("criado_em","")) >= _corte_tp
            ]

        # KPIs por status
        _contadores = {}
        _valores    = {}
        for _p in _todos_ped:
            _s = _p.get("status","")
            _contadores[_s] = _contadores.get(_s, 0) + 1
            _valores[_s]    = _valores.get(_s, 0.0) + \
                              float(_p.get("valor_total", 0) or 0)

        _status_order = ["rascunho","enviado","aprovado",
                         "em_separacao","faturado","cancelado"]
        _cols_kpi_tp  = st.columns(len(_status_order))
        for _ci, _s in enumerate(_status_order):
            _cnt = _contadores.get(_s, 0)
            _val = _valores.get(_s, 0.0)
            _cor_tp = {
                "rascunho":"#64748B","enviado":"#1D4ED8",
                "aprovado":"#15803D","em_separacao":"#C2410C",
                "faturado":"#14532D","cancelado":"#B91C1C",
            }.get(_s, "#1F4788")
            _erp_kpi(_cols_kpi_tp[_ci],
                     _s.replace("_"," ").title(),
                     f"{_cnt}  ·  R$ {formatar_numero_br(_val, 0)}", _cor_tp)

        st.markdown("---")

        # Exportar Excel
        if _todos_ped:
            _df_export_tp = pd.DataFrame([{
                "Número":     p.get("numero",""),
                "Status":     p.get("status",""),
                "Cliente":    p.get("cliente_razao_social",""),
                "Vendedor":   p.get("criado_por_nome",""),
                "Valor":      float(p.get("valor_total",0) or 0),
                "Frete":      p.get("tipo_frete",""),
                "Pagamento":  p.get("cond_pagto",""),
                "Data":       str(p.get("criado_em",""))[:10],
            } for p in _todos_ped])
            st.download_button(
                "📥 Exportar Excel",
                data=to_excel(_df_export_tp),
                file_name="todos_pedidos.xlsx",
                mime="application/vnd.ms-excel",
                key="dl_erp_todos_pedidos",
            )

        # Listagem
        for _pt in _todos_ped:
            _id_t   = _pt.get("id","")
            _num_t  = _pt.get("numero","—")
            _cli_t  = _pt.get("cliente_razao_social","—")
            _vnd_t  = _pt.get("criado_por_nome","—")
            _val_t  = float(_pt.get("valor_total",0) or 0)
            _sts_t  = _pt.get("status","")
            _dat_t  = str(_pt.get("criado_em",""))[:10]

            with st.expander(
                f"{_num_t}  ·  {_cli_t}  ·  {_vnd_t}  ·  "
                f"R$ {formatar_numero_br(_val_t, 2)}  ·  {_dat_t}",
                expanded=False
            ):
                st.markdown(_erp_badge(_sts_t), unsafe_allow_html=True)
                _at1, _at2, _at3, _at4 = st.columns(4)
                with _at1:
                    if _sts_t == "aprovado":
                        if st.button("📦 Em Separação",
                                     key=f"erp_sep_{_id_t}",
                                     use_container_width=True):
                            mudar_status_pedido(
                                _id_t, "em_separacao",
                                _erp_user_id, _erp_user_nome,
                                status_anterior="aprovado")
                            st.rerun()
                with _at2:
                    if _sts_t in ("aprovado","em_separacao"):
                        _nf_num = st.text_input("Nº NF (opcional)",
                                                key=f"erp_nf_{_id_t}")
                        if st.button("🧾 Marcar Faturado",
                                     key=f"erp_fat_{_id_t}",
                                     use_container_width=True,
                                     type="primary"):
                            mudar_status_pedido(
                                _id_t, "faturado",
                                _erp_user_id, _erp_user_nome,
                                observacao=f"NF: {_nf_num}" if _nf_num else "",
                                status_anterior=_sts_t)
                            if _nf_num:
                                supa_update("pedidos", _id_t,
                                            {"numero_nf": _nf_num})
                            st.rerun()
                with _at3:
                    if _sts_t not in ("faturado","cancelado"):
                        _mot_t = st.text_input("Motivo cancelamento",
                                               key=f"erp_mot_t_{_id_t}")
                        if st.button("❌ Cancelar", key=f"erp_ct_{_id_t}",
                                     use_container_width=True):
                            if not _mot_t:
                                st.error("Informe o motivo.")
                            else:
                                mudar_status_pedido(
                                    _id_t, "cancelado",
                                    _erp_user_id, _erp_user_nome,
                                    observacao=_mot_t,
                                    status_anterior=_sts_t)
                                st.rerun()
                with _at4:
                    _itens_t = supa_select("itens_pedido",
                                           filtros={"pedido_id": _id_t})
                    if _itens_t:
                        try:
                            _pdf_t = gerar_pdf_pedido(
                                {
                                    "representante": _pt.get("representante",""),
                                    "razao_social":  _pt.get("cliente_razao_social",""),
                                    "cnpj":          _pt.get("cliente_cpf_cnpj",""),
                                    "ie":            _pt.get("cliente_ie",""),
                                    "telefone":      _pt.get("cliente_telefone",""),
                                    "email":         _pt.get("cliente_email_nfe",""),
                                    "endereco":      _pt.get("cliente_endereco",""),
                                    "obs_cliente":   _pt.get("obs_cliente",""),
                                },
                                {
                                    "numero":          _num_t,
                                    "tabela_preco":    _pt.get("tabela_preco",""),
                                    "tipo_frete":      _pt.get("tipo_frete",""),
                                    "data_venda":      str(_pt.get("data_venda",""))[:10],
                                    "condicoes_pagto": _pt.get("cond_pagto",""),
                                },
                                [{
                                    "codigo":     i.get("codigo_produto",""),
                                    "descricao":  i.get("descricao",""),
                                    "peso":       i.get("gramatura",""),
                                    "cx_embarque":i.get("cx_embarque",""),
                                    "quantidade": i.get("quantidade",0),
                                    "valor_unit": i.get("valor_unit",0),
                                    "total":      i.get("total",0),
                                    "comissao":   i.get("comissao_perc",""),
                                } for i in _itens_t],
                                _pt.get("obs_pedido","")
                            )
                            st.download_button(
                                "📄 PDF",
                                data=_pdf_t,
                                file_name=f"Pedido_{_num_t}.pdf",
                                mime="application/pdf",
                                key=f"erp_dl_t_{_id_t}",
                            )
                        except Exception:
                            st.caption("PDF indisponível")

st.markdown("""
<hr style="border-color:#E9ECEF;margin-top:32px;margin-bottom:12px;">
<div style="text-align:center;color:#ADB5BD;font-size:0.78rem;padding-bottom:16px;">
    Dashboard BI Medtextil 2.0 &nbsp;·&nbsp; Desenvolvido com Streamlit
    &nbsp;·&nbsp; <span style="color:#4A7BC8;font-weight:600;">Medtextil Produtos Textil Hospitalares</span>
</div>
""", unsafe_allow_html=True)
